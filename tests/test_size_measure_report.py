import io
from datetime import date

import openpyxl
import pandas as pd

from size_report import (
    build_size_report_rows,
    compute_size_forecast,
    generate_size_measure_excel,
    validate_size_measurement_sequence,
)


def test_size_35_is_highlighted_as_near_harvest():
    result = compute_size_forecast({1: 35.0})

    assert result["highlight"] is True
    assert result["note"] == "Gần đạt size, tuần sau có thể thu"
    assert result["avg_growth"] is None


def test_average_growth_projects_next_measurement_to_harvest_size():
    result = compute_size_forecast({1: 32.0, 2: 34.0})

    assert result["avg_growth"] == 2.0
    assert result["projected_next"] == 36.0
    assert result["highlight"] is True
    assert result["note"] == "Dự kiến lần/tuần tới đạt size thu hoạch"


def test_completed_group_is_hidden_after_30_days():
    size_logs = pd.DataFrame(
        [
            {
                "dim_lo_id": 10,
                "farm": "Farm 126",
                "team": "NT1",
                "lo": "A8",
                "mau_day": "Cam",
                "lan_do": 1,
                "ngay_do": "2026-04-01",
                "size_cal": 32,
                "so_luong_mau": 10,
                "hang_kiem_tra": "H1",
            },
            {
                "dim_lo_id": 10,
                "farm": "Farm 126",
                "team": "NT1",
                "lo": "A8",
                "mau_day": "Cam",
                "lan_do": 2,
                "ngay_do": "2026-04-08",
                "size_cal": 34,
                "so_luong_mau": 10,
                "hang_kiem_tra": "H1",
            },
            {
                "dim_lo_id": 10,
                "farm": "Farm 126",
                "team": "NT1",
                "lo": "A8",
                "mau_day": "Cam",
                "lan_do": 3,
                "ngay_do": "2026-04-15",
                "size_cal": 36,
                "so_luong_mau": 10,
                "hang_kiem_tra": "H1",
            },
        ]
    )

    rows = build_size_report_rows(
        pd.DataFrame(),
        size_logs,
        pd.DataFrame(),
        today=date(2026, 6, 15),
    )

    assert rows.empty


def test_third_measurement_requires_second_measurement():
    existing = pd.DataFrame(
        [
            {"id": 1, "lan_do": 1, "hang_kiem_tra": "H1-H5", "is_deleted": False},
        ]
    )

    error = validate_size_measurement_sequence(existing, 3, "H1-H5")

    assert error == "Không thể đo Lần 3 khi chưa có Lần 2."


def test_measurements_in_same_group_must_use_same_test_row():
    existing = pd.DataFrame(
        [
            {"id": 1, "lan_do": 1, "hang_kiem_tra": "H1-H5", "is_deleted": False},
        ]
    )

    error = validate_size_measurement_sequence(existing, 2, "H6-H10")

    assert error == "Các lần đo cùng lô và màu dây phải dùng cùng Hàng kiểm tra: H1-H5."


def test_excel_matches_size_template_and_highlights_forecast_row():
    lots = pd.DataFrame(
        [
            {
                "id": 1,
                "dim_lo_id": 10,
                "farm": "Farm 126",
                "team": "NT1",
                "lo": "A8",
            }
        ]
    )
    size_logs = pd.DataFrame(
        [
            {
                "dim_lo_id": 10,
                "farm": "Farm 126",
                "team": "NT1",
                "lo": "A8",
                "mau_day": "Cam",
                "lan_do": 1,
                "ngay_do": "2026-06-01",
                "size_cal": 32,
                "so_luong_mau": 10,
                "hang_kiem_tra": "H1-H5",
            },
            {
                "dim_lo_id": 10,
                "farm": "Farm 126",
                "team": "NT1",
                "lo": "A8",
                "mau_day": "Cam",
                "lan_do": 2,
                "ngay_do": "2026-06-08",
                "size_cal": 34,
                "so_luong_mau": 10,
                "hang_kiem_tra": "H1-H5",
            },
        ]
    )
    stage_logs = pd.DataFrame(
        [
            {
                "dim_lo_id": 10,
                "farm": "Farm 126",
                "lo": "A8",
                "giai_doan": "Cắt bắp",
                "ngay_thuc_hien": "2026-05-25",
                "tuan": 22,
            }
        ]
    )
    ribbon_schedule = pd.DataFrame(
        [
            {
                "farm": "Farm 126",
                "year": 2026,
                "week_number": 22,
                "color_name": "cam",
            }
        ]
    )

    excel_bytes = generate_size_measure_excel(
        lots,
        size_logs,
        stage_logs,
        ribbon_schedule,
        farm_name="Farm 126",
        today=date(2026, 6, 15),
    )
    workbook = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    sheet = workbook["Báo cáo đo size"]

    assert sheet.cell(row=1, column=1).value == "BÁO CÁO KIỂM TRA ĐO SIZE THU HOẠCH CHUỐI"
    assert "36/37 cal" in sheet.cell(row=3, column=1).value
    assert sheet.cell(row=5, column=9).value == "Lần đo 1"
    assert sheet.cell(row=5, column=11).value == "Lần đo 3"
    assert sheet.cell(row=5, column=12).value == "Mức tăng size TB"
    assert sheet.cell(row=5, column=13).value == "Chú thích"

    data_row = 6
    assert sheet.cell(row=data_row, column=2).value == "A8"
    assert sheet.cell(row=data_row, column=7).value == 22
    assert sheet.cell(row=data_row, column=8).value == 29
    assert sheet.cell(row=data_row, column=12).value == 2
    assert sheet.cell(row=data_row, column=13).value == "Dự kiến lần/tuần tới đạt size thu hoạch"
    assert sheet.cell(row=data_row, column=1).fill.fill_type == "solid"
