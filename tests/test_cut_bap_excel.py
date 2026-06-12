import io

import openpyxl
import pandas as pd

import app


class _FakeResponse:
    def __init__(self, data):
        self.data = data


class _FakeTableQuery:
    def __init__(self, rows):
        self._rows = list(rows)

    def select(self, *_args, **_kwargs):
        return self

    def eq(self, key, value):
        self._rows = [row for row in self._rows if row.get(key) == value]
        return self

    def execute(self):
        return _FakeResponse(self._rows)


class _FakeSupabase:
    def __init__(self, ribbon_rows):
        self._ribbon_rows = ribbon_rows

    def table(self, table_name):
        assert table_name == "ribbon_schedule"
        return _FakeTableQuery(self._ribbon_rows)


def _load_cut_report(monkeypatch, df_lots, df_stg):
    monkeypatch.setattr(
        app,
        "get_farm_id_from_name",
        lambda farm_name: {"Farm 126": 1, "Farm 157": 2}.get(farm_name),
    )
    monkeypatch.setattr(
        app,
        "supabase",
        _FakeSupabase(
            [
                {
                    "farm_id": 1,
                    "year": 2026,
                    "week_number": 24,
                    "color_name": "Cam",
                    "is_deleted": False,
                },
                {
                    "farm_id": 2,
                    "year": 2026,
                    "week_number": 24,
                    "color_name": "Trắng",
                    "is_deleted": False,
                },
            ]
        ),
    )
    excel_bytes = app.generate_cut_bap_excel(
        df_lots,
        df_stg,
        pd.DataFrame(),
        pd.DataFrame(),
    )
    return openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)


def test_cut_report_single_farm_126_uses_plus_8_forecast(monkeypatch):
    df_lots = pd.DataFrame(
        [
            {
                "id": 1,
                "farm": "Farm 126",
                "lo": "A8",
                "ngay_trong": "2025-10-02",
            }
        ]
    )
    df_stg = pd.DataFrame(
        [
            {
                "giai_doan": "Cắt bắp",
                "ngay_thuc_hien": "2026-06-10",
                "tuan": 24,
                "base_lot_id": 1,
                "farm": "Farm 126",
                "lo": "A8",
                "so_luong": 88,
            }
        ]
    )

    wb = _load_cut_report(monkeypatch, df_lots, df_stg)
    ws = wb["2026"]

    assert ws.cell(row=1, column=1).value == "Lô"
    assert ws.cell(row=1, column=2).value == "31 (+8)"
    assert ws.cell(row=4, column=2).value == "Cam"


def test_cut_report_multi_farm_adds_farm_column_and_specific_forecasts(monkeypatch):
    df_lots = pd.DataFrame(
        [
            {
                "id": 1,
                "farm": "Farm 126",
                "lo": "A8",
                "ngay_trong": "2025-10-02",
            },
            {
                "id": 2,
                "farm": "Farm 157",
                "lo": "1A",
                "ngay_trong": "2025-10-15",
            },
        ]
    )
    df_stg = pd.DataFrame(
        [
            {
                "giai_doan": "Cắt bắp",
                "ngay_thuc_hien": "2026-06-10",
                "tuan": 24,
                "base_lot_id": 1,
                "farm": "Farm 126",
                "lo": "A8",
                "so_luong": 88,
            },
            {
                "giai_doan": "Cắt bắp",
                "ngay_thuc_hien": "2026-06-10",
                "tuan": 24,
                "base_lot_id": 2,
                "farm": "Farm 157",
                "lo": "1A",
                "so_luong": 100,
            },
        ]
    )

    wb = _load_cut_report(monkeypatch, df_lots, df_stg)
    ws = wb["2026"]

    assert ws.cell(row=1, column=1).value == "Farm"
    assert ws.cell(row=1, column=2).value == "Lô"
    assert ws.cell(row=1, column=3).value == "126: 31 (+8)\n157: 32 (+9)"
    assert ws.cell(row=4, column=3).value == "126: Cam\n157: Trắng"
    assert ws.cell(row=5, column=1).value == "Farm 126"
    assert ws.cell(row=6, column=1).value == "Farm 157"
