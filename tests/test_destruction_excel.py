import io

import openpyxl
import pandas as pd

import app


def _find_row(ws, farm, lot):
    for row_idx in range(3, ws.max_row + 1):
        if ws.cell(row_idx, 1).value == farm and ws.cell(row_idx, 2).value == lot:
            return row_idx
    raise AssertionError(f"Missing row {farm}/{lot}")


def test_destruction_report_is_horizontal_and_includes_all_stages():
    df_lots = pd.DataFrame(
        [
            {"id": 1, "farm": "Farm 126", "lo": "A8", "ngay_trong": "2025-10-01"},
            {"id": 2, "farm": "Farm 126", "lo": "D4", "ngay_trong": "2025-10-02"},
        ]
    )
    df_des = pd.DataFrame(
        [
            {
                "farm": "Farm 126", "lo": "A8", "base_lot_id": 1,
                "ngay_xuat_huy": "2026-05-25", "tuan": 22,
                "giai_doan": "Trước chích bắp", "so_luong": 2, "ly_do": "Bệnh",
            },
            {
                "farm": "Farm 126", "lo": "A8", "base_lot_id": 1,
                "ngay_xuat_huy": "2026-05-26", "tuan": 22,
                "giai_doan": "Trước cắt bắp", "so_luong": 3, "ly_do": "Gãy",
            },
            {
                "farm": "Farm 126", "lo": "A8", "base_lot_id": 1,
                "ngay_xuat_huy": "2026-06-05", "tuan": 23,
                "giai_doan": "Trước thu hoạch", "so_luong": 4, "ly_do": "Bệnh",
            },
            {
                "farm": "Farm 126", "lo": "D4", "base_lot_id": 2,
                "ngay_xuat_huy": "2026-06-05", "tuan": 23,
                "giai_doan": "Sau thu hoạch", "so_luong": 5, "ly_do": "Đổ ngã",
            },
        ]
    )

    workbook = openpyxl.load_workbook(
        io.BytesIO(app.generate_destruction_excel(df_lots, df_des)),
        data_only=True,
    )

    assert workbook.sheetnames == ["Theo lô", "Chi tiết"]
    ws = workbook["Theo lô"]
    assert ws["A1"].value == "Farm"
    assert ws["B1"].value == "Lô"
    assert ws["C1"].value == "Tuần 22/2026"
    assert ws["F1"].value == "Tuần 23/2026"
    assert ws["H1"].value == "Lũy kế"
    assert [ws.cell(2, col).value for col in range(3, 9)] == [
        "25/05", "26/05", "Tổng tuần", "05/06", "Tổng tuần", "Tổng xuất hủy",
    ]

    a8_row = _find_row(ws, "Farm 126", "A8")
    d4_row = _find_row(ws, "Farm 126", "D4")
    assert [ws.cell(a8_row, col).value for col in range(3, 9)] == [2, 3, 5, 4, 4, 9]
    assert [ws.cell(d4_row, col).value for col in range(3, 9)] == [None, None, None, 5, 5, 5]
    assert ws.freeze_panes == "C3"

    detail = workbook["Chi tiết"]
    assert [detail.cell(1, col).value for col in range(1, 8)] == [
        "Farm", "Lô", "Ngày xuất hủy", "Tuần", "Giai đoạn", "Số lượng", "Lý do",
    ]
    assert {detail.cell(row, 5).value for row in range(2, detail.max_row + 1)} == {
        "Trước chích bắp", "Trước cắt bắp", "Trước thu hoạch", "Sau thu hoạch",
    }
    assert detail["C2"].number_format == "dd/mm/yyyy"


def test_destruction_report_handles_empty_data():
    workbook = openpyxl.load_workbook(
        io.BytesIO(app.generate_destruction_excel(pd.DataFrame(), pd.DataFrame())),
        data_only=True,
    )

    assert workbook["Theo lô"]["A1"].value == "Chưa có dữ liệu Xuất hủy."
    assert workbook["Chi tiết"]["A1"].value == "Chưa có dữ liệu Xuất hủy."
