"""
=======================================================================
  ỨNG DỤNG NHẬP LIỆU QUẢN LÝ TIẾN ĐỘ SINH TRƯỞNG CHUỐI XUẤT KHẨU
  Công ty Trường Tồn  |  Streamlit + Supabase (RBAC Version 2.0)
=======================================================================
"""

import streamlit as st
import pandas as pd
import streamlit.components.v1 as components
from datetime import date, datetime, timezone, timedelta
import os
from dotenv import load_dotenv
from supabase import create_client, Client
import plotly.express as px
import plotly.graph_objects as go
import io
import uuid
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import html
import re
import unicodedata
import inspect

from cost_dashboard import render_cost_dashboard

try:
    import cost_lifecycle as _cost_lifecycle
except Exception:
    _cost_lifecycle = None


def _cl_money(value) -> float:
    try:
        if value is None or pd.isna(value):
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def _cl_normalize_label(value) -> str:
    text = str(value or "").strip().replace("Đ", "D").replace("đ", "d")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^A-Z0-9+]+", "", text.upper())


def _cl_to_timestamp(value):
    ts = pd.to_datetime(value, errors="coerce")
    if pd.isna(ts):
        return None
    return ts.tz_localize(None) if getattr(ts, "tzinfo", None) else ts


def _cl_group_rows_by_base_lot(rows):
    grouped = {}
    for row in rows or []:
        base_lot_id = row.get("base_lot_id")
        if base_lot_id is not None and not pd.isna(base_lot_id):
            grouped.setdefault(int(base_lot_id), []).append(dict(row))
    return grouped


def _fallback_build_harvest_rows_by_batch(harvest_rows):
    return _cl_group_rows_by_base_lot(harvest_rows)


def _fallback_build_stage_rows_by_batch(stage_rows):
    return _cl_group_rows_by_base_lot(stage_rows)


def _fallback_harvest_quantity_for_batch_until(harvest_rows_by_batch: dict, base_lot_id, cost_dt) -> float:
    cost_ts = _cl_to_timestamp(cost_dt)
    if base_lot_id is None or cost_ts is None:
        return 0.0
    total = 0.0
    for row in harvest_rows_by_batch.get(int(base_lot_id), []):
        harvest_ts = _cl_to_timestamp(row.get("ngay_thu_hoach"))
        if harvest_ts is not None and harvest_ts <= cost_ts:
            total += _cl_money(row.get("so_luong"))
    return total


def _fallback_stage_quantity_for_batch_until(stage_rows_by_batch: dict, base_lot_id, stage_name: str, cost_dt) -> float:
    cost_ts = _cl_to_timestamp(cost_dt)
    if base_lot_id is None or cost_ts is None:
        return 0.0
    total = 0.0
    for row in stage_rows_by_batch.get(int(base_lot_id), []):
        if str(row.get("giai_doan") or "") != stage_name:
            continue
        stage_ts = _cl_to_timestamp(row.get("ngay_thuc_hien"))
        if stage_ts is not None and stage_ts <= cost_ts:
            total += _cl_money(row.get("so_luong"))
    return total


def _fallback_build_batch_lifecycle(planting_rows, season_rows=None, harvest_rows=None, destruction_rows=None):
    seasons_by_batch = _cl_group_rows_by_base_lot(season_rows)
    harvest_by_batch = _cl_group_rows_by_base_lot(harvest_rows)
    destruction_by_batch = _cl_group_rows_by_base_lot(destruction_rows)
    processed = []
    for row in planting_rows or []:
        batch = dict(row)
        base_lot_id = batch.get("id") or batch.get("base_lot_id")
        if base_lot_id is None or pd.isna(base_lot_id):
            processed.append(batch)
            continue
        base_lot_id = int(base_lot_id)
        start_ts = _cl_to_timestamp(batch.get("ngay_trong"))
        batch["ngay_trong_ts"] = start_ts
        season_rows_for_batch = seasons_by_batch.get(base_lot_id, [])
        has_open_season = any(_cl_to_timestamp(s.get("ngay_ket_thuc_thuc_te")) is None for s in season_rows_for_batch)
        end_candidates = [
            _cl_to_timestamp(s.get("ngay_ket_thuc_thuc_te"))
            for s in season_rows_for_batch
            if _cl_to_timestamp(s.get("ngay_ket_thuc_thuc_te")) is not None
        ]
        active_until = None
        end_reason = ""
        if season_rows_for_batch and not has_open_season and end_candidates:
            active_until = max(end_candidates)
            end_reason = "season_end"
        if active_until is None and not has_open_season:
            events = []
            for h in harvest_by_batch.get(base_lot_id, []):
                events.append((_cl_to_timestamp(h.get("ngay_thu_hoach")), _cl_money(h.get("so_luong")), "harvest_full"))
            for d in destruction_by_batch.get(base_lot_id, []):
                events.append((_cl_to_timestamp(d.get("ngay_xuat_huy")), _cl_money(d.get("so_luong")), "destruction_full"))
            events = [(dt, qty, reason) for dt, qty, reason in events if dt is not None and (start_ts is None or dt >= start_ts)]
            events.sort(key=lambda item: item[0])
            total_removed = 0.0
            target_trees = _cl_money(batch.get("so_luong"))
            for event_dt, qty, reason in events:
                total_removed += qty
                if target_trees > 0 and total_removed >= target_trees:
                    active_until = event_dt
                    end_reason = reason
                    break
        batch["active_until_ts"] = active_until
        batch["active_until"] = active_until.date() if active_until is not None else None
        batch["lifecycle_end_reason"] = end_reason
        processed.append(batch)
    return processed


def _fallback_is_batch_active_on_date(batch: dict, cost_dt) -> bool:
    cost_ts = _cl_to_timestamp(cost_dt)
    start_ts = batch.get("ngay_trong_ts") or _cl_to_timestamp(batch.get("ngay_trong"))
    if cost_ts is None or start_ts is None or start_ts > cost_ts:
        return False
    active_until = batch.get("active_until_ts") or _cl_to_timestamp(batch.get("active_until"))
    return active_until is None or cost_ts <= active_until


def _fallback_is_harvest_related_cost(cost: dict) -> bool:
    fields = ("cong_doan", "category", "detail", "ma_cv", "ma_cv_chuan", "ma_khoan_muc_cp", "hang_muc", "scope_label")
    text = "|".join(_cl_normalize_label(cost.get(field)) for field in fields)
    return any(token in text for token in ("THUHOACH", "CATBAN", "BANBUONG", "RONGROC", "VANCHUYENTHUHOACH"))


def _fallback_cost_text(cost: dict) -> str:
    fields = ("cong_doan", "category", "detail", "ma_cv", "ma_cv_chuan", "ma_khoan_muc_cp", "hang_muc")
    return "|".join(_cl_normalize_label(cost.get(field)) for field in fields)


_BUNCH_CARE_REQUIRES_CUT_TOKENS = (
    "BAOBUONG",
    "BAOBUP",
    "BEHOA",
    "LATRAU",
    "CHAMSOCBUONG",
    "GOBAO",
    "SUABAO",
    "VESINHBUONG",
    "VENLA",
    "DOSIZECHUOI",
    "XOPLOTNAI",
    "DAYCHONGNGA",
    "GOBUONG",
    "DAYBUONG",
)

_PREHARVEST_CARE_TOKENS = (
    "BONPHAN",
    "PHAN",
    "DAM",
    "KALI",
    "URE",
    "LAN",
    "HCVS",
    "HC1",
    "HUUCO",
    "TRICHODERMA",
    "TRICODERMA",
    "CHAMSOCVUON",
    "LAMCO",
    "TUOINUOC",
    "THUOCBVTV",
    "PHUNXECAY",
    "THUOC",
    "BVTV",
    "DIETCO",
    "STREPTOMICIN",
    "BUPROFEZIN",
)

_GENERAL_OVERHEAD_TOKENS = (
    "DAUDO",
    "DAU",
    "COGIOI",
    "MAYCAY",
    "XECAY",
    "CHAYXECAY",
    "SUACHUA",
    "HETHONGTUOI",
    "DIENNUOC",
    "DIENPHANVANUOC",
)

_PLANTING_NURSERY_TOKENS = (
    "TRONGMOI",
    "CAYCHUOI",
    "CAYGIONG",
    "BAU",
    "DONGBAU",
    "VUONUOM",
    "UOM",
    "LOTHO",
    "DATBAU",
)


def _fallback_is_bunch_care_requiring_cut(cost: dict) -> bool:
    text = _fallback_cost_text(cost)
    return any(token in text for token in _BUNCH_CARE_REQUIRES_CUT_TOKENS)


def _fallback_is_stage_quantity_capped_cost(cost: dict) -> bool:
    text = _fallback_cost_text(cost)
    return any(token in text for token in _BUNCH_CARE_REQUIRES_CUT_TOKENS + ("CATBAP",))


def _fallback_is_preharvest_care_cost(cost: dict) -> bool:
    text = _fallback_cost_text(cost)
    return any(token in text for token in _PREHARVEST_CARE_TOKENS)


def _fallback_is_general_overhead_cost(cost: dict) -> bool:
    text = _fallback_cost_text(cost)
    return any(token in text for token in _GENERAL_OVERHEAD_TOKENS)


def _fallback_is_planting_or_nursery_cost(cost: dict) -> bool:
    text = _fallback_cost_text(cost)
    return any(token in text for token in _PLANTING_NURSERY_TOKENS)


def _fallback_lot_weight_on_date(lot_id, cost_dt, plantings_by_lot: dict, lot_meta_by_id: dict) -> float:
    active_rows = [
        row for row in plantings_by_lot.get(lot_id, [])
        if _fallback_is_batch_active_on_date(row, cost_dt)
    ]
    if not active_rows:
        return 0.0
    planted_area = sum(_cl_money(row.get("dien_tich_trong")) for row in active_rows)
    if planted_area > 0:
        return planted_area
    lot_area = _cl_money((lot_meta_by_id.get(lot_id) or {}).get("area_ha"))
    active_trees = sum(_cl_money(row.get("so_luong")) for row in active_rows)
    all_trees = sum(_cl_money(row.get("so_luong")) for row in plantings_by_lot.get(lot_id, []))
    if lot_area > 0 and active_trees > 0 and all_trees > 0:
        return lot_area * min(1.0, active_trees / all_trees)
    if lot_area > 0:
        return lot_area
    return active_trees


build_batch_lifecycle = getattr(_cost_lifecycle, "build_batch_lifecycle", _fallback_build_batch_lifecycle)
build_harvest_rows_by_batch = getattr(_cost_lifecycle, "build_harvest_rows_by_batch", _fallback_build_harvest_rows_by_batch)
build_stage_rows_by_batch = getattr(_cost_lifecycle, "build_stage_rows_by_batch", _fallback_build_stage_rows_by_batch)
harvest_quantity_for_batch_until = getattr(_cost_lifecycle, "harvest_quantity_for_batch_until", _fallback_harvest_quantity_for_batch_until)
is_bunch_care_requiring_cut = getattr(_cost_lifecycle, "is_bunch_care_requiring_cut", _fallback_is_bunch_care_requiring_cut)
is_batch_active_on_date = getattr(_cost_lifecycle, "is_batch_active_on_date", _fallback_is_batch_active_on_date)
is_harvest_related_cost = getattr(_cost_lifecycle, "is_harvest_related_cost", _fallback_is_harvest_related_cost)
is_stage_quantity_capped_cost = getattr(_cost_lifecycle, "is_stage_quantity_capped_cost", _fallback_is_stage_quantity_capped_cost)
is_preharvest_care_cost = getattr(_cost_lifecycle, "is_preharvest_care_cost", _fallback_is_preharvest_care_cost)
is_general_overhead_cost = getattr(_cost_lifecycle, "is_general_overhead_cost", _fallback_is_general_overhead_cost)
is_planting_or_nursery_cost = getattr(_cost_lifecycle, "is_planting_or_nursery_cost", _fallback_is_planting_or_nursery_cost)
lifecycle_lot_weight_on_date = getattr(_cost_lifecycle, "lot_weight_on_date", _fallback_lot_weight_on_date)
stage_quantity_for_batch_until = getattr(_cost_lifecycle, "stage_quantity_for_batch_until", _fallback_stage_quantity_for_batch_until)

FARM_MAP_COMPONENT_PATH = os.path.join(os.path.dirname(__file__), "components", "farm_map")
farm_map_component = components.declare_component("farm_map_component", path=FARM_MAP_COMPONENT_PATH)

try:
    from container_allocation import (
        DEFAULT_SKU_ROWS,
        OPTIMIZER_SKU_RULES,
        allocate_bunches_optimized,
        build_hand_weight_profile,
        calculate_min_bunches_for_container_plan,
        calculate_max_containers_by_market,
        get_optimizer_sku_rules,
    )
except ImportError:
    from container_allocation import DEFAULT_SKU_ROWS, allocate_bunches_by_hands

    OPTIMIZER_SKU_RULES = {
        "27CP": {
            "markets": ["Nhật"],
            "group": "Phần Ngọn",
            "description": "Quả to, mập nhất. Yêu cầu khắt khe nhất, ưu tiên cắt phần gần cuống đẹp nhất.",
            "ranges": [(1, 5)],
        },
        "8H": {
            "markets": ["Hàn"],
            "group": "Phần Ngọn",
            "description": "Lấy phần gần cuống tương tự 27CP nhưng dùng để ráp container Hàn.",
            "ranges": [(1, 4)],
        },
        "6H": {
            "markets": ["Nhật"],
            "group": "Khúc Giữa",
            "description": "Quả thon đều, cố định ở khúc giữa trên.",
            "ranges": [(5, 7)],
        },
        "30CP": {
            "markets": ["Nhật"],
            "group": "Khúc Giữa",
            "description": "Quả cỡ trung bình lớn, thường ráp nối tiếp sau khi 27CP đã lấy phần ngọn.",
            "ranges": [(1, 9)],
        },
        "5H": {
            "markets": ["Nhật"],
            "group": "Khúc Giữa",
            "description": "Cố định ở khúc giữa dưới, gần đuôi.",
            "ranges": [(8, 10)],
        },
        "5/6H": {
            "markets": ["Hàn"],
            "group": "Khúc Giữa",
            "description": "Cắt dải dài xuyên suốt khúc giữa buồng, tốc độ gom hàng nhanh.",
            "ranges": [(5, 9)],
        },
        "15CP": {
            "markets": ["Hàn"],
            "group": "Phần Đuôi",
            "description": "Mã tận dụng chiến lược, gom vét các nải nhỏ cuối buồng.",
            "ranges": [(10, 12)],
        },
        "12CP": {
            "markets": ["Hàn"],
            "group": "Phần Đuôi",
            "description": "Cùng nhóm quy cách Hàn với 15CP.",
            "ranges": [(10, 12)],
        },
        "10CP": {
            "markets": ["Hàn"],
            "group": "Phần Đuôi",
            "description": "Cùng nhóm quy cách Hàn với 15CP.",
            "ranges": [(10, 12)],
        },
    }

    def allocate_bunches_optimized(
        total_bunches,
        kg_per_bunch,
        hands_per_bunch,
        sku_rows,
        kg_per_box=13,
        boxes_per_container=1320,
        beam_width=None,
        hand_weights=None,
    ):
        fallback_rows = []
        for row in sku_rows:
            fallback_row = dict(row)
            sku = str(fallback_row.get("sku", "")).upper()
            ranges = OPTIMIZER_SKU_RULES.get(sku, {}).get("ranges", [])
            if ranges:
                fallback_row["hand_from"], fallback_row["hand_to"] = ranges[0]
            fallback_rows.append(fallback_row)
        result = allocate_bunches_by_hands(
            total_bunches,
            kg_per_bunch,
            hands_per_bunch,
            fallback_rows,
            kg_per_box=kg_per_box,
            boxes_per_container=boxes_per_container,
            hand_weights=hand_weights,
        )
        result.setdefault("loss", {})
        result["summary"].setdefault("optimizer_loss", 0.0)
        result["summary"].setdefault("active_bunches_estimated", 0)
        result["summary"].setdefault("segment_count", len(result.get("rows", [])))
        result["summary"].setdefault("solver_status", "APPROXIMATE")
        result["summary"].setdefault("solver_backend", "legacy_fallback")
        result.setdefault("solver_status", result["summary"]["solver_status"])
        result.setdefault("solver_backend", result["summary"]["solver_backend"])
        result.setdefault("detail_rows", result.get("rows", []))
        for row in result.get("rows", []):
            row.setdefault("range_label", f"{row.get('hand_from')}-{row.get('hand_to')}")
        return result

    def calculate_max_containers_by_market(
        total_bunches,
        kg_per_bunch,
        hands_per_bunch,
        market_order,
        kg_per_box=13,
        boxes_per_container=1320,
        hand_weights=None,
    ):
        source_kg = max(0, total_bunches) * max(0, kg_per_bunch)
        kg_per_hand = kg_per_bunch / hands_per_bunch if hands_per_bunch else 0
        return {
            "rows": [],
            "detail_rows": [],
            "remaining_hands": {},
            "summary": {
                "total_bunches": max(0, total_bunches),
                "source_kg": source_kg,
                "kg_per_hand": kg_per_hand,
                "source_boxes_capacity": int(source_kg // kg_per_box) if kg_per_box else 0,
                "source_cont_capacity": (int(source_kg // kg_per_box) / boxes_per_container) if boxes_per_container else 0,
                "fulfilled_boxes": 0,
                "fulfilled_containers": 0,
                "active_bunches_estimated": 0,
                "segment_count": 0,
                "solver_status": "APPROXIMATE",
                "solver_backend": "legacy_fallback",
            },
            "solver_status": "APPROXIMATE",
            "solver_backend": "legacy_fallback",
        }

    _FALLBACK_BASE_HAND_WEIGHT_PROFILES = {
        12: {
            1: 1.3,
            2: 1.5,
            3: 1.7,
            4: 2.0,
            5: 2.0,
            6: 2.3,
            7: 2.5,
            8: 2.6,
            9: 2.7,
            10: 3.0,
            11: 3.3,
            12: 3.5,
        },
        9: {
            1: 1.4,
            2: 1.5,
            3: 1.7,
            4: 1.9,
            5: 2.1,
            6: 2.3,
            7: 2.5,
            8: 2.7,
            9: 2.9,
        },
    }

    def build_hand_weight_profile(hands_per_bunch, target_kg_per_bunch):
        base_profile = _FALLBACK_BASE_HAND_WEIGHT_PROFILES.get(int(hands_per_bunch), {})
        base_total = sum(base_profile.values())
        if base_profile and base_total > 0:
            scale = float(target_kg_per_bunch) / base_total
            hand_weights = {hand: weight * scale for hand, weight in base_profile.items()}
        else:
            kg_per_hand = target_kg_per_bunch / hands_per_bunch if hands_per_bunch else 0
            hand_weights = {hand: kg_per_hand for hand in range(1, hands_per_bunch + 1)}
        kg_per_bunch = sum(hand_weights.values())
        kg_per_hand = kg_per_bunch / len(hand_weights) if hand_weights else 0
        return {
            "hands_per_bunch": len(hand_weights) or hands_per_bunch,
            "kg_per_bunch": kg_per_bunch,
            "kg_per_hand": kg_per_hand,
            "hand_weights": hand_weights,
        }

    def calculate_min_bunches_for_container_plan(
        sku_rows,
        kg_per_bunch,
        hands_per_bunch,
        kg_per_box=13,
        boxes_per_container=1320,
        hand_weights=None,
    ):
        requested_boxes = sum(int(float(row.get("demand", 0) or 0)) for row in sku_rows)
        high = 1
        best_result = None
        while high <= 1_000_000:
            result = allocate_bunches_optimized(
                high,
                kg_per_bunch,
                hands_per_bunch,
                sku_rows,
                kg_per_box=kg_per_box,
                boxes_per_container=boxes_per_container,
                hand_weights=hand_weights,
            )
            summary = result.get("summary", {})
            if int(summary.get("short_boxes", 0)) == 0 and int(summary.get("fulfilled_boxes", 0)) >= requested_boxes:
                best_result = result
                break
            high *= 2
        if best_result is None:
            return {
                "rows": [],
                "detail_rows": [],
                "remaining_hands": {},
                "summary": {
                    "total_bunches": 0,
                    "target_containers": requested_boxes // boxes_per_container,
                    "requested_boxes": requested_boxes,
                    "source_kg": 0,
                    "kg_per_hand": 0,
                    "hand_weights": hand_weights or {},
                    "source_boxes_capacity": 0,
                    "source_cont_capacity": 0,
                    "fulfilled_boxes": 0,
                    "fulfilled_containers": 0,
                    "short_boxes": requested_boxes,
                    "active_bunches_estimated": 0,
                    "segment_count": 0,
                    "kg_allocated": 0,
                    "extra_kg_from_rounding": 0,
                    "solver_status": "NO_SOLUTION",
                    "solver_backend": "legacy_fallback",
                },
                "solver_status": "NO_SOLUTION",
                "solver_backend": "legacy_fallback",
            }

        low = 0
        while low <= high:
            mid = (low + high) // 2
            result = allocate_bunches_optimized(
                mid,
                kg_per_bunch,
                hands_per_bunch,
                sku_rows,
                kg_per_box=kg_per_box,
                boxes_per_container=boxes_per_container,
                hand_weights=hand_weights,
            )
            summary = result.get("summary", {})
            if int(summary.get("short_boxes", 0)) == 0 and int(summary.get("fulfilled_boxes", 0)) >= requested_boxes:
                best_result = result
                high = mid - 1
            else:
                low = mid + 1
        active_bunches = int(best_result["summary"].get("active_bunches_estimated", best_result["summary"].get("total_bunches", 0)))
        kg_allocated = sum(float(row.get("kg_allocated", 0) or 0) for row in best_result.get("rows", []))
        best_result["summary"].update({
            "total_bunches": active_bunches,
            "target_containers": requested_boxes // boxes_per_container,
            "requested_boxes": requested_boxes,
            "source_kg": active_bunches * kg_per_bunch,
            "kg_allocated": kg_allocated,
            "extra_kg_from_rounding": max(0, kg_allocated - requested_boxes * kg_per_box),
            "solver_status": "APPROXIMATE",
            "solver_backend": "fallback_binary_search",
        })
        best_result["solver_status"] = "APPROXIMATE"
        best_result["solver_backend"] = "fallback_binary_search"
        return best_result

    def get_optimizer_sku_rules(hands_per_bunch):
        return OPTIMIZER_SKU_RULES

if hasattr(st, "dialog"):
    dialog_decorator = st.dialog
elif hasattr(st, "experimental_dialog"):
    dialog_decorator = st.experimental_dialog
else:
    # Fallback to function if strictly not supported
    def dialog_decorator(title, *args, **kwargs):
        def decorator(func):
            return func
        return decorator

def configured_dialog_decorator(title, *args, **kwargs):
    try:
        params = inspect.signature(dialog_decorator).parameters
        supports_kwargs = any(p.kind == inspect.Parameter.VAR_KEYWORD for p in params.values())
        if not supports_kwargs:
            kwargs = {key: value for key, value in kwargs.items() if key in params}
    except (TypeError, ValueError):
        pass
    return dialog_decorator(title, *args, **kwargs)

# =====================================================
# CẤU HÌNH BAN ĐẦU
# =====================================================

load_dotenv()
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

st.set_page_config(
    page_title="Trường Tồn - Quản lý Tiến độ Chuối",
    page_icon="🍌",
    layout="wide",
    initial_sidebar_state="expanded",
)

# =====================================================
# QUYỀN TRUY CẬP (RBAC) — ĐỌC TỪ DATABASE
# =====================================================
# Cấu trúc trả về: RBAC_DB[Farm][Team] = Password

@st.cache_data(ttl=300)  # Cache 5 phút
def fetch_rbac_from_db():
    """Đọc bảng user_roles từ Supabase, trả về dict {farm: {team: password}}."""
    try:
        res = supabase.table("user_roles").select("farm, team, password").eq("is_active", True).execute()
        rbac = {}
        for row in res.data:
            farm = row["farm"]
            team = row["team"]
            pw = row["password"]
            if farm not in rbac:
                rbac[farm] = {}
            rbac[farm][team] = pw
        return rbac
    except Exception as e:
        st.error(f"❌ Không thể tải dữ liệu phân quyền: {e}")
        return {}

RBAC_DB = fetch_rbac_from_db()
FARMS = list(RBAC_DB.keys())
TEAMS = ["NT1", "NT2", "Đội BVTV", "Đội Thu Hoạch", "Xưởng Đóng Gói", "Quản lý farm"]

# =====================================================
# CONFIG OPTIONS
# =====================================================
VU_OPTIONS = ["F0", "F1", "F2", "F3", "F4", "F5"]
LOAI_TRONG_OPTIONS = ["Trồng mới", "Trồng dặm"]
STAGE_NT_OPTIONS = ["Chích bắp", "Cắt bắp"]  # Không còn Thu Hoạch ở đây
DESTRUCTION_STAGE_OPTIONS = ["Trước chích bắp", "Trước cắt bắp", "Trước thu hoạch"]
PLANTING_DENSITY_TREES_PER_HA = 2190

# =====================================================
# BẢNG MÀU DÂY CHUẨN (Ribbon Colors)
# =====================================================
BASE_COLORS = ["Đỏ", "Cam", "Tím", "Đen", "Xanh lá", "Xanh dương", "Trắng"]

# =====================================================
# TIMELINE SINH TRƯỞNG CHUỐI (Constants)
# =====================================================
# F0 (từ ngày trồng): Trồng → +180d Chích → +14d Cắt → +70d Thu hoạch (tổng 264d)
# Fn (từ harvest trước): +90d Chích → +14d Cắt → +70d Thu hoạch (tổng 174d)
# Cây con Fn sinh trưởng 3 tháng dưới gốc mẹ trước khi mẹ thu hoạch
F0_DAYS_TO_CHICH = 180
F0_DAYS_TO_CAT = 194      # 180 + 14
F0_DAYS_TO_THU = 264       # 180 + 14 + 70
FN_CYCLE_DAYS = 174        # 90 + 14 + 70 (từ harvest trước)
FN_DAYS_CHICH_OFFSET = 90  # Từ harvest trước đến chích bắp Fn
FN_DAYS_CAT_OFFSET = 104   # 90 + 14
FN_DAYS_THU_OFFSET = 174   # 90 + 14 + 70
DAYS_CHICH_TO_THU = F0_DAYS_TO_THU - F0_DAYS_TO_CHICH  # 84 ngày: chích bắp → thu hoạch
DAYS_CAT_TO_THU   = F0_DAYS_TO_THU - F0_DAYS_TO_CAT    # 70 ngày: cắt bắp → thu hoạch
CONVERGENCE_DAYS = 15      # Ngưỡng hội tụ: 2 đợt chênh ≤15d → coi như gộp
MAX_GENERATION = 5         # Tính tối đa F0→F5
HARVEST_MIN_GROWTH_WEEKS = 18  # F1+ chỉ tính thu hoạch sau khi vụ đã đủ tuổi

# ── Micro-PDF cho Mốc ②③ (chích/cắt bắp → thu hoạch) ──
# Spread hẹp hơn Mốc ① vì gần ngày thu hơn → sai số nhỏ hơn
MICRO_WINDOW_HALF = 7      # ±7 ngày spread (fixed, không configurable)
MICRO_SIGMA = 3.0           # sigma cho mini Normal Distribution

# =====================================================
# SẢN LƯỢNG DỰ TOÁN (kg/buồng)
# =====================================================
KG_PER_TREE_F0 = 15        # F0: 15 kg/buồng
KG_PER_TREE_FN = 18        # Fn (F1+): 18 kg/buồng
KG_PER_BOX = 13            # 13 kg/thùng
BOXES_PER_CONTAINER = 1320 # 1320 thùng/container

# =====================================================
# TỈ LỆ HAO HỤT THEO GIAI ĐOẠN (so với số cây trồng gốc)
# =====================================================
# Trồng → Chích bắp: 5% loss | Chích bắp → Thu hoạch: +5% loss | Tổng: 10%
LOSS_RATE_TO_CHICH = 0.05   # 5% hao hụt từ trồng → chích bắp
LOSS_RATE_TO_THU   = 0.10   # 10% tổng hao hụt từ trồng → thu hoạch

def get_estimated_rate(stage: str) -> float:
    """
    Trả về tỉ lệ còn lại (1 - loss) theo giai đoạn.
    - Chích bắp: 0.95 (5% hao hụt)
    - Cắt bắp:  0.95 (cắt bắp chỉ cách chích 14 ngày, chưa chênh đáng kể)
    - Thu hoạch: 0.90 (10% hao hụt tổng cộng)
    """
    if stage in ("chich_bap", "cat_bap"):
        return 1 - LOSS_RATE_TO_CHICH  # 0.95
    elif stage == "thu_hoach":
        return 1 - LOSS_RATE_TO_THU    # 0.90
    return 1.0  # Trồng: không hao hụt

def get_kg_per_tree(vu: str) -> int:
    """Trả về kg/buồng tương ứng theo vụ. F0 = 15kg, Fn = 18kg."""
    return KG_PER_TREE_F0 if vu == "F0" else KG_PER_TREE_FN

# Stage offsets cho thuật toán matching
_STAGE_OFFSETS = {
    "Chích bắp": {"f0": F0_DAYS_TO_CHICH, "fn": FN_DAYS_CHICH_OFFSET},
    "Cắt bắp":   {"f0": F0_DAYS_TO_CAT,   "fn": FN_DAYS_CAT_OFFSET},
    "Thu hoạch":  {"f0": F0_DAYS_TO_THU,   "fn": FN_DAYS_THU_OFFSET},
}


# Map giai đoạn xuất hủy → stage tương ứng để dùng timeline matching
_DESTRUCTION_STAGE_MAP = {
    "Trước chích bắp": "Chích bắp",
    "Trước cắt bắp": "Cắt bắp",
    "Trước thu hoạch": "Thu hoạch",
}


def _safe_int_id(value):
    """Return an integer id when possible; otherwise None."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    try:
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            return int(float(value))
        return int(value)
    except (TypeError, ValueError, OverflowError):
        return None


def build_next_season_maps(df_seasons: pd.DataFrame, df_stg_all: pd.DataFrame):
    """Return next producing season boundaries keyed by (base_lot_id, vu)."""
    next_season = {}
    next_producing = set()
    if df_seasons.empty or "base_lot_id" not in df_seasons.columns:
        return next_season, next_producing

    season_rows = df_seasons[df_seasons["base_lot_id"].notna()].copy()
    season_rows["_base_lot_id_int"] = season_rows["base_lot_id"].apply(_safe_int_id)
    season_rows = season_rows[season_rows["_base_lot_id_int"].notna()]
    if (
        not df_stg_all.empty
        and "base_lot_id" in df_stg_all.columns
        and "giai_doan" in df_stg_all.columns
    ):
        stg_chich_all = df_stg_all[df_stg_all["giai_doan"] == "Chích bắp"]
        stg_chich_ids = pd.to_numeric(stg_chich_all["base_lot_id"], errors="coerce")
        for _, s_row in season_rows.iterrows():
            s_blid = int(s_row["_base_lot_id_int"])
            s_vu = s_row["vu"]
            s_start = pd.to_datetime(s_row["ngay_bat_dau"])
            chich_batch = stg_chich_all[stg_chich_ids == s_blid]
            if not chich_batch.empty and "ngay_thuc_hien" in chich_batch.columns:
                chich_in = chich_batch[
                    pd.to_datetime(chich_batch["ngay_thuc_hien"], errors="coerce") >= s_start
                ]
                if not chich_in.empty:
                    next_producing.add((s_blid, s_vu))

    for blid, blid_grp in season_rows.groupby("_base_lot_id_int"):
        blid_int = int(blid)
        sorted_s = blid_grp.sort_values("ngay_bat_dau").drop_duplicates("vu")
        vu_list = sorted_s[["vu", "ngay_bat_dau"]].values.tolist()
        for i, (vu_val, _start_dt) in enumerate(vu_list):
            if i + 1 < len(vu_list):
                next_vu = vu_list[i + 1][0]
                if (blid_int, next_vu) in next_producing:
                    next_season[(blid_int, vu_val)] = pd.to_datetime(vu_list[i + 1][1])
                else:
                    next_season[(blid_int, vu_val)] = None
            else:
                next_season[(blid_int, vu_val)] = None
    return next_season, next_producing


def build_batch_label_map(df_lots_trong_moi: pd.DataFrame):
    """Build display labels for lots with multiple planting batches."""
    labels = {}
    if df_lots_trong_moi.empty or "id" not in df_lots_trong_moi.columns or "lo" not in df_lots_trong_moi.columns:
        return labels

    for lo_name_grp, grp_df in df_lots_trong_moi.groupby("lo"):
        if len(grp_df) > 1:
            sorted_grp = grp_df.sort_values("ngay_trong") if "ngay_trong" in grp_df.columns else grp_df
            for i, (_, b_row) in enumerate(sorted_grp.iterrows(), 1):
                labels[b_row["id"]] = f"{lo_name_grp} (đợt {i})"
        else:
            for _, b_row in grp_df.iterrows():
                labels[b_row["id"]] = lo_name_grp
    return labels

def get_current_season_destruction(bid: int, giai_doan: str = None) -> int:
    """Tổng xuất hủy thuộc vụ hiện tại (season mới nhất) cho một base_lot.
    Nếu giai_doan được chỉ định, chỉ đếm destruction cùng giai_doan.
    """
    season_res = supabase.table("seasons").select("ngay_bat_dau") \
        .eq("base_lot_id", bid).eq("is_deleted", False) \
        .order("ngay_bat_dau", desc=True).limit(1).execute()
    if not season_res.data:
        return 0
    query = supabase.table("destruction_logs").select("so_luong") \
        .eq("base_lot_id", bid).eq("is_deleted", False) \
        .gte("ngay_xuat_huy", season_res.data[0]["ngay_bat_dau"])
    if giai_doan:
        query = query.eq("giai_doan", giai_doan)
    dest_res = query.execute()
    return sum(int(r["so_luong"]) for r in dest_res.data) if dest_res.data else 0


def resolve_base_lot_id(dim_lo_id, action_date, giai_doan, quantity=None, exclude_id=None):
    """
    Tự động suy luận base_lot_id dựa trên FIFO capacity.
    Đợt trồng cũ nhất (ngay_trong ascending) có remaining capacity > 0 được ưu tiên.
    Nếu có quantity, chỉ chọn batch còn đủ capacity cho toàn bộ record.
    Fallback: closest-match nếu tất cả đợt đều full hoặc không xác định được capacity.
    """
    res = supabase.table("base_lots") \
        .select("id, so_luong, ngay_trong") \
        .eq("dim_lo_id", dim_lo_id) \
        .eq("is_deleted", False) \
        .eq("loai_trong", "Trồng mới") \
        .order("ngay_trong", desc=False).execute()
    
    if not res.data:
        return None  # Lô chưa có đợt trồng nào

    effective_giai_doan = _DESTRUCTION_STAGE_MAP.get(giai_doan, giai_doan)
    requested_qty = None
    try:
        requested_qty = int(quantity) if quantity is not None else None
    except Exception:
        requested_qty = None

    # ── FIFO: đợt cũ nhất có remaining capacity > 0 ──
    for batch in res.data:
        bid = batch["id"]
        planted = int(batch.get("so_luong", 0)) - get_current_season_destruction(bid, "Trước chích bắp")

        if effective_giai_doan == "Chích bắp":
            used = get_current_season_used(bid, "stage", "Chích bắp", exclude_id=exclude_id)
            cap = planted - used
        elif effective_giai_doan == "Cắt bắp":
            total_chich = get_current_season_used(bid, "stage", "Chích bắp")
            total_cat = get_current_season_used(bid, "stage", "Cắt bắp", exclude_id=exclude_id)
            cap = total_chich - total_cat
        elif effective_giai_doan == "Thu hoạch":
            total_cat = get_current_season_used(bid, "stage", "Cắt bắp")
            total_har = get_current_season_used(bid, "harvest", exclude_id=exclude_id)
            cap = total_cat - total_har
        else:
            # Giai đoạn không xác định (destruction, etc.) → dùng planted
            cap = planted

        if requested_qty is not None and effective_giai_doan in ["Chích bắp", "Cắt bắp", "Thu hoạch"]:
            if cap >= requested_qty:
                return bid
        elif cap > 0:
            return bid

    # Với các stage có kiểm soát capacity, không fallback sang batch đã đầy.
    if requested_qty is not None and effective_giai_doan in ["Chích bắp", "Cắt bắp", "Thu hoạch"]:
        return None

    # Fallback: tất cả đợt đều full → chọn đợt trồng gần nhất trước ngày hành động
    action_dt = pd.to_datetime(action_date)
    fallback_id = None
    best_dist = float('inf')
    for batch in res.data:
        trong_dt = pd.to_datetime(batch["ngay_trong"])
        if trong_dt <= action_dt:
            dist = (action_dt - trong_dt).days
            if dist < best_dist:
                best_dist = dist
                fallback_id = batch["id"]
    return fallback_id or res.data[0]["id"]


# =====================================================
# STYLE TÙY CHỈNH (CSS)
# =====================================================
st.markdown("""
<style>
    /* --- Loại bỏ khoảng trắng mặc định ở đầu trang --- */
    .stMainBlockContainer { padding-top: 1rem !important; }
    header[data-testid="stHeader"] { height: 0 !important; min-height: 0 !important; padding: 0 !important; }
    .stDeployButton { display: none !important; }

    .main-title { text-align: center; color: #2E7D32; font-size: 2rem; font-weight: 700; margin-bottom: 0.2rem; }
    .sub-title { text-align: center; color: #6D6D6D; font-size: 1rem; margin-bottom: 1.5rem; }
    .farm-badge { display: inline-block; padding: 4px 14px; border-radius: 20px; background: #2E7D32; color: white; font-weight: 600; font-size: 0.9rem; }
    .team-badge { display: inline-block; padding: 4px 14px; border-radius: 20px; background: #FF9800; color: white; font-weight: 600; font-size: 0.9rem; margin-top: 5px; }
    .dataframe-header { font-size: 1.1rem; font-weight: 600; color: #1B5E20; margin: 1rem 0 0.5rem 0; padding-left: 0.5rem; border-left: 4px solid #2E7D32; }
    .section-divider { border: none; border-top: 2px solid #E8F5E9; margin: 2rem 0; }
</style>
""", unsafe_allow_html=True)


# =====================================================
# HÀM TIỆN ÍCH (SESSION & LOG)
# =====================================================
def init_session_state():
    if "logged_in" not in st.session_state:
        if st.query_params.get("logged_in") == "true":
            st.session_state["logged_in"] = True
            st.session_state["current_farm"] = st.query_params.get("farm")
            st.session_state["current_team"] = st.query_params.get("team")
        else:
            st.session_state["logged_in"] = False
            st.session_state["current_farm"] = None
            st.session_state["current_team"] = None
            
    # Bulk entry queues
    for k in ["queue_stg", "queue_des", "queue_har", "queue_sm", "queue_inv", "queue_bsr", "queue_fus"]:
        if k not in st.session_state: st.session_state[k] = []

def logout():
    st.session_state["logged_in"] = False
    st.session_state["current_farm"] = None
    st.session_state["current_team"] = None
    st.query_params.clear()


def _query_param_value(key: str, default=None):
    value = st.query_params.get(key, default)
    if isinstance(value, list):
        return value[0] if value else default
    return value


def _persisted_segmented_control(
    label: str,
    options: list,
    key: str,
    query_key: str,
    slugs: dict,
    default,
    label_visibility: str = "collapsed",
):
    slug_to_option = {slug: option for option, slug in slugs.items()}
    query_value = _query_param_value(query_key)
    default_option = slug_to_option.get(query_value, default)
    active_option = st.segmented_control(
        label,
        options,
        label_visibility=label_visibility,
        key=key,
        default=default_option,
    )
    if active_option is None:
        active_option = default_option

    active_slug = slugs.get(active_option)
    if active_slug and _query_param_value(query_key) != active_slug:
        st.query_params[query_key] = active_slug
    return active_option


def insert_access_log(farm: str, team: str, action: str):
    try:
        supabase.table("access_logs").insert({"farm": farm, "team": team, "action": action}).execute()
    except Exception as e:
        print(f"Lỗi log: {e}")

# =====================================================
# HÀM TƯƠNG TÁC DB (SUPABASE DATA FETCHERS)
# =====================================================
def get_lots_by_farm(farm: str) -> list:
    """Lấy danh sách Tên Lô gốc từ dim_lo (đã chuẩn hóa)."""
    res = supabase.table("dim_lo").select("lo_name, dim_farm!inner(farm_name)") \
        .eq("dim_farm.farm_name", farm).eq("is_active", True).order("lo_name").execute()
    if res.data:
        return list(dict.fromkeys(r["lo_name"] for r in res.data))  # unique, preserve order
    return []

def fetch_table_data(table_name: str, farm: str) -> pd.DataFrame:
    """Hàm chung lấy dữ liệu. Quản trị viên (Admin) sẽ lấy của tất cả các farm."""
    tables_with_lo = [
        "stage_logs", "harvest_logs", "destruction_logs", "bsr_logs", 
        "size_measure_logs", "tree_inventory_logs", "soil_ph_logs", 
        "fusarium_logs", "seasons", "base_lots"
    ]
    
    if table_name in tables_with_lo:
        # Use left join (not !inner) so records with dim_lo_id=NULL still appear
        query = supabase.table(table_name).select("*, dim_lo(lo_name, area_ha, is_active, dim_doi(doi_name), dim_farm(farm_name))").eq("is_deleted", False)
        if farm not in ["Admin", "Phòng Kinh doanh"] and farm:
            # For non-admin, we need dim_lo to exist for farm filtering to work
            # But we still show orphan records to avoid silent data loss
            query = supabase.table(table_name).select("*, dim_lo!inner(lo_name, area_ha, is_active, dim_doi(doi_name), dim_farm!inner(farm_name))").eq("is_deleted", False)
            query = query.eq("dim_lo.dim_farm.farm_name", farm)
        # Loại trừ các lô đã bị vô hiệu hóa (is_active = false), VD: lô "11" = 11A + 11B
        query = query.eq("dim_lo.is_active", True)
    else:
        query = supabase.table(table_name).select("*")
        if table_name != "user_roles":
            query = query.eq("is_deleted", False)
            if farm not in ["Admin", "Phòng Kinh doanh"] and farm:
                query = query.eq("farm", farm)

    if table_name != "user_roles":
        res = query.order("created_at", desc=True).execute()
    else:
        res = query.execute()

    if not res.data:
        return pd.DataFrame()
        
    df = pd.DataFrame(res.data)
    
    if table_name in tables_with_lo:
        df["farm"] = df["dim_lo"].apply(lambda x: (x.get("dim_farm") or {}).get("farm_name") if isinstance(x, dict) and x else None)
        df["team"] = df["dim_lo"].apply(lambda x: (x.get("dim_doi") or {}).get("doi_name") if isinstance(x, dict) and x else None)
        df["lo"] = df["dim_lo"].apply(lambda x: x.get("lo_name") if isinstance(x, dict) and x else None)
        df["dien_tich"] = df["dim_lo"].apply(lambda x: x.get("area_ha") if isinstance(x, dict) and x else None)
        df["lot_id"] = df["lo"]
            
        # Optional: we can drop dim_lo if needed, but it shouldn't hurt
        # df = df.drop(columns=["dim_lo"])

    return df
@st.cache_data(ttl=60)
def get_dim_lo_id(farm_name: str, lo_name: str):
    if not farm_name or not lo_name: return None
    res = supabase.table("dim_lo").select("lo_id, dim_farm!inner(farm_name)").eq("lo_name", lo_name).eq("dim_farm.farm_name", farm_name).limit(1).execute()
    return res.data[0]["lo_id"] if res.data else None

@st.cache_data(ttl=60)
def get_farm_id_from_name(farm_name: str):
    """Resolve farm_name → farm_id from dim_farm."""
    res = supabase.table("dim_farm").select("farm_id").eq("farm_name", farm_name).limit(1).execute()
    return res.data[0]["farm_id"] if res.data else None


MAP_COST_QUERY_KEYS = ("cost_farm", "cost_lot")
LOT_COST_CACHE_TTL_SECONDS = 900

FARM_157_LOT_TEAM_OVERRIDES = {
    "NT1": {
        "1A", "1B", "2A", "2B", "3A", "3B", "4", "5", "6",
        "7A", "7B", "A1",
    },
    "NT2": {
        "8A", "8B", "9", "10", "12", "12A", "12B",
        "14A", "14B", "15A", "15B",
    },
}

GENERAL_COST_SCOPE_KEYWORDS = (
    "FARM",
    "VUONUOM",
    "UOM",
    "NHADOI",
    "NHA",
    "COGIOI",
    "DIENNUOC",
    "DIEN",
    "NUOC",
    "BVTV",
    "THUHOACH",
    "TRONGMOI",
    "DUAN",
    "CONGTRINH",
    "HATANG",
    "XUONG",
    "TRAM",
    "XECUOC",
    "XECAY",
    "XEBEN",
    "XEXUC",
)

EXCLUDED_LOT_COST_SCOPE_KEYWORDS = (
    "XUONGDONGGOI",
    "XUONG",
    "DONGGOI",
    "KHOHOCMON",
    "KHO",
    "BANHANG",
    "VANPHONG",
    "CONGTRINH",
    "XAYDUNG",
    "NHAXUONG",
)

PREHARVEST_CARE_PER_TREE_LIMIT = 20_000
GENERAL_OVERHEAD_PER_TREE_LIMIT = 30_000
PREHARVEST_CARE_GROUP_PER_TREE_LIMIT = 80_000
GENERAL_OVERHEAD_GROUP_PER_TREE_LIMIT = 50_000
PLANTING_NURSERY_MAX_DAYS_FROM_PLANTING = 60


def _money(value) -> float:
    try:
        if value is None or pd.isna(value):
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def _format_money(value) -> str:
    return f"{_money(value):,.0f} đ"


def _format_money_compact(value) -> str:
    amount = _money(value)
    abs_amount = abs(amount)
    if abs_amount >= 1_000_000_000:
        return f"{amount / 1_000_000_000:,.1f} tỷ đ"
    if abs_amount >= 1_000_000:
        return f"{amount / 1_000_000:,.1f} triệu đ"
    return _format_money(amount)


def _format_int_commas(value) -> str:
    return f"{int(round(_money(value))):,}"


def _fetch_paginated_rows(table_name: str, select_cols: str, filter_fn=None, order_col: str = None, page_size: int = 1000):
    rows = []
    start = 0
    while True:
        query = supabase.table(table_name).select(select_cols)
        if filter_fn:
            query = filter_fn(query)
        if order_col:
            query = query.order(order_col, desc=False)
        res = query.range(start, start + page_size - 1).execute()
        batch = res.data or []
        rows.extend(batch)
        if len(batch) < page_size:
            break
        start += page_size
    return rows


def _normalize_cost_label(value) -> str:
    text = str(value or "").strip().replace("Đ", "D").replace("đ", "d")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^A-Z0-9+]+", "", text.upper())


def _lot_display_label(row: dict) -> str:
    if not isinstance(row, dict):
        return ""
    return str(row.get("lo_code") or row.get("lo_name") or "").strip()


def _team_group_from_scope_label(farm_name: str, label: str):
    norm = _normalize_cost_label(label)
    if not norm or farm_name == "Farm 195":
        return None

    if farm_name == "Farm 157":
        if norm in {"NT3", "NT4", "NT3+NT4", "NT3+4", "NT34", "NT3NT4", "NT2"}:
            return "NT2"
        if norm in {"NT1", "NT1+NT2", "NT1+2", "NT12", "NT1NT2"}:
            return "NT1"

    if norm in {"NT1", "NT1A", "NT1B"}:
        return "NT1"
    if norm in {"NT2", "NT2A", "NT2B"}:
        return "NT2"
    return None


def _is_general_cost_scope_label(label: str) -> bool:
    norm = _normalize_cost_label(label)
    if not norm:
        return True
    return any(keyword in norm for keyword in GENERAL_COST_SCOPE_KEYWORDS)


def _is_excluded_lot_cost_scope_label(label: str) -> bool:
    norm = _normalize_cost_label(label)
    if not norm:
        return False
    return any(keyword in norm for keyword in EXCLUDED_LOT_COST_SCOPE_KEYWORDS)


def _is_physical_lot_row(row: dict) -> bool:
    if not isinstance(row, dict) or not row.get("is_active", True):
        return False
    label = _lot_display_label(row)
    if not label:
        return False
    lo_type_norm = _normalize_cost_label(row.get("lo_type"))
    if "LOTHUC" in lo_type_norm:
        return True
    if row.get("area_ha") is None:
        return False
    if _team_group_from_scope_label("", label) or _is_general_cost_scope_label(label):
        return False
    return True


def _team_group_for_physical_lot(farm_name: str, lot_row: dict, doi_map: dict) -> str:
    label_norm = _normalize_cost_label(_lot_display_label(lot_row))
    if farm_name == "Farm 195":
        return ""
    if farm_name == "Farm 157":
        for team_group, lot_labels in FARM_157_LOT_TEAM_OVERRIDES.items():
            if label_norm in {_normalize_cost_label(x) for x in lot_labels}:
                return team_group
    doi = doi_map.get(lot_row.get("doi_id"), {})
    doi_label = doi.get("doi_code") or doi.get("doi_name") or ""
    return _team_group_from_scope_label(farm_name, doi_label) or ""


def _lot_weight_on_date(lot_id, cost_dt, plantings_by_lot: dict, lot_meta_by_id: dict) -> float:
    return lifecycle_lot_weight_on_date(lot_id, cost_dt, plantings_by_lot, lot_meta_by_id)


def _fetch_dimension_map(table_name: str, id_col: str, select_cols: str, ids):
    ids = sorted({int(x) for x in ids if x is not None and not pd.isna(x)})
    if not ids:
        return {}
    try:
        res = supabase.table(table_name).select(select_cols).in_(id_col, ids).execute()
        return {row.get(id_col): row for row in (res.data or [])}
    except Exception:
        return {}


def _clear_cost_query_params():
    for key in MAP_COST_QUERY_KEYS:
        if key in st.query_params:
            del st.query_params[key]
    for key in ("cost_dialog_farm", "cost_dialog_lot"):
        if key in st.session_state:
            del st.session_state[key]


def _on_lot_cost_dialog_dismiss():
    _clear_cost_query_params()


def _handle_map_cost_event(event, rerun_on_open: bool = False):
    if not isinstance(event, dict) or event.get("type") != "costClick":
        return False
    event_id = event.get("eventId")
    if event_id and st.session_state.get("_last_map_cost_event_id") == event_id:
        return False
    if event_id:
        st.session_state["_last_map_cost_event_id"] = event_id
    farm_name = str(event.get("farm") or "").strip()
    lo_name = str(event.get("lot") or "").strip()
    if not farm_name or not lo_name:
        return False
    st.session_state["cost_dialog_farm"] = farm_name
    st.session_state["cost_dialog_lot"] = lo_name
    if rerun_on_open:
        st.rerun()
    return True


@st.cache_data(ttl=LOT_COST_CACHE_TTL_SECONDS, show_spinner=False)
def _calculate_lot_cost_per_tree_direct_only(farm_name: str, lo_name: str) -> dict:
    return calculate_lot_cost_per_tree(farm_name, lo_name)
    farm_id = get_farm_id_from_name(farm_name)
    dim_lo_id = get_dim_lo_id(farm_name, lo_name)
    if not farm_id or not dim_lo_id:
        return {"error": "Không tìm thấy farm/lô trong dim_farm/dim_lo."}

    lot_res = supabase.table("dim_lo").select("lo_id, lo_name, area_ha, dim_farm!inner(farm_name)") \
        .eq("lo_id", dim_lo_id).limit(1).execute()
    lot_meta = (lot_res.data or [{}])[0]

    planting_rows = _fetch_paginated_rows(
        "base_lots",
        "id, ngay_trong, so_luong, dien_tich_trong, loai_trong, is_deleted, dim_lo_id",
        lambda q: q.eq("dim_lo_id", dim_lo_id)
                   .eq("is_deleted", False)
                   .eq("loai_trong", "Trồng mới"),
        order_col="ngay_trong",
    )
    if not planting_rows:
        return {
            "farm_name": farm_name,
            "lo_name": lo_name,
            "area_ha": lot_meta.get("area_ha"),
            "batches": pd.DataFrame(),
            "costs": pd.DataFrame(),
            "allocations": pd.DataFrame(),
            "unallocated": pd.DataFrame(),
            "summary": {
                "total_labor": 0,
                "total_material": 0,
                "total_cost": 0,
                "allocated_cost": 0,
                "avg_cost_per_tree": 0,
            },
            "warnings": ["Lô này chưa có đợt trồng mới nên chưa thể tính chi phí/cây."],
        }

    batch_rows = []
    for idx, row in enumerate(planting_rows, start=1):
        ngay_trong = pd.to_datetime(row.get("ngay_trong"), errors="coerce")
        trees = int(_money(row.get("so_luong")))
        batch_rows.append({
            "base_lot_id": int(row.get("id")),
            "dot": idx,
            "ngay_trong": ngay_trong.date() if not pd.isna(ngay_trong) else None,
            "ngay_trong_ts": ngay_trong,
            "so_cay": trees,
            "dien_tich_trong": _money(row.get("dien_tich_trong")),
            "labor_cost": 0.0,
            "material_cost": 0.0,
        })

    base_lot_ids = [b["base_lot_id"] for b in batch_rows]
    seasons_by_batch = {}
    if base_lot_ids:
        try:
            season_res = supabase.table("seasons").select("base_lot_id, vu, ngay_bat_dau, ngay_ket_thuc_thuc_te") \
                .eq("is_deleted", False).in_("base_lot_id", base_lot_ids).order("ngay_bat_dau", desc=True).execute()
            for row in season_res.data or []:
                blid = row.get("base_lot_id")
                if blid not in seasons_by_batch:
                    seasons_by_batch[blid] = row
                if row.get("ngay_ket_thuc_thuc_te") is None:
                    seasons_by_batch[blid] = row
        except Exception:
            seasons_by_batch = {}

    labor_rows = _fetch_paginated_rows(
        "fact_nhat_ky_san_xuat",
        "ngay, thanh_tien, cong_viec_id, hang_muc_du_toan_cong, ma_cv_chuan, lo_2",
        lambda q: q.eq("farm_id", farm_id).eq("lo_id", dim_lo_id).lte("ngay", date.today().isoformat()),
        order_col="ngay",
    )
    material_rows = _fetch_paginated_rows(
        "fact_vat_tu",
        "ngay, thanh_tien, vat_tu_id, cong_viec_id, hang_muc_du_toan_vat_tu, ma_khoan_muc_cp, ma_cv_chuan, lo_2",
        lambda q: q.eq("farm_id", farm_id).eq("lo_id", dim_lo_id).lte("ngay", date.today().isoformat()),
        order_col="ngay",
    )

    cong_viec_ids = [r.get("cong_viec_id") for r in labor_rows + material_rows]
    vat_tu_ids = [r.get("vat_tu_id") for r in material_rows]
    cong_viec_map = _fetch_dimension_map(
        "dim_cong_viec", "cong_viec_id", "cong_viec_id, ma_cv, ten_cong_viec, cong_doan", cong_viec_ids
    )
    vat_tu_map = _fetch_dimension_map(
        "dim_vat_tu", "vat_tu_id", "vat_tu_id, ma_vat_tu, ten_vat_tu, loai_vat_tu", vat_tu_ids
    )

    cost_rows = []
    for row in labor_rows:
        cv = cong_viec_map.get(row.get("cong_viec_id"), {})
        ten_cv = cv.get("ten_cong_viec") or row.get("ma_cv_chuan") or row.get("hang_muc_du_toan_cong") or f"CV {row.get('cong_viec_id')}"
        cost_rows.append({
            "source": "Nhân công",
            "ngay": row.get("ngay"),
            "amount": _money(row.get("thanh_tien")),
            "category": row.get("hang_muc_du_toan_cong") or cv.get("cong_doan") or row.get("ma_cv_chuan") or "Khác",
            "detail": ten_cv,
        })
    for row in material_rows:
        vt = vat_tu_map.get(row.get("vat_tu_id"), {})
        cv = cong_viec_map.get(row.get("cong_viec_id"), {})
        ten_vt = vt.get("ten_vat_tu") or row.get("ma_khoan_muc_cp") or row.get("ma_cv_chuan") or f"VT {row.get('vat_tu_id')}"
        cost_rows.append({
            "source": "Vật tư",
            "ngay": row.get("ngay"),
            "amount": _money(row.get("thanh_tien")),
            "category": row.get("hang_muc_du_toan_vat_tu") or vt.get("loai_vat_tu") or row.get("ma_khoan_muc_cp") or cv.get("ten_cong_viec") or "Khác",
            "detail": ten_vt,
        })

    cost_rows.sort(key=lambda r: (str(r.get("ngay") or ""), r.get("source") or ""))
    allocation_rows = []
    unallocated_rows = []
    for cost in cost_rows:
        amount = _money(cost.get("amount"))
        if amount == 0:
            continue
        cost_dt = pd.to_datetime(cost.get("ngay"), errors="coerce")
        if cost.get("is_harvest_related"):
            batch_weights = {
                b["base_lot_id"]: harvest_quantity_for_batch_until(
                    harvest_rows_by_batch, b["base_lot_id"], cost_dt
                )
                for b in batch_rows
            }
            active_batches = [
                b for b in batch_rows
                if batch_weights.get(b["base_lot_id"], 0) > 0
                and int(b.get("so_cay") or 0) > 0
            ]
            active_trees = sum(batch_weights.get(b["base_lot_id"], 0) for b in active_batches)
            if active_trees <= 0:
                blocked = dict(cost)
                blocked["unallocated_reason"] = "Chi phí thu hoạch chưa có thu hoạch tương ứng"
                unallocated_rows.append(blocked)
                continue
        else:
            batch_weights = {}
            active_batches = [
                b for b in batch_rows
                if is_batch_active_on_date(b, cost_dt)
                and int(b.get("so_cay") or 0) > 0
            ]
            active_trees = sum(int(b.get("so_cay") or 0) for b in active_batches)
            if active_trees <= 0:
                blocked = dict(cost)
                blocked["unallocated_reason"] = "Ngoài vòng đời đợt trồng hoặc thiếu ngày"
                unallocated_rows.append(blocked)
                continue
        for batch in active_batches:
            weight = batch_weights.get(batch["base_lot_id"], int(batch["so_cay"]))
            share = amount * weight / active_trees
            if cost["source"] == "Nhân công":
                batch["labor_cost"] += share
            else:
                batch["material_cost"] += share
            allocation_rows.append({
                "base_lot_id": batch["base_lot_id"],
                "dot": batch["dot"],
                "source": cost["source"],
                "ngay": cost["ngay"],
                "amount": share,
                "category": cost["category"],
                "detail": cost["detail"],
            })

    for batch in batch_rows:
        season = seasons_by_batch.get(batch["base_lot_id"], {})
        batch["vu"] = season.get("vu", "")
        batch["total_cost"] = batch["labor_cost"] + batch["material_cost"]
        batch["cost_per_tree"] = batch["total_cost"] / batch["so_cay"] if batch["so_cay"] > 0 else 0.0

    allocation_df = pd.DataFrame(allocation_rows)
    cost_df = allocation_df.copy()
    unallocated_df = pd.DataFrame(unallocated_rows)
    batch_df = pd.DataFrame(batch_rows)

    allocated_cost = float(batch_df["total_cost"].sum()) if not batch_df.empty else 0.0
    allocated_tree_denominator = int(batch_df.loc[batch_df["total_cost"] > 0, "so_cay"].sum()) if not batch_df.empty else 0
    if allocated_tree_denominator <= 0 and not batch_df.empty:
        allocated_tree_denominator = int(batch_df["so_cay"].sum())
    total_labor = sum(_money(r.get("amount")) for r in cost_rows if r.get("source") == "Nhân công")
    total_material = sum(_money(r.get("amount")) for r in cost_rows if r.get("source") == "Vật tư")
    total_labor = float(batch_df["labor_cost"].sum()) if not batch_df.empty else 0.0
    total_material = float(batch_df["material_cost"].sum()) if not batch_df.empty else 0.0
    total_labor = float(batch_df["labor_cost"].sum()) if not batch_df.empty else 0.0
    total_material = float(batch_df["material_cost"].sum()) if not batch_df.empty else 0.0
    warnings = []
    if len(batch_rows) > 1:
        warnings.append("Lô có nhiều đợt trồng: chi phí từng ngày được chia theo tỷ lệ số cây của các đợt đang active.")
    if not unallocated_df.empty:
        warnings.append("Một số chi phí không phù hợp vòng đời đợt trồng đã được tách khỏi chi phí/cây.")

    return {
        "farm_name": farm_name,
        "lo_name": lo_name,
        "area_ha": lot_meta.get("area_ha"),
        "batches": batch_df,
        "costs": cost_df,
        "allocations": allocation_df,
        "unallocated": unallocated_df,
        "summary": {
            "total_labor": total_labor,
            "total_material": total_material,
            "total_cost": total_labor + total_material,
            "allocated_cost": allocated_cost,
            "unallocated_cost": float(unallocated_df["amount"].sum()) if not unallocated_df.empty else 0.0,
            "avg_cost_per_tree": allocated_cost / allocated_tree_denominator if allocated_tree_denominator > 0 else 0.0,
            "tree_denominator": allocated_tree_denominator,
        },
        "warnings": warnings,
    }


@st.cache_data(ttl=LOT_COST_CACHE_TTL_SECONDS, show_spinner=False)
def _fetch_lot_cost_context(farm_id, today_iso: str) -> dict:
    dim_lot_rows = _fetch_paginated_rows(
        "dim_lo",
        "lo_id, farm_id, lo_name, lo_code, area_ha, lo_type, doi_id, is_active",
        lambda q: q.eq("farm_id", farm_id).eq("is_active", True),
    )
    doi_ids = [row.get("doi_id") for row in dim_lot_rows]
    doi_map = _fetch_dimension_map("dim_doi", "doi_id", "doi_id, doi_name, doi_code", doi_ids)

    physical_lot_rows = [row for row in dim_lot_rows if _is_physical_lot_row(row)]
    physical_ids = [row.get("lo_id") for row in physical_lot_rows if row.get("lo_id") is not None]
    all_planting_rows = []
    if physical_ids:
        all_planting_rows = _fetch_paginated_rows(
            "base_lots",
            "id, ngay_trong, so_luong, dien_tich_trong, loai_trong, is_deleted, dim_lo_id",
            lambda q: q.in_("dim_lo_id", physical_ids)
                       .eq("is_deleted", False)
                       .eq("loai_trong", "Trồng mới"),
            order_col="ngay_trong",
        )

    all_base_lot_ids = [
        row.get("id") for row in all_planting_rows
        if row.get("id") is not None and not pd.isna(row.get("id"))
    ]
    season_rows = []
    harvest_lifecycle_rows = []
    destruction_lifecycle_rows = []
    stage_lifecycle_rows = []
    if all_base_lot_ids:
        season_rows = _fetch_paginated_rows(
            "seasons",
            "base_lot_id, vu, ngay_bat_dau, ngay_ket_thuc_thuc_te, is_deleted",
            lambda q: q.in_("base_lot_id", all_base_lot_ids).eq("is_deleted", False),
            order_col="ngay_bat_dau",
        )
        harvest_lifecycle_rows = _fetch_paginated_rows(
            "harvest_logs",
            "base_lot_id, ngay_thu_hoach, so_luong, is_deleted",
            lambda q: q.in_("base_lot_id", all_base_lot_ids).eq("is_deleted", False),
            order_col="ngay_thu_hoach",
        )
        destruction_lifecycle_rows = _fetch_paginated_rows(
            "destruction_logs",
            "base_lot_id, ngay_xuat_huy, so_luong, is_deleted",
            lambda q: q.in_("base_lot_id", all_base_lot_ids).eq("is_deleted", False),
            order_col="ngay_xuat_huy",
        )
        stage_lifecycle_rows = _fetch_paginated_rows(
            "stage_logs",
            "base_lot_id, giai_doan, ngay_thuc_hien, so_luong, is_deleted",
            lambda q: q.in_("base_lot_id", all_base_lot_ids).eq("is_deleted", False),
            order_col="ngay_thuc_hien",
        )

    labor_rows = _fetch_paginated_rows(
        "fact_nhat_ky_san_xuat",
        "nhat_ky_id, ngay, thanh_tien, klcv, cong_viec_id, hang_muc_du_toan_cong, ma_cv_chuan, lo_id, lo_2",
        lambda q: q.eq("farm_id", farm_id).lte("ngay", today_iso),
        order_col="ngay",
    )
    material_rows = _fetch_paginated_rows(
        "fact_vat_tu",
        "vat_tu_fact_id, ngay, thanh_tien, so_luong, vat_tu_id, cong_viec_id, hang_muc_du_toan_vat_tu, ma_khoan_muc_cp, ma_cv_chuan, lo_id, lo_2",
        lambda q: q.eq("farm_id", farm_id).lte("ngay", today_iso),
        order_col="ngay",
    )

    cong_viec_ids = [r.get("cong_viec_id") for r in labor_rows + material_rows]
    vat_tu_ids = [r.get("vat_tu_id") for r in material_rows]
    row_lot_ids = [r.get("lo_id") for r in labor_rows + material_rows]
    global_lot_map = _fetch_dimension_map(
        "dim_lo", "lo_id", "lo_id, farm_id, lo_name, lo_code, area_ha, lo_type, doi_id, is_active", row_lot_ids
    )
    cong_viec_map = _fetch_dimension_map(
        "dim_cong_viec", "cong_viec_id", "cong_viec_id, ma_cv, ten_cong_viec, cong_doan", cong_viec_ids
    )
    vat_tu_map = _fetch_dimension_map(
        "dim_vat_tu", "vat_tu_id", "vat_tu_id, ma_vat_tu, ten_vat_tu, loai_vat_tu", vat_tu_ids
    )

    return {
        "dim_lot_rows": dim_lot_rows,
        "doi_map": doi_map,
        "all_planting_rows": all_planting_rows,
        "season_rows": season_rows,
        "harvest_lifecycle_rows": harvest_lifecycle_rows,
        "destruction_lifecycle_rows": destruction_lifecycle_rows,
        "stage_lifecycle_rows": stage_lifecycle_rows,
        "labor_rows": labor_rows,
        "material_rows": material_rows,
        "global_lot_map": global_lot_map,
        "cong_viec_map": cong_viec_map,
        "vat_tu_map": vat_tu_map,
    }


@st.cache_data(ttl=LOT_COST_CACHE_TTL_SECONDS, show_spinner=False)
def calculate_lot_cost_per_tree(farm_name: str, lo_name: str) -> dict:
    farm_id = get_farm_id_from_name(farm_name)
    dim_lo_id = get_dim_lo_id(farm_name, lo_name)
    if not farm_id or not dim_lo_id:
        return {"error": "Không tìm thấy farm/lô trong dim_farm/dim_lo."}

    today_iso = date.today().isoformat()
    cost_context = _fetch_lot_cost_context(farm_id, today_iso)
    dim_lot_rows = cost_context["dim_lot_rows"]
    lot_meta_by_id = {row.get("lo_id"): row for row in dim_lot_rows}
    lot_meta = lot_meta_by_id.get(dim_lo_id, {})
    if not lot_meta:
        return {"error": "Không tìm thấy farm/lô trong dim_lo."}

    doi_map = cost_context["doi_map"]
    physical_lot_rows = [row for row in dim_lot_rows if _is_physical_lot_row(row)]
    physical_lot_by_id = {row.get("lo_id"): row for row in physical_lot_rows}
    physical_lot_by_label = {}
    for row in physical_lot_rows:
        for label in (row.get("lo_code"), row.get("lo_name")):
            norm = _normalize_cost_label(label)
            if norm:
                physical_lot_by_label[norm] = row

    target_team_group = _team_group_for_physical_lot(farm_name, lot_meta, doi_map)
    lot_team_group_by_id = {
        row.get("lo_id"): _team_group_for_physical_lot(farm_name, row, doi_map)
        for row in physical_lot_rows
    }

    physical_ids = list(physical_lot_by_id.keys()) or [dim_lo_id]
    all_planting_rows = cost_context["all_planting_rows"] or _fetch_paginated_rows(
        "base_lots",
        "id, ngay_trong, so_luong, dien_tich_trong, loai_trong, is_deleted, dim_lo_id",
        lambda q: q.in_("dim_lo_id", physical_ids)
                   .eq("is_deleted", False)
                   .eq("loai_trong", "Trồng mới"),
        order_col="ngay_trong",
    )
    all_planting_rows = build_batch_lifecycle(
        all_planting_rows,
        cost_context.get("season_rows", []),
        cost_context.get("harvest_lifecycle_rows", []),
        cost_context.get("destruction_lifecycle_rows", []),
    )
    harvest_rows_by_batch = build_harvest_rows_by_batch(cost_context.get("harvest_lifecycle_rows", []))
    stage_rows_by_batch = build_stage_rows_by_batch(cost_context.get("stage_lifecycle_rows", []))
    plantings_by_lot = {}
    for row in all_planting_rows:
        row = dict(row)
        if row.get("ngay_trong_ts") is None:
            row["ngay_trong_ts"] = pd.to_datetime(row.get("ngay_trong"), errors="coerce")
        plantings_by_lot.setdefault(row.get("dim_lo_id"), []).append(row)

    planting_rows = sorted(
        plantings_by_lot.get(dim_lo_id, []),
        key=lambda row: str(row.get("ngay_trong") or ""),
    )
    if not planting_rows:
        return {
            "farm_name": farm_name,
            "lo_name": lo_name,
            "area_ha": lot_meta.get("area_ha"),
            "batches": pd.DataFrame(),
            "costs": pd.DataFrame(),
            "allocations": pd.DataFrame(),
            "unallocated": pd.DataFrame(),
            "summary": {
                "total_labor": 0,
                "total_material": 0,
                "total_cost": 0,
                "allocated_cost": 0,
                "avg_cost_per_tree": 0,
            },
            "warnings": ["Lô này chưa có đợt trồng mới nên chưa thể tính chi phí/cây."],
        }

    batch_rows = []
    for idx, row in enumerate(planting_rows, start=1):
        ngay_trong = row.get("ngay_trong_ts")
        trees = int(_money(row.get("so_luong")))
        batch_rows.append({
            "base_lot_id": int(row.get("id")),
            "dot": idx,
            "ngay_trong": ngay_trong.date() if not pd.isna(ngay_trong) else None,
            "ngay_trong_ts": ngay_trong,
            "active_until": row.get("active_until"),
            "active_until_ts": row.get("active_until_ts"),
            "lifecycle_end_reason": row.get("lifecycle_end_reason", ""),
            "so_cay": trees,
            "dien_tich_trong": _money(row.get("dien_tich_trong")),
            "labor_cost": 0.0,
            "material_cost": 0.0,
        })

    base_lot_ids = [b["base_lot_id"] for b in batch_rows]
    seasons_by_batch = {}
    if base_lot_ids:
        try:
            season_res = supabase.table("seasons").select("base_lot_id, vu, ngay_bat_dau, ngay_ket_thuc_thuc_te") \
                .eq("is_deleted", False).in_("base_lot_id", base_lot_ids).order("ngay_bat_dau", desc=True).execute()
            for row in season_res.data or []:
                blid = row.get("base_lot_id")
                if blid not in seasons_by_batch:
                    seasons_by_batch[blid] = row
                if row.get("ngay_ket_thuc_thuc_te") is None:
                    seasons_by_batch[blid] = row
        except Exception:
            seasons_by_batch = {}

    labor_rows = cost_context["labor_rows"]
    material_rows = cost_context["material_rows"]
    global_lot_map = cost_context["global_lot_map"]
    cong_viec_map = cost_context["cong_viec_map"]
    vat_tu_map = cost_context["vat_tu_map"]

    def classify_cost_scope(row):
        row_lo_id = row.get("lo_id")
        local_lot = lot_meta_by_id.get(row_lo_id)
        global_lot = global_lot_map.get(row_lo_id)
        fallback_label = row.get("lo_2")
        fallback_norm = _normalize_cost_label(fallback_label)

        if local_lot and local_lot.get("lo_id") in physical_lot_by_id:
            return {
                "scope": "direct",
                "target_lot_id": local_lot.get("lo_id"),
                "scope_label": _lot_display_label(local_lot),
            }

        if not local_lot and fallback_norm in physical_lot_by_label:
            fallback_lot = physical_lot_by_label[fallback_norm]
            return {
                "scope": "direct",
                "target_lot_id": fallback_lot.get("lo_id"),
                "scope_label": _lot_display_label(fallback_lot),
            }

        scope_label = _lot_display_label(local_lot) or _lot_display_label(global_lot) or str(fallback_label or "")
        if _is_excluded_lot_cost_scope_label(scope_label):
            return {"scope": "excluded", "scope_label": scope_label or farm_name}

        team_group = _team_group_from_scope_label(farm_name, scope_label)
        if team_group:
            return {"scope": "team", "team_group": team_group, "scope_label": scope_label}

        if row_lo_id is None or _is_general_cost_scope_label(scope_label):
            return {"scope": "farm", "scope_label": scope_label or farm_name}
        return {"scope": "unallocated", "scope_label": scope_label or farm_name}

    def lot_harvest_weight_on_date(lot_id, cost_dt):
        return sum(
            harvest_quantity_for_batch_until(harvest_rows_by_batch, row.get("id"), cost_dt)
            for row in plantings_by_lot.get(lot_id, [])
        )

    def share_cost_to_target_lot(cost, cost_dt):
        amount = _money(cost.get("amount"))
        scope = cost.get("scope")
        if amount == 0:
            return None
        if scope in {"unallocated", "excluded"}:
            return None

        if scope == "direct":
            if cost.get("target_lot_id") != dim_lo_id:
                return None
            shared = dict(cost)
            shared["amount"] = amount
            shared["original_amount"] = amount
            shared["original_quantity"] = _money(cost.get("quantity"))
            return shared

        if scope == "team":
            if not target_team_group or cost.get("team_group") != target_team_group:
                return None
            recipient_ids = [
                lot_id for lot_id, team_group in lot_team_group_by_id.items()
                if team_group == cost.get("team_group")
            ]
        else:
            recipient_ids = list(physical_lot_by_id.keys())

        if cost.get("is_harvest_related"):
            weights = {
                lot_id: lot_harvest_weight_on_date(lot_id, cost_dt)
                for lot_id in recipient_ids
            }
        else:
            weights = {
                lot_id: _lot_weight_on_date(lot_id, cost_dt, plantings_by_lot, lot_meta_by_id)
                for lot_id in recipient_ids
            }
        total_weight = sum(weight for weight in weights.values() if weight > 0)
        target_weight = weights.get(dim_lo_id, 0.0)
        if total_weight <= 0 or target_weight <= 0:
            return None

        shared = dict(cost)
        shared["original_amount"] = amount
        ratio = target_weight / total_weight
        shared["amount"] = amount * ratio
        shared["original_quantity"] = _money(cost.get("quantity"))
        shared["quantity"] = _money(cost.get("quantity")) * ratio
        return shared

    raw_cost_rows = []
    for row in labor_rows:
        cv = cong_viec_map.get(row.get("cong_viec_id"), {})
        ten_cv = cv.get("ten_cong_viec") or row.get("ma_cv_chuan") or row.get("hang_muc_du_toan_cong") or f"CV {row.get('cong_viec_id')}"
        cost = {
            "source": "Nhân công",
            "source_row_id": row.get("nhat_ky_id"),
            "ngay": row.get("ngay"),
            "amount": _money(row.get("thanh_tien")),
            "quantity": _money(row.get("klcv")),
            "category": row.get("hang_muc_du_toan_cong") or cv.get("cong_doan") or row.get("ma_cv_chuan") or "Khác",
            "detail": ten_cv,
            "cong_doan": cv.get("cong_doan"),
            "ma_cv": cv.get("ma_cv"),
            "ma_cv_chuan": row.get("ma_cv_chuan"),
            "hang_muc": row.get("hang_muc_du_toan_cong"),
        }
        cost["is_harvest_related"] = is_harvest_related_cost(cost)
        cost["requires_cut_done"] = is_bunch_care_requiring_cut(cost)
        cost["stage_quantity_capped"] = is_stage_quantity_capped_cost(cost)
        cost["is_preharvest_care"] = is_preharvest_care_cost(cost)
        cost["is_general_overhead"] = is_general_overhead_cost(cost)
        cost["is_planting_or_nursery"] = is_planting_or_nursery_cost(cost)
        cost.update(classify_cost_scope(row))
        cost["is_harvest_related"] = cost["is_harvest_related"] or is_harvest_related_cost(cost)
        cost["requires_cut_done"] = cost["requires_cut_done"] or is_bunch_care_requiring_cut(cost)
        cost["stage_quantity_capped"] = cost["stage_quantity_capped"] or is_stage_quantity_capped_cost(cost)
        cost["is_preharvest_care"] = cost["is_preharvest_care"] or is_preharvest_care_cost(cost)
        cost["is_general_overhead"] = cost["is_general_overhead"] or is_general_overhead_cost(cost)
        cost["is_planting_or_nursery"] = cost["is_planting_or_nursery"] or is_planting_or_nursery_cost(cost)
        raw_cost_rows.append(cost)
    for row in material_rows:
        vt = vat_tu_map.get(row.get("vat_tu_id"), {})
        cv = cong_viec_map.get(row.get("cong_viec_id"), {})
        ten_vt = vt.get("ten_vat_tu") or row.get("ma_khoan_muc_cp") or row.get("ma_cv_chuan") or f"VT {row.get('vat_tu_id')}"
        cost = {
            "source": "Vật tư",
            "source_row_id": row.get("vat_tu_fact_id"),
            "ngay": row.get("ngay"),
            "amount": _money(row.get("thanh_tien")),
            "quantity": _money(row.get("so_luong")),
            "category": row.get("hang_muc_du_toan_vat_tu") or vt.get("loai_vat_tu") or row.get("ma_khoan_muc_cp") or cv.get("ten_cong_viec") or "Khác",
            "detail": ten_vt,
            "cong_doan": cv.get("cong_doan"),
            "ma_cv": cv.get("ma_cv"),
            "ma_cv_chuan": row.get("ma_cv_chuan"),
            "ma_khoan_muc_cp": row.get("ma_khoan_muc_cp"),
            "hang_muc": row.get("hang_muc_du_toan_vat_tu"),
        }
        cost["is_harvest_related"] = is_harvest_related_cost(cost)
        cost["requires_cut_done"] = is_bunch_care_requiring_cut(cost)
        cost["stage_quantity_capped"] = False
        cost["is_preharvest_care"] = is_preharvest_care_cost(cost)
        cost["is_general_overhead"] = is_general_overhead_cost(cost)
        cost["is_planting_or_nursery"] = is_planting_or_nursery_cost(cost)
        cost.update(classify_cost_scope(row))
        cost["is_harvest_related"] = cost["is_harvest_related"] or is_harvest_related_cost(cost)
        cost["requires_cut_done"] = cost["requires_cut_done"] or is_bunch_care_requiring_cut(cost)
        cost["is_preharvest_care"] = cost["is_preharvest_care"] or is_preharvest_care_cost(cost)
        cost["is_general_overhead"] = cost["is_general_overhead"] or is_general_overhead_cost(cost)
        cost["is_planting_or_nursery"] = cost["is_planting_or_nursery"] or is_planting_or_nursery_cost(cost)
        raw_cost_rows.append(cost)

    cost_rows = []
    for cost in raw_cost_rows:
        cost_dt = pd.to_datetime(cost.get("ngay"), errors="coerce")
        shared_cost = share_cost_to_target_lot(cost, cost_dt)
        if shared_cost and _money(shared_cost.get("amount")) != 0:
            cost_rows.append(shared_cost)

    cost_rows.sort(key=lambda r: (str(r.get("ngay") or ""), r.get("source") or ""))
    allocation_rows = []
    unallocated_rows = []
    stage_quantity_usage = {}

    def _stage_work_key(cost: dict) -> str:
        for field in ("ma_cv_chuan", "ma_cv", "detail", "category"):
            norm = _normalize_cost_label(cost.get(field))
            if norm:
                return norm
        return "UNKNOWN"

    def _add_unallocated(cost: dict, reason: str, amount_override=None, quantity_override=None):
        blocked = dict(cost)
        if amount_override is not None:
            blocked["amount"] = amount_override
        if quantity_override is not None:
            blocked["quantity"] = quantity_override
        blocked["unallocated_reason"] = reason
        unallocated_rows.append(blocked)

    def _add_allocation(cost: dict, batch: dict, share: float, quantity_share=None):
        if share <= 0:
            return
        if cost["source"] == "Nhân công":
            batch["labor_cost"] += share
        else:
            batch["material_cost"] += share
        allocation_rows.append({
            "base_lot_id": batch["base_lot_id"],
            "dot": batch["dot"],
            "source": cost["source"],
            "ngay": cost["ngay"],
            "amount": share,
            "quantity": quantity_share,
            "category": cost["category"],
            "detail": cost["detail"],
            "scope": cost.get("scope"),
            "scope_label": cost.get("scope_label"),
            "is_harvest_related": cost.get("is_harvest_related", False),
            "requires_cut_done": cost.get("requires_cut_done", False),
            "stage_quantity_capped": cost.get("stage_quantity_capped", False),
            "is_preharvest_care": cost.get("is_preharvest_care", False),
            "is_general_overhead": cost.get("is_general_overhead", False),
            "is_planting_or_nursery": cost.get("is_planting_or_nursery", False),
        })

    def _days_from_nearest_planting(active_batches, cost_dt):
        cost_ts = pd.to_datetime(cost_dt, errors="coerce")
        if pd.isna(cost_ts):
            return None
        day_gaps = []
        for batch in active_batches:
            planting_ts = batch.get("ngay_trong_ts") or pd.to_datetime(batch.get("ngay_trong"), errors="coerce")
            if pd.isna(planting_ts):
                continue
            day_gaps.append(abs((cost_ts - planting_ts).days))
        return min(day_gaps) if day_gaps else None

    def _active_cost_exclusion_reason(cost: dict, amount: float, active_batches, active_trees, cost_dt):
        if active_trees <= 0:
            return ""
        amount_per_tree = amount / active_trees
        if cost.get("is_preharvest_care") and amount_per_tree >= PREHARVEST_CARE_PER_TREE_LIMIT:
            return (
                "Chi phí phân bón/chăm sóc cây vượt ngưỡng theo số cây active "
                f"({amount_per_tree:,.0f} đ/cây)"
            )
        if cost.get("is_general_overhead") and amount_per_tree >= GENERAL_OVERHEAD_PER_TREE_LIMIT:
            return (
                "Chi phí cơ giới/điện nước/dầu DO vượt ngưỡng theo số cây active "
                f"({amount_per_tree:,.0f} đ/cây)"
            )
        if cost.get("is_planting_or_nursery"):
            day_gap = _days_from_nearest_planting(active_batches, cost_dt)
            if day_gap is not None and day_gap > PLANTING_NURSERY_MAX_DAYS_FROM_PLANTING:
                return (
                    "Chi phí trồng mới/vườn ươm phát sinh xa ngày trồng active "
                    f"({day_gap} ngày)"
                )
        return ""
    for cost in cost_rows:
        amount = _money(cost.get("amount"))
        if amount == 0:
            continue
        cost_dt = pd.to_datetime(cost.get("ngay"), errors="coerce")
        if cost.get("is_harvest_related"):
            batch_weights = {
                b["base_lot_id"]: harvest_quantity_for_batch_until(
                    harvest_rows_by_batch, b["base_lot_id"], cost_dt
                )
                for b in batch_rows
            }
            active_batches = [
                b for b in batch_rows
                if batch_weights.get(b["base_lot_id"], 0) > 0
                and int(b.get("so_cay") or 0) > 0
            ]
            active_trees = sum(batch_weights.get(b["base_lot_id"], 0) for b in active_batches)
            if active_trees <= 0:
                blocked = dict(cost)
                blocked["unallocated_reason"] = "Chi phí thu hoạch chưa có thu hoạch tương ứng"
                unallocated_rows.append(blocked)
                continue
        else:
            batch_weights = {}
            active_batches = [
                b for b in batch_rows
                if is_batch_active_on_date(b, cost_dt)
                and int(b.get("so_cay") or 0) > 0
            ]
            active_trees = sum(int(b.get("so_cay") or 0) for b in active_batches)
            if active_trees <= 0:
                blocked = dict(cost)
                blocked["unallocated_reason"] = "Ngoài vòng đời đợt trồng hoặc thiếu ngày"
                unallocated_rows.append(blocked)
                continue
            if cost.get("requires_cut_done") or cost.get("stage_quantity_capped"):
                gated_batches = []
                for batch in active_batches:
                    cut_qty = stage_quantity_for_batch_until(
                        stage_rows_by_batch, batch["base_lot_id"], "Cắt bắp", cost_dt
                    )
                    if cut_qty > 0:
                        gated_batches.append(batch)
                        batch_weights[batch["base_lot_id"]] = min(cut_qty, int(batch.get("so_cay") or 0))
                active_batches = gated_batches
                active_trees = sum(batch_weights.get(b["base_lot_id"], 0) for b in active_batches)
                if active_trees <= 0:
                    _add_unallocated(cost, "Chưa có mốc Cắt bắp tương ứng cho hạng mục chăm sóc buồng")
                    continue
            exclusion_reason = _active_cost_exclusion_reason(cost, amount, active_batches, active_trees, cost_dt)
            if exclusion_reason:
                _add_unallocated(cost, exclusion_reason)
                continue
            quantity = _money(cost.get("quantity"))
            if cost.get("stage_quantity_capped") and quantity > 0:
                work_key = _stage_work_key(cost)
                remaining_qty = quantity
                remaining_amount = amount
                allocated_amount = 0.0
                for batch in sorted(active_batches, key=lambda b: int(b.get("dot") or 0)):
                    cut_qty = min(
                        stage_quantity_for_batch_until(stage_rows_by_batch, batch["base_lot_id"], "Cắt bắp", cost_dt),
                        int(batch.get("so_cay") or 0),
                    )
                    usage_key = (batch["base_lot_id"], work_key)
                    used_qty = stage_quantity_usage.get(usage_key, 0.0)
                    available_qty = max(0.0, cut_qty - used_qty)
                    if available_qty <= 0:
                        continue
                    qty_share = min(remaining_qty, available_qty)
                    share = amount * qty_share / quantity
                    _add_allocation(cost, batch, share, qty_share)
                    stage_quantity_usage[usage_key] = used_qty + qty_share
                    remaining_qty -= qty_share
                    remaining_amount -= share
                    allocated_amount += share
                    if remaining_qty <= 1e-9:
                        break
                if allocated_amount <= 0:
                    _add_unallocated(cost, "Vượt KL Cắt bắp lũy kế hoặc số cây của đợt")
                elif remaining_qty > 1e-9 and remaining_amount > 0:
                    _add_unallocated(
                        cost,
                        "Phần KL vượt Cắt bắp lũy kế hoặc số cây của đợt",
                        amount_override=remaining_amount,
                        quantity_override=remaining_qty,
                    )
                continue
        for batch in active_batches:
            weight = batch_weights.get(batch["base_lot_id"], int(batch["so_cay"]))
            share = amount * weight / active_trees
            _add_allocation(cost, batch, share)

    batch_tree_by_id = {row["base_lot_id"]: int(row.get("so_cay") or 0) for row in batch_rows}

    def _exclude_large_allocation_groups():
        nonlocal allocation_rows
        if not allocation_rows:
            return
        allocation_df_tmp = pd.DataFrame(allocation_rows)
        if allocation_df_tmp.empty:
            return
        drop_indices = set()
        group_cols = ["source", "scope", "scope_label", "category", "detail"]
        for _, group in allocation_df_tmp.groupby(group_cols, dropna=False):
            affected_batch_ids = set(group["base_lot_id"].dropna().astype(int).tolist())
            affected_trees = sum(batch_tree_by_id.get(batch_id, 0) for batch_id in affected_batch_ids)
            if affected_trees <= 0:
                continue
            group_amount = float(group["amount"].sum())
            amount_per_tree = group_amount / affected_trees
            is_preharvest_group = bool(group.get("is_preharvest_care", pd.Series([False])).fillna(False).any())
            is_general_group = bool(group.get("is_general_overhead", pd.Series([False])).fillna(False).any())
            reason = ""
            if is_preharvest_group and amount_per_tree >= PREHARVEST_CARE_GROUP_PER_TREE_LIMIT:
                reason = (
                    "Nhóm phân bón/chăm sóc cây vượt ngưỡng theo số cây active "
                    f"({amount_per_tree:,.0f} đ/cây)"
                )
            elif is_general_group and amount_per_tree >= GENERAL_OVERHEAD_GROUP_PER_TREE_LIMIT:
                reason = (
                    "Nhóm cơ giới/điện nước/dầu DO vượt ngưỡng theo số cây active "
                    f"({amount_per_tree:,.0f} đ/cây)"
                )
            if not reason:
                continue
            for idx, row in group.iterrows():
                blocked = row.to_dict()
                blocked["unallocated_reason"] = reason
                unallocated_rows.append(blocked)
                drop_indices.add(idx)
        if not drop_indices:
            return
        allocation_df_tmp = allocation_df_tmp.drop(index=list(drop_indices))
        allocation_rows = allocation_df_tmp.to_dict("records")
        for batch in batch_rows:
            batch["labor_cost"] = 0.0
            batch["material_cost"] = 0.0
        for row in allocation_rows:
            batch = next((b for b in batch_rows if b["base_lot_id"] == row.get("base_lot_id")), None)
            if not batch:
                continue
            if row.get("source") == "Nhân công":
                batch["labor_cost"] += _money(row.get("amount"))
            else:
                batch["material_cost"] += _money(row.get("amount"))

    _exclude_large_allocation_groups()

    for batch in batch_rows:
        season = seasons_by_batch.get(batch["base_lot_id"], {})
        batch["vu"] = season.get("vu", "")
        batch["total_cost"] = batch["labor_cost"] + batch["material_cost"]
        batch["cost_per_tree"] = batch["total_cost"] / batch["so_cay"] if batch["so_cay"] > 0 else 0.0

    allocation_df = pd.DataFrame(allocation_rows)
    cost_df = allocation_df.copy()
    unallocated_df = pd.DataFrame(unallocated_rows)
    batch_df = pd.DataFrame(batch_rows)

    allocated_cost = float(batch_df["total_cost"].sum()) if not batch_df.empty else 0.0
    allocated_tree_denominator = int(batch_df.loc[batch_df["total_cost"] > 0, "so_cay"].sum()) if not batch_df.empty else 0
    if allocated_tree_denominator <= 0 and not batch_df.empty:
        allocated_tree_denominator = int(batch_df["so_cay"].sum())
    total_labor = sum(_money(r.get("amount")) for r in cost_rows if r.get("source") == "Nhân công")
    total_material = sum(_money(r.get("amount")) for r in cost_rows if r.get("source") == "Vật tư")
    total_labor = float(batch_df["labor_cost"].sum()) if not batch_df.empty else 0.0
    total_material = float(batch_df["material_cost"].sum()) if not batch_df.empty else 0.0
    warnings = []
    if len(batch_rows) > 1:
        warnings.append("Lô có nhiều đợt trồng: chi phí từng ngày được chia theo tỷ lệ số cây của các đợt đang active.")
    if not unallocated_df.empty:
        reasons = unallocated_df.get("unallocated_reason", pd.Series([""] * len(unallocated_df))).fillna("")
        if reasons.str.contains("Chưa có mốc Cắt bắp|Cắt bắp lũy kế", case=False, regex=True).any():
            warnings.append("Một số chi phí chăm sóc buồng không phù hợp mốc Cắt bắp đã được tách khỏi chi phí/cây.")
        if reasons.str.contains("thu hoạch", case=False, regex=False).any():
            warnings.append("Một số chi phí thu hoạch chưa có thu hoạch tương ứng đã được tách khỏi chi phí/cây.")
        if reasons.str.contains("Ngoài vòng đời", regex=False).any():
            warnings.append("Một số chi phí ngoài vòng đời đợt trồng đã được tách khỏi chi phí/cây.")

    return {
        "farm_name": farm_name,
        "lo_name": lo_name,
        "area_ha": lot_meta.get("area_ha"),
        "batches": batch_df,
        "costs": cost_df,
        "allocations": allocation_df,
        "unallocated": unallocated_df,
        "summary": {
            "total_labor": total_labor,
            "total_material": total_material,
            "total_cost": total_labor + total_material,
            "allocated_cost": allocated_cost,
            "unallocated_cost": float(unallocated_df["amount"].sum()) if not unallocated_df.empty else 0.0,
            "avg_cost_per_tree": allocated_cost / allocated_tree_denominator if allocated_tree_denominator > 0 else 0.0,
            "tree_denominator": allocated_tree_denominator,
        },
        "warnings": warnings,
    }


@configured_dialog_decorator("Dashboard chi phí/cây", width="large", on_dismiss=_on_lot_cost_dialog_dismiss)
def _render_lot_cost_dialog(farm_name: str, lo_name: str):
    st.markdown("""
    <style>
    div[role="dialog"][aria-modal="true"] {
        width: min(92vw, 1400px) !important;
        max-width: 92vw !important;
    }
    </style>
    """, unsafe_allow_html=True)

    with st.spinner("Đang tính chi phí/cây cho lô này..."):
        result = calculate_lot_cost_per_tree(farm_name, lo_name)
    if result.get("error"):
        st.error(result["error"])
        if st.button("Đóng", key="close_lot_cost_dialog_error"):
            _clear_cost_query_params()
            st.rerun()
        return False

    st.markdown(f"#### {farm_name} · Lô {lo_name}")
    area_ha = result.get("area_ha")
    batch_df = result.get("batches", pd.DataFrame())
    if area_ha is not None:
        st.caption(f"Diện tích lô: {float(area_ha):.2f} ha · Số đợt trồng: {len(batch_df) if isinstance(batch_df, pd.DataFrame) else 0}")
    else:
        st.caption(f"Số đợt trồng: {len(batch_df) if isinstance(batch_df, pd.DataFrame) else 0}")

    summary = result.get("summary", {})
    m1, m2, m3 = st.columns([1.4, 1, 1])
    m1.metric("Chi phí/cây TB", _format_money_compact(summary.get("avg_cost_per_tree", 0)))
    m2.metric("Tổng chi phí tính vào cây", _format_money_compact(summary.get("allocated_cost", 0)))
    m3.metric("Tổng cây tính", _format_int_commas(summary.get("tree_denominator", 0)))
    st.caption("Chỉ tính các khoản chi phí phù hợp với vòng đời đợt trồng và mốc sinh trưởng của lô.")

    if not isinstance(batch_df, pd.DataFrame) or batch_df.empty:
        st.info("Lô này chưa có đợt trồng mới để tính chi phí/cây.")
    else:
        batch_view = batch_df.copy().sort_values("dot")
        batch_display = pd.DataFrame({
            "Đợt": batch_view["dot"].apply(lambda x: f"Đợt {int(x)}"),
            "Ngày trồng": batch_view["ngay_trong"].astype(str),
            "Số cây": batch_view["so_cay"].apply(_format_int_commas),
            "Chi phí/cây": batch_view["cost_per_tree"].apply(_format_int_commas),
            "Tổng chi phí": batch_view["total_cost"].apply(_format_int_commas),
        })
        st.markdown("##### Theo từng đợt trồng")
        st.dataframe(batch_display, use_container_width=True, hide_index=True)

    if st.button("Đóng", key=f"close_lot_cost_dialog_{farm_name}_{lo_name}"):
        _clear_cost_query_params()
        st.rerun()


def _maybe_render_lot_cost_dialog(current_farm: str):
    farm_name = st.session_state.get("cost_dialog_farm") or _query_param_value("cost_farm")
    lo_name = st.session_state.get("cost_dialog_lot") or _query_param_value("cost_lot")
    if not farm_name or not lo_name:
        return False
    if current_farm not in ["Admin", "Phòng Kinh doanh", farm_name]:
        st.warning("Bạn không có quyền xem chi phí của farm này.")
        _clear_cost_query_params()
        return True
    _render_lot_cost_dialog(str(farm_name), str(lo_name))
    return True


def lookup_ribbon(farm_id, year, week):
    """Lookup ribbon color for (farm, year, week). Returns color_name or None."""
    res = supabase.table("ribbon_schedule").select("color_name") \
        .eq("farm_id", farm_id).eq("year", year).eq("week_number", week) \
        .eq("is_deleted", False).maybe_single().execute()
    return res.data["color_name"] if res.data else None

def get_or_create_ribbon(farm_id, year, week, color_name):
    """Lookup or create ribbon. Returns (color, None) or (None, error_msg)."""
    existing = lookup_ribbon(farm_id, year, week)
    if existing:
        if existing != color_name:
            return None, f"❌ Tuần {week}/{year} đã có màu dây '{existing}', không thể nhập '{color_name}'"
        return existing, None
    # Auto-create
    supabase.table("ribbon_schedule").insert({
        "farm_id": farm_id, "year": year,
        "week_number": week, "color_name": color_name
    }).execute()
    return color_name, None

def get_all_ribbons_for_farm(farm_id):
    """Get all ribbon colors for a farm (for selectbox options)."""
    res = supabase.table("ribbon_schedule").select("color_name, year, week_number") \
        .eq("farm_id", farm_id).eq("is_deleted", False) \
        .order("year", desc=True).order("week_number", desc=True).execute()
    return res.data or []

def build_color_selectbox(key_prefix, default_color=None):
    """Render dual selectbox (primary + secondary color). Returns standardized color string."""
    c1, c2 = st.columns(2)
    # Determine default indices
    primary_default = 0
    secondary_default = 0
    if default_color:
        parts = [p.strip() for p in default_color.split("-")]
        if len(parts) >= 1 and parts[0] in BASE_COLORS:
            primary_default = BASE_COLORS.index(parts[0])
        if len(parts) >= 2 and parts[1] in BASE_COLORS:
            # secondary options exclude primary, so index shifts
            pass  # handled below

    with c1:
        primary = st.selectbox("🎨 Màu chính", options=BASE_COLORS, index=primary_default, key=f"{key_prefix}_pri")
    secondary_opts = ["Không"] + [c for c in BASE_COLORS if c != primary]
    sec_default = 0
    if default_color and " - " in default_color:
        sec_part = default_color.split(" - ")[1].strip()
        if sec_part in secondary_opts:
            sec_default = secondary_opts.index(sec_part)
    with c2:
        secondary = st.selectbox("🎨 Màu phụ (nếu có)", options=secondary_opts, index=sec_default, key=f"{key_prefix}_sec")

    return primary if secondary == "Không" else f"{primary} - {secondary}"


def get_or_create_dim_lo(farm_name: str, lo_name: str, team_name: str = None):
    """Tìm dim_lo_id. Nếu lô chưa tồn tại → tự động tạo mới trong dim_lo."""
    existing = get_dim_lo_id(farm_name, lo_name)
    if existing:
        return existing
    
    # Lô chưa có → tạo mới
    # 1. Tìm farm_id
    farm_res = supabase.table("dim_farm").select("farm_id").eq("farm_name", farm_name).limit(1).execute()
    if not farm_res.data:
        st.error(f"❌ Không tìm thấy Farm '{farm_name}' trong hệ thống.")
        return None
    farm_id = farm_res.data[0]["farm_id"]
    
    # 2. Tìm doi_id (team)
    doi_id = None
    if team_name:
        doi_res = supabase.table("dim_doi").select("doi_id").eq("farm_id", farm_id).eq("doi_name", team_name).limit(1).execute()
        if doi_res.data:
            doi_id = doi_res.data[0]["doi_id"]
    
    # 3. Insert vào dim_lo
    new_lo = {
        "farm_id": farm_id,
        "lo_code": lo_name,
        "lo_name": lo_name,
        "lo_type": "Lô thực",
        "is_active": True
    }
    if doi_id:
        new_lo["doi_id"] = doi_id
    
    try:
        res = supabase.table("dim_lo").insert(new_lo).execute()
        if res.data:
            new_id = res.data[0]["lo_id"]
            # Clear cache so subsequent calls see the new lot
            get_dim_lo_id.clear()
            return new_id
    except Exception as e:
        st.error(f"❌ Lỗi khi tạo Lô mới trong danh mục: {e}")
    return None

def insert_to_db(table_name: str, data: dict) -> bool:
    try:
        tables_with_lo = [
            "stage_logs", "harvest_logs", "destruction_logs", "bsr_logs", 
            "size_measure_logs", "tree_inventory_logs", "soil_ph_logs", 
            "fusarium_logs", "seasons", "base_lots"
        ]
        if table_name in tables_with_lo:
            dim_id = None
            if "dim_lo_id" not in data:
                if "lot_id" in data and "farm" in data:
                    dim_id = get_dim_lo_id(data["farm"], data["lot_id"])
                elif "lo" in data and "farm" in data:
                    dim_id = get_dim_lo_id(data["farm"], data["lo"])
                if dim_id:
                    data["dim_lo_id"] = dim_id
            
            # Remove denormalized fields so insert to Supabase doesn't fail
            for col in ["farm", "team", "lot_id", "lo"]:
                data.pop(col, None)
            
            # GUARD: Block insert if dim_lo_id is still missing
            if not data.get("dim_lo_id"):
                st.error("❌ Lỗi hệ thống: Không tìm thấy Lô trong danh mục (dim_lo_id = null). Vui lòng kiểm tra tên Lô.")
                return False
            
            # Auto-resolve base_lot_id cho stage/harvest/destruction logs
            auto_resolve_tables = ["stage_logs", "harvest_logs", "destruction_logs"]
            if table_name in auto_resolve_tables and not data.get("base_lot_id"):
                date_col_map = {"stage_logs": "ngay_thuc_hien", "harvest_logs": "ngay_thu_hoach", "destruction_logs": "ngay_xuat_huy"}
                giai_doan_map = {"stage_logs": data.get("giai_doan", ""), "harvest_logs": "Thu hoạch", "destruction_logs": data.get("giai_doan", "")}
                action_date = data.get(date_col_map.get(table_name, ""))
                giai_doan = giai_doan_map.get(table_name, "")
                if action_date and data.get("dim_lo_id"):
                    resolve_qty = data.get("so_luong") if table_name in ["stage_logs", "harvest_logs"] else None
                    resolved_id = resolve_base_lot_id(data["dim_lo_id"], action_date, giai_doan, quantity=resolve_qty)
                    if resolved_id:
                        data["base_lot_id"] = resolved_id

        supabase.table(table_name).insert(data).execute()
        return True
    except Exception as e:
        if 'duplicate key value violates unique constraint' in str(e):
            if table_name == "base_lots":
                dim_lo_id = data.get("dim_lo_id")
                check_res = supabase.table("base_lots").select("is_deleted").eq("dim_lo_id", dim_lo_id).execute()
                if check_res.data and check_res.data[0].get("is_deleted") is True:
                    data["is_deleted"] = False
                    supabase.table("base_lots").update(data).eq("dim_lo_id", dim_lo_id).execute()
                    return True
            st.error(f"❌ Mã Lô đã tồn tại trong hệ thống. Vui lòng kiểm tra lại!")
        else:
            st.error(f"❌ Lỗi khi lưu vào {table_name}: {e}")
        return False

def get_current_season_used(bid: int, log_type: str, giai_doan: str = None, exclude_id: int = None) -> int:
    """Lấy tổng số lượng đã sử dụng trong vụ MỞ hiện tại của một base_lot."""
    season_res = supabase.table("seasons").select("ngay_bat_dau") \
        .eq("base_lot_id", bid).eq("is_deleted", False).is_("ngay_ket_thuc_thuc_te", "null") \
        .order("ngay_bat_dau", desc=True).limit(1).execute()
    if not season_res.data:
        return 0
    start_date = season_res.data[0]["ngay_bat_dau"]
    
    if log_type == "stage":
        q = supabase.table("stage_logs").select("id, so_luong").eq("base_lot_id", bid).eq("is_deleted", False).gte("ngay_thuc_hien", start_date)
        if giai_doan:
            q = q.eq("giai_doan", giai_doan)
    elif log_type == "harvest":
        q = supabase.table("harvest_logs").select("id, so_luong").eq("base_lot_id", bid).eq("is_deleted", False).gte("ngay_thu_hoach", start_date)
    elif log_type == "destruction":
        q = supabase.table("destruction_logs").select("id, so_luong").eq("base_lot_id", bid).eq("is_deleted", False).gte("ngay_xuat_huy", start_date)
        if giai_doan:
            q = q.eq("giai_doan", giai_doan)
    else:
        return 0
        
    res = q.execute()
    return sum(int(r["so_luong"]) for r in res.data if r["id"] != exclude_id) if res.data else 0

def _iso_year_week(value):
    dt = pd.to_datetime(value)
    iso = dt.isocalendar()
    return int(iso.year), int(iso.week)

def _shift_iso_week(year: int, week: int, week_offset: int):
    from datetime import date as _date, timedelta as _timedelta
    shifted = _date.fromisocalendar(int(year), int(week), 1) + _timedelta(weeks=int(week_offset))
    iso = shifted.isocalendar()
    return int(iso.year), int(iso.week)

def _report_clean_farm_name(value):
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()

def _report_farm_sort_key(farm_name):
    farm_name = _report_clean_farm_name(farm_name)
    m = re.search(r"(\d+)", farm_name)
    return (int(m.group(1)) if m else 9999, farm_name)

def _report_forecast_offsets_for_farm(farm_name):
    farm_name = _report_clean_farm_name(farm_name)
    if farm_name == "Farm 126":
        return [8]
    if farm_name == "Farm 157":
        return [9]
    return [8, 9]

def _report_expected_harvest_from_cut(cut_value):
    if not isinstance(cut_value, (int, float)) or cut_value <= 0:
        return 0
    return int(float(cut_value) * 0.97 + 0.5)

def get_ribbon_cut_week_candidates_for_harvest(farm_id, mau_day, harvest_date):
    """Trả về các tuần cắt bắp ứng với màu dây thu hoạch.

    Ưu tiên tuần cắt có dự báo +8/+9 trùng tuần thu hoạch thực tế.
    Fallback về mọi tuần có cùng màu dây để dữ liệu vẫn nhập được khi lệch lịch.
    """
    if not farm_id or not mau_day or not harvest_date:
        return []
    ribbon_res = supabase.table("ribbon_schedule").select("year, week_number") \
        .eq("farm_id", farm_id).eq("color_name", mau_day).eq("is_deleted", False).execute()
    candidates = [
        (int(r["year"]), int(r["week_number"]))
        for r in (ribbon_res.data or [])
        if r.get("year") is not None and r.get("week_number") is not None
    ]
    if not candidates:
        return []
    harvest_pair = _iso_year_week(harvest_date)
    exact = [
        pair for pair in candidates
        if _shift_iso_week(pair[0], pair[1], 7) == harvest_pair
        or _shift_iso_week(pair[0], pair[1], 8) == harvest_pair
    ]
    return exact or candidates

def get_cut_quantity_for_ribbon_candidates(base_lot_id, candidate_pairs):
    """Tổng Cắt bắp của một base_lot thuộc các tuần/màu dây nguồn đã chọn."""
    candidate_set = {(int(y), int(w)) for y, w in (candidate_pairs or [])}
    if not candidate_set:
        return 0
    res = supabase.table("stage_logs").select("so_luong, ngay_thuc_hien, tuan") \
        .eq("base_lot_id", base_lot_id).eq("giai_doan", "Cắt bắp").eq("is_deleted", False).execute()
    total = 0
    for row in (res.data or []):
        row_date = row.get("ngay_thuc_hien")
        if not row_date:
            continue
        row_year, row_week = _iso_year_week(row_date)
        if row.get("tuan") is not None:
            row_week = int(row["tuan"])
        if (row_year, row_week) in candidate_set:
            total += int(row.get("so_luong") or 0)
    return total

def get_available_capacity_for_lot(farm_name, lot_id_str, log_type, giai_doan=None, exclude_id=None):
    dim_id = get_dim_lo_id(farm_name, lot_id_str)
    if not dim_id: return 0, 0
    
    res_lot = supabase.table("base_lots").select("id, so_luong").eq("dim_lo_id", dim_id).eq("is_deleted", False).eq("loai_trong", "Trồng mới").execute()
    if not res_lot.data: return 0, 0
    
    # Chỉ lấy các base_lot đang có vụ MỞ
    active_bids = []
    for row in res_lot.data:
        s_res = supabase.table("seasons").select("id").eq("base_lot_id", row["id"]).eq("is_deleted", False).is_("ngay_ket_thuc_thuc_te", "null").execute()
        if s_res.data:
            active_bids.append(row)
            
    if not active_bids: return 0, 0

    total_planted = sum(
        max(0, int(row.get("so_luong", 0)) - get_current_season_destruction(row["id"], "Trước chích bắp"))
        for row in active_bids
    )

    if log_type == "stage" and giai_doan == "Cắt bắp":
        max_allowed = sum(get_current_season_used(row["id"], "stage", "Chích bắp") for row in active_bids)
    elif log_type == "harvest":
        max_allowed = sum(get_current_season_used(row["id"], "stage", "Cắt bắp") for row in active_bids)
    else:
        max_allowed = total_planted

    total_used = sum(get_current_season_used(row["id"], log_type, giai_doan, exclude_id) for row in active_bids)
        
    return max_allowed, total_used


def allocate_fifo_quantity(farm_name, lo_name, new_sl, log_type, target_date, action_type, giai_doan=None, mau_day=None):
    """
    Phân bổ số lượng theo FIFO (First-In-First-Out) theo ngày trồng.
    Đợt trồng cũ nhất được ưu tiên phân bổ trước, tràn thì chuyển sang đợt tiếp.
    Mỗi allocation kèm base_lot_id để ghi chính xác vào DB.
    """
    dim_id = get_dim_lo_id(farm_name, lo_name)
    if not dim_id:
        return False, f"❌ Lỗi: Không tìm thấy lô {lo_name} hoặc đã bị xóa.", []
        
    is_time_valid, msg_time = validate_timeline_logic(farm_name, lo_name, target_date, action_type)
    if not is_time_valid:
        return False, msg_time, []

    # ── Destruction: delegate to specialized FIFO ──
    if log_type == "destruction" and giai_doan:
        return allocate_destruction_fifo(dim_id, lo_name, int(new_sl), giai_doan, target_date, mau_day)

    # Lấy tất cả base_lots active của lô, sort FIFO theo ngày trồng (cũ nhất trước)
    res_bl = supabase.table("base_lots").select("id, so_luong, ngay_trong") \
        .eq("dim_lo_id", dim_id).eq("is_deleted", False) \
        .eq("loai_trong", "Trồng mới") \
        .order("ngay_trong", desc=False).execute()
    
    if not res_bl.data:
        return False, f"❌ Lô {lo_name} chưa có đợt trồng nào.", []

    # Lọc base_lots đang có vụ MỞ
    active_bids = []
    for batch in res_bl.data:
        s_res = supabase.table("seasons").select("id").eq("base_lot_id", batch["id"]).eq("is_deleted", False).is_("ngay_ket_thuc_thuc_te", "null").execute()
        if s_res.data:
            active_bids.append(batch)

    if not active_bids:
        return False, f"❌ Lô {lo_name} không có vụ (season) nào đang mở để nhận dữ liệu.", []

    quantity_left = int(new_sl)
    allocations = []

    for batch in active_bids:
        if quantity_left <= 0:
            break
        bid = batch["id"]
        planted = int(batch.get("so_luong", 0)) - get_current_season_destruction(bid, "Trước chích bắp")

        if log_type == "stage" and giai_doan == "Chích bắp":
            used = get_current_season_used(bid, "stage", "Chích bắp")
            cap = planted - used

        elif log_type == "stage" and giai_doan == "Cắt bắp":
            total_chich = get_current_season_used(bid, "stage", "Chích bắp")
            total_cat = get_current_season_used(bid, "stage", "Cắt bắp")
            cap = total_chich - total_cat

        elif log_type == "harvest":
            if mau_day:
                farm_id = get_farm_id_from_name(farm_name)
                candidate_pairs = get_ribbon_cut_week_candidates_for_harvest(farm_id, mau_day, target_date)
                total_cat = get_cut_quantity_for_ribbon_candidates(bid, candidate_pairs)
                har_q = supabase.table("harvest_logs").select("so_luong") \
                    .eq("base_lot_id", bid).eq("mau_day", mau_day).eq("is_deleted", False).execute()
                total_har = sum(int(r["so_luong"]) for r in (har_q.data or []))
            else:
                total_cat = get_current_season_used(bid, "stage", "Cắt bắp")
                total_har = get_current_season_used(bid, "harvest")
            cap = total_cat - total_har

        else:
            # Fallback: dùng tổng planted
            cap = planted

        if cap <= 0:
            continue

        alloc_qty = min(quantity_left, cap)
        allocations.append({
            "dim_lo_id": dim_id,
            "base_lot_id": bid,
            "so_luong": alloc_qty,
            "lot_id": lo_name
        })
        quantity_left -= alloc_qty

    if quantity_left > 0:
        total_cap = sum(a["so_luong"] for a in allocations)
        return False, f"❌ Yêu cầu {new_sl} nhưng tổng capacity của lô '{lo_name}' chỉ còn {total_cap}.", []

    return True, "", allocations


def allocate_destruction_fifo(dim_lo_id, lo_name, quantity, giai_doan, ngay_xuat_huy, mau_day=None):
    """
    Phân bổ xuất hủy theo 3 chiến lược FIFO dựa trên giai đoạn.
    - "Trước chích bắp": FIFO by ngay_trong, cap = planted - đã_chích - hủy_trước_chích
    - "Trước cắt bắp": Record-level FIFO across chích records
    - "Trước thu hoạch": Match mau_day → closest week → FIFO
    """
    # Lấy tất cả batches
    res_bl = supabase.table("base_lots").select("id, so_luong, ngay_trong") \
        .eq("dim_lo_id", dim_lo_id).eq("is_deleted", False) \
        .eq("loai_trong", "Trồng mới") \
        .order("ngay_trong", desc=False).execute()
    
    if not res_bl.data:
        return False, f"❌ Lô {lo_name} chưa có đợt trồng nào.", []

    # ━━━━ Strategy 1: Trước chích bắp ━━━━
    if giai_doan == "Trước chích bắp":
        quantity_left = quantity
        allocations = []
        for batch in res_bl.data:
            if quantity_left <= 0:
                break
            bid = batch["id"]
            # Cap = planted - đã_chích - hủy_trước_chích_vụ_HT
            planted = int(batch.get("so_luong", 0))
            chich_res = supabase.table("stage_logs").select("so_luong") \
                .eq("base_lot_id", bid).eq("giai_doan", "Chích bắp").eq("is_deleted", False).execute()
            da_chich = sum(int(r["so_luong"]) for r in chich_res.data) if chich_res.data else 0
            huy_truoc_chich = get_current_season_destruction(bid, "Trước chích bắp")
            cap = max(0, planted - da_chich - huy_truoc_chich)
            if cap <= 0:
                continue
            alloc_qty = min(quantity_left, cap)
            allocations.append({"dim_lo_id": dim_lo_id, "base_lot_id": bid, "so_luong": alloc_qty, "lot_id": lo_name})
            quantity_left -= alloc_qty
        if quantity_left > 0:
            total_cap = sum(a["so_luong"] for a in allocations)
            return False, f"❌ Số lượng xuất hủy vượt quá số cây khả dụng (còn {total_cap}).", []
        return True, "", allocations

    # ━━━━ Strategy 2: Trước cắt bắp ━━━━
    elif giai_doan == "Trước cắt bắp":
        # Record-level FIFO: gộp chích records từ mọi batch, sort by ngay_thuc_hien, tiebreak by ngay_trong
        batch_map = {b["id"]: b for b in res_bl.data}
        all_chich_records = []
        for batch in res_bl.data:
            bid = batch["id"]
            chich_res = supabase.table("stage_logs").select("id, so_luong, ngay_thuc_hien") \
                .eq("base_lot_id", bid).eq("giai_doan", "Chích bắp").eq("is_deleted", False).execute()
            for r in (chich_res.data or []):
                all_chich_records.append({
                    "bid": bid,
                    "ngay_trong": batch["ngay_trong"],
                    "ngay_thuc_hien": r["ngay_thuc_hien"],
                    "so_luong": int(r["so_luong"]),
                })
        if not all_chich_records:
            return False, "❌ Không có đợt nào đã chích bắp để phân bổ.", []
        
        # Sort: ngay_thuc_hien ASC, tiebreak ngay_trong ASC
        all_chich_records.sort(key=lambda x: (x["ngay_thuc_hien"], x["ngay_trong"]))
        
        # Calculate existing destruction "Trước cắt bắp" per batch
        existing_dest = {}
        for batch in res_bl.data:
            bid = batch["id"]
            dest_res = supabase.table("destruction_logs").select("so_luong") \
                .eq("base_lot_id", bid).eq("giai_doan", "Trước cắt bắp").eq("is_deleted", False).execute()
            existing_dest[bid] = sum(int(r["so_luong"]) for r in dest_res.data) if dest_res.data else 0
        
        # Calculate remaining capacity per batch: đã_chích - existing_dest_trước_cắt
        batch_chich_total = {}
        for r in all_chich_records:
            batch_chich_total[r["bid"]] = batch_chich_total.get(r["bid"], 0) + r["so_luong"]
        batch_remaining = {}
        for bid, total in batch_chich_total.items():
            batch_remaining[bid] = max(0, total - existing_dest.get(bid, 0))
        
        quantity_left = quantity
        alloc_by_bid = {}
        for rec in all_chich_records:
            if quantity_left <= 0:
                break
            bid = rec["bid"]
            if batch_remaining.get(bid, 0) <= 0:
                continue
            alloc_qty = min(quantity_left, rec["so_luong"], batch_remaining[bid])
            alloc_by_bid[bid] = alloc_by_bid.get(bid, 0) + alloc_qty
            batch_remaining[bid] -= alloc_qty
            quantity_left -= alloc_qty
        
        if quantity_left > 0:
            total_cap = sum(alloc_by_bid.values())
            return False, f"❌ Số lượng xuất hủy vượt quá số cây đã chích (còn {total_cap}).", []
        
        allocations = [{"dim_lo_id": dim_lo_id, "base_lot_id": bid, "so_luong": qty, "lot_id": lo_name}
                       for bid, qty in alloc_by_bid.items()]
        return True, "", allocations

    # ━━━━ Strategy 3: Trước thu hoạch ━━━━
    elif giai_doan == "Trước thu hoạch":
        if not mau_day:
            return False, "❌ Vui lòng nhập màu dây cho giai đoạn Trước thu hoạch.", []
        
        # Get farm_id from dim_lo_id via dim_lo table
        lo_res = supabase.table("dim_lo").select("farm_id").eq("lo_id", dim_lo_id).limit(1).execute()
        if not lo_res.data:
            return False, "❌ Không tìm thấy thông tin farm cho lô này.", []
        farm_id = lo_res.data[0]["farm_id"]
        
        # Find all weeks with this color from ribbon_schedule
        ribbon_res = supabase.table("ribbon_schedule").select("week_number, year") \
            .eq("farm_id", farm_id).eq("color_name", mau_day).eq("is_deleted", False).execute()
        if not ribbon_res.data:
            return False, f"❌ Không tìm thấy tuần nào với màu dây '{mau_day}' trong ribbon_schedule.", []
        
        target_weeks = [(r["year"], r["week_number"]) for r in ribbon_res.data]
        
        # Find cắt bắp records matching those weeks
        all_cat_records = []
        for (yr, wk) in target_weeks:
            cat_res = supabase.table("stage_logs").select("id, base_lot_id, ngay_thuc_hien, so_luong, tuan") \
                .eq("dim_lo_id", dim_lo_id).eq("giai_doan", "Cắt bắp").eq("tuan", wk) \
                .eq("is_deleted", False).execute()
            if cat_res.data:
                all_cat_records.extend(cat_res.data)
        
        if not all_cat_records:
            return False, f"❌ Không tìm thấy record cắt bắp với màu dây '{mau_day}'.", []
        
        # Tính khoảng cách ngày, chọn tuần gần nhất
        ngay_huy = pd.to_datetime(ngay_xuat_huy)
        for r in all_cat_records:
            r["distance"] = abs((pd.to_datetime(r["ngay_thuc_hien"]) - ngay_huy).days)
        
        min_dist = min(r["distance"] for r in all_cat_records)
        # Tiebreaker equidistant: chọn tuần CŨ hơn
        eligible = [r for r in all_cat_records if r["distance"] == min_dist]
        if not eligible:
            eligible = all_cat_records  # fallback
        
        # Lấy ngay_trong cho tiebreaker
        batch_map = {b["id"]: b["ngay_trong"] for b in res_bl.data}
        
        # Sort eligible: ngay_thuc_hien ASC, ngay_trong ASC
        eligible.sort(key=lambda r: (r["ngay_thuc_hien"], batch_map.get(r["base_lot_id"], "9999-12-31")))
        
        # Trừ existing destruction + harvest cùng batch
        batch_cat_qty = {}
        for r in eligible:
            bid = r["base_lot_id"]
            batch_cat_qty[bid] = batch_cat_qty.get(bid, 0) + int(r["so_luong"])
        
        batch_remaining = {}
        for bid, cat_total in batch_cat_qty.items():
            # Existing destruction "Trước thu hoạch" cùng batch
            ex_dest = supabase.table("destruction_logs").select("so_luong") \
                .eq("base_lot_id", bid).eq("giai_doan", "Trước thu hoạch") \
                .eq("is_deleted", False).execute()
            ex_dest_total = sum(int(r["so_luong"]) for r in ex_dest.data) if ex_dest.data else 0
            # Existing harvest cùng batch
            ex_har = supabase.table("harvest_logs").select("so_luong") \
                .eq("base_lot_id", bid).eq("is_deleted", False).execute()
            ex_har_total = sum(int(r["so_luong"]) for r in ex_har.data) if ex_har.data else 0
            batch_remaining[bid] = max(0, cat_total - ex_dest_total - ex_har_total)
        
        total_available = sum(batch_remaining.values())
        if total_available <= 0:
            return False, f"❌ Không còn cây chờ thu hoạch với màu dây '{mau_day}'.", []
        if quantity > total_available:
            return False, f"❌ Số lượng xuất hủy ({quantity}) vượt quá số cây cắt bắp cùng màu dây ({total_available}).", []
        
        quantity_left = quantity
        alloc_by_bid = {}
        for r in eligible:
            if quantity_left <= 0:
                break
            bid = r["base_lot_id"]
            if batch_remaining.get(bid, 0) <= 0:
                continue
            alloc_qty = min(quantity_left, int(r["so_luong"]), batch_remaining[bid])
            alloc_by_bid[bid] = alloc_by_bid.get(bid, 0) + alloc_qty
            batch_remaining[bid] -= alloc_qty
            quantity_left -= alloc_qty
        
        allocations = [{"dim_lo_id": dim_lo_id, "base_lot_id": bid, "so_luong": qty, "lot_id": lo_name}
                       for bid, qty in alloc_by_bid.items()]
        return True, "", allocations

    # Fallback: giai_doan không xác định
    else:
        return False, f"❌ Giai đoạn xuất hủy không hợp lệ: '{giai_doan}'.", []


def check_quantity_limit(farm_name, lot_id_str, new_sl, log_type, giai_doan=None, exclude_id=None):
    max_allowed, total_used = get_available_capacity_for_lot(farm_name, lot_id_str, log_type, giai_doan, exclude_id)
    unit = "buồng" if log_type == "harvest" else "cây"

    if log_type == "destruction":
        if int(new_sl) > (max_allowed - total_used):
            return False, f"❌ Mã lứa này chỉ còn {max_allowed - total_used} cây sống, không thể xuất hủy {new_sl} cây."
        return True, ""

    if total_used + int(new_sl) > max_allowed:
        remain = max_allowed - total_used
        return False, f"❌ Bạn đã nhập {new_sl} {unit}, nhưng lứa này chỉ có {remain} {unit} cho phép."
    return True, ""

def validate_timeline_logic(farm_name, lot_id_str, target_date, action_type):
    dim_id = get_dim_lo_id(farm_name, lot_id_str)
    if not dim_id: return False, "Lỗi không tìm thấy lô."
    
    target_dt = pd.to_datetime(target_date).tz_localize(None)
    
    if action_type == "Chích bắp":
        res = supabase.table("base_lots").select("ngay_trong").eq("dim_lo_id", dim_id).eq("is_deleted", False).execute()
        if res.data and any(r.get("ngay_trong") for r in res.data):
            valid_dates = [pd.to_datetime(r["ngay_trong"]).tz_localize(None) for r in res.data if r.get("ngay_trong")]
            if valid_dates:
                earliest_trong = min(valid_dates)
                if target_dt < earliest_trong:
                    return False, f"❌ Ngày Chích bắp ({target_dt.date()}) không thể trước Ngày Trồng ({earliest_trong.date()})."
                
    elif action_type == "Cắt bắp":
        res = supabase.table("stage_logs").select("ngay_thuc_hien").eq("dim_lo_id", dim_id).eq("giai_doan", "Chích bắp").eq("is_deleted", False).execute()
        if res.data:
            earliest_cb = min([pd.to_datetime(r["ngay_thuc_hien"]).tz_localize(None) for r in res.data])
            if target_dt < earliest_cb:
                return False, f"❌ Ngày Cắt bắp ({target_dt.date()}) không thể trước Ngày Chích bắp sớm nhất ({earliest_cb.date()})."
        else:
            return False, "❌ Lô này chưa được ghi nhận Chích bắp, không thể Cắt bắp!"
            
    elif action_type == "Thu hoạch":
        res = supabase.table("stage_logs").select("ngay_thuc_hien").eq("dim_lo_id", dim_id).eq("giai_doan", "Cắt bắp").eq("is_deleted", False).execute()
        if res.data:
            earliest_cut = min([pd.to_datetime(r["ngay_thuc_hien"]).tz_localize(None) for r in res.data])
            if target_dt < earliest_cut:
                return False, f"❌ Ngày Thu hoạch ({target_dt.date()}) không thể trước Ngày Cắt bắp sớm nhất ({earliest_cut.date()})."
        else:
            return False, "❌ Lô này chưa được ghi nhận Cắt bắp, không thể Thu hoạch!"
            
    return True, ""


@dialog_decorator("⚠️ Xác nhận")
def confirm_action_dialog(action, table_name, rec_id_or_none, data_dict, success_msg):
    st.warning("Vui lòng kiểm tra kỹ trước khi thực hiện!")
    if not data_dict and action == "DELETE":
        st.error("Xóa dữ liệu vĩnh viễn?")
    
    # Guard flag chống double-submit
    processing_flag = f"_processing_confirm_{action}_{table_name}"
    is_processing = st.session_state.get(processing_flag, False)

    col1, col2 = st.columns(2)
    with col1:
        if st.button("✅ Tôi đã kiểm tra kỹ", use_container_width=True, disabled=is_processing):
            if not st.session_state.get(processing_flag, False):
                st.session_state[processing_flag] = True
                try:
                    success = False
                    if action == "INSERT":
                        success = insert_to_db(table_name, data_dict)
                    elif action == "INSERT_FIFO":
                        db_data = data_dict["base_data"]
                        allocations = data_dict["allocations"]
                        success = True
                        for alloc in allocations:
                            row_data = db_data.copy()
                            row_data["lot_id"] = alloc["lot_id"]
                            row_data["so_luong"] = alloc["so_luong"]
                            try:
                                supabase.table(table_name).insert(row_data).execute()
                            except Exception as e:
                                st.error(f"❌ Lỗi ghi hệ thống FIFO nhánh {alloc['lot_id']}: {e}")
                                success = False
                    elif action == "INSERT_BASE":
                        db_data, season_data = data_dict
                        success = insert_to_db("base_lots", db_data)
                        if success:
                            try:
                                supabase.table("seasons").insert(season_data).execute()
                            except Exception as e:
                                st.error(f"❌ Lỗi ghi vụ: {e}")
                    elif action == "UPDATE":
                        try:
                            supabase.table(table_name).update(data_dict).eq("id", rec_id_or_none).execute()
                            success = True
                        except Exception as e:
                            st.error(f"❌ Lỗi cập nhật: {e}")
                    elif action == "DELETE":
                        try:
                            supabase.table(table_name).update({"is_deleted": True}).eq("id", rec_id_or_none).execute()
                            success = True
                        except Exception as e:
                            st.error(f"❌ Lỗi xóa: {e}")
                    
                    if success:
                        st.cache_data.clear()
                        st.session_state["toast"] = success_msg
                        st.rerun()
                finally:
                    st.session_state[processing_flag] = False
    with col2:
        if st.button("❌ Quay lại", use_container_width=True):
            st.rerun()

def get_editing_row(table_name, df):
    idx_list = st.session_state.get(f"sel_{table_name}", {}).get("selection", {}).get("rows", [])
    if idx_list and len(idx_list) > 0 and idx_list[0] < len(df):
        row = df.iloc[idx_list[0]].to_dict()
        created_at = pd.to_datetime(row["created_at"], utc=True)
        is_within_48h = created_at > (pd.Timestamp.utcnow() - pd.Timedelta(hours=48))
        return row, is_within_48h
    return None, False

def render_team_dataframe(table_name, df, display_cols):
    if df.empty:
        st.info("Chưa có dữ liệu.")
        return
    st.dataframe(
        df[display_cols],
        use_container_width=True,
        hide_index=True,
        on_select="rerun",
        selection_mode="single-row",
        key=f"sel_{table_name}"
    )

def render_queue_ui(queue_key, display_cols, process_func):
    """Render the queue and action buttons for bulk data entry."""
    queue = st.session_state[queue_key]
    if not queue:
        return
        
    st.markdown("#### 📋 Danh sách chờ duyệt")
    df_queue = pd.DataFrame(queue)
    df_queue.insert(0, "Xóa", False) # Thêm cột checkbox
    
    edited_df = st.data_editor(
        df_queue[["Xóa"] + display_cols],
        hide_index=True,
        use_container_width=True,
        disabled=display_cols,
        key=f"editor_{queue_key}"
    )
    
    to_delete_idxs = edited_df.index[edited_df["Xóa"] == True].tolist()
    
    # Guard flag chống double-submit
    processing_flag = f"_processing_{queue_key}"
    is_processing = st.session_state.get(processing_flag, False)
    
    col_q1, col_q2 = st.columns(2)
    with col_q1:
        if st.button("🚀 Lưu toàn bộ lên Hệ thống", type="primary", use_container_width=True,
                      key=f"btn_sb_{queue_key}", disabled=is_processing):
            if not st.session_state.get(processing_flag, False):
                st.session_state[processing_flag] = True
                try:
                    process_func()
                finally:
                    st.session_state[processing_flag] = False
    with col_q2:
        if to_delete_idxs:
            if st.button("🗑️ Xóa dòng đã chọn", use_container_width=True, key=f"btn_del_sel_{queue_key}"):
                st.session_state[queue_key] = [item for i, item in enumerate(queue) if i not in to_delete_idxs]
                st.rerun()
        else:
            if st.button("🗑️ Xóa toàn bộ danh sách", use_container_width=True, key=f"btn_del_all_{queue_key}"):
                st.session_state[queue_key] = []
                st.rerun()

# =====================================================
# CÁC DIALOG CHỈNH SỬA
# =====================================================
@dialog_decorator("✏️ Chỉnh sửa Lô Trồng")
def edit_base_lot_dialog(editing_row):
    def_lo = editing_row["lo"]
    def_ngay = pd.to_datetime(editing_row["ngay_trong"]).date()
    def_sl = int(editing_row["so_luong"])

    with st.container(border=True):
        col_a, col_b = st.columns(2)
        with col_a:
            st.text_input("🏷️ Tên Lô", value=def_lo, key="dlg_lo_base", disabled=True)
            st.info("⚠️ Không thể sửa thông tin Lô/Loại trồng. Nếu sai hãy Xóa và tạo mới Lô.")
        with col_b:
            col_b1, col_b2 = st.columns([2, 1])
            with col_b1:
                ngay_trong = st.date_input("📆 Ngày trồng", value=def_ngay, key="dlg_dt_base")
            with col_b2:
                st.text_input("📍 Tuần", value=str(ngay_trong.isocalendar()[1]), disabled=True, key=f"dlg_w_base_{ngay_trong}")
            so_luong = st.number_input("🔢 Số lượng", min_value=0, step=100, value=def_sl, key="dlg_sl_base")

        if st.button("✅ Cập nhật", key="btn_edit_base", use_container_width=True, type="primary"):
            if so_luong <= 0: st.error("❌ Cần nhập số lượng.")
            else:
                data = {
                    "ngay_trong": ngay_trong.isoformat(), "so_luong": so_luong,
                    "tuan": ngay_trong.isocalendar()[1]
                }
                supabase.table("base_lots").update(data).eq("id", editing_row["id"]).execute()
                
                # Cập nhật thêm cho season đang chạy nếu cần (tùy chọn)
                supabase.table("seasons").update({"ngay_bat_dau": ngay_trong.isoformat()}).eq("lo", def_lo).eq("vu", "F0").execute()
                
                st.session_state["toast"] = f"✅ Cập nhật {def_lo} thành công!"
                st.rerun()

@dialog_decorator("✏️ Chỉnh sửa Tiến độ")
def edit_stage_log_dialog(editing_row, available_lots, c_team):
    gd_ops = ["Chích bắp"] if c_team == "Đội BVTV" else ["Cắt bắp"]
    def_lot = available_lots.index(editing_row["lot_id"]) if editing_row["lot_id"] in available_lots else 0
    def_gd = gd_ops.index(editing_row["giai_doan"]) if editing_row["giai_doan"] in gd_ops else 0
    def_ngay = pd.to_datetime(editing_row["ngay_thuc_hien"]).date()
    def_sl = int(editing_row["so_luong"])

    with st.container(border=True):
        col_a, col_b = st.columns(2)
        with col_a:
            st.text_input("🏷️ Lứa (Mã hệ thống)", value=editing_row["lot_id"], disabled=True, key="dlg_lot_stg")
            lot_id = editing_row["lot_id"]
            giai_doan = st.radio("📌 Giai đoạn", options=gd_ops, index=def_gd, horizontal=True, key="dlg_gd_stg")
        with col_b:
            col_b1, col_b2 = st.columns([2, 1])
            with col_b1:
                ngay_th = st.date_input("📆 Ngày thực hiện", value=def_ngay, key="dlg_dt_stg")
            with col_b2:
                st.text_input("📍 Tuần", value=str(ngay_th.isocalendar()[1]), disabled=True, key=f"dlg_w_stg_{ngay_th}")
            sl = st.number_input("🔢 Số lượng cây", min_value=0, step=100, value=def_sl, key="dlg_sl_stg")

        # Ribbon color (only for Cắt bắp)
        mau_day_color = None
        if giai_doan != "Chích bắp":
            farm_id = get_farm_id_from_name(editing_row["farm"])
            iso = ngay_th.isocalendar()
            existing_ribbon = lookup_ribbon(farm_id, iso[0], iso[1]) if farm_id else None
            if existing_ribbon:
                st.info(f"🎨 Tuần {iso[1]}: **{existing_ribbon}**")
                mau_day_color = existing_ribbon
            else:
                mau_day_color = build_color_selectbox("dlg_stg")
        
        if st.button("✅ Cập nhật", key="btn_edit_stg", use_container_width=True, type="primary"):
            if sl <= 0: st.error("❌ Cần nhập số lượng.")
            elif giai_doan != "Chích bắp" and not mau_day_color: st.error("❌ Phải chọn màu dây cho Cắt bắp.")
            else:
                # Auto-create/validate ribbon
                if giai_doan != "Chích bắp" and mau_day_color:
                    farm_id = get_farm_id_from_name(editing_row["farm"])
                    iso = ngay_th.isocalendar()
                    _, err = get_or_create_ribbon(farm_id, iso[0], iso[1], mau_day_color)
                    if err:
                        st.error(err)
                        st.stop()
                is_valid, msg = check_quantity_limit(editing_row["farm"], lot_id, sl, "stage", giai_doan=giai_doan, exclude_id=editing_row["id"])
                if not is_valid: st.error(msg)
                else:
                    resolved_blid = resolve_base_lot_id(editing_row["dim_lo_id"], ngay_th.isoformat(), giai_doan, quantity=sl, exclude_id=editing_row["id"])
                    if not resolved_blid:
                        st.error("❌ Không còn đợt trồng nào đủ capacity cho số lượng này. Vui lòng giảm số lượng hoặc nhập lại qua form thêm mới để hệ thống tự tách FIFO.")
                        st.stop()
                    data = {
                        "giai_doan": giai_doan, 
                        "ngay_thuc_hien": ngay_th.isoformat(), "so_luong": sl,
                        "tuan": ngay_th.isocalendar()[1],
                        "base_lot_id": resolved_blid
                    }
                    supabase.table("stage_logs").update(data).eq("id", editing_row["id"]).execute()
                    st.session_state["toast"] = f"✅ Cập nhật tiến độ: {lot_id}!"
                    st.rerun()


@dialog_decorator("✏️ Chỉnh sửa Báo cáo Xuất Hủy")
def edit_destruction_log_dialog(editing_row, available_lots):
    gd_ops = DESTRUCTION_STAGE_OPTIONS
    def_lot = available_lots.index(editing_row["lot_id"]) if editing_row["lot_id"] in available_lots else 0
    def_gd = gd_ops.index(editing_row["giai_doan"]) if editing_row["giai_doan"] in gd_ops else 0
    def_ly_do = str(editing_row["ly_do"])
    def_ngay = pd.to_datetime(editing_row["ngay_xuat_huy"]).date()
    def_sl = int(editing_row["so_luong"])
    
    with st.container(border=True):
        col_a, col_b = st.columns(2)
        with col_a:
            st.text_input("🏷️ Lứa (Mã hệ thống)", value=editing_row["lot_id"], disabled=True, key="dlg_lot_des")
            lot_id = editing_row["lot_id"]
            gxh = st.selectbox("⏱️ Giai đoạn", options=gd_ops, index=def_gd, key="dlg_gxh_des")
            
            # Ribbon color — only visible for "Trước thu hoạch"
            mau_day_color = None
            if gxh == "Trước thu hoạch":
                farm_id = get_farm_id_from_name(editing_row["farm"])
                if farm_id:
                    ribbons = get_all_ribbons_for_farm(farm_id)
                    ribbon_opts = list({r["color_name"] for r in ribbons})
                    if ribbon_opts:
                        mau_day_color = st.selectbox("🎨 Màu dây", options=sorted(ribbon_opts), key="dlg_mau_des_sel")
                    else:
                        st.warning("⚠️ Farm chưa có màu dây nào trong ribbon_schedule.")
            
            predefined_reasons = ["Bệnh", "Đổ Ngã", "Khác"]
            matched_reason = "Khác"
            if def_ly_do in ["Bệnh", "Đổ Ngã"]:
                matched_reason = def_ly_do
            
            if hasattr(st, "pills"):
                selected_reason = st.pills("📝 Nhóm lý do", options=predefined_reasons, default=matched_reason, key="dlg_des_reason_group")
                if not selected_reason:
                    selected_reason = "Khác"
            else:
                selected_reason = st.radio("📝 Nhóm lý do", options=predefined_reasons, index=predefined_reasons.index(matched_reason), horizontal=True, key="dlg_des_reason_group")
            
            if selected_reason == "Khác":
                ly_do = st.text_area("📝 Chi tiết lý do", height=80, value=def_ly_do if matched_reason == "Khác" else "", key="dlg_lydo_des")
            else:
                ly_do = selected_reason
                
        with col_b:
            col_b1, col_b2 = st.columns([2, 1])
            with col_b1:
                ngay = st.date_input("📆 Ngày xuất hủy", value=def_ngay, key="dlg_dt_des")
            with col_b2:
                st.text_input("📍 Tuần", value=str(ngay.isocalendar()[1]), disabled=True, key=f"dlg_w_des_{ngay}")
            sl = st.number_input("🔢 Số lượng xuất hủy", min_value=0, step=10, value=def_sl, key="dlg_sl_des")

        if st.button("✅ Cập nhật", key="btn_edit_des", use_container_width=True, type="primary"):
            if sl <= 0: st.error("❌ Cần nhập số lượng.")
            elif not ly_do.strip(): st.error("❌ Cần ghi rõ lý do chi tiết.")
            else:
                is_valid, msg = check_quantity_limit(editing_row["farm"], lot_id, sl, "destruction", exclude_id=editing_row["id"])
                if not is_valid: st.error(msg)
                else:
                    resolved_blid = resolve_base_lot_id(editing_row["dim_lo_id"], ngay.isoformat(), gxh)
                    data = {"ngay_xuat_huy": ngay.isoformat(), "giai_doan": gxh, "ly_do": ly_do.strip(), "so_luong": sl, "tuan": ngay.isocalendar()[1], "base_lot_id": resolved_blid}
                    supabase.table("destruction_logs").update(data).eq("id", editing_row["id"]).execute()
                    st.session_state["toast"] = "✅ Đã cập nhật!"
                    st.rerun()


@dialog_decorator("✏️ Chỉnh sửa Nhật ký Thu Hoạch")
def edit_harvest_log_dialog(editing_row, available_lots):
    def_lot = available_lots.index(editing_row["lot_id"]) if editing_row["lot_id"] in available_lots else 0
    
    hinh_thuc_opts = ["Bằng xe cày", "Bằng ròng rọc"]
    def_hinh_thuc = hinh_thuc_opts.index(editing_row.get("hinh_thuc_thu_hoach", "")) if editing_row.get("hinh_thuc_thu_hoach") in hinh_thuc_opts else 0
    
    def_ngay = pd.to_datetime(editing_row["ngay_thu_hoach"]).date()
    def_sl = int(editing_row["so_luong"])

    with st.container(border=True):
        col_a, col_b = st.columns(2)
        with col_a:
            st.text_input("🏷️ Lứa (Mã hệ thống)", value=editing_row["lot_id"], disabled=True, key="dlg_lot_har")
            lot_id = editing_row["lot_id"]
            # Ribbon color from schedule
            farm_id = get_farm_id_from_name(editing_row["farm"])
            ribbons = get_all_ribbons_for_farm(farm_id) if farm_id else []
            ribbon_opts = sorted({r["color_name"] for r in ribbons})
            if ribbon_opts:
                def_mau = str(editing_row.get("mau_day", ""))
                mau_idx = ribbon_opts.index(def_mau) if def_mau in ribbon_opts else 0
                mau_day_color = st.selectbox("🎨 Màu dây", options=ribbon_opts, index=mau_idx, key="dlg_mau_har_sel")
            else:
                st.warning("⚠️ Farm chưa có màu dây trong ribbon_schedule.")
                mau_day_color = None
            hinh_thuc_thu_hoach = st.selectbox("🚜 Hình thức thu hoạch", options=hinh_thuc_opts, index=def_hinh_thuc, key="dlg_hinh_thuc_har")
        with col_b:
            col_b1, col_b2 = st.columns([2, 1])
            with col_b1:
                ngay = st.date_input("📆 Ngày thu hoạch", value=def_ngay, key="dlg_dt_har")
            with col_b2:
                st.text_input("📍 Tuần", value=str(ngay.isocalendar()[1]), disabled=True, key=f"dlg_w_har_{ngay}")
            sl = st.number_input("🍌 Số lượng buồng", min_value=0, step=50, value=def_sl, key="dlg_sl_har")

        if st.button("✅ Cập nhật", key="btn_edit_har", use_container_width=True, type="primary"):
            if not mau_day_color: st.error("❌ Cần chọn Màu dây.")
            elif sl <= 0: st.error("❌ Số lượng buồng phải > 0")
            else:
                is_valid, msg = check_quantity_limit(editing_row["farm"], lot_id, sl, "harvest", exclude_id=editing_row["id"])
                if not is_valid: st.error(msg)
                else:
                    resolved_blid = resolve_base_lot_id(editing_row["dim_lo_id"], ngay.isoformat(), "Thu hoạch", quantity=sl, exclude_id=editing_row["id"])
                    if not resolved_blid:
                        st.error("❌ Không còn đợt trồng nào đủ capacity để ghi nhận thu hoạch này. Vui lòng giảm số lượng hoặc nhập lại qua form thêm mới để hệ thống tự tách FIFO.")
                        st.stop()
                    candidate_pairs = get_ribbon_cut_week_candidates_for_harvest(farm_id, mau_day_color, ngay.isoformat())
                    color_cap = get_cut_quantity_for_ribbon_candidates(resolved_blid, candidate_pairs)
                    har_same_color = supabase.table("harvest_logs").select("id, so_luong") \
                        .eq("base_lot_id", resolved_blid).eq("mau_day", mau_day_color).eq("is_deleted", False).execute()
                    used_same_color = sum(
                        int(r["so_luong"]) for r in (har_same_color.data or [])
                        if r.get("id") != editing_row["id"]
                    )
                    if sl > max(0, color_cap - used_same_color):
                        st.error(f"❌ Màu dây `{mau_day_color}` chỉ còn {max(0, color_cap - used_same_color)} buồng có thể thu hoạch.")
                        st.stop()
                    data = {
                        "ngay_thu_hoach": ngay.isoformat(), "so_luong": sl, 
                        "hinh_thuc_thu_hoach": hinh_thuc_thu_hoach, "tuan": ngay.isocalendar()[1],
                        "base_lot_id": resolved_blid, "mau_day": mau_day_color
                    }
                    supabase.table("harvest_logs").update(data).eq("id", editing_row["id"]).execute()
                    st.session_state["toast"] = f"✅ Lưu thu hoạch {lot_id} thành công!"
                    st.rerun()


@dialog_decorator("✏️ Chỉnh sửa BSR")
def edit_bsr_log_dialog(editing_row, available_lots):
    def_lot = available_lots.index(editing_row["lot_id"]) if editing_row["lot_id"] in available_lots else 0
    def_ngay = pd.to_datetime(editing_row["ngay_nhap"]).date()
    def_bsr = float(editing_row["bsr"])
    
    with st.container(border=True):
        col_a, col_b = st.columns(2)
        with col_a:
            st.text_input("🏷️ Lứa (Mã hệ thống)", value=editing_row["lot_id"], disabled=True, key="dlg_lot_bsr")
            lot_id = editing_row["lot_id"]
        with col_b:
            col_b1, col_b2 = st.columns([2, 1])
            with col_b1:
                ngay = st.date_input("📆 Ngày đóng gói", value=def_ngay, key="dlg_dt_bsr")
            with col_b2:
                st.text_input("📍 Tuần", value=str(ngay.isocalendar()[1]), disabled=True, key=f"dlg_w_bsr_{ngay}")
            bsr_val = st.number_input("📐 Tỷ lệ BSR", min_value=0.0, step=0.1, value=def_bsr, format="%.2f", key="dlg_v_bsr")

        if st.button("✅ Cập nhật", key="btn_edit_bsr", use_container_width=True, type="primary"):
            if bsr_val <= 0: st.error("❌ Tỷ lệ BSR phải > 0")
            else:
                data = {"ngay_nhap": ngay.isoformat(), "bsr": bsr_val, "tuan": ngay.isocalendar()[1]}
                supabase.table("bsr_logs").update(data).eq("id", editing_row["id"]).execute()
                st.session_state["toast"] = f"✅ Lưu BSR lô {lot_id} thành công!"
                st.rerun()

@dialog_decorator("✏️ Chỉnh sửa Đo Size")
def edit_size_measure_dialog(editing_row, available_lots):
    def_lot = available_lots.index(editing_row["lot_id"]) if editing_row["lot_id"] in available_lots else 0
    def_mau = str(editing_row.get("mau_day", ""))
    def_lan_do = editing_row["lan_do"]
    def_ngay = pd.to_datetime(editing_row["ngay_do"]).date()
    def_sl = int(editing_row["so_luong_mau"])
    def_hkt = str(editing_row.get("hang_kiem_tra", ""))
    def_cal = float(editing_row.get("size_cal", 0.0))
    
    with st.container(border=True):
        col_a, col_b = st.columns(2)
        with col_a:
            st.text_input("🏷️ Lứa (Mã hệ thống)", value=editing_row["lot_id"], disabled=True, key="dlg_sm_lot")
            lot_id = editing_row["lot_id"]
            mau_day_color = build_color_selectbox("dlg_sm", default_color=def_mau)
            lan_do = st.radio("📏 Lần đo", options=[1, 2], index=def_lan_do-1, horizontal=True, key="dlg_sm_lando")
        with col_b:
            col_b1, col_b2 = st.columns([2, 1])
            with col_b1:
                ngay = st.date_input("📆 Ngày đo", value=def_ngay, key="dlg_sm_ngay")
            with col_b2:
                st.text_input("📍 Tuần", value=str(ngay.isocalendar()[1]), disabled=True, key=f"dlg_w_sm_{ngay}")
            col_b3, col_b4 = st.columns(2)
            with col_b3:
                hang_kiem_tra = st.text_input("📏 Hàng kiểm tra", value=def_hkt, key="dlg_sm_hkt")
            with col_b4:
                size_cal = st.number_input("📏 Size (Cal)", min_value=0.0, step=0.1, value=def_cal, key="dlg_sm_cal")
            sl = st.number_input("🔢 Số lượng buồng mẫu", min_value=0, step=10, value=def_sl, key="dlg_sm_sl")

        if st.button("✅ Cập nhật", key="btn_edit_sm", use_container_width=True, type="primary"):
            if not mau_day_color: st.error("❌ Cần chọn màu dây")
            elif sl <= 0: st.error("❌ Số lượng buồng mẫu > 0")
            else:
                # Auto-create/validate ribbon
                farm_id = get_farm_id_from_name(editing_row["farm"])
                iso = ngay.isocalendar()
                _, err = get_or_create_ribbon(farm_id, iso[0], iso[1], mau_day_color)
                if err:
                    st.error(err)
                    st.stop()
                data = {
                    "mau_day": mau_day_color, "lan_do": lan_do,
                    "ngay_do": ngay.isoformat(), "so_luong_mau": sl, "tuan": ngay.isocalendar()[1],
                    "hang_kiem_tra": hang_kiem_tra.strip(), "size_cal": size_cal
                }
                supabase.table("size_measure_logs").update(data).eq("id", editing_row["id"]).execute()
                st.session_state["toast"] = f"✅ Sửa đo size {lot_id} thành công!"
                st.rerun()


@dialog_decorator("✏️ Chỉnh sửa Kiểm kê cây")
def edit_tree_inventory_dialog(editing_row, available_lots):
    def_lot = available_lots.index(editing_row["lot_id"]) if editing_row["lot_id"] in available_lots else 0
    def_ngay = pd.to_datetime(editing_row["ngay_kiem_ke"]).date()
    def_sl = int(editing_row["so_luong_cay_thuc_te"])
    
    with st.container(border=True):
        col_a, col_b = st.columns(2)
        with col_a:
            st.text_input("🏷️ Lứa (Mã hệ thống)", value=editing_row["lot_id"], disabled=True, key="dlg_inv_lot")
            lot_id = editing_row["lot_id"]
        with col_b:
            col_b1, col_b2 = st.columns([2, 1])
            with col_b1:
                ngay = st.date_input("📆 Ngày kiểm kê", value=def_ngay, key="dlg_inv_ngay")
            with col_b2:
                st.text_input("📍 Tuần", value=str(ngay.isocalendar()[1]), disabled=True, key=f"dlg_w_inv_{ngay}")
            sl = st.number_input("🔢 Số lượng cây thực tế", min_value=0, step=100, value=def_sl, key="dlg_inv_sl")

        if st.button("✅ Cập nhật", key="btn_edit_inv", use_container_width=True, type="primary"):
            if sl <= 0: st.error("❌ Cần nhập Số lượng cây lớn hơn 0.")
            else:
                data = {"ngay_kiem_ke": ngay.isoformat(), "so_luong_cay_thuc_te": sl, "tuan": ngay.isocalendar()[1]}
                supabase.table("tree_inventory_logs").update(data).eq("id", editing_row["id"]).execute()
                st.session_state["toast"] = f"✅ Lưu kiểm kê cây lô {lot_id} thành công!"
                st.rerun()

@dialog_decorator("✏️ Chỉnh sửa Đo pH Đất")
def edit_soil_ph_dialog(editing_row, available_lots):
    def_lot = available_lots.index(editing_row["lot_id"]) if editing_row["lot_id"] in available_lots else 0
    def_lan_do = int(editing_row["lan_do"])
    def_ngay = pd.to_datetime(editing_row["ngay_do"]).date()
    def_val = float(editing_row["ph_value"])
    
    with st.container(border=True):
        col_a, col_b = st.columns(2)
        with col_a:
            st.text_input("🏷️ Lứa (Mã hệ thống)", value=editing_row["lot_id"], disabled=True, key="dlg_ph_lot")
            lot_id = editing_row["lot_id"]
            st.number_input("Lần đo", value=def_lan_do, disabled=True, key="dlg_ph_lando")
        with col_b:
            col_b1, col_b2 = st.columns([2, 1])
            with col_b1:
                ngay = st.date_input("📆 Ngày đo", value=def_ngay, key="dlg_ph_ngay")
            with col_b2:
                st.text_input("📍 Tuần", value=str(ngay.isocalendar()[1]), disabled=True, key=f"dlg_w_ph_{ngay}")
            val = st.number_input("pH", min_value=0.0, max_value=14.0, step=0.1, value=def_val, key="dlg_ph_val")

        if st.button("✅ Cập nhật", key="btn_edit_ph", use_container_width=True, type="primary"):
            if val <= 0: st.error("❌ Cần nhập giá trị pH hợp lệ.")
            else:
                data = {"ngay_do": ngay.isoformat(), "ph_value": val, "tuan": ngay.isocalendar()[1]}
                supabase.table("soil_ph_logs").update(data).eq("id", editing_row["id"]).execute()
                st.session_state["toast"] = f"✅ Lưu kết quả pH lô {lot_id} thành công!"
                st.rerun()

@dialog_decorator("✏️ Chỉnh sửa Kiểm tra Fusarium")
def edit_fusarium_log_dialog(editing_row, available_lots):
    def_lot = available_lots.index(editing_row["lot_id"]) if editing_row["lot_id"] in available_lots else 0
    def_ngay = pd.to_datetime(editing_row["ngay_kiem_tra"]).date()
    def_sl = int(editing_row["so_cay_fusarium"])
    
    with st.container(border=True):
        col_a, col_b = st.columns(2)
        with col_a:
            st.text_input("🏷️ Lứa (Mã hệ thống)", value=editing_row["lot_id"], disabled=True, key="dlg_fus_lot")
            lot_id = editing_row["lot_id"]
        with col_b:
            col_b1, col_b2 = st.columns([2, 1])
            with col_b1:
                ngay = st.date_input("📆 Ngày kiểm tra", value=def_ngay, key="dlg_fus_ngay")
            with col_b2:
                st.text_input("📍 Tuần", value=str(ngay.isocalendar()[1]), disabled=True, key=f"dlg_w_fus_{ngay}")
            sl = st.number_input("🔢 Số cây bị Fusarium", min_value=0, step=1, value=def_sl, key="dlg_fus_sl")

        if st.button("✅ Cập nhật", key="btn_edit_fus", use_container_width=True, type="primary"):
            if sl < 0: st.error("❌ Cần nhập số cây lớn hơn hoặc bằng 0.")
            else:
                data = {"ngay_kiem_tra": ngay.isoformat(), "so_cay_fusarium": sl, "tuan": ngay.isocalendar()[1]}
                supabase.table("fusarium_logs").update(data).eq("id", editing_row["id"]).execute()
                st.session_state["toast"] = f"✅ Chi tiết Fusarium lô {lot_id} đã được lưu thành công!"
                st.rerun()

# MÀN HÌNH ĐĂNG NHẬP
# =====================================================
def render_login():
    col_logo1, col_logo2, col_logo3 = st.columns([1, 1, 1])
    with col_logo2:
        if os.path.exists("logo.png"): st.image("logo.png", use_container_width=True)

    st.markdown('<p class="main-title">🍌 Trường Tồn Banana Tracker</p>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1.2, 1])
    with col2:
        st.markdown("<h3 style='text-align: center;'>🔐 Đăng nhập hệ thống</h3>", unsafe_allow_html=True)
        st.divider()

        # Chọn Farm
        selected_farm = st.selectbox("🏗️ Chọn Farm", options=FARMS, index=None, placeholder="Vui lòng chọn Farm...", key="login_farm")
        
        # Chọn Team thuộc Farm
        av_teams = list(RBAC_DB.get(selected_farm, {}).keys()) if selected_farm else []
        selected_team = st.selectbox("👥 Chọn Đội / Vai trò", options=av_teams, index=None, placeholder="Vui lòng chọn Đội / Vai trò...", key="login_team")

        # Nhập MK
        password = st.text_input("🔑 Mật khẩu", type="password", key="login_pass", placeholder="Nhập mật khẩu...")

        st.markdown("")
        if st.button("🚀 Đăng nhập", use_container_width=True, type="primary"):
            if not selected_farm:
                st.warning("⚠️ Vui lòng chọn Farm.")
            elif not selected_team:
                st.warning("⚠️ Vui lòng chọn Đội / Vai trò.")
            elif not password:
                st.warning("⚠️ Vui lòng nhập mật khẩu.")
            else:
                correct_pass = RBAC_DB.get(selected_farm, {}).get(selected_team, "")
                if password.strip() == str(correct_pass).strip():
                    st.session_state["logged_in"] = True
                    st.session_state["current_farm"] = selected_farm
                    st.session_state["current_team"] = selected_team
                    
                    # Đăng nhập thành công -> lưu vào URL để tránh mất session khi nhấn F5
                    st.query_params["logged_in"] = "true"
                    st.query_params["farm"] = selected_farm
                    st.query_params["team"] = selected_team
                    
                    insert_access_log(selected_farm, selected_team, "Đăng nhập thành công")
                    st.success(f"✅ Đăng nhập {selected_team} - {selected_farm}!")
                    st.rerun()
                else:
                    insert_access_log(selected_farm, selected_team, "Sai mật khẩu")
                    st.error("❌ Mật khẩu không đúng.")

        st.divider()
        st.markdown("<p style='text-align: center; color: #888888; font-size: 0.85rem;'>💡 Vui lòng chọn đúng vai trò của mình để thao tác đúng nghiệp vụ.</p>", unsafe_allow_html=True)

def generate_chich_bap_excel(df_lots, df_stg) -> bytes:
    """Tạo file Excel báo cáo Chích bắp theo ngày, chia sheet theo năm.
    Mỗi sheet = 1 năm. Cột = từng ngày nhóm theo tuần, Hàng = Lô (đợt trồng)."""
    import re as _re_chich

    # ── Styles ──
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    week_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    subtotal_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    grand_total_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ── Filter chích bắp ──
    df_chich = df_stg[df_stg["giai_doan"] == "Chích bắp"].copy() if not df_stg.empty and "giai_doan" in df_stg.columns else pd.DataFrame()

    # Chỉ lấy đợt trồng mới
    if not df_lots.empty and "loai_trong" in df_lots.columns:
        valid_blids = set(df_lots[df_lots["loai_trong"] == "Trồng mới"]["id"].dropna().astype(int).tolist())
        if not df_chich.empty and "base_lot_id" in df_chich.columns:
            df_chich = df_chich[df_chich["base_lot_id"].isin(valid_blids)]

    # Loại bỏ records trước ngày trồng
    if not df_chich.empty and not df_lots.empty and "ngay_trong" in df_lots.columns:
        planting_date_map = {}
        for _, lr in df_lots.iterrows():
            if pd.notna(lr.get("id")) and pd.notna(lr.get("ngay_trong")):
                planting_date_map[int(lr["id"])] = pd.to_datetime(lr["ngay_trong"])
        def _is_after_planting(row):
            blid = row.get("base_lot_id")
            if pd.isna(blid): return True
            plant_dt = planting_date_map.get(int(blid))
            if plant_dt is None: return True
            return pd.to_datetime(row["ngay_thuc_hien"]) >= plant_dt
        df_chich = df_chich[df_chich.apply(_is_after_planting, axis=1)]

    wb = Workbook()

    if df_chich.empty:
        ws = wb.active
        ws.title = "Báo cáo Chích bắp"
        ws.cell(row=1, column=1, value="Chưa có dữ liệu Chích bắp.")
        output = io.BytesIO(); wb.save(output); output.seek(0)
        return output.getvalue()

    # Parse dates & year
    df_chich["ngay_thuc_hien"] = pd.to_datetime(df_chich["ngay_thuc_hien"])
    df_chich["tuan"] = df_chich["tuan"].fillna(
        df_chich["ngay_thuc_hien"].dt.isocalendar().week
    ).astype(int)
    df_chich["_year"] = df_chich["ngay_thuc_hien"].dt.year

    # ── Build lot display names ──
    lot_batch_keys = []
    if not df_lots.empty and "id" in df_lots.columns:
        lot_groups = df_lots.groupby("lo")
        for lo_name, grp in lot_groups:
            if len(grp) > 1:
                sorted_grp = grp.sort_values("ngay_trong") if "ngay_trong" in grp.columns else grp
                for i, (_, b_row) in enumerate(sorted_grp.iterrows(), 1):
                    lot_batch_keys.append((lo_name, b_row["id"], f"{lo_name} (đợt {i})"))
            else:
                for _, b_row in grp.iterrows():
                    lot_batch_keys.append((lo_name, b_row["id"], lo_name))
    active_blids = set(df_chich["base_lot_id"].dropna().astype(int).unique())
    lot_batch_keys = [x for x in lot_batch_keys if x[1] in active_blids]
    mapped_blids = {x[1] for x in lot_batch_keys}
    for blid in active_blids - mapped_blids:
        lo_name = df_chich[df_chich["base_lot_id"] == blid]["lo"].iloc[0] if "lo" in df_chich.columns else f"Lot_{blid}"
        lot_batch_keys.append((lo_name, blid, lo_name))

    def _nat_sort(item):
        name = item[2]
        m_batch = _re_chich.match(r"^(.+?)\s*\(đợt\s*(\d+)\)$", name)
        base_name, batch_num = (m_batch.group(1), int(m_batch.group(2))) if m_batch else (name, 0)
        m_num = _re_chich.match(r"^(\d+)(.*)", base_name)
        return (int(m_num.group(1)), m_num.group(2), batch_num) if m_num else (9999, base_name, batch_num)
    lot_batch_keys.sort(key=_nat_sort)

    # ── Build one sheet per year ──
    years = sorted(df_chich["_year"].unique())
    wb.remove(wb.active)  # remove default empty sheet

    for year in years:
        ws = wb.create_sheet(title=str(year))
        df_yr = df_chich[df_chich["_year"] == year]

        weeks = sorted(df_yr["tuan"].unique())
        week_dates = {}
        for wk in weeks:
            week_dates[wk] = sorted(df_yr[df_yr["tuan"] == wk]["ngay_thuc_hien"].dt.date.unique())

        # Lọc lot_batch_keys chỉ giữ lô có data năm này
        yr_blids = set(df_yr["base_lot_id"].dropna().astype(int).unique())
        yr_lots = [x for x in lot_batch_keys if x[1] in yr_blids]

        # === HEADER ===
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        c_lo = ws.cell(row=1, column=1, value="Lô")
        c_lo.font = Font(bold=True, size=11); c_lo.fill = header_fill
        c_lo.alignment = center_align; c_lo.border = thin_border
        ws.cell(row=2, column=1).border = thin_border

        col = 2
        week_col_map = {}
        for wk in weeks:
            dates = week_dates[wk]
            start_col = col
            end_col = col + len(dates)  # dates + subtotal
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            c_wk = ws.cell(row=1, column=start_col, value=f"Tuần {wk}")
            c_wk.font = Font(bold=True, size=11); c_wk.fill = week_fill
            c_wk.alignment = center_align; c_wk.border = thin_border

            date_cols = {}
            for d in dates:
                c_d = ws.cell(row=2, column=col, value=d.strftime("%d/%m"))
                c_d.font = Font(size=9); c_d.fill = header_fill
                c_d.alignment = center_align; c_d.border = thin_border
                date_cols[d] = col; col += 1

            c_sub = ws.cell(row=2, column=col, value="Σ tuần")
            c_sub.font = Font(bold=True, size=9); c_sub.fill = subtotal_fill
            c_sub.alignment = center_align; c_sub.border = thin_border
            week_col_map[wk] = {"date_cols": date_cols, "subtotal_col": col}
            col += 1

        # Lũy kế column
        c_gt = ws.cell(row=1, column=col, value="Lũy kế")
        c_gt.font = Font(bold=True, size=11); c_gt.fill = grand_total_fill
        c_gt.alignment = center_align; c_gt.border = thin_border
        c_gt2 = ws.cell(row=2, column=col, value="Tổng")
        c_gt2.font = Font(bold=True, size=9); c_gt2.fill = grand_total_fill
        c_gt2.alignment = center_align; c_gt2.border = thin_border
        grand_total_col = col; total_col_end = col

        for r in range(1, 3):
            for ci in range(1, total_col_end + 1):
                ws.cell(row=r, column=ci).border = thin_border

        # === DATA ROWS ===
        data_start_row = 3
        for li, (_, blid, display_name) in enumerate(yr_lots):
            row_idx = data_start_row + li
            c = ws.cell(row=row_idx, column=1, value=display_name)
            c.font = Font(bold=True); c.border = thin_border; c.alignment = center_align
            lot_total = 0
            for wk in weeks:
                wm = week_col_map[wk]; week_sum = 0
                for d, d_col in wm["date_cols"].items():
                    mask = (df_yr["base_lot_id"] == blid) & (df_yr["ngay_thuc_hien"].dt.date == d)
                    val = int(df_yr[mask]["so_luong"].sum())
                    cell = ws.cell(row=row_idx, column=d_col, value=val if val > 0 else "")
                    cell.border = thin_border; cell.alignment = center_align
                    week_sum += val
                c_ws = ws.cell(row=row_idx, column=wm["subtotal_col"], value=week_sum if week_sum > 0 else "")
                c_ws.font = Font(bold=True); c_ws.fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
                c_ws.border = thin_border; c_ws.alignment = center_align
                lot_total += week_sum
            c_gt_val = ws.cell(row=row_idx, column=grand_total_col, value=lot_total if lot_total > 0 else "")
            c_gt_val.font = Font(bold=True); c_gt_val.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
            c_gt_val.border = thin_border; c_gt_val.alignment = center_align

        # === TOTAL ROW ===
        total_row = data_start_row + len(yr_lots)
        c = ws.cell(row=total_row, column=1, value="TỔNG")
        c.font = Font(bold=True, size=11); c.fill = grand_total_fill
        c.border = thin_border; c.alignment = center_align
        for ci in range(2, total_col_end + 1):
            col_sum = sum(int(ws.cell(row=r, column=ci).value) for r in range(data_start_row, total_row) if isinstance(ws.cell(row=r, column=ci).value, (int, float)))
            c_tot = ws.cell(row=total_row, column=ci, value=col_sum if col_sum > 0 else "")
            c_tot.font = Font(bold=True); c_tot.fill = grand_total_fill
            c_tot.border = thin_border; c_tot.alignment = center_align

        ws.column_dimensions[get_column_letter(1)].width = 16
        for ci in range(2, total_col_end + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 9
        ws.freeze_panes = "B3"

    output = io.BytesIO(); wb.save(output); output.seek(0)
    return output.getvalue()


def generate_cut_bap_excel(df_lots, df_stg, df_des=None, df_har=None) -> bytes:
    """Tạo file Excel báo cáo Cắt bắp theo tuần.
    Mỗi sheet = 1 năm. Mỗi tuần = 4 cột (CẮT BẮP | XUẤT HỦY | Thu hoạch | Tồn trên lô).
    Cột Thu hoạch dùng harvest_logs.mau_day để quy về tuần cắt bắp nguồn."""
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    total_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _clean_farm_name(value):
        if value is None or pd.isna(value):
            return ""
        return str(value).strip()

    def _farm_sort_key(farm_name):
        m = re.search(r"(\d+)", _clean_farm_name(farm_name))
        return (int(m.group(1)) if m else 9999, _clean_farm_name(farm_name))

    def _farm_short_label(farm_name):
        farm_name = _clean_farm_name(farm_name)
        m = re.search(r"(\d+)", farm_name)
        return m.group(1) if m else farm_name

    def _forecast_offsets_for_farm(farm_name):
        farm_name = _clean_farm_name(farm_name)
        if farm_name == "Farm 126":
            return [8]
        if farm_name == "Farm 157":
            return [9]
        return [8, 9]

    def _farm_names_from_frames(*frames):
        names = []
        for frame in frames:
            if frame is None or frame.empty or "farm" not in frame.columns:
                continue
            for value in frame["farm"].dropna().unique().tolist():
                farm_name = _clean_farm_name(value)
                if farm_name and farm_name not in names:
                    names.append(farm_name)
        return sorted(names, key=_farm_sort_key)

    COLOR_MAP = {
        "đỏ": "FFB3B3", "cam": "FFD9B3", "vàng": "FFFFB3",
        "xanh lá": "B3FFB3", "xanh dương": "B3D9FF", "tím": "D9B3FF",
        "đen": "D9D9D9", "trắng": "F5F5F5", "hồng": "FFB3D9", "nâu": "D9C4B3",
    }
    def get_mau_day_fill(mau_day_name):
        base = mau_day_name.split("-")[0].strip().lower() if mau_day_name else ""
        hex_color = COLOR_MAP.get(base)
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid") if hex_color else None

    # Filter data
    df_cut = df_stg[df_stg["giai_doan"] == "Cắt bắp"].copy() if not df_stg.empty and "giai_doan" in df_stg.columns else pd.DataFrame()
    df_xh = pd.DataFrame()
    if df_des is not None and not df_des.empty and "giai_doan" in df_des.columns:
        df_xh = df_des[df_des["giai_doan"].isin(["Trước thu hoạch", "Sau thu hoạch"])].copy()
    df_hv = df_har.copy() if df_har is not None and not df_har.empty else pd.DataFrame()
    report_farms = _farm_names_from_frames(df_cut, df_xh, df_hv, df_lots)
    include_farm_col = len(_farm_names_from_frames(df_cut, df_xh, df_hv)) > 1

    wb = Workbook()

    if df_cut.empty:
        ws = wb.active; ws.title = "Báo cáo Cắt bắp"
        ws.cell(row=1, column=1, value="Chưa có dữ liệu Cắt bắp.")
        output = io.BytesIO(); wb.save(output); output.seek(0)
        return output.getvalue()

    # Parse dates, cast types
    df_cut["ngay_thuc_hien"] = pd.to_datetime(df_cut["ngay_thuc_hien"])
    df_cut["_year"] = df_cut["ngay_thuc_hien"].dt.year.astype(int)
    df_cut["tuan"] = pd.to_numeric(df_cut["tuan"], errors="coerce").fillna(
        df_cut["ngay_thuc_hien"].dt.isocalendar().week
    ).astype(int)

    if not df_xh.empty:
        df_xh["ngay_xuat_huy"] = pd.to_datetime(df_xh["ngay_xuat_huy"])
        df_xh["_year"] = df_xh["ngay_xuat_huy"].dt.year.astype(int)
        df_xh["tuan"] = pd.to_numeric(df_xh["tuan"], errors="coerce").fillna(
            df_xh["ngay_xuat_huy"].dt.isocalendar().week
        ).astype(int)

    if not df_hv.empty and "ngay_thu_hoach" in df_hv.columns:
        df_hv["ngay_thu_hoach"] = pd.to_datetime(df_hv["ngay_thu_hoach"])
        df_hv["_year"] = df_hv["ngay_thu_hoach"].dt.year.astype(int)
        df_hv["tuan"] = pd.to_numeric(df_hv["tuan"], errors="coerce").fillna(
            df_hv["ngay_thu_hoach"].dt.isocalendar().week
        ).astype(int)

    # ── Build lot display names (chia đợt giống chích bắp) ──
    import re as _re_cut
    lot_batch_keys = []
    base_lot_farm = {}
    if not df_lots.empty and "id" in df_lots.columns:
        group_cols = ["farm", "lo"] if "farm" in df_lots.columns else ["lo"]
        lot_groups = df_lots.groupby(group_cols, dropna=False)
        for group_key, grp in lot_groups:
            if "farm" in df_lots.columns:
                farm_name, lo_name = group_key if isinstance(group_key, tuple) else ("", group_key)
                farm_name = _clean_farm_name(farm_name)
            else:
                farm_name, lo_name = "", group_key
            if len(grp) > 1:
                sorted_grp = grp.sort_values("ngay_trong") if "ngay_trong" in grp.columns else grp
                for i, (_, b_row) in enumerate(sorted_grp.iterrows(), 1):
                    blid = int(b_row["id"])
                    base_lot_farm[blid] = farm_name
                    lot_batch_keys.append((farm_name, lo_name, blid, f"{lo_name} (đợt {i})"))
            else:
                for _, b_row in grp.iterrows():
                    blid = int(b_row["id"])
                    base_lot_farm[blid] = farm_name
                    lot_batch_keys.append((farm_name, lo_name, blid, lo_name))
    # Also include lots from df_cut/df_xh that might not be in df_lots
    active_blids_cut = set()
    if not df_cut.empty and "base_lot_id" in df_cut.columns:
        active_blids_cut.update(df_cut["base_lot_id"].dropna().astype(int).unique())
    if not df_xh.empty and "base_lot_id" in df_xh.columns:
        active_blids_cut.update(df_xh["base_lot_id"].dropna().astype(int).unique())
    # Keep only lots that have data
    lot_batch_keys = [x for x in lot_batch_keys if x[2] in active_blids_cut]
    mapped_blids = {x[2] for x in lot_batch_keys}
    for blid in active_blids_cut - mapped_blids:
        lo_name = f"Lot_{blid}"
        farm_name = ""
        if "lo" in df_cut.columns and "base_lot_id" in df_cut.columns:
            _m = df_cut[df_cut["base_lot_id"] == blid]["lo"]
            if not _m.empty:
                lo_name = _m.iloc[0]
            if "farm" in df_cut.columns:
                _f = df_cut[df_cut["base_lot_id"] == blid]["farm"]
                if not _f.empty:
                    farm_name = _clean_farm_name(_f.iloc[0])
        base_lot_farm[int(blid)] = farm_name
        lot_batch_keys.append((farm_name, lo_name, int(blid), lo_name))

    def _nat_sort_cut(item):
        farm_name, _, _, name = item
        m_batch = _re_cut.match(r"^(.+?)\s*\(đợt\s*(\d+)\)$", name)
        base_name, batch_num = (m_batch.group(1), int(m_batch.group(2))) if m_batch else (name, 0)
        m_num = _re_cut.match(r"^(\d+)(.*)", base_name)
        lot_key = (int(m_num.group(1)), m_num.group(2), batch_num) if m_num else (9999, base_name, batch_num)
        return (_farm_sort_key(farm_name), *lot_key)
    lot_batch_keys.sort(key=_nat_sort_cut)

    farm_id_by_name = {
        farm_name: get_farm_id_from_name(farm_name)
        for farm_name in report_farms
        if farm_name
    }

    years = set(df_cut["_year"].dropna().astype(int).unique())
    if not df_xh.empty and "_year" in df_xh.columns:
        years.update(df_xh["_year"].dropna().astype(int).unique())
    years = sorted(years)
    wb.remove(wb.active)

    for year in years:
        ws = wb.create_sheet(title=str(year))
        df_cut_yr = df_cut[df_cut["_year"] == year]
        df_xh_yr = df_xh[df_xh["_year"] == year] if not df_xh.empty and "_year" in df_xh.columns else pd.DataFrame()
        df_xh_before_yr = df_xh_yr[df_xh_yr["giai_doan"] == "Trước thu hoạch"] if not df_xh_yr.empty and "giai_doan" in df_xh_yr.columns else pd.DataFrame()
        df_xh_after_yr = df_xh_yr[df_xh_yr["giai_doan"] == "Sau thu hoạch"] if not df_xh_yr.empty and "giai_doan" in df_xh_yr.columns else pd.DataFrame()
        df_hv_yr = df_hv[df_hv["_year"] == year] if not df_hv.empty and "_year" in df_hv.columns else pd.DataFrame()

        # Week-color map
        year_farms = _farm_names_from_frames(df_cut_yr, df_xh_yr, df_hv_yr) or report_farms
        week_color_by_farm = {farm_name: {} for farm_name in year_farms}
        all_weeks = set()
        data_weeks_by_farm = {farm_name: set() for farm_name in year_farms}
        for farm_name in year_farms:
            farm_id = farm_id_by_name.get(farm_name)
            if farm_id:
                _rb_res = supabase.table("ribbon_schedule").select("week_number, color_name") \
                    .eq("farm_id", farm_id).eq("year", int(year)).eq("is_deleted", False).execute()
                for _rb in (_rb_res.data or []):
                    week_number = int(_rb["week_number"])
                    week_color_by_farm.setdefault(farm_name, {})[week_number] = _rb["color_name"]
                    all_weeks.add(week_number)
        if "farm" in df_cut_yr.columns:
            for farm_name, grp in df_cut_yr.groupby("farm"):
                farm_name = _clean_farm_name(farm_name)
                data_weeks_by_farm.setdefault(farm_name, set()).update(int(t) for t in grp["tuan"].dropna().unique())
        for t in df_cut_yr["tuan"].dropna().unique(): all_weeks.add(int(t))
        if not df_xh_yr.empty:
            if "farm" in df_xh_yr.columns:
                for farm_name, grp in df_xh_yr.groupby("farm"):
                    farm_name = _clean_farm_name(farm_name)
                    data_weeks_by_farm.setdefault(farm_name, set()).update(int(t) for t in grp["tuan"].dropna().unique())
            for t in df_xh_yr["tuan"].dropna().unique(): all_weeks.add(int(t))

        weeks = sorted(all_weeks)
        if not weeks:
            ws.cell(row=1, column=1, value=f"Chưa có dữ liệu năm {year}.")
            continue

        def _farms_for_week(week):
            farms = [
                farm_name
                for farm_name in year_farms
                if week in week_color_by_farm.get(farm_name, {})
                or week in data_weeks_by_farm.get(farm_name, set())
            ]
            return farms or year_farms

        def _week_color_label(week):
            entries = [
                (farm_name, week_color_by_farm.get(farm_name, {}).get(week, ""))
                for farm_name in _farms_for_week(week)
            ]
            entries = [(farm, color) for farm, color in entries if color]
            if not entries:
                return ""
            if include_farm_col and len(entries) > 1:
                return "\n".join(f"{_farm_short_label(farm)}: {color}" for farm, color in entries)
            return entries[0][1]

        def _week_color_fill(week):
            colors = [
                color for farm_name in _farms_for_week(week)
                for color in [week_color_by_farm.get(farm_name, {}).get(week, "")]
                if color
            ]
            base_colors = {str(color).split("-")[0].strip().lower() for color in colors}
            if len(base_colors) == 1:
                return get_mau_day_fill(colors[0])
            return None

        harvest_by_blid_week = {}
        if not df_hv_yr.empty and {"base_lot_id", "mau_day", "ngay_thu_hoach", "so_luong"}.issubset(df_hv_yr.columns):
            color_to_weeks_by_farm = {}
            for farm_name, farm_week_color in week_color_by_farm.items():
                farm_color_to_weeks = {}
                for wk, color in farm_week_color.items():
                    if color:
                        farm_color_to_weeks.setdefault(str(color).strip(), []).append(int(wk))
                color_to_weeks_by_farm[farm_name] = farm_color_to_weeks
            for _, h_row in df_hv_yr.iterrows():
                blid = h_row.get("base_lot_id")
                mau_day = str(h_row.get("mau_day") or "").strip()
                ngay_har = h_row.get("ngay_thu_hoach")
                if pd.isna(blid) or not mau_day or pd.isna(ngay_har):
                    continue
                farm_name = _clean_farm_name(h_row.get("farm")) if "farm" in df_hv_yr.columns else ""
                if not farm_name:
                    farm_name = base_lot_farm.get(int(blid), "")
                candidate_weeks = color_to_weeks_by_farm.get(farm_name, {}).get(mau_day, [])
                if not candidate_weeks:
                    continue
                har_pair = _iso_year_week(ngay_har)
                forecast_week_offsets = [offset - 1 for offset in _forecast_offsets_for_farm(farm_name)]
                exact_weeks = [
                    wk for wk in candidate_weeks
                    if any(_shift_iso_week(year, wk, offset) == har_pair for offset in forecast_week_offsets)
                ]
                # Nếu màu dây chỉ xuất hiện một lần trong sheet, cho phép map dù thu hoạch lệch vài ngày.
                target_weeks = exact_weeks or (candidate_weeks if len(candidate_weeks) == 1 else [])
                if not target_weeks:
                    continue
                qty = int(pd.to_numeric(pd.Series([h_row.get("so_luong")]), errors="coerce").fillna(0).iloc[0])
                key = (int(blid), int(target_weeks[0]))
                harvest_by_blid_week[key] = harvest_by_blid_week.get(key, 0) + qty

        # === Helper: forecast harvest label ===
        def _shifted_week_label(cut_week, cut_year, inclusive_weeks):
            from datetime import date as _date
            dec_28 = _date(cut_year, 12, 28)
            max_week = dec_28.isocalendar()[1]  # 52 or 53

            target_week = int(cut_week) + int(inclusive_weeks) - 1
            target_year = int(cut_year)

            if target_week > max_week:
                target_week -= max_week
                target_year += 1

            week_str = f"{target_week}" if target_year == cut_year else f"{target_week}-{target_year}"
            return f"{week_str} (+{inclusive_weeks})"

        def _forecast_harvest_label(cut_week, cut_year, farm_names):
            """Tạo label dự báo thu hoạch theo offset riêng từng farm."""
            labels = []
            for farm_name in farm_names:
                offsets = _forecast_offsets_for_farm(farm_name)
                farm_label = "/".join(_shifted_week_label(cut_week, cut_year, offset) for offset in offsets)
                if include_farm_col and len(farm_names) > 1:
                    farm_label = f"{_farm_short_label(farm_name)}: {farm_label}"
                labels.append(farm_label)
            return "\n".join(labels)

        forecast_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

        # === HEADER: 4 rows ===
        # Row 1: "Farm" + "Lô" (khi gộp farm) | Dự báo thu hoạch merged 4 cols
        # Row 2:                              | "Tuần X" merged across 4 cols
        # Row 3:                              | "CẮT BẮP" | "XUẤT HỦY" | "Thu hoạch" | "Tồn trên lô"
        # Row 4:                              | màu dây    | màu dây    | màu dây     | màu dây
        label_col_count = 2 if include_farm_col else 1
        if include_farm_col:
            ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=1)
            c_farm_header = ws.cell(row=1, column=1, value="Farm")
            c_farm_header.font = Font(bold=True, size=10); c_farm_header.fill = header_fill
            c_farm_header.border = thin_border; c_farm_header.alignment = center_align
            ws.merge_cells(start_row=1, start_column=2, end_row=4, end_column=2)
            c_lo = ws.cell(row=1, column=2, value="Lô")
        else:
            ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=1)
            c_lo = ws.cell(row=1, column=1, value="Lô")
        c_lo.font = Font(bold=True, size=10); c_lo.fill = header_fill
        c_lo.border = thin_border; c_lo.alignment = center_align

        col = label_col_count + 1
        week_cut_col = {}   # week -> col index of CẮT BẮP
        week_des_col = {}   # week -> col index of XUẤT HỦY
        week_har_col = {}   # week -> col index of Thu hoạch
        week_rem_col = {}   # week -> col index of Tồn trên lô
        for week in weeks:
            cut_col = col
            des_col = col + 1
            har_col = col + 2
            rem_col = col + 3
            week_cut_col[week] = cut_col
            week_des_col[week] = des_col
            week_har_col[week] = har_col
            week_rem_col[week] = rem_col
            color_name = _week_color_label(week)
            color_fill_cell = _week_color_fill(week)

            # Row 1: Dự báo thu hoạch merged
            forecast_label = _forecast_harvest_label(week, year, _farms_for_week(week))
            ws.merge_cells(start_row=1, start_column=cut_col, end_row=1, end_column=rem_col)
            c_fc = ws.cell(row=1, column=cut_col, value=forecast_label)
            c_fc.font = Font(bold=True, size=9, italic=True); c_fc.fill = forecast_fill
            c_fc.alignment = center_align; c_fc.border = thin_border
            for c_i in range(cut_col, rem_col + 1):
                ws.cell(row=1, column=c_i).border = thin_border

            # Row 2: "Tuần X" merged
            ws.merge_cells(start_row=2, start_column=cut_col, end_row=2, end_column=rem_col)
            c_w = ws.cell(row=2, column=cut_col, value=f"Tuần {week}")
            c_w.font = Font(bold=True, size=11); c_w.fill = header_fill
            c_w.alignment = center_align; c_w.border = thin_border
            for c_i in range(cut_col, rem_col + 1):
                ws.cell(row=2, column=c_i).border = thin_border

            # Row 3: "CẮT BẮP" | "XUẤT HỦY" | "Thu hoạch" | "Tồn trên lô"
            c_cut = ws.cell(row=3, column=cut_col, value="CẮT BẮP")
            c_cut.font = Font(bold=True, size=9); c_cut.fill = white_fill
            c_cut.alignment = center_align; c_cut.border = thin_border
            c_des = ws.cell(row=3, column=des_col, value="XUẤT HỦY")
            c_des.font = Font(bold=True, size=9); c_des.fill = white_fill
            c_des.alignment = center_align; c_des.border = thin_border
            c_har = ws.cell(row=3, column=har_col, value="Thu hoạch")
            c_har.font = Font(bold=True, size=9, color="C00000"); c_har.fill = white_fill
            c_har.alignment = center_align; c_har.border = thin_border
            c_rem = ws.cell(row=3, column=rem_col, value="Tồn trên lô")
            c_rem.font = Font(bold=True, size=9, color="C00000"); c_rem.fill = white_fill
            c_rem.alignment = center_align; c_rem.border = thin_border

            # Row 4: màu dây
            for c_i in range(cut_col, rem_col + 1):
                c_r4 = ws.cell(row=4, column=c_i, value=color_name)
                c_r4.font = Font(bold=True, size=8)
                c_r4.fill = color_fill_cell if color_fill_cell else white_fill
                c_r4.alignment = center_align; c_r4.border = thin_border

            col += 4

        # Lũy kế: 4 cols (CẮT + HỦY + THU + TỒN)
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col + 3)
        c_lk = ws.cell(row=1, column=col, value="Lũy kế")
        c_lk.font = Font(bold=True, size=11); c_lk.fill = total_fill
        c_lk.alignment = center_align; c_lk.border = thin_border
        for c_i in range(col, col + 4):
            ws.cell(row=1, column=c_i).border = thin_border
            ws.cell(row=2, column=c_i).border = thin_border
        lk_cut_col = col
        lk_des_col = col + 1
        lk_har_col = col + 2
        lk_rem_col = col + 3
        c_lk_c = ws.cell(row=3, column=lk_cut_col, value="CẮT"); c_lk_c.font = Font(bold=True, size=9)
        c_lk_c.fill = white_fill; c_lk_c.alignment = center_align; c_lk_c.border = thin_border
        c_lk_d = ws.cell(row=3, column=lk_des_col, value="HỦY"); c_lk_d.font = Font(bold=True, size=9)
        c_lk_d.fill = white_fill; c_lk_d.alignment = center_align; c_lk_d.border = thin_border
        c_lk_h = ws.cell(row=3, column=lk_har_col, value="THU"); c_lk_h.font = Font(bold=True, size=9, color="C00000")
        c_lk_h.fill = white_fill; c_lk_h.alignment = center_align; c_lk_h.border = thin_border
        c_lk_r = ws.cell(row=3, column=lk_rem_col, value="TỒN"); c_lk_r.font = Font(bold=True, size=9, color="C00000")
        c_lk_r.fill = white_fill; c_lk_r.alignment = center_align; c_lk_r.border = thin_border
        for c_i in [lk_cut_col, lk_des_col, lk_har_col, lk_rem_col]:
            ws.cell(row=4, column=c_i).border = thin_border
            ws.cell(row=4, column=c_i).fill = total_fill
        last_col = lk_rem_col

        # === DATA ===
        # Filter lot_batch_keys to only include lots with data this year
        yr_blids = set()
        if "base_lot_id" in df_cut_yr.columns:
            yr_blids.update(df_cut_yr["base_lot_id"].dropna().astype(int).unique())
        if not df_xh_yr.empty and "base_lot_id" in df_xh_yr.columns:
            yr_blids.update(df_xh_yr["base_lot_id"].dropna().astype(int).unique())
        yr_lots = [x for x in lot_batch_keys if x[2] in yr_blids]

        data_start_row = 5
        for li, (farm_name, lo_name, blid, display_name) in enumerate(yr_lots):
            row_idx = data_start_row + li
            if include_farm_col:
                c_farm = ws.cell(row=row_idx, column=1, value=farm_name)
                c_farm.font = Font(bold=True); c_farm.border = thin_border; c_farm.alignment = center_align
                c = ws.cell(row=row_idx, column=2, value=display_name)
            else:
                c = ws.cell(row=row_idx, column=1, value=display_name)
            c.font = Font(bold=True); c.border = thin_border; c.alignment = center_align
            lot_cut_total = 0
            lot_des_total = 0
            lot_har_total = 0
            for week in weeks:
                # CẮT BẮP — filter by base_lot_id
                if "base_lot_id" in df_cut_yr.columns:
                    cut_mask = (df_cut_yr["base_lot_id"] == blid) & (df_cut_yr["tuan"] == week)
                else:
                    cut_mask = (df_cut_yr["lo"] == lo_name) & (df_cut_yr["tuan"] == week)
                    if "farm" in df_cut_yr.columns:
                        cut_mask = cut_mask & (df_cut_yr["farm"] == farm_name)
                val_cut = int(df_cut_yr[cut_mask]["so_luong"].sum())
                cell_c = ws.cell(row=row_idx, column=week_cut_col[week], value=val_cut if val_cut > 0 else "")
                cell_c.border = thin_border; cell_c.alignment = center_align
                lot_cut_total += val_cut
                # XUẤT HỦY — filter by base_lot_id
                val_des_before = 0
                val_des_after = 0
                if not df_xh_before_yr.empty and "base_lot_id" in df_xh_before_yr.columns:
                    des_mask = (df_xh_before_yr["base_lot_id"] == blid) & (df_xh_before_yr["tuan"] == week)
                    val_des_before = int(df_xh_before_yr[des_mask]["so_luong"].sum())
                elif not df_xh_before_yr.empty:
                    des_mask = (df_xh_before_yr["lo"] == lo_name) & (df_xh_before_yr["tuan"] == week)
                    if "farm" in df_xh_before_yr.columns:
                        des_mask = des_mask & (df_xh_before_yr["farm"] == farm_name)
                    val_des_before = int(df_xh_before_yr[des_mask]["so_luong"].sum())
                if not df_xh_after_yr.empty and "base_lot_id" in df_xh_after_yr.columns:
                    des_after_mask = (df_xh_after_yr["base_lot_id"] == blid) & (df_xh_after_yr["tuan"] == week)
                    val_des_after = int(df_xh_after_yr[des_after_mask]["so_luong"].sum())
                elif not df_xh_after_yr.empty:
                    des_after_mask = (df_xh_after_yr["lo"] == lo_name) & (df_xh_after_yr["tuan"] == week)
                    if "farm" in df_xh_after_yr.columns:
                        des_after_mask = des_after_mask & (df_xh_after_yr["farm"] == farm_name)
                    val_des_after = int(df_xh_after_yr[des_after_mask]["so_luong"].sum())
                val_des = val_des_before + val_des_after
                cell_d = ws.cell(row=row_idx, column=week_des_col[week], value=val_des if val_des > 0 else "")
                cell_d.border = thin_border; cell_d.alignment = center_align
                if val_des_after > 0:
                    cell_d.comment = Comment(
                        f"Trước thu hoạch: {val_des_before}\nSau thu hoạch: {val_des_after}",
                        "Truong Ton App",
                    )
                lot_des_total += val_des

                # THU HOẠCH — map từ harvest_logs.mau_day về tuần cắt bắp nguồn.
                val_har_gross = int(harvest_by_blid_week.get((int(blid), int(week)), 0))
                val_har = max(0, val_har_gross - val_des_after)
                cell_h = ws.cell(row=row_idx, column=week_har_col[week], value=val_har if val_har > 0 else "")
                cell_h.font = Font(color="C00000")
                cell_h.border = thin_border; cell_h.alignment = center_align
                if val_des_after > 0 and val_har_gross > 0:
                    cell_h.comment = Comment(
                        f"Thu hoạch gốc: {val_har_gross}\n"
                        f"Trừ xuất hủy sau thu hoạch: {val_des_after}\n"
                        f"Thu hoạch ròng hiển thị: {val_har}",
                        "Truong Ton App",
                    )
                lot_har_total += val_har

                val_rem = val_cut - val_des - val_har
                rem_value = val_rem if (val_cut > 0 or val_des > 0 or val_har > 0) else ""
                cell_r = ws.cell(row=row_idx, column=week_rem_col[week], value=rem_value)
                cell_r.font = Font(color="C00000")
                cell_r.border = thin_border; cell_r.alignment = center_align
            # Lũy kế
            c_lc = ws.cell(row=row_idx, column=lk_cut_col, value=lot_cut_total if lot_cut_total > 0 else "")
            c_lc.font = Font(bold=True); c_lc.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
            c_lc.border = thin_border; c_lc.alignment = center_align
            c_ld = ws.cell(row=row_idx, column=lk_des_col, value=lot_des_total if lot_des_total > 0 else "")
            c_ld.font = Font(bold=True); c_ld.fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
            c_ld.border = thin_border; c_ld.alignment = center_align
            c_lh = ws.cell(row=row_idx, column=lk_har_col, value=lot_har_total if lot_har_total > 0 else "")
            c_lh.font = Font(bold=True, color="C00000"); c_lh.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            c_lh.border = thin_border; c_lh.alignment = center_align
            lot_rem_total = lot_cut_total - lot_des_total - lot_har_total
            c_lr = ws.cell(row=row_idx, column=lk_rem_col, value=lot_rem_total if (lot_cut_total > 0 or lot_des_total > 0 or lot_har_total > 0) else "")
            c_lr.font = Font(bold=True, color="C00000"); c_lr.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            c_lr.border = thin_border; c_lr.alignment = center_align

        # === TOTAL ROW ===
        total_row = data_start_row + len(yr_lots)
        if include_farm_col:
            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
        c = ws.cell(row=total_row, column=1, value="Tổng")
        c.font = Font(bold=True, size=11); c.fill = total_fill
        c.border = thin_border; c.alignment = center_align
        if include_farm_col:
            ws.cell(row=total_row, column=2).fill = total_fill
            ws.cell(row=total_row, column=2).border = thin_border
            ws.cell(row=total_row, column=2).alignment = center_align
        for ci in range(label_col_count + 1, last_col + 1):
            values = [
                int(ws.cell(row=r, column=ci).value)
                for r in range(data_start_row, total_row)
                if isinstance(ws.cell(row=r, column=ci).value, (int, float))
            ]
            col_sum = sum(values)
            c_tot = ws.cell(row=total_row, column=ci, value=col_sum if values else "")
            c_tot.font = Font(bold=True); c_tot.fill = total_fill
            c_tot.border = thin_border; c_tot.alignment = center_align

        # === EXPECTED HARVEST ROW ===
        # Dự kiến thu hoạch = round(97% tổng cắt bắp), chưa trừ xuất hủy/thu hoạch.
        expected_row = total_row + 1
        if include_farm_col:
            ws.merge_cells(start_row=expected_row, start_column=1, end_row=expected_row, end_column=2)
        c_exp_label = ws.cell(row=expected_row, column=1, value="Dự kiến thu\nhoạch")
        c_exp_label.font = Font(bold=True, size=11)
        c_exp_label.fill = total_fill
        c_exp_label.border = thin_border
        c_exp_label.alignment = center_align
        if include_farm_col:
            ws.cell(row=expected_row, column=2).fill = total_fill
            ws.cell(row=expected_row, column=2).border = thin_border
            ws.cell(row=expected_row, column=2).alignment = center_align
        ws.row_dimensions[expected_row].height = 36

        def _expected_harvest_from_cut(cut_value):
            if not isinstance(cut_value, (int, float)) or cut_value <= 0:
                return "-"
            return int(float(cut_value) * 0.97 + 0.5)

        for week in weeks:
            cut_total = ws.cell(row=total_row, column=week_cut_col[week]).value
            for ci in (week_cut_col[week], week_des_col[week], week_rem_col[week]):
                c_dash = ws.cell(row=expected_row, column=ci, value="-")
                c_dash.font = Font(bold=True)
                c_dash.fill = white_fill
                c_dash.border = thin_border
                c_dash.alignment = center_align

            c_exp = ws.cell(row=expected_row, column=week_har_col[week], value=_expected_harvest_from_cut(cut_total))
            c_exp.font = Font(bold=True)
            c_exp.fill = total_fill
            c_exp.border = thin_border
            c_exp.alignment = center_align

        lk_expected = _expected_harvest_from_cut(ws.cell(row=total_row, column=lk_cut_col).value)
        for ci in (lk_cut_col, lk_des_col, lk_rem_col):
            c_dash = ws.cell(row=expected_row, column=ci, value="-")
            c_dash.font = Font(bold=True)
            c_dash.fill = white_fill
            c_dash.border = thin_border
            c_dash.alignment = center_align
        c_lk_exp = ws.cell(row=expected_row, column=lk_har_col, value=lk_expected)
        c_lk_exp.font = Font(bold=True)
        c_lk_exp.fill = total_fill
        c_lk_exp.border = thin_border
        c_lk_exp.alignment = center_align

        ws.column_dimensions[get_column_letter(1)].width = 14
        if include_farm_col:
            ws.column_dimensions[get_column_letter(2)].width = 14
        for ci in range(label_col_count + 1, last_col + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 11
        ws.freeze_panes = "C5" if include_farm_col else "B5"

    output = io.BytesIO(); wb.save(output); output.seek(0)
    return output.getvalue()

def generate_harvest_forecast_excel(df_lots, df_stg) -> bytes:
    """Tạo file dự báo thu hoạch từ dữ liệu Cắt bắp.

    Farm 126 dùng +8 tuần, Farm 157 dùng +9 tuần. Cách đếm inclusive:
    cắt tuần 20, +8 sẽ rơi vào tuần thu hoạch dự báo 27.
    Số dự báo = round(số cắt bắp * 97%).
    """
    wb = Workbook()
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if df_stg is None or df_stg.empty or "giai_doan" not in df_stg.columns:
        ws = wb.active
        ws.title = "Dự báo Thu hoạch"
        ws.cell(row=1, column=1, value="Chưa có dữ liệu Cắt bắp.")
        output = io.BytesIO(); wb.save(output); output.seek(0)
        return output.getvalue()

    df_cut = df_stg[df_stg["giai_doan"] == "Cắt bắp"].copy()
    if df_cut.empty:
        ws = wb.active
        ws.title = "Dự báo Thu hoạch"
        ws.cell(row=1, column=1, value="Chưa có dữ liệu Cắt bắp.")
        output = io.BytesIO(); wb.save(output); output.seek(0)
        return output.getvalue()

    lot_lookup = {}
    if df_lots is not None and not df_lots.empty and "id" in df_lots.columns:
        for _, lot_row in df_lots.iterrows():
            raw_id = lot_row.get("id")
            if pd.isna(raw_id):
                continue
            blid = int(raw_id)
            lot_lookup[blid] = {
                "farm": _report_clean_farm_name(lot_row.get("farm")) if "farm" in df_lots.columns else "",
                "lo": lot_row.get("lo") if "lo" in df_lots.columns else "",
            }

    parsed_rows = []
    farm_years = {}
    for _, row in df_cut.iterrows():
        ngay_cut = pd.to_datetime(row.get("ngay_thuc_hien"), errors="coerce")
        if pd.isna(ngay_cut):
            continue
        iso = ngay_cut.isocalendar()
        cut_year = int(iso.year)
        cut_week = int(row.get("tuan")) if pd.notna(row.get("tuan")) else int(iso.week)
        blid = row.get("base_lot_id")
        blid_int = int(blid) if pd.notna(blid) else None
        fallback_lot = lot_lookup.get(blid_int, {}) if blid_int is not None else {}
        farm_name = _report_clean_farm_name(row.get("farm")) if "farm" in df_cut.columns else ""
        if not farm_name:
            farm_name = fallback_lot.get("farm", "")
        lo_name = row.get("lo") if "lo" in df_cut.columns else ""
        if not lo_name:
            lo_name = fallback_lot.get("lo", "")
        cut_qty = pd.to_numeric(pd.Series([row.get("so_luong")]), errors="coerce").fillna(0).iloc[0]
        cut_qty = int(cut_qty)
        if cut_qty <= 0:
            continue
        parsed_rows.append({
            "farm": farm_name,
            "cut_year": cut_year,
            "cut_week": cut_week,
            "lo": lo_name,
            "base_lot_id": blid_int,
            "cut_qty": cut_qty,
        })
        if farm_name:
            farm_years.setdefault(farm_name, set()).add(cut_year)

    if not parsed_rows:
        ws = wb.active
        ws.title = "Dự báo Thu hoạch"
        ws.cell(row=1, column=1, value="Chưa có dữ liệu Cắt bắp hợp lệ.")
        output = io.BytesIO(); wb.save(output); output.seek(0)
        return output.getvalue()

    ribbon_lookup = {}
    for farm_name, years in farm_years.items():
        farm_id = get_farm_id_from_name(farm_name)
        if not farm_id:
            continue
        for year in sorted(years):
            res = supabase.table("ribbon_schedule").select("week_number, color_name") \
                .eq("farm_id", farm_id).eq("year", int(year)).eq("is_deleted", False).execute()
            for rb in (res.data or []):
                ribbon_lookup[(farm_name, int(year), int(rb["week_number"]))] = rb.get("color_name") or ""

    detail_rows = []
    for row in parsed_rows:
        offsets = _report_forecast_offsets_for_farm(row["farm"])
        for inclusive_offset in offsets:
            forecast_year, forecast_week = _shift_iso_week(row["cut_year"], row["cut_week"], inclusive_offset - 1)
            forecast_qty = _report_expected_harvest_from_cut(row["cut_qty"])
            detail_rows.append({
                "Năm TH dự báo": forecast_year,
                "Tuần TH dự báo": forecast_week,
                "Farm": row["farm"],
                "Năm cắt bắp": row["cut_year"],
                "Tuần cắt bắp": row["cut_week"],
                "Màu dây": ribbon_lookup.get((row["farm"], row["cut_year"], row["cut_week"]), ""),
                "Lô": row["lo"],
                "Base lot": row["base_lot_id"] if row["base_lot_id"] is not None else "",
                "Số cắt bắp": row["cut_qty"],
                "Dự kiến thu hoạch 97%": forecast_qty,
                "Cách dự báo": f"+{inclusive_offset}",
            })

    detail_rows.sort(key=lambda r: (
        int(r["Năm TH dự báo"]),
        int(r["Tuần TH dự báo"]),
        _report_farm_sort_key(r["Farm"]),
        str(r["Màu dây"]),
        str(r["Lô"]),
    ))
    farms = sorted({r["Farm"] for r in detail_rows if r["Farm"]}, key=_report_farm_sort_key)

    ws_summary = wb.active
    ws_summary.title = "Tổng hợp"
    summary_headers = ["Năm TH dự báo", "Tuần TH dự báo", "Tổng dự kiến", *farms, "Ghi chú"]
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    summary_map = {}
    for row in detail_rows:
        key = (row["Năm TH dự báo"], row["Tuần TH dự báo"])
        summary_map.setdefault(key, {"Tổng dự kiến": 0, **{farm: 0 for farm in farms}})
        summary_map[key]["Tổng dự kiến"] += row["Dự kiến thu hoạch 97%"]
        if row["Farm"] in farms:
            summary_map[key][row["Farm"]] += row["Dự kiến thu hoạch 97%"]

    for row_idx, (key, values) in enumerate(sorted(summary_map.items()), 2):
        forecast_year, forecast_week = key
        row_values = [forecast_year, forecast_week, values["Tổng dự kiến"]]
        row_values.extend(values.get(farm, 0) or "" for farm in farms)
        row_values.append("Đã nhân 97% từ số cắt bắp")
        for col_idx, value in enumerate(row_values, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_align
    ws_summary.freeze_panes = "A2"

    ws_detail = wb.create_sheet("Chi tiết nguồn")
    detail_headers = [
        "Năm TH dự báo", "Tuần TH dự báo", "Farm", "Năm cắt bắp", "Tuần cắt bắp",
        "Màu dây", "Lô", "Base lot", "Số cắt bắp", "Dự kiến thu hoạch 97%", "Cách dự báo"
    ]
    for col_idx, header in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    for row_idx, row in enumerate(detail_rows, 2):
        for col_idx, header in enumerate(detail_headers, 1):
            cell = ws_detail.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
            cell.border = thin_border
            cell.alignment = center_align
    ws_detail.freeze_panes = "A2"

    for ws in (ws_summary, ws_detail):
        for col_idx in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_len = max(len(str(ws.cell(row=row_idx, column=col_idx).value or "")) for row_idx in range(1, ws.max_row + 1))
            ws.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 28)
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                if row_idx > 1 and col_idx in (3, *range(4, 4 + len(farms))):
                    ws.cell(row=row_idx, column=col_idx).number_format = "#,##0"
        if ws.max_row >= 2:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=ws.max_row, column=col_idx).border = thin_border

    output = io.BytesIO(); wb.save(output); output.seek(0)
    return output.getvalue()

def generate_planting_excel(df_lots, df_seasons):
    """Tạo file Excel báo cáo Trồng mới, chia sheet theo năm."""
    wb = Workbook()
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    headers = ["Ngày trồng", "Lô", "Số lượng cây", "Loại trồng"]

    if df_lots.empty:
        ws = wb.active; ws.title = "Báo cáo Trồng mới"
        ws.cell(row=1, column=1, value="Chưa có dữ liệu Trồng mới.")
        output = io.BytesIO(); wb.save(output); output.seek(0)
        return output.getvalue()

    df_plot = df_lots[['ngay_trong', 'farm', 'lo', 'so_luong']].copy()
    df_plot.dropna(subset=['ngay_trong'], inplace=True)
    df_plot['ngay_trong'] = pd.to_datetime(df_plot['ngay_trong'])
    df_plot.sort_values(by='ngay_trong', inplace=True)

    # loai_trong already available in base_lots
    if 'loai_trong' in df_lots.columns:
        df_plot['loai_trong'] = df_lots['loai_trong']
    elif not df_seasons.empty and 'loai_trong' in df_seasons.columns and 'lo' in df_seasons.columns:
        seasons_map = df_seasons.drop_duplicates(subset=['farm', 'lo'])[['farm', 'lo', 'loai_trong']]
        df_plot = pd.merge(df_plot, seasons_map, on=['farm', 'lo'], how='left')
    else:
        df_plot['loai_trong'] = None
    df_plot['loai_trong'] = df_plot['loai_trong'].fillna('Trồng mới')
    df_plot['_year'] = df_plot['ngay_trong'].dt.year

    years = sorted(df_plot['_year'].unique())
    wb.remove(wb.active)

    for year in years:
        ws = wb.create_sheet(title=str(int(year)))
        # Header row
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header_title)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        df_yr = df_plot[df_plot['_year'] == year]
        for r_idx, row in df_yr.iterrows():
            ngay_str = row['ngay_trong'].strftime('%d/%m/%Y')
            loai = row.get('loai_trong', 'Trồng mới') or 'Trồng mới'
            ws.append([ngay_str, row['lo'], row['so_luong'], loai])
            current_row = ws.max_row
            for col_idx in range(1, 5):
                ws.cell(row=current_row, column=col_idx).border = thin_border

        # Auto column widths
        for col in ws.columns:
            max_length = 0
            column = [cell for cell in col]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(cell.value)
                except: pass
            ws.column_dimensions[get_column_letter(column[0].column)].width = max_length + 2

    output = io.BytesIO(); wb.save(output); output.seek(0)
    return output.getvalue()

def render_global_data_tab(c_farm):
    st.markdown("### 🌐 Bảng dữ liệu Toàn cục Farm")
    st.caption("Khám phá dữ liệu tổng quan bằng các Biểu đồ phân tích và Bộ lọc.")
    
    if _maybe_render_lot_cost_dialog(c_farm):
        st.stop()

    # Fetch all data
    df_lots_all = fetch_table_data("base_lots", c_farm)
    df_stg_all = fetch_table_data("stage_logs", c_farm)
    df_des_all = fetch_table_data("destruction_logs", c_farm)
    df_har_all = fetch_table_data("harvest_logs", c_farm)
    df_bsr_all = fetch_table_data("bsr_logs", c_farm)
    df_tree_inv_all = fetch_table_data("tree_inventory_logs", c_farm)
    df_seasons = fetch_table_data("seasons", c_farm)

    # ─── Phân loại Trồng mới vs Trồng dặm ───
    # loai_trong nằm trực tiếp trong base_lots (sau migration add_loai_trong_to_base_lots)
    # Trồng dặm KHÔNG phải đợt trồng độc lập → tách riêng khỏi forecast & bảng chi tiết
    if not df_lots_all.empty and 'loai_trong' not in df_lots_all.columns:
        df_lots_all['loai_trong'] = 'Trồng mới'  # fallback nếu cột chưa có
    
    df_lots_trong_moi = df_lots_all[df_lots_all['loai_trong'] == 'Trồng mới'] if not df_lots_all.empty else pd.DataFrame()
    df_lots_trong_dam = df_lots_all[df_lots_all['loai_trong'] == 'Trồng dặm'] if not df_lots_all.empty else pd.DataFrame()

    # Nút Xuất Báo cáo Excel 
    import base64
    st.markdown("""
    <style>
    .custom-dl-btn {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        font-weight: 600;
        padding: 0.35rem 0.75rem;
        border-radius: 0.5rem;
        line-height: 1.6;
        width: 100%;
        min-height: 64px;
        text-align: center;
        text-decoration: none !important;
        transition: opacity 0.2s ease, filter 0.2s ease;
        box-sizing: border-box;
    }
    .custom-dl-btn:hover {
        opacity: 0.85;
        filter: brightness(0.95);
    }

    .btn-forecast { background-color: #e3f2fd; color: #1565c0 !important; border: 1px solid #90caf9; }
    .btn-chich { background-color: #fff8e1; color: #f57f17 !important; border: 1px solid #ffe082; }
    .btn-cat   { background-color: #ffebee; color: #c62828 !important; border: 1px solid #ef9a9a; }
    .btn-trong { background-color: #e8f5e9; color: #2e7d32 !important; border: 1px solid #a5d6a7; }
    /* Popover buttons cho Admin/KD — match original colors */

    .st-key-pop_chich button {
        background-color: #fff8e1 !important; color: #f57f17 !important;
        border: 1px solid #ffe082 !important; font-weight: 600;
        min-height: 64px; border-radius: 0.5rem;
    }
    .st-key-pop_cat button {
        background-color: #ffebee !important; color: #c62828 !important;
        border: 1px solid #ef9a9a !important; font-weight: 600;
        min-height: 64px; border-radius: 0.5rem;
    }
    .st-key-pop_trong button {
        background-color: #e8f5e9 !important; color: #2e7d32 !important;
        border: 1px solid #a5d6a7 !important; font-weight: 600;
        min-height: 64px; border-radius: 0.5rem;
    }
    .st-key-pop_chich button:hover,
    .st-key-pop_cat button:hover,
    .st-key-pop_trong button:hover {
        opacity: 0.85 !important; filter: brightness(0.95) !important;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # ── Farm selector cho Admin/KD (dùng chung cho download buttons) ──
    is_multi_farm = c_farm in ["Admin", "Phòng Kinh doanh"]
    available_farms = sorted(df_lots_all["farm"].dropna().unique().tolist()) if is_multi_farm and not df_lots_all.empty else []

    def _filter_by_farm(df, farm_name):
        """Filter DataFrame theo farm. Nếu không phải multi-farm thì trả nguyên."""
        if not is_multi_farm or df.empty or "farm" not in df.columns:
            return df
        return df[df["farm"] == farm_name]

    def _gen_dl_link(data_bytes, filename, css_class, label):
        """Generate HTML download link."""
        b64 = base64.b64encode(data_bytes if isinstance(data_bytes, bytes) else data_bytes).decode()
        return f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}" class="custom-dl-btn {css_class}">{label}</a>'

    col_t1, col_t2, col_t3, col_t4, col_t5 = st.columns([1.8, 1, 1, 1, 1])
    report_generated_at = datetime.now()
    report_stamp = report_generated_at.strftime("%Y%m%d_%H%M%S")

    with col_t1:
        if st.button("🔄 Làm mới báo cáo", key="global_reports_refresh", use_container_width=True):
            st.cache_data.clear()
            st.rerun()

    if is_multi_farm and available_farms:
        # ── Admin/KD: mỗi nút download là popover chọn farm ──
        # Dùng session_state để đảm bảo radio value không bị reset khi download rerun
        for _k in ["pop_farm_chich", "pop_farm_cat", "pop_farm_trong"]:
            if _k not in st.session_state:
                st.session_state[_k] = available_farms[0]

        with col_t2:
            forecast_excel = generate_harvest_forecast_excel(df_lots_all, df_stg_all)
            fn = f"Bao_cao_du_bao_thu_hoach_tat_ca_farm_{report_stamp}.xlsx"
            href = _gen_dl_link(forecast_excel, fn, "btn-forecast", "Dự báo Thu hoạch")
            st.markdown(href, unsafe_allow_html=True)

        with col_t3:
            with st.popover("Báo cáo Chích bắp", use_container_width=True, key="pop_chich"):
                sel_chich = st.radio("Chọn Farm", available_farms, key="pop_farm_chich", horizontal=True)
                fl = _filter_by_farm(df_lots_all, sel_chich)
                fs = _filter_by_farm(df_stg_all, sel_chich)
                chich_excel = generate_chich_bap_excel(fl, fs)
                fn = f"Bao_cao_chich_bap_{sel_chich}_{report_stamp}.xlsx"
                st.download_button("⬇️ Tải về", data=chich_excel, file_name=fn, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

        with col_t4:
            with st.popover("Báo cáo Cắt bắp", use_container_width=True, key="pop_cat"):
                sel_cat = st.radio("Chọn Farm", available_farms, key="pop_farm_cat", horizontal=True)
                fl = _filter_by_farm(df_lots_all, sel_cat)
                fs = _filter_by_farm(df_stg_all, sel_cat)
                fd = _filter_by_farm(df_des_all, sel_cat)
                fh = _filter_by_farm(df_har_all, sel_cat)
                cut_excel = generate_cut_bap_excel(fl, fs, fd, fh)
                fn = f"Bao_cao_cat_bap_{sel_cat}_{report_stamp}.xlsx"
                st.download_button("⬇️ Tải về", data=cut_excel, file_name=fn, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

        with col_t5:
            with st.popover("Báo cáo Trồng mới", use_container_width=True, key="pop_trong"):
                sel_trong = st.radio("Chọn Farm", available_farms, key="pop_farm_trong", horizontal=True)
                fl = _filter_by_farm(df_lots_all, sel_trong)
                f_seasons = _filter_by_farm(df_seasons, sel_trong)
                plant_excel = generate_planting_excel(fl, f_seasons)
                fn = f"Bao_cao_trong_moi_{sel_trong}_{report_stamp}.xlsx"
                st.download_button("⬇️ Tải về", data=plant_excel, file_name=fn, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    else:
        # ── User thường: download trực tiếp (giữ nguyên) ──
        with col_t2:
            forecast_excel = generate_harvest_forecast_excel(df_lots_all, df_stg_all)
            fn = f"Bao_cao_du_bao_thu_hoach_{c_farm}_{report_stamp}.xlsx"
            href = _gen_dl_link(forecast_excel, fn, "btn-forecast", "Dự báo Thu hoạch")
            st.markdown(href, unsafe_allow_html=True)

        with col_t3:
            chich_excel = generate_chich_bap_excel(df_lots_all, df_stg_all)
            fn = f"Bao_cao_chich_bap_{c_farm}_{report_stamp}.xlsx"
            href = _gen_dl_link(chich_excel, fn, "btn-chich", "Báo cáo Chích bắp")
            st.markdown(href, unsafe_allow_html=True)

        with col_t4:
            cut_excel = generate_cut_bap_excel(df_lots_all, df_stg_all, df_des_all, df_har_all)
            fn = f"Bao_cao_cat_bap_{c_farm}_{report_stamp}.xlsx"
            href = _gen_dl_link(cut_excel, fn, "btn-cat", "Báo cáo Cắt bắp")
            st.markdown(href, unsafe_allow_html=True)

        with col_t5:
            plant_excel = generate_planting_excel(df_lots_all, df_seasons)
            fn = f"Bao_cao_trong_moi_{c_farm}_{report_stamp}.xlsx"
            href = _gen_dl_link(plant_excel, fn, "btn-trong", "Báo cáo Trồng mới")
            st.markdown(href, unsafe_allow_html=True)

    st.divider()

    # Filter helpers
    farms_all = ["Tất cả"] + list(df_lots_all["farm"].dropna().unique()) if not df_lots_all.empty else ["Tất cả"]
    teams_all = ["Tất cả"] + list(df_lots_all["team"].dropna().unique()) if not df_lots_all.empty else ["Tất cả"]
    lots_all = ["Tất cả"] + list(df_lots_all["lo"].dropna().unique()) if not df_lots_all.empty else ["Tất cả"]
    seasons_all = ["Tất cả"] + list(df_seasons["vu"].dropna().unique()) if not df_seasons.empty else ["Tất cả"]

    def get_dynamic_lot_options(df_lots, df_seasons, f_farm, f_team, f_vu):
        df_filtered = df_lots.copy()
        if not df_filtered.empty:
            if f_farm != "Tất cả" and "farm" in df_filtered.columns:
                df_filtered = df_filtered[df_filtered["farm"] == f_farm]
            if f_team != "Tất cả" and "team" in df_filtered.columns:
                df_filtered = df_filtered[df_filtered["team"] == f_team]
            if f_vu != "Tất cả" and not df_seasons.empty:
                valid_lots = df_seasons[df_seasons["vu"] == f_vu]["lo"].tolist()
                # we don't have lot_id in df_seasons directly linked sometimes, so match by lot ID
                # Actually, df_seasons has 'lo' or 'lot_id', but in app.py we saw: `valid_lots = df_seasons[df_seasons["vu"] == f_vu]["lo"].tolist()`. Let's match df_lots_all's 'lo' with that. 
                # Oh wait, earlier the code did: df_filtered["lot_id"].isin(valid_lots_season) but valid_lots_season was a list of "lo". Wait, that might be a bug in the old code. base_lots 'lo' should match seasons 'lo'.
                if "lo" in df_filtered.columns:
                    df_filtered = df_filtered[df_filtered["lo"].isin(valid_lots)]
        return ["Tất cả"] + list(df_filtered["lo"].dropna().unique()) if not df_filtered.empty else ["Tất cả"]

    def _int_id_set(series):
        if series is None:
            return set()
        vals = pd.to_numeric(series.dropna(), errors="coerce").dropna()
        return {int(v) for v in vals}

    def _filter_by_base_lot_ids(df, valid_ids, name):
        if not valid_ids or df.empty:
            return df
        if name == "lots" and "id" in df.columns:
            ids = pd.to_numeric(df["id"], errors="coerce")
            return df[ids.isin(valid_ids)]
        if "base_lot_id" in df.columns:
            ids = pd.to_numeric(df["base_lot_id"], errors="coerce")
            return df[ids.isin(valid_ids)]
        return df

    def apply_filters_local(f_farm, f_vu, f_team, f_lot, f_date, df_dict):
        """Helper function to apply filters locally to a set of dataframes"""
        res = {}
        # Apply Season format
        valid_lots_season = None
        valid_base_lot_ids_season = None
        season_window_by_base_lot = {}
        if f_vu != "Tất cả" and not df_seasons.empty:
            season_scope = df_seasons[df_seasons["vu"] == f_vu]
            valid_lots_season = season_scope["lo"].dropna().astype(str).tolist() if "lo" in season_scope.columns else []
            valid_base_lot_ids_season = _int_id_set(season_scope["base_lot_id"]) if "base_lot_id" in season_scope.columns else set()
            if "base_lot_id" in season_scope.columns and "ngay_bat_dau" in season_scope.columns:
                for _, season_row in season_scope.iterrows():
                    blid = season_row.get("base_lot_id")
                    blid_int = _safe_int_id(blid)
                    if blid_int is None:
                        continue
                    start_dt = pd.to_datetime(season_row.get("ngay_bat_dau"), errors="coerce")
                    next_start = None
                    try:
                        next_start = _map_next_season.get((blid_int, f_vu))
                    except NameError:
                        next_start = None
                    season_window_by_base_lot[blid_int] = {
                        "start": start_dt if pd.notna(start_dt) else None,
                        "next_start": next_start,
                    }

        for name, df in df_dict.items():
            if df.empty:
                res[name] = df
                continue
            
            df_filtered = df.copy()
            if f_farm != "Tất cả" and "farm" in df_filtered.columns:
                df_filtered = df_filtered[df_filtered["farm"] == f_farm]
            if valid_base_lot_ids_season is not None:
                before_season_filter = len(df_filtered)
                df_filtered = _filter_by_base_lot_ids(df_filtered, valid_base_lot_ids_season, name)
                if len(df_filtered) == before_season_filter and valid_lots_season is not None and "lot_id" in df_filtered.columns:
                    df_filtered = df_filtered[df_filtered["lot_id"].isin(valid_lots_season)]
            if (
                f_vu != "Tất cả"
                and valid_base_lot_ids_season
                and season_window_by_base_lot
                and "base_lot_id" in df_filtered.columns
                and name in ["stg", "har", "des"]
            ):
                date_col_map = {
                    "stg": "ngay_thuc_hien",
                    "har": "ngay_thu_hoach",
                    "des": "ngay_xuat_huy",
                }
                date_col = date_col_map.get(name)
                if date_col in df_filtered.columns:
                    ids = pd.to_numeric(df_filtered["base_lot_id"], errors="coerce")
                    dates = pd.to_datetime(df_filtered[date_col], errors="coerce")
                    keep = pd.Series(False, index=df_filtered.index)
                    for blid, window in season_window_by_base_lot.items():
                        start_dt = window.get("start")
                        next_start = window.get("next_start")
                        id_mask = ids == blid
                        if start_dt is None:
                            date_mask = id_mask
                        else:
                            min_dt = start_dt + pd.Timedelta(weeks=HARVEST_MIN_GROWTH_WEEKS) if name == "har" and f_vu != "F0" else start_dt
                            date_mask = id_mask & (dates >= min_dt)
                        if next_start is not None:
                            upper_dt = pd.Timestamp(next_start)
                            if name == "har":
                                upper_dt = upper_dt + pd.Timedelta(weeks=HARVEST_MIN_GROWTH_WEEKS)
                            date_mask = date_mask & (dates < upper_dt)
                        keep = keep | date_mask
                    df_filtered = df_filtered[keep]
            elif (
                f_vu != "Tất cả"
                and valid_lots_season
                and season_window_by_base_lot
                and "lot_id" in df_filtered.columns
                and name in ["inv"]
            ):
                # Inventory is lot-level, not batch-level. Use all matching season windows
                # for the lot as a best-effort date guard.
                date_col = "ngay_kiem_ke" if "ngay_kiem_ke" in df_filtered.columns else None
                if date_col:
                    dates = pd.to_datetime(df_filtered[date_col], errors="coerce")
                    keep = pd.Series(False, index=df_filtered.index)
                    for _, season_row in season_scope.iterrows():
                        lo_val = season_row.get("lo")
                        if pd.isna(lo_val):
                            continue
                        start_dt = pd.to_datetime(season_row.get("ngay_bat_dau"), errors="coerce")
                        lot_mask = df_filtered["lot_id"].astype(str) == str(lo_val)
                        date_mask = lot_mask
                        if pd.notna(start_dt):
                            date_mask = date_mask & (dates >= start_dt)
                        blid = season_row.get("base_lot_id")
                        next_start = None
                        blid_int = _safe_int_id(blid)
                        if blid_int is not None:
                            try:
                                next_start = _map_next_season.get((blid_int, f_vu))
                            except NameError:
                                next_start = None
                        if next_start is not None:
                            date_mask = date_mask & (dates < pd.Timestamp(next_start))
                        keep = keep | date_mask
                    df_filtered = df_filtered[keep]
            if f_team != "Tất cả" and "team" in df_filtered.columns:
                df_filtered = df_filtered[df_filtered["team"] == f_team]
            if f_lot != "Tất cả" and "lot_id" in df_filtered.columns:
                valid_ids = df_lots_all[df_lots_all["lo"] == f_lot]["lot_id"].tolist() if not df_lots_all.empty else []
                df_filtered = df_filtered[df_filtered["lot_id"].isin(valid_ids)]
            
            # Apply Date Range filter
            if f_date and len(f_date) == 2:
                start_date, end_date = f_date
                date_col = None
                if name == "lots" and "ngay_trong" in df_filtered.columns:
                    date_col = "ngay_trong"
                elif name == "stg" and "ngay_thuc_hien" in df_filtered.columns:
                    date_col = "ngay_thuc_hien"
                elif name == "des" and "ngay_xuat_huy" in df_filtered.columns:
                    date_col = "ngay_xuat_huy"
                elif name == "har" and "ngay_thu_hoach" in df_filtered.columns:
                    date_col = "ngay_thu_hoach"
                elif name == "inv" and "ngay_kiem_ke" in df_filtered.columns:
                    date_col = "ngay_kiem_ke"
                
                if date_col:
                    # Convert to datetime and then normalize to date for comparison
                    df_filtered[date_col] = pd.to_datetime(df_filtered[date_col])
                    df_filtered = df_filtered[
                        (df_filtered[date_col].dt.date >= start_date) & 
                        (df_filtered[date_col].dt.date <= end_date)
                    ]
            
            res[name] = df_filtered
        return res

    def show_season_info(f_vu, f_lot):
        """Hiển thị thông tin khoảng thời gian vụ khi chọn cả Vụ và Lô cụ thể."""
        if f_vu == "Tất cả" or f_lot == "Tất cả" or df_seasons.empty:
            return
        matched = df_seasons[(df_seasons["vu"] == f_vu) & (df_seasons["lo"] == f_lot)]
        if matched.empty:
            return
        row = matched.iloc[0]
        start = pd.to_datetime(row.get("ngay_bat_dau"))
        end_actual = row.get("ngay_ket_thuc_thuc_te")
        end_planned = row.get("ngay_ket_thuc_du_kien")
        end = pd.to_datetime(end_actual) if pd.notna(end_actual) else pd.to_datetime(end_planned) if pd.notna(end_planned) else None
        loai = row.get("loai_trong", "")
        
        start_str = start.strftime("%d/%m/%Y") if pd.notna(start) else "—"
        if end:
            end_label = "Kết thúc" if pd.notna(end_actual) else "Dự kiến KT"
            end_str = end.strftime("%d/%m/%Y")
        else:
            end_label, end_str = "Kết thúc", "Chưa xác định"
        
        loai_str = f" · Loại: **{loai}**" if loai else ""
        st.info(f"📅 Vụ **{f_vu}** — Lô **{f_lot}**{loai_str} · Bắt đầu: **{start_str}** · {end_label}: **{end_str}**", icon="ℹ️")

    def render_chart_filters(prefix: str, include_date: bool = False, use_dynamic_lots: bool = True):
        """
        Render bộ lọc chuẩn (Farm, Vụ, Đội, Lô, [Date]) và trả về giá trị filter.
        Args:
            prefix: tiền tố key duy nhất cho session_state (vd: "dt", "ek", "lc")
            include_date: hiển thị ô chọn khoảng thời gian
            use_dynamic_lots: True = lọc lô theo farm/team/vụ, False = hiện tất cả lô
        Returns: (farm, vu, team, lot, date_or_None)
        """
        if c_farm in ["Admin", "Phòng Kinh doanh"]:
            cols = st.columns([1, 1, 1, 1, 1.5] if include_date else [1, 1, 1, 1])
            col_idx = 0
            with cols[col_idx]:
                f_farm = st.selectbox("Lọc theo Farm", options=farms_all, key=f"{prefix}_farm")
            col_idx += 1
        else:
            f_farm = c_farm
            cols = st.columns([1, 1, 1, 1.5] if include_date else [1, 1, 1])
            col_idx = 0

        with cols[col_idx]:
            f_vu = st.selectbox("Lọc theo Vụ", options=seasons_all, key=f"{prefix}_vu")
        col_idx += 1
        with cols[col_idx]:
            f_team = st.selectbox("Lọc theo Đội", options=teams_all, key=f"{prefix}_team")
        col_idx += 1
        with cols[col_idx]:
            if use_dynamic_lots:
                lot_opts = get_dynamic_lot_options(df_lots_all, df_seasons, f_farm, f_team, f_vu)
            else:
                lot_opts = lots_all
            f_lot = st.selectbox("Lọc theo Lô", options=lot_opts, key=f"{prefix}_lot")

        f_date = None
        if include_date:
            col_idx += 1
            with cols[col_idx]:
                f_date = st.date_input("Khoảng thời gian", value=(), key=f"{prefix}_date")

        return f_farm, f_vu, f_team, f_lot, f_date

    def get_filtered_dfs(farm, vu, team, lot, date_range, data_dict):
        """Show season info và apply bộ lọc chuẩn. Trả về dict filtered DataFrames."""
        show_season_info(vu, lot)
        return apply_filters_local(farm, vu, team, lot, date_range, data_dict)

    st.divider()

    # ═══════════════════════════════════════════════════════════════════
    # 🗺️ SHARED MAP CONSTANTS & HELPERS  (dùng cho cả Farm 157 & 195)
    # ═══════════════════════════════════════════════════════════════════
    _MAP_STAGE_COLORS = {
        "Đang sinh trưởng": "#00b894",
        "Chích bắp": "#fdcb6e",
        "Cắt bắp": "#e17055",
        "Thu hoạch": "#0984e3",
    }
    _MAP_STAGE_COLORS_JSON = json.dumps(_MAP_STAGE_COLORS, ensure_ascii=False)
    _MAP_DEFAULT_COLOR = "#636e72"

    def _map_point_in_polygon(px, py, pts):
        """Ray casting algorithm."""
        n = len(pts)
        inside = False
        j = n - 1
        for i in range(n):
            xi, yi = pts[i]
            xj, yj = pts[j]
            if ((yi > py) != (yj > py)) and (px < (xj - xi) * (py - yi) / (yj - yi) + xi):
                inside = not inside
            j = i
        return inside

    def _map_geometric_centroid(pts):
        """Shoelace-based area centroid — trọng tâm diện tích chính xác."""
        n = len(pts)
        if n < 3:
            return sum(p[0] for p in pts)/n, sum(p[1] for p in pts)/n
        signed_area = 0
        cx = cy = 0
        for i in range(n):
            x0, y0 = pts[i]
            x1, y1 = pts[(i + 1) % n]
            cross = x0 * y1 - x1 * y0
            signed_area += cross
            cx += (x0 + x1) * cross
            cy += (y0 + y1) * cross
        signed_area *= 0.5
        if abs(signed_area) < 1e-10:
            return sum(p[0] for p in pts)/n, sum(p[1] for p in pts)/n
        cx /= (6 * signed_area)
        cy /= (6 * signed_area)
        return cx, cy

    def _map_pole_of_inaccessibility(pts):
        """Simple grid-search for the 'visual center'."""
        xs = [p[0] for p in pts]
        ys = [p[1] for p in pts]
        min_x, max_x = min(xs), max(xs)
        min_y, max_y = min(ys), max(ys)
        best = (sum(xs)/len(xs), sum(ys)/len(ys))
        best_dist = 0
        steps = 12
        for xi in range(steps + 1):
            for yi in range(steps + 1):
                px = min_x + (max_x - min_x) * xi / steps
                py = min_y + (max_y - min_y) * yi / steps
                if _map_point_in_polygon(px, py, pts):
                    min_edge_dist = float('inf')
                    n = len(pts)
                    for i in range(n):
                        x1, y1 = pts[i]
                        x2, y2 = pts[(i + 1) % n]
                        dx, dy = x2 - x1, y2 - y1
                        if dx == 0 and dy == 0:
                            d = ((px - x1)**2 + (py - y1)**2)**0.5
                        else:
                            t = max(0, min(1, ((px - x1)*dx + (py - y1)*dy) / (dx*dx + dy*dy)))
                            d = ((px - (x1 + t*dx))**2 + (py - (y1 + t*dy))**2)**0.5
                        min_edge_dist = min(min_edge_dist, d)
                    if min_edge_dist > best_dist:
                        best_dist = min_edge_dist
                        best = (px, py)
        return best

    def _map_best_label_pos(pts):
        """Hybrid: geometric centroid → fallback pole of inaccessibility."""
        cx, cy = _map_geometric_centroid(pts)
        if _map_point_in_polygon(cx, cy, pts):
            return cx, cy
        return _map_pole_of_inaccessibility(pts)

    # ─── Shared: compute_batch_stats (dùng cho Map, Table và chart theo vụ) ───
    def compute_batch_stats(lo_name, base_lot_id, vu="F0", season_start=None,
                            season_end=None, next_season_start=None, next_vu_producing=False,
                            stage_df=None, harvest_df=None):
        """Tính (giai_doan, so_chich, so_cat, so_thu) cho 1 batch/season."""
        so_chich, so_cat, so_thu = 0, 0, 0
        season_start_ts = pd.Timestamp(season_start) if season_start is not None else None
        stg_source = stage_df if stage_df is not None else df_stg_all
        har_source = harvest_df if harvest_df is not None else df_har_all
        if not stg_source.empty and "lo" in stg_source.columns and "base_lot_id" in stg_source.columns:
            stg = stg_source[(stg_source["lo"] == lo_name) & (stg_source["base_lot_id"] == base_lot_id)]
            if season_start_ts is not None and not stg.empty and "ngay_thuc_hien" in stg.columns:
                stg_dates = pd.to_datetime(stg["ngay_thuc_hien"], errors="coerce")
                stg = stg[stg_dates >= season_start_ts]
            if next_season_start is not None and next_vu_producing and not stg.empty and "ngay_thuc_hien" in stg.columns:
                stg_dates = pd.to_datetime(stg["ngay_thuc_hien"], errors="coerce")
                stg = stg[stg_dates < pd.Timestamp(next_season_start)]
            if not stg.empty:
                c = stg[stg["giai_doan"] == "Chích bắp"]
                k = stg[stg["giai_doan"] == "Cắt bắp"]
                so_chich = int(c["so_luong"].sum()) if not c.empty else 0
                so_cat = int(k["so_luong"].sum()) if not k.empty else 0
        if not har_source.empty and "lo" in har_source.columns and "base_lot_id" in har_source.columns:
            har = har_source[(har_source["lo"] == lo_name) & (har_source["base_lot_id"] == base_lot_id)]
            if not har.empty and "ngay_thu_hoach" in har.columns:
                har_dates = pd.to_datetime(har["ngay_thu_hoach"], errors="coerce")
                if season_start_ts is not None:
                    har = har[har_dates >= season_start_ts]
                    har_dates = pd.to_datetime(har["ngay_thu_hoach"], errors="coerce")
                if vu != "F0" and season_start is not None:
                    harvest_min = pd.Timestamp(season_start) + pd.Timedelta(weeks=HARVEST_MIN_GROWTH_WEEKS)
                    har = har[har_dates >= harvest_min]
                    har_dates = pd.to_datetime(har["ngay_thu_hoach"], errors="coerce")
                if next_season_start is not None and next_vu_producing:
                    harvest_upper = pd.Timestamp(next_season_start) + pd.Timedelta(weeks=HARVEST_MIN_GROWTH_WEEKS)
                    har = har[har_dates < harvest_upper]
            so_thu = int(har["so_luong"].sum()) if not har.empty else 0
        if vu != "F0" and so_chich == 0:
            so_thu = 0
        gd = "Đang sinh trưởng"
        if so_thu > 0: gd = "Thu hoạch"
        elif so_cat > 0: gd = "Cắt bắp"
        elif so_chich > 0: gd = "Chích bắp"
        return gd, so_chich, so_cat, so_thu

    _map_next_season, _map_next_producing = build_next_season_maps(df_seasons, df_stg_all)
    batch_label_map = build_batch_label_map(df_lots_trong_moi)

    def _map_component_initial_height(img_w, img_h):
        aspect = img_h / img_w if img_w else 0.5625
        return min(max(560, int(aspect * 1700) + 110), 1500)

    def _compute_map_area_metrics(lot_info_map, total_farm_area=None):
        """Tính panel diện tích bản đồ theo mốc lũy kế, dùng chung cho mọi farm."""
        if total_farm_area is None:
            total_farm_area = sum(
                max(float(info.get("area_ha") or 0), float(info.get("dien_tich_trong") or 0))
                for info in lot_info_map.values()
            )

        metrics = {
            "total_farm_area": float(total_farm_area or 0),
            "area_planted": 0.0,
            "area_growing": 0.0,
            "area_chich": 0.0,
            "area_cat": 0.0,
            "area_harvest": 0.0,
        }
        counted_blids = set()
        for info in lot_info_map.values():
            for batch in info.get("batches", []):
                blid = batch.get("base_lot_id")
                if blid is None or blid in counted_blids:
                    continue
                counted_blids.add(blid)
                dt = float(batch.get("dien_tich_trong") or 0)
                total_cay = int(batch.get("so_cay") or 0)
                if dt <= 0 or total_cay <= 0:
                    continue

                n_harvest = min(max(int(batch.get("thu") or 0), 0), total_cay)
                n_cat = min(max(int(batch.get("cat") or 0), n_harvest), total_cay)
                n_chich = min(max(int(batch.get("chich") or 0), n_cat), total_cay)
                ratio = dt / total_cay

                metrics["area_planted"] += dt
                metrics["area_harvest"] += n_harvest * ratio
                metrics["area_cat"] += n_cat * ratio
                metrics["area_chich"] += n_chich * ratio
                metrics["area_growing"] += max(total_cay - n_chich, 0) * ratio
        return metrics

    def _build_map_info_panel_html(metrics):
        info_panel_html = ""
        for label, value, color in [
            ("Tổng DT lô", f"{metrics['total_farm_area']:.2f} ha", "#94a3b8"),
            ("Đã trồng", f"{metrics['area_planted']:.2f} ha", "#a8e6cf"),
            ("Sinh trưởng", f"{metrics['area_growing']:.2f} ha", "#00b894"),
            ("Chích bắp", f"{metrics['area_chich']:.2f} ha", "#fdcb6e"),
            ("Cắt bắp", f"{metrics['area_cat']:.2f} ha", "#e17055"),
            ("Thu hoạch", f"{metrics['area_harvest']:.2f} ha", "#0984e3"),
        ]:
            info_panel_html += f'<div class="info-row"><span class="info-label">{label}</span><span class="info-value" style="color:{color}">{value}</span></div>'
        return info_panel_html

    def _render_generic_farm_map(farm_name, polygon_filename, default_width, default_height, map_zoom=None):
        polygon_path = os.path.join(os.path.dirname(__file__), polygon_filename)
        if not os.path.exists(polygon_path) or c_farm not in [farm_name, "Admin", "Phòng Kinh doanh"]:
            return

        st.markdown(f"#### 🗺️ Bản đồ {farm_name}")
        st.caption("Di chuột vào từng lô để xem thông tin chi tiết. Màu sắc thể hiện giai đoạn hiện tại.")

        with open(polygon_path, "r", encoding="utf-8") as f:
            polygon_data = json.load(f)

        dim_lo_area_map = {}
        total_farm_area = 0.0
        try:
            dim_lo_res = supabase.table("dim_lo").select("lo_name, lo_code, area_ha, dim_farm!inner(farm_name)").eq("is_active", True).eq("dim_farm.farm_name", farm_name).execute()
            if dim_lo_res.data:
                for dl in dim_lo_res.data:
                    lo_name_db = dl.get("lo_name")
                    lo_code = dl.get("lo_code")
                    area_ha = dl.get("area_ha")
                    if area_ha is not None:
                        area_val = float(area_ha)
                        total_farm_area += area_val
                        if lo_code:
                            dim_lo_area_map[lo_code] = area_val
                        if lo_name_db:
                            dim_lo_area_map[lo_name_db] = area_val
        except Exception:
            pass

        lot_info_map = {}
        df_lots_farm = df_lots_trong_moi[df_lots_trong_moi["farm"] == farm_name] if not df_lots_trong_moi.empty and "farm" in df_lots_trong_moi.columns else df_lots_trong_moi
        df_seasons_farm = df_seasons[df_seasons["farm"] == farm_name] if not df_seasons.empty and "farm" in df_seasons.columns else df_seasons
        polygon_lots = [lot["name"] for lot in polygon_data.get("lots", [])]

        if not df_seasons_farm.empty and not df_lots_farm.empty and "lo" in df_lots_farm.columns:
            for lo_name in df_lots_farm[df_lots_farm["lo"].isin(polygon_lots)]["lo"].dropna().unique():
                lo_seasons = df_seasons_farm[
                    (df_seasons_farm["lo"] == lo_name) & (df_seasons_farm["loai_trong"] != "Trồng dặm")
                ].sort_values("ngay_bat_dau", ascending=False)
                if lo_seasons.empty:
                    continue

                batches = []
                total_batches = len(lo_seasons)
                for _, s_row in lo_seasons.iterrows():
                    vu = s_row.get("vu", "?")
                    ngay_bd_raw = s_row.get("ngay_bat_dau")
                    ngay_bd = str(ngay_bd_raw)[:10] if ngay_bd_raw is not None else ""
                    blid = s_row.get("base_lot_id")
                    blid_int = _safe_int_id(blid)
                    season_start = None
                    if ngay_bd_raw is not None:
                        try:
                            season_start = pd.Timestamp(ngay_bd_raw).date() if not isinstance(ngay_bd_raw, date) else ngay_bd_raw
                        except Exception:
                            pass

                    if blid_int is not None and not df_lots_farm.empty and "id" in df_lots_farm.columns:
                        lot_ids = pd.to_numeric(df_lots_farm["id"], errors="coerce")
                        batch_lot = df_lots_farm[lot_ids == blid_int]
                        so_cay = int(batch_lot["so_luong"].sum()) if not batch_lot.empty else 0
                        dt_trong = float(batch_lot["dien_tich_trong"].dropna().iloc[0]) if not batch_lot.empty and "dien_tich_trong" in batch_lot.columns and not batch_lot["dien_tich_trong"].dropna().empty else 0.0
                    else:
                        so_cay = 0
                        dt_trong = 0.0

                    if blid_int is not None:
                        next_s = _map_next_season.get((blid_int, vu))
                        next_prod = (blid_int, vu) in _map_next_producing if next_s else False
                        next_s_date = next_s.date() if next_s is not None else None
                        gd, chich, cat, thu = compute_batch_stats(
                            lo_name, blid_int, vu=vu, season_start=season_start,
                            next_season_start=next_s_date, next_vu_producing=next_prod
                        )
                    else:
                        gd, chich, cat, thu = "Đang sinh trưởng", 0, 0, 0

                    display_label = batch_label_map.get(blid_int, batch_label_map.get(blid, lo_name)) if blid_int is not None else lo_name
                    is_multi = total_batches > 1
                    dot_num = 0
                    if is_multi and "đợt " in display_label:
                        dot_raw = display_label.split("đợt ")[-1].rstrip(")")
                        dot_num = int(dot_raw) if dot_raw.isdigit() else 0
                    batches.append({
                        "base_lot_id": blid_int,
                        "vu": vu, "ngay_bd": ngay_bd, "so_cay": so_cay,
                        "dien_tich_trong": dt_trong, "gd": gd,
                        "chich": chich, "cat": cat, "thu": thu,
                        "dot": dot_num, "multi": is_multi,
                    })

                if not batches:
                    continue

                dominant = max(batches, key=lambda b: b["so_cay"])
                dt_trong_total = 0.0
                if not df_lots_farm.empty and "dien_tich_trong" in df_lots_farm.columns:
                    lo_batches_f0 = df_lots_farm[
                        (df_lots_farm["lo"] == lo_name) & (df_lots_farm["loai_trong"] != "Trồng dặm")
                    ]
                    dt_vals = lo_batches_f0["dien_tich_trong"].dropna()
                    if not dt_vals.empty:
                        dt_trong_total = float(dt_vals.sum())

                area_ha = dim_lo_area_map.get(lo_name)
                if area_ha is not None:
                    area_limit = float(area_ha)
                    batch_dt_sum = sum(float(b.get("dien_tich_trong") or 0) for b in batches if b.get("base_lot_id") is not None)
                    missing_dt_batches = [
                        b for b in batches
                        if b.get("base_lot_id") is not None
                        and float(b.get("dien_tich_trong") or 0) <= 0
                        and int(b.get("so_cay") or 0) > 0
                    ]
                    if missing_dt_batches:
                        for b in missing_dt_batches:
                            b["dien_tich_trong"] = int(b.get("so_cay") or 0) / PLANTING_DENSITY_TREES_PER_HA
                        batch_dt_sum = sum(float(b.get("dien_tich_trong") or 0) for b in batches if b.get("base_lot_id") is not None)

                    if dt_trong_total > 0:
                        dt_trong_total = min(dt_trong_total, area_limit)
                    if batch_dt_sum > 0:
                        dt_trong_total = min(max(dt_trong_total, batch_dt_sum), area_limit)
                    if batch_dt_sum > area_limit:
                        area_scale = area_limit / batch_dt_sum
                        for b in batches:
                            b["dien_tich_trong"] = float(b.get("dien_tich_trong") or 0) * area_scale
                        dt_trong_total = area_limit
                elif dt_trong_total > 0:
                    # Some Farm 126 lots are missing dim_lo.area_ha. Use planted area as
                    # a conservative display fallback so total lot area never appears
                    # smaller than planted area.
                    area_ha = dt_trong_total

                lot_info_map[lo_name] = {
                    "dominant_gd": dominant["gd"],
                    "area_ha": area_ha,
                    "dien_tich_trong": dt_trong_total,
                    "total_cay": sum(b["so_cay"] for b in batches if b["vu"] == "F0"),
                    "batches": batches,
                }

        for lot in polygon_data.get("lots", []):
            name = lot["name"]
            if name not in lot_info_map and name in dim_lo_area_map:
                lot_info_map[name] = {
                    "dominant_gd": "Chưa có dữ liệu",
                    "area_ha": dim_lo_area_map.get(name),
                    "dien_tich_trong": 0.0,
                    "total_cay": 0,
                    "batches": [],
                }

        stage_colors = _MAP_STAGE_COLORS
        default_color = _MAP_DEFAULT_COLOR
        img_w = polygon_data.get("image_width", default_width)
        img_h = polygon_data.get("image_height", default_height)

        svg_polygons = ""
        for lot in polygon_data.get("lots", []):
            name = lot["name"]
            points_str = " ".join(f'{p["x"]},{p["y"]}' for p in lot["points"])
            info = lot_info_map.get(name, {})
            giai_doan = info.get("dominant_gd", "Chưa có dữ liệu")
            fill = stage_colors.get(giai_doan, default_color)
            area_ha = info.get("area_ha")
            area_ha_str = f'{area_ha:.2f} ha' if area_ha else "—"
            dt_trong = info.get("dien_tich_trong", 0)
            dt_trong_str = f'{dt_trong:.2f} ha' if dt_trong > 0 else "—"
            total_cay = info.get("total_cay", 0)
            batches_json = json.dumps(info.get("batches", []), ensure_ascii=False).replace('"', '&quot;')
            cost_farm = html.escape(farm_name, quote=True)
            cost_lot = html.escape(name, quote=True)
            poly_pts = [(p["x"], p["y"]) for p in lot["points"]]
            cx, cy = _map_best_label_pos(poly_pts)
            svg_polygons += f'''
            <polygon class="lot-poly" points="{points_str}" fill="{fill}"
                data-name="{name}" data-area-ha="{area_ha_str}" data-dt-trong="{dt_trong_str}" data-total="{total_cay}"
                data-gd="{giai_doan}" data-batches="{batches_json}" data-cost-farm="{cost_farm}" data-cost-lot="{cost_lot}" />
            <text x="{cx:.0f}" y="{cy:.0f}" class="lot-label">{name}</text>
            '''

        legend_html = ""
        for stage, color in stage_colors.items():
            legend_html += f'<span class="legend-item"><span class="legend-dot" style="background:{color}"></span>{stage}</span>'
        legend_html += f'<span class="legend-item"><span class="legend-dot" style="background:{default_color}"></span>Chưa có dữ liệu</span>'

        info_panel_html = _build_map_info_panel_html(
            _compute_map_area_metrics(lot_info_map, total_farm_area=total_farm_area or None)
        )

        from map_template import build_farm_map_html
        html_content = build_farm_map_html(
            svg_polygons=svg_polygons,
            legend_html=legend_html,
            info_panel_html=info_panel_html,
            img_w=img_w,
            img_h=img_h,
            stage_colors_json=_MAP_STAGE_COLORS_JSON,
            map_zoom=map_zoom,
            map_id=farm_name,
        )

        map_event = farm_map_component(
            html=html_content,
            initial_height=_map_component_initial_height(img_w, img_h),
            img_w=img_w,
            img_h=img_h,
            key=f"farm_map_component_{farm_name}",
            default=None,
        )
        _handle_map_cost_event(map_event, rerun_on_open=True)

    for _farm_name, _polygon_file, _default_w, _default_h, _map_zoom in [
        ("Farm 126", "farm_126_polygons.json", 2382, 1684, 1.0),
        ("Farm 157", "farm_157_polygons.json", 4000, 2250, None),
        ("Farm 195", "farm_195_polygons.json", 1683, 1190, None),
    ]:
        _render_generic_farm_map(
            _farm_name,
            _polygon_file,
            _default_w,
            _default_h,
            map_zoom=_map_zoom,
        )

    st.divider()

    # --- BẢNG CHI TIẾT THÔNG TIN CÁC LÔ ---
    # KG dự toán: sử dụng get_kg_per_tree(vu) — F0=15, Fn=18
    st.markdown("#### 📋 Bảng chi tiết thông tin các lô (Theo Vụ)")
    st.caption("Xem thông tin chi tiết từng lô phân loại theo vụ (Season). Các cột dữ liệu dự toán và thực tế được tính toán trong phạm vi khoảng thời gian của mục tiêu.")
    st.caption(f"📉 Hao hụt: Trồng → Chích bắp **{LOSS_RATE_TO_CHICH*100:.0f}%** · Chích bắp → Thu hoạch **{LOSS_RATE_TO_CHICH*100:.0f}%** · Tổng **{LOSS_RATE_TO_THU*100:.0f}%** · Sản lượng: **F0 = {KG_PER_TREE_F0} kg/buồng**, **Fn = {KG_PER_TREE_FN} kg/buồng**")

    dt_farm, dt_vu, dt_team, dt_lot, _ = render_chart_filters("dt")
    dt_season_status = st.radio(
        "Trạng thái vụ",
        ["Chưa kết thúc vụ", "Tất cả"],
        index=0,
        horizontal=True,
        key="dt_season_status"
    )

    df_dt_seasons = df_seasons.copy()
    if not df_dt_seasons.empty:
        # Loại trồng dặm khỏi bảng chi tiết (trồng dặm hiện ở bảng riêng bên dưới)
        if 'loai_trong' in df_dt_seasons.columns:
            df_dt_seasons = df_dt_seasons[df_dt_seasons['loai_trong'] != 'Trồng dặm']
        # Filter vụ đã kết thúc nếu chọn "Chưa kết thúc vụ"
        if dt_season_status == "Chưa kết thúc vụ" and "ngay_ket_thuc_thuc_te" in df_dt_seasons.columns:
            df_dt_seasons = df_dt_seasons[df_dt_seasons["ngay_ket_thuc_thuc_te"].isna()]
        if dt_farm != "Tất cả" and "farm" in df_dt_seasons.columns:
            df_dt_seasons = df_dt_seasons[df_dt_seasons["farm"] == dt_farm]
        if dt_vu != "Tất cả" and "vu" in df_dt_seasons.columns:
            df_dt_seasons = df_dt_seasons[df_dt_seasons["vu"] == dt_vu]
        if dt_team != "Tất cả" and "team" in df_dt_seasons.columns:
            df_dt_seasons = df_dt_seasons[df_dt_seasons["team"] == dt_team]
        if dt_lot != "Tất cả" and "lo" in df_dt_seasons.columns:
            df_dt_seasons = df_dt_seasons[df_dt_seasons["lo"] == dt_lot]
        
        # ── Merge diện tích trồng thực tế từ base_lots ──
        # dien_tich_trong (per-batch) thay thế area_ha (per-lot max)
        if not df_lots_trong_moi.empty and "id" in df_lots_trong_moi.columns and "dien_tich_trong" in df_lots_trong_moi.columns:
            _dt_map = df_lots_trong_moi.set_index("id")["dien_tich_trong"].to_dict()
            df_dt_seasons["dien_tich_trong"] = df_dt_seasons["base_lot_id"].map(_dt_map)

    if not df_dt_seasons.empty:
        detail_rows_by_vu = {}
        
        # ─── batch_label_map đã được build ở trên (shared với Map tooltip) ───

        # ─── Dedup: bỏ season trùng (vu, base_lot_id) ───
        # Sort: ưu tiên season đã kết thúc (có ngay_ket_thuc_thuc_te), rồi ngay_bat_dau mới nhất
        df_dt_seasons = df_dt_seasons.copy()
        df_dt_seasons["_has_end"] = df_dt_seasons["ngay_ket_thuc_thuc_te"].notna().astype(int)
        df_dt_seasons = df_dt_seasons.sort_values(["_has_end", "ngay_bat_dau"], ascending=[False, False])
        seen_vu_blid = set()
        
        for idx, row in df_dt_seasons.iterrows():
            f_vu = row.get("vu")
            lo_name = row.get("lo")
            lot_id = row.get("lot_id") or row.get("dim_lo_id")
            season_blid = row.get("base_lot_id")
            season_blid_int = _safe_int_id(season_blid)
            dien_tich_trong = row.get("dien_tich_trong")
            dien_tich_fallback = row.get("dien_tich", 0)
            # Ưu tiên diện tích trồng thực tế (per-batch), fallback diện tích lô tối đa (per-lot)
            dien_tich = float(dien_tich_trong) if pd.notna(dien_tich_trong) else (float(dien_tich_fallback) if pd.notna(dien_tich_fallback) else 0.0)

            # Skip duplicate (vu, base_lot_id) - giữ dòng đầu tiên
            if season_blid_int is not None:
                dedup_key = (f_vu, season_blid_int)
                if dedup_key in seen_vu_blid:
                    continue
                seen_vu_blid.add(dedup_key)

            start = pd.to_datetime(row.get("ngay_bat_dau"))
            end_actual = row.get("ngay_ket_thuc_thuc_te")
            end_planned = row.get("ngay_ket_thuc_du_kien")
            end = pd.to_datetime(end_actual) if pd.notna(end_actual) else pd.to_datetime(end_planned) if pd.notna(end_planned) else None

            start_str = start.strftime("%d/%m/%Y") if pd.notna(start) else "—"
            if pd.notna(end):
                end_str = end.strftime("%d/%m/%Y")
            else:
                end_str = "Hiện tại"
            thoi_gian_vu = f"{start_str} - {end_str}"

            # ─── Lọc dữ liệu theo base_lot_id (chính xác theo đợt trồng) ───
            has_blid_col = (not df_stg_all.empty and "base_lot_id" in df_stg_all.columns) or \
                           (not df_har_all.empty and "base_lot_id" in df_har_all.columns)
            
            if season_blid_int is not None and has_blid_col:
                if "id" in df_lots_all.columns:
                    sub_lots = df_lots_all[pd.to_numeric(df_lots_all["id"], errors="coerce") == season_blid_int]
                else:
                    sub_lots = pd.DataFrame()
                if not df_des_all.empty and "base_lot_id" in df_des_all.columns:
                    sub_des = df_des_all[pd.to_numeric(df_des_all["base_lot_id"], errors="coerce") == season_blid_int]
                else:
                    sub_des = pd.DataFrame()
            else:
                sub_lots = df_lots_all[df_lots_all["lot_id"] == lot_id] if not df_lots_all.empty else pd.DataFrame()
                sub_des = df_des_all[df_des_all["lot_id"] == lot_id] if not df_des_all.empty else pd.DataFrame()

            # Filter destruction theo date range
            if pd.notna(start):
                if not sub_des.empty and "ngay_xuat_huy" in sub_des.columns:
                    sub_des = sub_des[pd.to_datetime(sub_des["ngay_xuat_huy"]).dt.date >= start.date()]
            if pd.notna(end):
                if not sub_des.empty and "ngay_xuat_huy" in sub_des.columns:
                    sub_des = sub_des[pd.to_datetime(sub_des["ngay_xuat_huy"]).dt.date <= end.date()]

            so_luong_trong = int(sub_lots["so_luong"].sum()) if not sub_lots.empty else 0

            # ─── Dùng shared compute_batch_stats (đồng bộ logic với Map) ───
            if season_blid_int is not None and lo_name:
                next_start = _map_next_season.get((season_blid_int, f_vu))
                next_producing = next_start is not None
                next_s_date = next_start.date() if next_start is not None else None
                season_start_date = start.date() if pd.notna(start) else None
                _, so_chich_bap, so_cat_bap, so_thu_hoach = compute_batch_stats(
                    lo_name, season_blid_int, vu=f_vu, season_start=season_start_date,
                    next_season_start=next_s_date, next_vu_producing=next_producing
                )
            else:
                so_chich_bap, so_cat_bap, so_thu_hoach = 0, 0, 0

            # Tên lô: gắn "(đợt X)" nếu lô có nhiều đợt trồng
            display_lo = batch_label_map.get(season_blid_int, batch_label_map.get(season_blid, lo_name)) if season_blid_int is not None else lo_name

            if f_vu not in detail_rows_by_vu:
                detail_rows_by_vu[f_vu] = []

            dt_chich_est = int(round(so_luong_trong * get_estimated_rate("chich_bap")))
            dt_cat_est   = int(round(so_luong_trong * get_estimated_rate("cat_bap")))
            dt_thu_est   = int(round(so_luong_trong * get_estimated_rate("thu_hoach")))

            # Vụ đã chốt = có ngay_ket_thuc_thuc_te
            _season_completed = pd.notna(end_actual)
            detail_rows_by_vu[f_vu].append({
                ("Thông tin", "Thời gian vụ"): thoi_gian_vu,
                ("Thông tin", "Tên lô"): display_lo,
                ("Thông tin", "DT trồng (ha)"): f"{dien_tich:.2f}",
                ("Thông tin", "Cây đã trồng"): so_luong_trong,
                ("Chích bắp", "Dự toán"): dt_chich_est,
                ("Chích bắp", "Thực tế"): so_chich_bap,
                ("Cắt bắp", "Dự toán"): dt_cat_est,
                ("Cắt bắp", "Thực tế"): so_cat_bap,
                ("Thu hoạch", "Dự toán"): dt_thu_est,
                ("Thu hoạch", "Thực tế"): so_thu_hoach,
                ("Tổng khối lượng (kg)", "Dự toán"): dt_thu_est * get_kg_per_tree(f_vu),
                ("Tổng khối lượng (kg)", "Thực tế"): so_thu_hoach * get_kg_per_tree(f_vu),
                "_completed": _season_completed,
                "_sort_start_date": start if pd.notna(start) else pd.Timestamp.min,
            })
            
        # ─── Sort bảng theo tên lô tự nhiên (3B < 4A < 12A) rồi đợt trồng ───
        import re
        def _lo_sort_key(row_dict):
            name = row_dict.get(("Thông tin", "Tên lô"), "")
            # Tách: "3B (đợt 2)" → base="3B", batch=2
            m_batch = re.match(r"^(.+?)\s*\(đợt\s*(\d+)\)$", name)
            if m_batch:
                base_name, batch_num = m_batch.group(1), int(m_batch.group(2))
            else:
                base_name, batch_num = name, 0
            # Natural sort: "3B" → (3, "B"), "12A" → (12, "A")
            m_num = re.match(r"^(\d+)(.*)", base_name)
            if m_num:
                return (int(m_num.group(1)), m_num.group(2), batch_num)
            return (9999, base_name, batch_num)

        # ─── Sort controls ───
        _sort_options = {
            "Tên lô (mặc định)": None,
            "Thời gian vụ": ("Thông tin", "Thời gian vụ"),
            "DT trồng (ha)": ("Thông tin", "DT trồng (ha)"),
            "Cây đã trồng": ("Thông tin", "Cây đã trồng"),
            "Chích bắp - Dự toán": ("Chích bắp", "Dự toán"),
            "Chích bắp - Thực tế": ("Chích bắp", "Thực tế"),
            "Cắt bắp - Dự toán": ("Cắt bắp", "Dự toán"),
            "Cắt bắp - Thực tế": ("Cắt bắp", "Thực tế"),
            "Thu hoạch - Dự toán": ("Thu hoạch", "Dự toán"),
            "Thu hoạch - Thực tế": ("Thu hoạch", "Thực tế"),
            "Tổng KG - Dự toán": ("Tổng khối lượng (kg)", "Dự toán"),
            "Tổng KG - Thực tế": ("Tổng khối lượng (kg)", "Thực tế"),
        }
        _sc1, _sc2 = st.columns([3, 1])
        with _sc1:
            _sort_col_label = st.selectbox("Sắp xếp theo", list(_sort_options.keys()), index=0, key="dt_sort_col")
        with _sc2:
            _sort_dir = st.radio("Thứ tự", ["↑ Tăng", "↓ Giảm"], index=0, horizontal=True, key="dt_sort_dir")
        _sort_col = _sort_options[_sort_col_label]
        _sort_asc = _sort_dir == "↑ Tăng"

        # ─── Apply sort ───
        for vu_key in detail_rows_by_vu:
            if _sort_col is None:
                # Mặc định: sort theo tên lô tự nhiên
                detail_rows_by_vu[vu_key].sort(key=_lo_sort_key)
            elif _sort_col == ("Thông tin", "Thời gian vụ"):
                # Sort theo ngày bắt đầu thực tế (đã lưu trong _sort_start_date)
                def _date_sort_key(row_dict):
                    return row_dict.get("_sort_start_date", pd.Timestamp.min)
                detail_rows_by_vu[vu_key].sort(key=_date_sort_key, reverse=not _sort_asc)
            else:
                def _custom_sort_key(row_dict, col=_sort_col):
                    val = row_dict.get(col, 0)
                    try:
                        return float(str(val).replace(",", ""))
                    except (ValueError, TypeError):
                        return str(val)
                detail_rows_by_vu[vu_key].sort(key=_custom_sort_key, reverse=not _sort_asc)

        with st.expander("📋 Xem toàn bộ thông tin", expanded=True):
            if detail_rows_by_vu:
                st.markdown("""
                    <style>
                    .centered-table-wrapper table th, .centered-table-wrapper table td {
                        text-align: center !important;
                    }
                    .centered-table-wrapper table {
                        width: 100%;
                    }
                    </style>
                """, unsafe_allow_html=True)

                for vu_val, rows in detail_rows_by_vu.items():
                    st.markdown(f"##### 🌿 Vụ {vu_val}")
                    df_detail = pd.DataFrame(rows)
                    # Trích cờ ẩn trước khi chuyển MultiIndex
                    _completed_series = df_detail.pop("_completed") if "_completed" in df_detail.columns else pd.Series([False] * len(df_detail))
                    if "_sort_start_date" in df_detail.columns:
                        df_detail.pop("_sort_start_date")
                    if not isinstance(df_detail.columns, pd.MultiIndex):
                        df_detail.columns = pd.MultiIndex.from_tuples(df_detail.columns)
                    
                    # Tính tổng các cột trước khi format chuỗi
                    _area_col = df_detail[("Thông tin", "DT trồng (ha)")].astype(float)
                    total_dien_tich = _area_col.sum()
                    total_row = {
                        ("Thông tin", "Thời gian vụ"): "",
                        ("Thông tin", "Tên lô"): "<b>TỔNG</b>",
                        ("Thông tin", "DT trồng (ha)"): f"<b>{total_dien_tich:.2f}</b>",
                        ("Thông tin", "Cây đã trồng"): f"<b>{df_detail[('Thông tin', 'Cây đã trồng')].sum():,}</b>",
                        ("Chích bắp", "Dự toán"): f"<b>{df_detail[('Chích bắp', 'Dự toán')].sum():,}</b>",
                        ("Chích bắp", "Thực tế"): f"<b>{df_detail[('Chích bắp', 'Thực tế')].sum():,}</b>",
                        ("Cắt bắp", "Dự toán"): f"<b>{df_detail[('Cắt bắp', 'Dự toán')].sum():,}</b>",
                        ("Cắt bắp", "Thực tế"): f"<b>{df_detail[('Cắt bắp', 'Thực tế')].sum():,}</b>",
                        ("Thu hoạch", "Dự toán"): f"<b>{df_detail[('Thu hoạch', 'Dự toán')].sum():,}</b>",
                        ("Thu hoạch", "Thực tế"): f"<b>{df_detail[('Thu hoạch', 'Thực tế')].sum():,}</b>",
                        ("Tổng khối lượng (kg)", "Dự toán"): f"<b>{df_detail[('Tổng khối lượng (kg)', 'Dự toán')].sum():,}</b>",
                        ("Tổng khối lượng (kg)", "Thực tế"): f"<b>{df_detail[('Tổng khối lượng (kg)', 'Thực tế')].sum():,}</b>"
                    }
                    
                    # Format số nguyên có dấu phẩy
                    for c in df_detail.columns:
                        if df_detail[c].dtype.kind in 'iuf' and c[1] != "DT trồng (ha)":
                            df_detail[c] = df_detail[c].apply(lambda x: f"{int(x):,}")
                            
                    # Thêm dòng tổng
                    df_detail = pd.concat([df_detail, pd.DataFrame([total_row])], ignore_index=True)
                    
                    # ── Highlight hàng đã thu hoạch xong (vụ đã chốt) ──
                    _completed_rows = set()
                    for _ri in range(min(len(_completed_series), len(df_detail) - 1)):
                        if _completed_series.iloc[_ri]:
                            _completed_rows.add(_ri)

                    def _highlight_completed_rows(row):
                        if row.name in _completed_rows:
                            return ["background-color: rgba(0, 184, 148, 0.12)"] * len(row)
                        return [""] * len(row)

                    # Căn giữa MultiIndex Header và các ô
                    styled_df = df_detail.style.set_properties(**{'text-align': 'center'})
                    styled_df = styled_df.apply(_highlight_completed_rows, axis=1)
                    styled_df = styled_df.set_table_styles([
                        {"selector": "th", "props": [("text-align", "center")]},
                        {"selector": "th.col_heading", "props": [("text-align", "center")]},
                        {"selector": "td", "props": [("text-align", "center")]}
                    ])
                    styled_df = styled_df.hide(axis="index")
                    
                    table_html = styled_df.to_html(escape=False)
                    st.markdown(f'<div class="centered-table-wrapper" style="overflow-x: auto; margin-bottom: 2rem;">{table_html}</div>', unsafe_allow_html=True)
            else:
                st.info("Chưa có cấu hình Vụ/Lô nào để hiển thị bảng chi tiết.")
    else:
        st.info("Chưa có cấu hình Vụ/Lô nào để hiển thị bảng chi tiết.")
    # ─── Bảng Lịch sử Trồng dặm ───
    if not df_lots_trong_dam.empty:
        import re as _re_dam
        # Filter theo farm hiện tại (dt_farm) nếu có
        df_dam_display = df_lots_trong_dam.copy()
        if dt_farm != "Tất cả" and "farm" in df_dam_display.columns:
            df_dam_display = df_dam_display[df_dam_display["farm"] == dt_farm]
        
        if not df_dam_display.empty:
            with st.expander(f"📋 Lịch sử Trồng dặm ({len(df_dam_display)} đợt · {df_dam_display['so_luong'].sum():,} cây)", expanded=False):
                # Sort tự nhiên theo lô, rồi theo ngày
                def _nat_sort_dam(name):
                    m = _re_dam.match(r"^(\d+)(.*)", str(name))
                    return (int(m.group(1)), m.group(2)) if m else (9999, str(name))
                
                df_dam_tbl = df_dam_display[["lo", "farm", "ngay_trong", "so_luong"]].copy()
                df_dam_tbl["ngay_trong"] = pd.to_datetime(df_dam_tbl["ngay_trong"]).dt.strftime("%d/%m/%Y")
                df_dam_tbl["_sort"] = df_dam_tbl["lo"].apply(_nat_sort_dam)
                df_dam_tbl = df_dam_tbl.sort_values(["_sort", "ngay_trong"]).drop(columns=["_sort"])
                df_dam_tbl.columns = ["Lô", "Farm", "Ngày trồng dặm", "Số cây"]
                
                # Tổng dặm theo lô
                summary_dam = df_dam_display.groupby("lo")["so_luong"].agg(["sum", "count"]).reset_index()
                summary_dam.columns = ["Lô", "Tổng cây dặm", "Số đợt"]
                
                col_dam1, col_dam2 = st.columns([2, 1])
                with col_dam1:
                    st.caption("Chi tiết từng đợt trồng dặm")
                    styled_dam = df_dam_tbl.style.set_properties(**{'text-align': 'center', 'font-size': '0.85rem'})
                    styled_dam = styled_dam.set_table_styles([
                        {"selector": "th", "props": [("text-align", "center"), ("font-size", "0.85rem")]},
                    ]).hide(axis="index")
                    st.markdown(f'<div style="overflow-x:auto;">{styled_dam.to_html(escape=False)}</div>', unsafe_allow_html=True)
                with col_dam2:
                    st.caption("Tổng hợp theo lô")
                    styled_sum = summary_dam.style.set_properties(**{'text-align': 'center', 'font-size': '0.85rem'})
                    styled_sum = styled_sum.set_table_styles([
                        {"selector": "th", "props": [("text-align", "center"), ("font-size", "0.85rem")]},
                    ]).hide(axis="index")
                    st.markdown(f'<div style="overflow-x:auto;">{styled_sum.to_html(escape=False)}</div>', unsafe_allow_html=True)

    st.divider()

    # --- LỊCH THU HOẠCH DỰ KIẾN (Normal Distribution Model) ---
    st.markdown("#### 📅 Lịch Thu hoạch Dự kiến")
    st.caption(
        "① **Từ Trồng**: dự báo từ ngày trồng (lý thuyết) · "
        "② **Từ Chích bắp**: dự báo từ số cây đã chích bắp thực tế · "
        "③ **Từ Cắt bắp**: dự báo từ số cây đã cắt bắp thực tế · "
        "④ **Thực tế**: số buồng đã thu hoạch thực tế"
    )
    st.caption(f"Hao hụt từ Trồng → Thu hoạch: **{LOSS_RATE_TO_THU*100:.0f}%/vụ** · Sản lượng: **F0 = {KG_PER_TREE_F0} kg/buồng**, **Fn = {KG_PER_TREE_FN} kg/buồng**")

    if not df_lots_trong_moi.empty and "ngay_trong" in df_lots_trong_moi.columns and "lo" in df_lots_trong_moi.columns:
        from scipy.stats import norm as scipy_norm
        import numpy as np
        import re as _re
        
        LOSS_RATE = LOSS_RATE_TO_THU  # Sử dụng constant trung tâm
        FORECAST_GENERATIONS = 4  # F0, F1, F2, F3
        # Mặc định khoảng thời gian (ngày) — có thể tùy chỉnh bởi user
        DEFAULT_DAYS_BOI = 14   # 14 ngày thu bói
        DEFAULT_DAYS_RO  = 26   # 26 ngày thu rộ
        DEFAULT_DAYS_VET = 14   # 14 ngày thu vét
        
        # ─── Tùy chỉnh tỷ lệ phân phối thu hoạch ───
        with st.expander("⚙️ Tùy chỉnh tỷ lệ phân phối thu hoạch", expanded=False):
            st.caption("Mặc định: Thu bói 10% · Thu rộ 80% · Thu vét 10%. Thay đổi tỷ lệ để xem kịch bản khác. Tổng phải = 100%.")
            col_boi, col_ro, col_vet = st.columns(3)
            with col_boi:
                pct_boi = st.number_input("Thu bói (%)", min_value=0, max_value=100, 
                                           value=10, step=1, key="pct_thu_boi")
            with col_ro:
                pct_ro = st.number_input("Thu rộ (%)", min_value=0, max_value=100, 
                                          value=80, step=1, key="pct_thu_ro")
            with col_vet:
                pct_vet = st.number_input("Thu vét (%)", min_value=0, max_value=100, 
                                           value=10, step=1, key="pct_thu_vet")
            
            total_pct = pct_boi + pct_ro + pct_vet
            if total_pct != 100:
                st.warning(f"⚠️ Tổng tỷ lệ = {total_pct}%, cần = 100%. Đang dùng mặc định 10/80/10.")
                pct_boi, pct_ro, pct_vet = 10, 80, 10

            st.divider()
            st.caption(f"Mặc định: Thu bói {DEFAULT_DAYS_BOI} ngày · Thu rộ {DEFAULT_DAYS_RO} ngày · Thu vét {DEFAULT_DAYS_VET} ngày (tổng {DEFAULT_DAYS_BOI + DEFAULT_DAYS_RO + DEFAULT_DAYS_VET} ngày). Thay đổi để điều chỉnh cửa sổ thu hoạch.")
            col_d_boi, col_d_ro, col_d_vet = st.columns(3)
            with col_d_boi:
                days_boi = st.number_input("Thu bói (ngày)", min_value=1, max_value=60, 
                                            value=DEFAULT_DAYS_BOI, step=1, key="days_thu_boi")
            with col_d_ro:
                days_ro = st.number_input("Thu rộ (ngày)", min_value=1, max_value=120, 
                                           value=DEFAULT_DAYS_RO, step=1, key="days_thu_ro")
            with col_d_vet:
                days_vet = st.number_input("Thu vét (ngày)", min_value=1, max_value=60, 
                                            value=DEFAULT_DAYS_VET, step=1, key="days_thu_vet")
            
            total_days = days_boi + days_ro + days_vet
            st.caption(f"📐 Tổng cửa sổ: **{total_days} ngày** (Thu bói {days_boi} + Thu rộ {days_ro} + Thu vét {days_vet})")
        
        # Tính lại DAYS_RO_HALF, WINDOW_HALF từ input user
        DAYS_RO_HALF  = days_ro // 2   # Nửa cửa sổ thu rộ (VD: 26 → 13)
        DAYS_BOI_VET  = max(days_boi, days_vet)  # Lấy max để window cân đối
        WINDOW_HALF   = DAYS_RO_HALF + days_boi  # Tổng nửa window = nửa rộ + bói
        # Đảm bảo window đủ chứa cả vét: nếu vét > bói thì mở rộng phía sau
        WINDOW_HALF_RIGHT = DAYS_RO_HALF + days_vet
        
        # σ tính từ: P(|X| ≤ DAYS_RO_HALF) = 0.80 → σ = DAYS_RO_HALF / Φ⁻¹(0.90)
        SIGMA = DAYS_RO_HALF / scipy_norm.ppf(0.90) if DAYS_RO_HALF > 0 else 10.14
        
        # Tính trọng số PDF cho cửa sổ thu hoạch (bất đối xứng nếu bói ≠ vét)
        day_offsets = np.arange(-WINDOW_HALF, WINDOW_HALF_RIGHT + 1)  # [-bói-rộ/2 .. +vét+rộ/2]
        pdf_weights = scipy_norm.pdf(day_offsets, loc=0, scale=SIGMA)
        pdf_weights /= pdf_weights.sum()  # Normalize tổng = 1.0
        
        # Xác định loại thu cho mỗi offset
        def _classify_phase(offset):
            if offset < -DAYS_RO_HALF:
                return "Thu bói"
            elif offset <= DAYS_RO_HALF:
                return "Thu rộ"
            else:
                return "Thu vét"
        
        day_phases = [_classify_phase(d) for d in day_offsets]
        
        # ─── Rescale PDF weights theo tỷ lệ custom ───
        # Giữ nguyên shape Normal Distribution trong mỗi phase,
        # nhưng scale tổng trọng số mỗi phase khớp với % user nhập.
        phase_arr = np.array(day_phases)
        mask_boi = phase_arr == "Thu bói"
        mask_ro  = phase_arr == "Thu rộ"
        mask_vet = phase_arr == "Thu vét"
        
        raw_sum_boi = pdf_weights[mask_boi].sum()
        raw_sum_ro  = pdf_weights[mask_ro].sum()
        raw_sum_vet = pdf_weights[mask_vet].sum()
        
        target_boi = pct_boi / 100.0
        target_ro  = pct_ro / 100.0
        target_vet = pct_vet / 100.0
        
        # Scale mỗi phase: giữ shape tương đối bên trong, thay tổng
        if raw_sum_boi > 0:
            pdf_weights[mask_boi] *= (target_boi / raw_sum_boi)
        if raw_sum_ro > 0:
            pdf_weights[mask_ro]  *= (target_ro / raw_sum_ro)
        if raw_sum_vet > 0:
            pdf_weights[mask_vet] *= (target_vet / raw_sum_vet)
        
        # Đảm bảo tổng = 1.0 (phòng floating point)
        pdf_weights /= pdf_weights.sum()
        
        # ─── Pre-compute: micro-PDF weights cho Mốc ②③ ───
        # Mỗi điểm chích/cắt bắp spread ±7d Normal Distribution (σ=3)
        micro_offsets = np.arange(-MICRO_WINDOW_HALF, MICRO_WINDOW_HALF + 1)  # [-7..+7]
        micro_weights = scipy_norm.pdf(micro_offsets, loc=0, scale=MICRO_SIGMA)
        micro_weights /= micro_weights.sum()  # normalize tổng = 1.0
        
        # ─── Pre-compute: ribbon lookup for resolving mau_day from tuan ───
        _forecast_farm_id = get_farm_id_from_name(c_farm)
        _ribbon_lookup = {}  # {(year, week_number): color_name}
        if _forecast_farm_id:
            _ribbon_res = supabase.table("ribbon_schedule").select("year, week_number, color_name") \
                .eq("farm_id", _forecast_farm_id).eq("is_deleted", False).execute()
            for _rb in (_ribbon_res.data or []):
                _ribbon_lookup[(_rb["year"], _rb["week_number"])] = _rb["color_name"]
        
        def _resolve_ribbon_color(row):
            """Resolve mau_day from ribbon_schedule using tuan and date year."""
            tuan = row.get("tuan")
            if not tuan or not pd.notna(tuan):
                return None
            tuan = int(tuan)
            # Try to derive year from the date column
            for date_col in ["ngay_xuat_huy", "ngay_thuc_hien", "ngay_thu_hoach"]:
                d = row.get(date_col)
                if d and pd.notna(d):
                    yr = pd.to_datetime(d, errors="coerce").year
                    color = _ribbon_lookup.get((yr, tuan))
                    if color:
                        return color
            return None
        
        # ─── Pre-compute: xuất hủy & cắt bắp & thu hoạch per (base_lot_id) ───
        # Giữ raw destruction records để match vào generation per-lot (closest midpoint)
        # Tách theo giai_doan: mỗi giai đoạn chỉ ảnh hưởng 1 mốc dự báo
        des_records_direct = []  # [(blid, giai_doan, ngay, qty, mau_day), ...]
        # Proportional: records thiếu base_lot_id, tách theo (dim_lo_id, giai_doan)
        des_lo_by_stage = {}  # {(dim_lo_id, giai_doan): total_qty}
        if not df_des_all.empty:
            for _, d_row in df_des_all.iterrows():
                blid = d_row.get("base_lot_id")
                dlid = d_row.get("dim_lo_id")
                gd = d_row.get("giai_doan", "")
                ngay_h = pd.to_datetime(d_row.get("ngay_xuat_huy"), errors="coerce")
                qty = int(d_row.get("so_luong", 0)) if pd.notna(d_row.get("so_luong")) else 0
                md = _resolve_ribbon_color(d_row)
                if pd.notna(blid):
                    des_records_direct.append((blid, gd, ngay_h, qty, md))
                elif pd.notna(dlid):
                    key = (dlid, gd)
                    des_lo_by_stage[key] = des_lo_by_stage.get(key, 0) + qty
        
        # Pre-compute tổng cây trồng mới theo dim_lo_id (cho phân bổ tỉ lệ)
        lot_trees_by_lo = {}  # {dim_lo_id: total_trees_trong_moi}
        if not df_lots_trong_moi.empty:
            for _, lr in df_lots_trong_moi.iterrows():
                dlid = lr.get("dim_lo_id")
                qty = int(lr.get("so_luong", 0)) if pd.notna(lr.get("so_luong")) else 0
                if pd.notna(dlid):
                    lot_trees_by_lo[dlid] = lot_trees_by_lo.get(dlid, 0) + qty
        
        # Cắt bắp: giữ raw records (base_lot_id, ngày, qty, mau_day) để match theo generation
        cat_bap_records = []  # [(base_lot_id, ngay, so_luong, mau_day), ...]
        if not df_stg_all.empty:
            df_cat = df_stg_all[df_stg_all["giai_doan"] == "Cắt bắp"]
            if not df_cat.empty:
                for _, s_row in df_cat.iterrows():
                    blid = s_row.get("base_lot_id")
                    ngay = pd.to_datetime(s_row.get("ngay_thuc_hien"), errors="coerce")
                    qty = int(s_row.get("so_luong", 0)) if pd.notna(s_row.get("so_luong")) else 0
                    md = _resolve_ribbon_color(s_row)
                    if pd.notna(blid) and pd.notna(ngay):
                        cat_bap_records.append((blid, ngay, qty, md))
        
        # Chích bắp: giữ raw records (base_lot_id, ngày, qty) để match theo generation
        chich_bap_records = []  # [(base_lot_id, ngay, so_luong), ...]
        if not df_stg_all.empty:
            df_chich = df_stg_all[df_stg_all["giai_doan"] == "Chích bắp"]
            if not df_chich.empty:
                for _, s_row in df_chich.iterrows():
                    blid = s_row.get("base_lot_id")
                    ngay = pd.to_datetime(s_row.get("ngay_thuc_hien"), errors="coerce")
                    qty = int(s_row.get("so_luong", 0)) if pd.notna(s_row.get("so_luong")) else 0
                    if pd.notna(blid) and pd.notna(ngay):
                        chich_bap_records.append((blid, ngay, qty))
        
        # Thu hoạch thực tế: giữ raw records để match theo generation
        harvest_records = []  # [(base_lot_id, ngay, so_luong), ...]
        if not df_har_all.empty:
            for _, h_row in df_har_all.iterrows():
                blid = h_row.get("base_lot_id")
                ngay = pd.to_datetime(h_row.get("ngay_thu_hoach"), errors="coerce")
                qty = int(h_row.get("so_luong", 0)) if pd.notna(h_row.get("so_luong")) else 0
                if pd.notna(blid) and pd.notna(ngay):
                    harvest_records.append((blid, ngay, qty))
        
        # ─── Tạo daily harvest data (3 mốc) ───
        daily_rows = []
        shift_chich_rows = []  # Shift-based Mốc ② (chích bắp)
        shift_cat_rows = []    # Shift-based Mốc ③ (cắt bắp)
        lot_gen_midpoints = {}  # {(base_lot_id, gen_index): midpoint_date}
        for _, lot_row in df_lots_trong_moi.iterrows():
            ngay_trong_raw = lot_row.get("ngay_trong")
            lo_name = lot_row.get("lo", "")
            so_luong = int(lot_row.get("so_luong", 0)) if pd.notna(lot_row.get("so_luong")) else 0
            farm_name = lot_row.get("farm", "")
            base_lot_id = lot_row.get("id")
            dim_lo_id = lot_row.get("dim_lo_id")
            
            if pd.isna(ngay_trong_raw) or so_luong == 0:
                continue
            
            ngay_trong = pd.to_datetime(ngay_trong_raw)
            
            # ── Pre-compute midpoints cho tất cả generations ──
            all_midpoints = []
            mp = ngay_trong + timedelta(days=F0_DAYS_TO_THU)
            for g in range(FORECAST_GENERATIONS):
                all_midpoints.append(mp)
                mp = mp + timedelta(days=FN_CYCLE_DAYS)
            
            # ── Match destruction records vào generation (closest midpoint) ──
            # Tách theo giai_doan → mỗi giai đoạn chỉ ảnh hưởng 1 mốc dự báo
            des_gen_chich = {}   # {gen: qty} — "Trước chích bắp" → Mốc ①
            des_gen_cat = {}     # {gen: qty} — "Trước cắt bắp" → Mốc ②
            des_gen_thu = {}     # {gen: {mau_day: qty}} — "Trước thu hoạch" → Mốc ③
            des_gen_thu_total = {}  # {gen: total_qty} — tổng hủy TH per gen (cho fallback)
            for d_blid, d_gd, d_ngay, d_qty, d_md in des_records_direct:
                if d_blid != base_lot_id or pd.isna(d_ngay):
                    continue
                closest_gen = min(range(len(all_midpoints)),
                                  key=lambda g: abs((all_midpoints[g] - d_ngay).days))
                if d_gd == "Trước chích bắp":
                    des_gen_chich[closest_gen] = des_gen_chich.get(closest_gen, 0) + d_qty
                elif d_gd == "Trước cắt bắp":
                    des_gen_cat[closest_gen] = des_gen_cat.get(closest_gen, 0) + d_qty
                elif d_gd == "Trước thu hoạch":
                    des_gen_thu.setdefault(closest_gen, {})
                    if d_md:
                        des_gen_thu[closest_gen][d_md] = des_gen_thu[closest_gen].get(d_md, 0) + d_qty
                    else:
                        des_gen_thu[closest_gen]["__no_md__"] = des_gen_thu[closest_gen].get("__no_md__", 0) + d_qty
                    des_gen_thu_total[closest_gen] = des_gen_thu_total.get(closest_gen, 0) + d_qty
            
            # Proportional allocation per giai_doan (records thiếu base_lot_id)
            lot_ratio_prop = 0
            if pd.notna(dim_lo_id):
                total_lo_trees = lot_trees_by_lo.get(dim_lo_id, so_luong)
                lot_ratio_prop = so_luong / total_lo_trees if total_lo_trees > 0 else 0
            for (dlid, gd_key), gd_qty in des_lo_by_stage.items():
                if dlid != dim_lo_id or not pd.notna(dim_lo_id):
                    continue
                prop_qty = int(round(gd_qty * lot_ratio_prop))
                if prop_qty <= 0:
                    continue
                # Phân bổ tỉ lệ vào gen 0 (F0) — proportional records thiếu ngày nên mặc định F0
                if gd_key == "Trước chích bắp":
                    des_gen_chich[0] = des_gen_chich.get(0, 0) + prop_qty
                elif gd_key == "Trước cắt bắp":
                    des_gen_cat[0] = des_gen_cat.get(0, 0) + prop_qty
                elif gd_key == "Trước thu hoạch":
                    des_gen_thu_total[0] = des_gen_thu_total.get(0, 0) + prop_qty
                    des_gen_thu.setdefault(0, {})
                    des_gen_thu[0]["__no_md__"] = des_gen_thu[0].get("__no_md__", 0) + prop_qty
            
            # Tổng destruction (tất cả giai đoạn, tất cả gen) — cho UI info
            total_des_all = sum(d_qty for d_blid, _, _, d_qty, _ in des_records_direct if d_blid == base_lot_id)
            if pd.notna(dim_lo_id):
                for (dlid, _), gd_qty in des_lo_by_stage.items():
                    if dlid == dim_lo_id:
                        total_des_all += int(round(gd_qty * lot_ratio_prop))
            
            # ── Match cắt bắp records vào generation gần nhất — giữ daily detail + mau_day ──
            cat_bap_by_gen = {}       # {gen_index: total_qty}
            cat_daily_by_gen = {}     # {gen_index: [(ngay, qty, mau_day), ...]}
            for blid, ngay, qty, md in cat_bap_records:
                if blid != base_lot_id:
                    continue
                closest_gen = min(range(len(all_midpoints)),
                                  key=lambda g: abs((all_midpoints[g] - ngay).days))
                cat_bap_by_gen[closest_gen] = cat_bap_by_gen.get(closest_gen, 0) + qty
                cat_daily_by_gen.setdefault(closest_gen, []).append((ngay, qty, md))
            
            # ── Match chích bắp records vào generation gần nhất — giữ daily detail ──
            chich_bap_by_gen = {}     # {gen_index: total_qty}
            chich_daily_by_gen = {}   # {gen_index: [(ngay, qty), ...]}
            for blid, ngay, qty in chich_bap_records:
                if blid != base_lot_id:
                    continue
                closest_gen = min(range(len(all_midpoints)),
                                  key=lambda g: abs((all_midpoints[g] - ngay).days))
                chich_bap_by_gen[closest_gen] = chich_bap_by_gen.get(closest_gen, 0) + qty
                chich_daily_by_gen.setdefault(closest_gen, []).append((ngay, qty))
            
            harvest_midpoint = ngay_trong + timedelta(days=F0_DAYS_TO_THU)
            for gen in range(FORECAST_GENERATIONS):
                vu_label = f"F{gen}"
                lot_gen_midpoints[(base_lot_id, gen)] = harvest_midpoint
                
                # ── Mốc ①: Dự báo từ Trồng (trừ chỉ 'Trước chích bắp' per gen) ──
                des_chich_gen = des_gen_chich.get(gen, 0)
                so_luong_sau_huy = max(so_luong - des_chich_gen, 0)
                so_thu_after_loss = so_luong_sau_huy * (1 - LOSS_RATE)
                
                # ── Mốc ② (Chích bắp) & Mốc ③ (Cắt bắp) cho ĐÚNG generation này ──
                so_chich_bap_gen = chich_bap_by_gen.get(gen, None)  # None = chưa chích bắp cho vụ này
                so_cat_bap_gen = cat_bap_by_gen.get(gen, None)  # None = chưa cắt bắp cho vụ này
                
                # Window boundaries (hỗ trợ bất đối xứng: bói ≠ vét)
                win_start = harvest_midpoint - timedelta(days=WINDOW_HALF)
                thu_boi_start = win_start
                thu_boi_end   = harvest_midpoint - timedelta(days=DAYS_RO_HALF + 1)
                thu_ro_start  = harvest_midpoint - timedelta(days=DAYS_RO_HALF)
                thu_ro_end    = harvest_midpoint + timedelta(days=DAYS_RO_HALF)
                thu_vet_start = harvest_midpoint + timedelta(days=DAYS_RO_HALF + 1)
                thu_vet_end   = harvest_midpoint + timedelta(days=WINDOW_HALF_RIGHT)
                
                # Tạo row cho mỗi ngày
                for idx, (offset, weight, phase) in enumerate(zip(day_offsets, pdf_weights, day_phases)):
                    actual_date = harvest_midpoint + timedelta(days=int(offset))
                    daily_qty = so_thu_after_loss * weight
                    
                    # Mốc ②③: giờ dùng shift-based (tính riêng bên dưới)
                    daily_qty_chich = None
                    daily_qty_cat = None
                    
                    # Window label cho phase này
                    if phase == "Thu bói":
                        wlabel = f"{thu_boi_start.strftime('%d/%m')} – {thu_boi_end.strftime('%d/%m/%Y')}"
                    elif phase == "Thu rộ":
                        wlabel = f"{thu_ro_start.strftime('%d/%m')} – {thu_ro_end.strftime('%d/%m/%Y')}"
                    else:
                        wlabel = f"{thu_vet_start.strftime('%d/%m')} – {thu_vet_end.strftime('%d/%m/%Y')}"
                    
                    daily_rows.append({
                        "farm": farm_name,
                        "lo": lo_name,
                        "base_lot_id": base_lot_id,
                        "vu": vu_label,
                        "loai_thu": phase,
                        "ngay": actual_date,
                        "thang": actual_date.strftime("%m/%Y"),
                        "year": actual_date.year,
                        "so_luong_trong": so_luong,
                        "so_xuat_huy": total_des_all,
                        "daily_qty": daily_qty,
                        "daily_qty_chich": daily_qty_chich,
                        "daily_qty_cat": daily_qty_cat,

                        "window_label": wlabel,
                    })
                
                harvest_midpoint = harvest_midpoint + timedelta(days=FN_CYCLE_DAYS)
            
            def _build_shift_rows(daily_by_gen, days_shift, col_name, des_by_gen_total=None, des_by_gen_mau_day=None):
                """Tạo shift rows cho chích hoặc cắt bắp — Micro-PDF approach.
                
                Mỗi record → shift +days_shift → spread ±7d Normal Distribution
                → gộp tất cả mini-PDFs → xác định phase bằng diện tích tích lũy 10/80/10
                với boundary-day splitting để đảm bảo tỷ lệ chính xác.
                
                des_by_gen_total: {gen: qty} — Aggregate Ratio (Mốc ②)
                des_by_gen_mau_day: {gen: {mau_day: qty}} — Pro-rata mau_day (Mốc ③)
                """
                rows = []
                for gen_idx in range(FORECAST_GENERATIONS):
                    records = sorted(daily_by_gen.get(gen_idx, []), key=lambda x: x[0])
                    if not records:
                        continue
                    
                    # ── Destruction deduction (GIỮ NGUYÊN) ──
                    # Pro-rata mau_day: trừ destruction từ đúng records cùng mau_day
                    if des_by_gen_mau_day is not None:
                        des_md_map = des_by_gen_mau_day.get(gen_idx, {})
                        des_total_gen = des_gen_thu_total.get(gen_idx, 0)
                        des_no_md = des_md_map.get("__no_md__", 0)
                        # Tính tổng qty per mau_day cho cắt bắp
                        cat_by_md = {}  # {mau_day: total_cat_qty}
                        total_qty_gen = 0
                        for rec in records:
                            r_md = rec[2] if len(rec) > 2 else None
                            r_qty = rec[1]
                            total_qty_gen += r_qty
                            if r_md:
                                cat_by_md[r_md] = cat_by_md.get(r_md, 0) + r_qty
                        
                        # Tính adjusted qty per record
                        adjusted_records = []
                        for rec in records:
                            ngay_src = rec[0]
                            qty = rec[1]
                            r_md = rec[2] if len(rec) > 2 else None
                            
                            if r_md and r_md in des_md_map:
                                # Pro-rata: trừ destruction cùng mau_day
                                total_cat_md = cat_by_md.get(r_md, qty)
                                if total_cat_md > 0:
                                    ratio_md = qty / total_cat_md
                                    des_this = int(round(des_md_map[r_md] * ratio_md))
                                else:
                                    des_this = 0
                                qty = max(0, qty - des_this)
                            elif r_md is None and des_no_md > 0 and total_qty_gen > 0:
                                # Fallback Aggregate Ratio cho records thiếu mau_day
                                ratio_agg = max(0, 1 - des_no_md / total_qty_gen)
                                qty = max(0, int(round(qty * ratio_agg)))
                            
                            adjusted_records.append((ngay_src, qty))
                    # Aggregate Ratio: nhân đều tỷ lệ cho tất cả records
                    elif des_by_gen_total is not None:
                        des_qty = des_by_gen_total.get(gen_idx, 0)
                        total_qty_gen = sum(r[1] for r in records)
                        ratio = max(0, 1 - des_qty / total_qty_gen) if total_qty_gen > 0 else 1
                        adjusted_records = [(r[0], max(0, int(round(r[1] * ratio)))) for r in records]
                    else:
                        adjusted_records = [(r[0], r[1]) for r in records]
                    
                    # ── Bước 1: Spread ±7d Normal Distribution cho từng record ──
                    daily_harvest = {}  # {day_ordinal: float_qty}
                    for ngay_src, qty in adjusted_records:
                        if qty <= 0:
                            continue
                        mid = ngay_src + timedelta(days=days_shift)
                        for offset, weight in zip(micro_offsets, micro_weights):
                            day = mid + timedelta(days=int(offset))
                            day_key = day.toordinal()
                            daily_harvest[day_key] = daily_harvest.get(day_key, 0) + qty * weight
                    
                    if not daily_harvest:
                        continue
                    
                    # ── Bước 2: Sort theo ngày, tính tổng ──
                    sorted_days = sorted(daily_harvest.items())  # [(ordinal, qty), ...]
                    total_all = sum(v for _, v in sorted_days)
                    if total_all <= 0:
                        continue
                    
                    # ── Bước 3: Phase assignment với boundary-day splitting ──
                    # Dùng diện tích tích lũy: 10% bói, 80% rộ, 10% vét
                    cum_target_boi = total_all * target_boi
                    cum_target_boi_ro = total_all * (target_boi + target_ro)
                    cum = 0
                    
                    for day_ord, qty in sorted_days:
                        actual_date = date.fromordinal(day_ord)
                        thang = actual_date.strftime("%m/%Y")
                        
                        if cum + qty <= cum_target_boi:
                            # Toàn bộ ngày này thuộc bói
                            rows.append({
                                "base_lot_id": base_lot_id,
                                "vu": f"F{gen_idx}",
                                "loai_thu": "Thu bói",
                                "thang": thang,
                                col_name: qty,
                            })
                            cum += qty
                        elif cum < cum_target_boi:
                            # Boundary split: bói + rộ (có thể + vét nếu qty rất lớn)
                            boi_part = cum_target_boi - cum
                            rest = qty - boi_part
                            rows.append({
                                "base_lot_id": base_lot_id,
                                "vu": f"F{gen_idx}",
                                "loai_thu": "Thu bói",
                                "thang": thang,
                                col_name: boi_part,
                            })
                            if cum + qty > cum_target_boi_ro:
                                # Edge case: ngày này chứa cả bói, rộ, vét
                                ro_part = cum_target_boi_ro - cum_target_boi
                                vet_part = qty - boi_part - ro_part
                                rows.append({
                                    "base_lot_id": base_lot_id,
                                    "vu": f"F{gen_idx}",
                                    "loai_thu": "Thu rộ",
                                    "thang": thang,
                                    col_name: ro_part,
                                })
                                rows.append({
                                    "base_lot_id": base_lot_id,
                                    "vu": f"F{gen_idx}",
                                    "loai_thu": "Thu vét",
                                    "thang": thang,
                                    col_name: vet_part,
                                })
                            else:
                                rows.append({
                                    "base_lot_id": base_lot_id,
                                    "vu": f"F{gen_idx}",
                                    "loai_thu": "Thu rộ",
                                    "thang": thang,
                                    col_name: rest,
                                })
                            cum += qty
                        elif cum + qty <= cum_target_boi_ro:
                            # Toàn bộ ngày này thuộc rộ
                            rows.append({
                                "base_lot_id": base_lot_id,
                                "vu": f"F{gen_idx}",
                                "loai_thu": "Thu rộ",
                                "thang": thang,
                                col_name: qty,
                            })
                            cum += qty
                        elif cum < cum_target_boi_ro:
                            # Boundary split: rộ + vét
                            ro_part = cum_target_boi_ro - cum
                            vet_part = qty - ro_part
                            rows.append({
                                "base_lot_id": base_lot_id,
                                "vu": f"F{gen_idx}",
                                "loai_thu": "Thu rộ",
                                "thang": thang,
                                col_name: ro_part,
                            })
                            rows.append({
                                "base_lot_id": base_lot_id,
                                "vu": f"F{gen_idx}",
                                "loai_thu": "Thu vét",
                                "thang": thang,
                                col_name: vet_part,
                            })
                            cum += qty
                        else:
                            # Toàn bộ ngày này thuộc vét
                            rows.append({
                                "base_lot_id": base_lot_id,
                                "vu": f"F{gen_idx}",
                                "loai_thu": "Thu vét",
                                "thang": thang,
                                col_name: qty,
                            })
                            cum += qty
                return rows
            
            shift_chich_rows.extend(_build_shift_rows(
                chich_daily_by_gen, DAYS_CHICH_TO_THU, "so_thu_chich_bap",
                des_by_gen_total=des_gen_cat  # Mốc ②: Aggregate Ratio "Trước cắt bắp"
            ))
            shift_cat_rows.extend(_build_shift_rows(
                cat_daily_by_gen, DAYS_CAT_TO_THU, "so_thu_cat_bap",
                des_by_gen_mau_day=des_gen_thu  # Mốc ③: Pro-rata mau_day "Trước thu hoạch"
            ))
        
        if daily_rows:
            df_daily = pd.DataFrame(daily_rows)
            
            # Gom theo tháng + lô + vụ + loại thu (từ daily → monthly) — 3 mốc
            agg_dict = {
                "so_thu_hoach_dk": ("daily_qty", "sum"),
            }
            df_harvest = df_daily.groupby(
                ["farm", "lo", "base_lot_id", "vu", "loai_thu", "thang", "year", 
                 "so_luong_trong", "so_xuat_huy", "window_label"],
                as_index=False
            ).agg(**agg_dict)
            
            # Mốc ②: Chích bắp — shift-based aggregation
            if shift_chich_rows:
                df_shift_chich = pd.DataFrame(shift_chich_rows)
                chich_agg = df_shift_chich.groupby(
                    ["base_lot_id", "vu", "loai_thu", "thang"],
                    as_index=False
                ).agg(so_thu_chich_bap=("so_thu_chich_bap", "sum"))
                df_harvest = df_harvest.merge(chich_agg, on=["base_lot_id", "vu", "loai_thu", "thang"], how="left")
            else:
                df_harvest["so_thu_chich_bap"] = None
            
            # Mốc ③: Cắt bắp — shift-based aggregation
            if shift_cat_rows:
                df_shift_cat = pd.DataFrame(shift_cat_rows)
                cat_agg = df_shift_cat.groupby(
                    ["base_lot_id", "vu", "loai_thu", "thang"],
                    as_index=False
                ).agg(so_thu_cat_bap=("so_thu_cat_bap", "sum"))
                df_harvest = df_harvest.merge(cat_agg, on=["base_lot_id", "vu", "loai_thu", "thang"], how="left")
            else:
                df_harvest["so_thu_cat_bap"] = None
            
            # Largest Remainder Method: làm tròn mà đảm bảo tổng mỗi (lô, vụ) chính xác
            # Áp dụng cho cả 3 mốc — Mốc ②③ bây giờ cũng là float từ micro-PDF spread
            for (lo_k, bid_k, vu_k), grp_idx in df_harvest.groupby(["lo", "base_lot_id", "vu"]).groups.items():
                for col in ["so_thu_hoach_dk", "so_thu_chich_bap", "so_thu_cat_bap"]:
                    vals = df_harvest.loc[grp_idx, col].values.copy()
                    # Skip nếu toàn None/NaN
                    if pd.isna(vals).all():
                        continue
                    vals = np.where(pd.isna(vals), 0, vals).astype(float)
                    target_total = int(round(vals.sum()))
                    floors = np.floor(vals).astype(int)
                    remainders = vals - floors
                    deficit = target_total - floors.sum()
                    if deficit > 0:
                        top_idx = np.argsort(-remainders)[:deficit]
                        floors[top_idx] += 1
                    df_harvest.loc[grp_idx, col] = floors
            
            df_harvest["so_thu_hoach_dk"] = df_harvest["so_thu_hoach_dk"].astype(int)
            # so_thu_chich_bap: keep as int or None
            df_harvest["so_thu_chich_bap"] = df_harvest["so_thu_chich_bap"].apply(
                lambda x: int(x) if pd.notna(x) and x != 0 else None)
            # so_thu_cat_bap: keep as int or None
            df_harvest["so_thu_cat_bap"] = df_harvest["so_thu_cat_bap"].apply(
                lambda x: int(x) if pd.notna(x) and x != 0 else None)
            # ── Mốc ④: Thu hoạch thực tế — match từng record theo (gen, phase, tháng) ──
            # harvest_logs có từng ngày riêng lẻ, cần xác định mỗi record
            # thuộc generation nào + phase nào (Thu bói/rộ/vét) + tháng nào.
            actual_harvest_rows = []
            for blid, ngay, qty in harvest_records:
                # Tìm tất cả midpoints của lot này
                lot_mps = [(g, mp) for (lid, g), mp in lot_gen_midpoints.items() if lid == blid]
                if not lot_mps:
                    continue
                # Match vào generation gần nhất
                closest_gen, closest_mp = min(lot_mps, key=lambda x: abs((x[1] - ngay).days))
                # Xác định phase dựa trên khoảng cách đến midpoint
                days_from_mp = (ngay - closest_mp).days
                if days_from_mp < -(DAYS_RO_HALF):
                    phase = "Thu bói"
                elif days_from_mp <= DAYS_RO_HALF:
                    phase = "Thu rộ"
                else:
                    phase = "Thu vét"
                actual_harvest_rows.append({
                    "base_lot_id": blid,
                    "vu": f"F{closest_gen}",
                    "loai_thu": phase,
                    "thang": ngay.strftime("%m/%Y"),
                    "so_thu_thuc_te": qty
                })
            if actual_harvest_rows:
                df_actual = pd.DataFrame(actual_harvest_rows)
                df_actual = df_actual.groupby(
                    ["base_lot_id", "vu", "loai_thu", "thang"], as_index=False
                ).agg(so_thu_thuc_te=("so_thu_thuc_te", "sum"))
                df_actual["so_thu_thuc_te"] = df_actual["so_thu_thuc_te"].astype(int)
                df_harvest = df_harvest.merge(
                    df_actual, on=["base_lot_id", "vu", "loai_thu", "thang"], how="left")
            else:
                df_harvest["so_thu_thuc_te"] = None
            df_harvest.rename(columns={"thang": "thang_thu_hoach"}, inplace=True)
            
            # ─── Bộ lọc: Farm + Năm + Tháng ───
            year_options = sorted(df_harvest["year"].unique())
            current_year = date.today().year
            default_year_idx = year_options.index(current_year) + 1 if current_year in year_options else 0

            if c_farm in ["Admin", "Phòng Kinh doanh"]:
                hcf0, hcf1, hcf2 = st.columns([1, 1, 1])
                with hcf0:
                    hv_farm = st.selectbox("Lọc theo Farm", options=farms_all, key="hv_farm_sched")
                with hcf1:
                    hv_year = st.selectbox("Lọc theo Năm",
                                          options=["Tất cả"] + [str(y) for y in year_options],
                                          index=default_year_idx, key="hv_year_sched")
                with hcf2:
                    _df_tmp = df_harvest.copy()
                    if hv_year != "Tất cả":
                        _df_tmp = _df_tmp[_df_tmp["year"] == int(hv_year)]
                    all_months = sorted(_df_tmp["thang_thu_hoach"].unique(),
                                       key=lambda x: pd.to_datetime(x, format="%m/%Y"))
                    hv_month = st.selectbox("Lọc theo tháng thu hoạch",
                                          options=["Tất cả"] + list(all_months), key="hv_month_sched")
            else:
                hv_farm = c_farm
                hcf1, hcf2 = st.columns([1, 1])
                with hcf1:
                    hv_year = st.selectbox("Lọc theo Năm",
                                          options=["Tất cả"] + [str(y) for y in year_options],
                                          index=default_year_idx, key="hv_year_sched")
                with hcf2:
                    _df_tmp = df_harvest.copy()
                    if hv_year != "Tất cả":
                        _df_tmp = _df_tmp[_df_tmp["year"] == int(hv_year)]
                    all_months = sorted(_df_tmp["thang_thu_hoach"].unique(),
                                       key=lambda x: pd.to_datetime(x, format="%m/%Y"))
                    hv_month = st.selectbox("Lọc theo tháng thu hoạch",
                                           options=["Tất cả"] + list(all_months), key="hv_month_sched")
            
            df_hv = df_harvest.copy()
            if hv_farm != "Tất cả":
                df_hv = df_hv[df_hv["farm"] == hv_farm]
            if hv_year != "Tất cả":
                df_hv = df_hv[df_hv["year"] == int(hv_year)]
            if hv_month != "Tất cả":
                df_hv = df_hv[df_hv["thang_thu_hoach"] == hv_month]
            
            if not df_hv.empty:
                # ─── @st.dialog: chi tiết breakdown theo tháng ───
                @st.dialog("📊 Chi tiết thu hoạch", width="large")
                def _show_harvest_detail(month_key, df_src):
                    df_month = df_src[df_src["thang_thu_hoach"] == month_key].copy()
                    
                    def _nat_key_pop(name):
                        _m = _re.match(r"^(\d+)(.*)", str(name))
                        return (int(_m.group(1)), _m.group(2)) if _m else (9999, str(name))
                    
                    df_month["_sort"] = df_month["lo"].apply(_nat_key_pop)
                    df_month = df_month.sort_values(["_sort", "vu", "loai_thu"]).drop(columns=["_sort"])
                    
                    # Header — Mốc ①
                    total_buong = df_month["so_thu_hoach_dk"].sum()
                    df_month["_kg"] = df_month.apply(lambda r: r["so_thu_hoach_dk"] * get_kg_per_tree(r["vu"]), axis=1)
                    total_kg = df_month["_kg"].sum()
                    so_thung = int(total_kg // KG_PER_BOX)
                    so_container = so_thung / BOXES_PER_CONTAINER
                    st.markdown(f"### Tháng {month_key}")
                    
                    # 4 mốc summary
                    col_m1, col_m2, col_m3, col_m4 = st.columns(4)
                    with col_m1:
                        st.metric("① Từ Trồng", f"{total_buong:,} buồng", f"≈ {total_kg:,.0f} kg")
                    with col_m2:
                        chich_total = df_month["so_thu_chich_bap"].sum() if df_month["so_thu_chich_bap"].notna().any() else None
                        if chich_total is not None:
                            chich_kg = df_month.apply(lambda r: r["so_thu_chich_bap"] * get_kg_per_tree(r["vu"]) if pd.notna(r.get("so_thu_chich_bap")) else 0, axis=1).sum()
                            st.metric("② Chích bắp", f"{int(chich_total):,} buồng", f"≈ {chich_kg:,.0f} kg")
                        else:
                            st.metric("② Chích bắp", "Chưa có TT", "")
                    with col_m3:
                        cat_total = df_month["so_thu_cat_bap"].sum() if df_month["so_thu_cat_bap"].notna().any() else None
                        if cat_total is not None:
                            cat_kg = df_month.apply(lambda r: r["so_thu_cat_bap"] * get_kg_per_tree(r["vu"]) if pd.notna(r.get("so_thu_cat_bap")) else 0, axis=1).sum()
                            st.metric("③ Từ Cắt bắp", f"{int(cat_total):,} buồng", f"≈ {cat_kg:,.0f} kg")
                        else:
                            st.metric("③ Từ Cắt bắp", "Chưa có TT", "")
                    with col_m4:
                        # Thực tế: distinct per (base_lot_id, vu)
                        tt_df = df_month.drop_duplicates(subset=["base_lot_id", "vu"])
                        tt_total = tt_df["so_thu_thuc_te"].sum() if tt_df["so_thu_thuc_te"].notna().any() else None
                        if tt_total is not None:
                            st.metric("④ Thực tế", f"{int(tt_total):,} buồng", "")
                        else:
                            st.metric("④ Thực tế", "Chưa có TT", "")
                    
                    # Container summary — all milestones
                    cont_parts = [f"① ~{so_container:,.1f} cont"]
                    if chich_total is not None:
                        _cont_chich = int(chich_kg // KG_PER_BOX) / BOXES_PER_CONTAINER
                        cont_parts.append(f"② ~{_cont_chich:,.1f}")
                    if cat_total is not None:
                        _cont_cat = int(cat_kg // KG_PER_BOX) / BOXES_PER_CONTAINER
                        cont_parts.append(f"③ ~{_cont_cat:,.1f}")
                    if tt_total is not None:
                        _tt_kg = tt_df.apply(lambda r: r["so_thu_thuc_te"] * get_kg_per_tree(r["vu"]) if pd.notna(r.get("so_thu_thuc_te")) else 0, axis=1).sum()
                        _cont_tt = int(_tt_kg // KG_PER_BOX) / BOXES_PER_CONTAINER
                        cont_parts.append(f"④ ~{_cont_tt:,.1f}")
                    st.markdown(f"🚛 **Container**: {' · '.join(cont_parts)}  _(13 kg/thùng · 1320 thùng/cont)_")
                    
                    # Tổng hợp theo loại thu
                    summary_by_type = df_month.groupby("loai_thu")["so_thu_hoach_dk"].sum()
                    type_parts = []
                    for lt in ["Thu bói", "Thu rộ", "Thu vét"]:
                        if lt in summary_by_type.index:
                            type_parts.append(f"**{lt}**: {summary_by_type[lt]:,}")
                    st.markdown(" · ".join(type_parts))
                    st.markdown("---")
                    
                    # Bảng chi tiết — 4 mốc
                    df_pop = df_month[["lo", "vu", "loai_thu", "so_thu_hoach_dk", "so_thu_chich_bap", "so_thu_cat_bap", "so_thu_thuc_te", "window_label"]].copy()
                    df_pop["so_thu_chich_bap"] = df_pop["so_thu_chich_bap"].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "—")
                    df_pop["so_thu_cat_bap"] = df_pop["so_thu_cat_bap"].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "—")
                    df_pop["so_thu_thuc_te"] = df_pop["so_thu_thuc_te"].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "—")
                    df_pop.columns = ["Lô", "Vụ", "Loại thu", "① Từ Trồng", "② Chích bắp", "③ Cắt bắp", "④ Thực tế", "Khoảng TG"]
                    
                    styled_pop = df_pop.style.set_properties(**{'text-align': 'center', 'font-size': '0.85rem'})
                    styled_pop = styled_pop.set_table_styles([
                        {"selector": "th", "props": [("text-align", "center"), ("font-size", "0.85rem")]},
                        {"selector": "td", "props": [("text-align", "center")]}
                    ]).hide(axis="index")
                    st.markdown(f'<div style="overflow-x:auto;">{styled_pop.to_html(escape=False)}</div>',
                               unsafe_allow_html=True)
                
                # ─── CSS: style button thành card (scoped qua container key) ───
                st.markdown("""
                <style>
                .st-key-harvest-cards button {
                    background: linear-gradient(135deg, #b7e4c7 0%, #95d5b2 100%) !important;
                    color: #1b4332 !important;
                    border: 1px solid #74c69d !important;
                    border-radius: 12px !important;
                    padding: 1.2rem 0.5rem !important;
                    width: 100%;
                    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
                    cursor: pointer;
                    transition: transform 0.15s, box-shadow 0.15s;
                    min-height: 180px;
                }
                .st-key-harvest-cards button:hover {
                    transform: translateY(-2px);
                    box-shadow: 0 4px 16px rgba(0,0,0,0.12);
                    background: linear-gradient(135deg, #95d5b2 0%, #74c69d 100%) !important;
                    color: #1b4332 !important;
                    border: 1px solid #52b788 !important;
                }
                .st-key-harvest-cards button:active {
                    transform: translateY(0px);
                }
                .st-key-harvest-cards button p {
                    color: #1b4332 !important;
                    margin: 0;
                    font-size: 0.85rem;
                }
                </style>
                """, unsafe_allow_html=True)
                
                # ─── Metric cards (buttons) trong container scoped ───
                # Tính kg theo từng dòng trước khi gộp (vì mỗi vụ có kg/buồng khác nhau)
                df_hv["_kg_est"] = df_hv.apply(lambda r: r["so_thu_hoach_dk"] * get_kg_per_tree(r["vu"]), axis=1)
                df_hv["_kg_chich"] = df_hv.apply(
                    lambda r: r["so_thu_chich_bap"] * get_kg_per_tree(r["vu"]) if pd.notna(r.get("so_thu_chich_bap")) else None, axis=1)
                df_hv["_kg_cat"] = df_hv.apply(
                    lambda r: r["so_thu_cat_bap"] * get_kg_per_tree(r["vu"]) if pd.notna(r.get("so_thu_cat_bap")) else None, axis=1)
                monthly_summary = df_hv.groupby("thang_thu_hoach").agg(
                    tong_cay=("so_thu_hoach_dk", "sum"),
                    kg_est=("_kg_est", "sum"),
                    so_lo=("lo", "nunique")
                ).reset_index()
                
                # Mốc ②: tổng chích bắp theo tháng
                chich_month = df_hv.groupby("thang_thu_hoach").agg(
                    tong_chich=("so_thu_chich_bap", lambda x: int(x.sum()) if x.notna().any() else None),
                    kg_chich=("_kg_chich", lambda x: x.sum() if x.notna().any() else None)
                ).reset_index()
                monthly_summary = monthly_summary.merge(chich_month, on="thang_thu_hoach", how="left")
                
                # Mốc ③: tổng cắt bắp theo tháng (None nếu chưa có data nào)
                cat_month = df_hv.groupby("thang_thu_hoach").agg(
                    tong_cat=("so_thu_cat_bap", lambda x: int(x.sum()) if x.notna().any() else None),
                    kg_cat=("_kg_cat", lambda x: x.sum() if x.notna().any() else None)
                ).reset_index()
                monthly_summary = monthly_summary.merge(cat_month, on="thang_thu_hoach", how="left")
                
                # Mốc ④: tổng thực tế theo tháng — dùng unique lot-level values
                # so_thu_thuc_te là per-lot, cần distinct trước khi sum
                _df_tt_dedup = df_hv.drop_duplicates(subset=["base_lot_id", "vu", "thang_thu_hoach"])
                _df_tt_dedup["_kg_tt"] = _df_tt_dedup.apply(
                    lambda r: r["so_thu_thuc_te"] * get_kg_per_tree(r["vu"]) if pd.notna(r.get("so_thu_thuc_te")) else None, axis=1)
                actual_by_month = _df_tt_dedup.groupby("thang_thu_hoach").agg(
                    tong_thuc_te=("so_thu_thuc_te", lambda x: int(x.sum()) if x.notna().any() else None),
                    kg_tt=("_kg_tt", lambda x: x.sum() if x.notna().any() else None)
                ).reset_index()
                monthly_summary = monthly_summary.merge(actual_by_month, on="thang_thu_hoach", how="left")
                
                monthly_summary = monthly_summary.sort_values("thang_thu_hoach",
                    key=lambda x: pd.to_datetime(x, format="%m/%Y"))
                
                month_list = monthly_summary.to_dict("records")
                with st.container(key="harvest-cards"):
                    for i in range(0, len(month_list), 4):
                        cols = st.columns(min(4, len(month_list) - i))
                        for j, col in enumerate(cols):
                            if i + j < len(month_list):
                                m = month_list[i + j]
                                kg_est = m["kg_est"]
                                month_key = m["thang_thu_hoach"]
                                
                                with col:
                                    so_thung_card = int(kg_est // KG_PER_BOX)
                                    so_cont_card = so_thung_card / BOXES_PER_CONTAINER
                                    
                                    # Mốc ① — Từ Trồng
                                    line1 = f"① Trồng: **{m['tong_cay']:,}** buồng ≈ {kg_est:,.0f} kg · ~{so_cont_card:,.1f} cont"
                                    
                                    # Mốc ② — Từ Chích bắp
                                    tong_chich = m.get("tong_chich")
                                    if pd.notna(tong_chich) and tong_chich is not None:
                                        kg_chich_val = m.get("kg_chich", 0) or 0
                                        cont_chich = int(kg_chich_val // KG_PER_BOX) / BOXES_PER_CONTAINER
                                        line2 = f"② Chích: **{int(tong_chich):,}** buồng ≈ {kg_chich_val:,.0f} kg · ~{cont_chich:,.1f} cont"
                                    else:
                                        line2 = "② Chích: _Chưa có TT_"
                                    
                                    # Mốc ③ — Từ Cắt bắp
                                    tong_cat = m.get("tong_cat")
                                    if pd.notna(tong_cat) and tong_cat is not None:
                                        kg_cat_val = m.get("kg_cat", 0) or 0
                                        cont_cat = int(kg_cat_val // KG_PER_BOX) / BOXES_PER_CONTAINER
                                        line3 = f"③ Cắt: **{int(tong_cat):,}** buồng ≈ {kg_cat_val:,.0f} kg · ~{cont_cat:,.1f} cont"
                                    else:
                                        line3 = "③ Cắt: _Chưa có TT_"
                                    
                                    # Mốc ④ — Thực tế
                                    tong_tt = m.get("tong_thuc_te")
                                    if pd.notna(tong_tt) and tong_tt is not None:
                                        kg_tt_val = m.get("kg_tt", 0) or 0
                                        cont_tt = int(kg_tt_val // KG_PER_BOX) / BOXES_PER_CONTAINER
                                        line4 = f"④ TT: **{int(tong_tt):,}** buồng ≈ {kg_tt_val:,.0f} kg · ~{cont_tt:,.1f} cont"
                                    else:
                                        line4 = "④ TT: _Chưa có TT_"
                                    
                                    btn_label = f"📅 Tháng {month_key}\n\n{line1}\n\n{line2}\n\n{line3}\n\n{line4}"
                                    if st.button(btn_label, key=f"hv_card_{month_key}",
                                               use_container_width=True):
                                        _show_harvest_detail(month_key, df_hv)
                
                # ─── Bảng tổng hợp (expander) ───
                with st.expander("📋 Bảng tổng hợp lịch thu hoạch", expanded=False):
                    def _nat_key(name):
                        m = _re.match(r"^(\d+)(.*)", str(name))
                        return (int(m.group(1)), m.group(2)) if m else (9999, str(name))
                    
                    df_display = df_hv[["lo", "vu", "loai_thu", "thang_thu_hoach",
                                        "so_luong_trong", "so_xuat_huy", "so_thu_hoach_dk",
                                        "so_thu_chich_bap", "so_thu_cat_bap", "so_thu_thuc_te", "window_label"]].copy()
                    df_display["so_thu_chich_bap"] = df_display["so_thu_chich_bap"].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "—")
                    df_display["so_thu_cat_bap"] = df_display["so_thu_cat_bap"].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "—")
                    df_display["so_thu_thuc_te"] = df_display["so_thu_thuc_te"].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "—")
                    df_display["_sort"] = df_display["lo"].apply(lambda x: _nat_key(x))
                    df_display = df_display.sort_values(["_sort", "vu", "loai_thu"]).drop(columns=["_sort"])
                    df_display.columns = ["Lô", "Vụ", "Loại thu", "Tháng TH",
                                         "Trồng", "Xuất hủy", "① Từ Trồng",
                                         "② Chích bắp", "③ Cắt bắp", "④ Thực tế", "Khoảng TG"]
                    
                    styled = df_display.style.set_properties(**{'text-align': 'center'})
                    styled = styled.set_table_styles([
                        {"selector": "th", "props": [("text-align", "center")]},
                        {"selector": "td", "props": [("text-align", "center")]}
                    ]).hide(axis="index")
                    
                    st.markdown(f'<div class="centered-table-wrapper" style="overflow-x: auto;">'
                               f'{styled.to_html(escape=False)}</div>', unsafe_allow_html=True)
            else:
                st.info("Không có dữ liệu thu hoạch phù hợp với bộ lọc.")
    else:
        st.info("Chưa có dữ liệu lô trồng để tính lịch thu hoạch.")

    st.divider()

    # --- DỰ TOÁN SẢN LƯỢNG THU HOẠCH (KG) ---
    st.markdown("#### ⚖️ Dự toán Sản lượng Thu hoạch (Kg)")
    st.caption(f"Ước tính sản lượng dựa trên số cây ở giai đoạn gần nhất × **{KG_PER_TREE_F0} kg/cây (F0)** hoặc **{KG_PER_TREE_FN} kg/cây (Fn)**.")

    ek_farm, ek_vu, ek_team, ek_lot, _ = render_chart_filters("ek")
    filtered_ek_dfs = get_filtered_dfs(ek_farm, ek_vu, ek_team, ek_lot, None, {
        "lots": df_lots_all, "stg": df_stg_all, "des": df_des_all, "har": df_har_all
    })

    ek_lots_df = filtered_ek_dfs["lots"]
    ek_stg_df = filtered_ek_dfs["stg"]
    ek_har_df = filtered_ek_dfs["har"]
    ek_des_df = filtered_ek_dfs["des"]

    total_cay_da_trong = int(ek_lots_df["so_luong"].sum()) if not ek_lots_df.empty else 0
    total_du_toan_thu = int(round(total_cay_da_trong * get_estimated_rate("thu_hoach")))
    total_da_thu = int(ek_har_df["so_luong"].sum()) if not ek_har_df.empty else 0
    total_xuat_huy = int(ek_des_df["so_luong"].sum()) if not ek_des_df.empty else 0
    total_con_lai = max(total_du_toan_thu - total_da_thu - total_xuat_huy, 0)

    # Metric cards
    m1, m2, m3, m4 = st.columns(4)
    with m1:
        st.metric("🌱 Dự toán thu hoạch", f"{total_du_toan_thu:,} buồng",
                  delta=f"Trồng: {total_cay_da_trong:,} · Hao hụt: {LOSS_RATE_TO_THU*100:.0f}%")
    with m2:
        _kg_rate = get_kg_per_tree(ek_vu if ek_vu != "Tất cả" else "Fn")
        st.metric("✅ Đã thu hoạch", f"{total_da_thu:,} buồng", delta=f"{total_da_thu * _kg_rate:,.0f} kg")
    with m3:
        st.metric("🗑️ Xuất hủy", f"{total_xuat_huy:,} cây")
    with m4:
        st.metric("📦 Kg dự toán còn lại", f"{total_con_lai * _kg_rate:,.0f} kg")

    st.divider()

    # --- BIỂU ĐỒ PHỄU TIẾN ĐỘ THEO LÔ ---
    st.markdown("#### 📊 Biểu đồ Phễu Tiến độ theo Đợt trồng (Pipeline Funnel)")
    st.caption("So sánh tương quan Mức độ Hao hụt và Năng suất từ lúc Xuống giống đến khi Thu hoạch — phân chia theo từng đợt trồng.")
    
    # 1. Pipeline Chart Filters
    pf_farm, pf_vu, pf_team, pf_lot, pf_date = render_chart_filters("pf", include_date=True)
    filtered_pipe_dfs = get_filtered_dfs(pf_farm, pf_vu, pf_team, pf_lot, pf_date, {
        "lots": df_lots_all, "stg": df_stg_all, "des": df_des_all, "har": df_har_all
    })
    
    pipe_lots_df = filtered_pipe_dfs["lots"]
    pipe_stg_df = filtered_pipe_dfs["stg"]
    pipe_har_df = filtered_pipe_dfs["har"]
    pipe_des_df = filtered_pipe_dfs["des"]
    pipe_seasons_df = df_seasons.copy()
    if not pipe_seasons_df.empty:
        if pf_farm != "Tất cả" and "farm" in pipe_seasons_df.columns:
            pipe_seasons_df = pipe_seasons_df[pipe_seasons_df["farm"] == pf_farm]
        if pf_vu != "Tất cả" and "vu" in pipe_seasons_df.columns:
            pipe_seasons_df = pipe_seasons_df[pipe_seasons_df["vu"] == pf_vu]
        if pf_team != "Tất cả" and "team" in pipe_seasons_df.columns:
            pipe_seasons_df = pipe_seasons_df[pipe_seasons_df["team"] == pf_team]
        if pf_lot != "Tất cả" and "lo" in pipe_seasons_df.columns:
            pipe_seasons_df = pipe_seasons_df[pipe_seasons_df["lo"] == pf_lot]

    def _pipe_date_or_none(value):
        if value is None or pd.isna(value):
            return None
        try:
            return pd.Timestamp(value).date()
        except Exception:
            return None

    def _sum_pipe_destruction(base_lot_id, season_start=None, next_season_start=None, next_vu_producing=False):
        if pipe_des_df.empty or "base_lot_id" not in pipe_des_df.columns:
            return 0
        des = pipe_des_df[pd.to_numeric(pipe_des_df["base_lot_id"], errors="coerce") == int(base_lot_id)]
        if des.empty or "ngay_xuat_huy" not in des.columns:
            return int(des["so_luong"].sum()) if not des.empty else 0
        des_dates = pd.to_datetime(des["ngay_xuat_huy"], errors="coerce")
        if season_start is not None:
            des = des[des_dates >= pd.Timestamp(season_start)]
            des_dates = pd.to_datetime(des["ngay_xuat_huy"], errors="coerce")
        if next_season_start is not None and next_vu_producing:
            des = des[des_dates < pd.Timestamp(next_season_start)]
        return int(des["so_luong"].sum()) if not des.empty else 0

    # Gom dữ liệu theo từng đợt trồng (base_lots.id) để vẽ grouped/stacked bar chart
    if not pipe_lots_df.empty:
        pipe_lots_merged = pipe_lots_df

        # Dùng "id" (PK) của base_lots làm key unique cho mỗi đợt trồng
        batch_ids = pipe_lots_merged["id"].unique()
        pipeline_data = []
        for bid in batch_ids:
            bid_int = int(bid)
            batch_row = pipe_lots_merged[pipe_lots_merged["id"] == bid]
            lo_name = batch_row["lo"].iloc[0] if not batch_row.empty else str(bid)
            ngay_trong = batch_row["ngay_trong"].iloc[0] if "ngay_trong" in batch_row.columns and not batch_row.empty else ""
            # Dùng batch_label_map (đã build ở trên): "3B (đợt 1)" hoặc "3B" nếu chỉ 1 đợt
            label = batch_label_map.get(bid, lo_name)

            season_start_date = None
            next_s_date = None
            next_producing = False
            season_vu = pf_vu
            if pf_vu != "Tất cả":
                if pipe_seasons_df.empty or "base_lot_id" not in pipe_seasons_df.columns:
                    continue
                season_rows = pipe_seasons_df[
                    pd.to_numeric(pipe_seasons_df["base_lot_id"], errors="coerce") == bid_int
                ].sort_values("ngay_bat_dau")
                if season_rows.empty:
                    continue
                season_row = season_rows.iloc[-1]
                season_vu = season_row.get("vu", pf_vu)
                season_start_date = _pipe_date_or_none(season_row.get("ngay_bat_dau"))
                next_s = _map_next_season.get((bid_int, season_vu))
                next_producing = next_s is not None
                next_s_date = next_s.date() if next_s is not None else None
            
            # 1. Trồng mới và Trồng dặm — mỗi batch chỉ có 1 loại
            sl_trong_moi = batch_row[batch_row["loai_trong"] == "Trồng mới"]["so_luong"].sum()
            sl_trong_dam = batch_row[batch_row["loai_trong"] == "Trồng dặm"]["so_luong"].sum()
            # Diện tích trồng: ưu tiên dien_tich_trong (per-batch), fallback dien_tich (lot max)
            dt = 0
            if "dien_tich_trong" in batch_row.columns:
                dt_val = batch_row["dien_tich_trong"].iloc[0]
                if pd.notna(dt_val):
                    dt = float(dt_val)
            if dt == 0 and "dien_tich" in batch_row.columns:
                dt_val = batch_row["dien_tich"].iloc[0]
                if pd.notna(dt_val):
                    dt = float(dt_val)
            
            hover_base = f"<b>{label}</b><br>DT trồng: {dt:.2f} ha"
            pipeline_data.append({"Đợt trồng": label, "Giai đoạn": "1a. Trồng mới", "Số lượng": sl_trong_moi, "hover": hover_base})
            pipeline_data.append({"Đợt trồng": label, "Giai đoạn": "1b. Trồng dặm", "Số lượng": sl_trong_dam, "hover": hover_base})
            
            # Stage/Harvest/Destruction match via base_lot_id == base_lots.id
            if pf_vu != "Tất cả":
                _, sl_cb, sl_cut, sl_har = compute_batch_stats(
                    lo_name, bid_int, vu=season_vu, season_start=season_start_date,
                    next_season_start=next_s_date, next_vu_producing=next_producing,
                    stage_df=pipe_stg_df, harvest_df=pipe_har_df,
                )
                sl_des = _sum_pipe_destruction(
                    bid_int, season_start=season_start_date,
                    next_season_start=next_s_date, next_vu_producing=next_producing,
                )
            else:
                if not pipe_stg_df.empty and "base_lot_id" in pipe_stg_df.columns:
                    pipe_bid_stg = pipe_stg_df[pd.to_numeric(pipe_stg_df["base_lot_id"], errors="coerce") == bid_int]
                    sl_cb = pipe_bid_stg[pipe_bid_stg["giai_doan"] == "Chích bắp"]["so_luong"].sum()
                    sl_cut = pipe_bid_stg[pipe_bid_stg["giai_doan"] == "Cắt bắp"]["so_luong"].sum()
                else:
                    sl_cb, sl_cut = 0, 0
                if not pipe_har_df.empty and "base_lot_id" in pipe_har_df.columns:
                    sl_har = pipe_har_df[
                        pd.to_numeric(pipe_har_df["base_lot_id"], errors="coerce") == bid_int
                    ]["so_luong"].sum()
                else:
                    sl_har = 0
                sl_des = _sum_pipe_destruction(bid_int)

            pipeline_data.append({"Đợt trồng": label, "Giai đoạn": "2. Chích bắp", "Số lượng": int(sl_cb), "hover": hover_base})
            pipeline_data.append({"Đợt trồng": label, "Giai đoạn": "3. Cắt bắp", "Số lượng": int(sl_cut), "hover": hover_base})
            pipeline_data.append({"Đợt trồng": label, "Giai đoạn": "4. Thu hoạch", "Số lượng": int(sl_har), "hover": hover_base})
            pipeline_data.append({"Đợt trồng": label, "Giai đoạn": "5. Xuất hủy", "Số lượng": int(sl_des), "hover": hover_base})
            
        df_pipeline = pd.DataFrame(pipeline_data)
        
        # Build hybrid chart: Trồng mới + Trồng dặm stacked, rest clustered
        lots_list = list(df_pipeline["Đợt trồng"].unique())
        
        fig_pipe = go.Figure()
        
        # --- Stacked pair: Trồng mới + Trồng dặm (same offsetgroup) ---
        df_tm = df_pipeline[df_pipeline["Giai đoạn"] == "1a. Trồng mới"]
        fig_pipe.add_trace(go.Bar(
            name="1a. Trồng mới", x=df_tm["Đợt trồng"], y=df_tm["Số lượng"],
            marker_color="#4CAF50", offsetgroup="trong",
            customdata=df_tm["hover"], hovertemplate="%{customdata}<br>Giai đoạn: 1a. Trồng mới<br>Số lượng: %{y:,}<extra></extra>"
        ))
        df_td = df_pipeline[df_pipeline["Giai đoạn"] == "1b. Trồng dặm"]
        fig_pipe.add_trace(go.Bar(
            name="1b. Trồng dặm", x=df_td["Đợt trồng"], y=df_td["Số lượng"],
            marker_color="#8BC34A", offsetgroup="trong", base=df_tm["Số lượng"].values,
            customdata=df_td["hover"], hovertemplate="%{customdata}<br>Giai đoạn: 1b. Trồng dặm<br>Số lượng: %{y:,}<extra></extra>"
        ))
        
        # --- Clustered bars: each gets its own offsetgroup ---
        cluster_stages = [
            ("2. Chích bắp", "#FFC107"),
            ("3. Cắt bắp", "#FF9800"),
            ("4. Thu hoạch", "#2196F3"),
            ("5. Xuất hủy", "#F44336"),
        ]
        for stage_name, color in cluster_stages:
            df_s = df_pipeline[df_pipeline["Giai đoạn"] == stage_name]
            fig_pipe.add_trace(go.Bar(
                name=stage_name, x=df_s["Đợt trồng"], y=df_s["Số lượng"],
                marker_color=color, offsetgroup=stage_name,
                customdata=df_s["hover"], hovertemplate=f"%{{customdata}}<br>Giai đoạn: {stage_name}<br>Số lượng: %{{y:,}}<extra></extra>"
            ))
        
        fig_pipe.update_layout(
            barmode="group",
            plot_bgcolor="rgba(0,0,0,0)",
            yaxis={"showgrid": True, "gridcolor": "rgba(0,0,0,0.1)", "title": "Số lượng cây / buồng"},
            xaxis={"title": "Đợt trồng (Lô + Ngày trồng)", "tickangle": -30},
            legend_title_text="Tiến trình"
        )
        st.plotly_chart(fig_pipe, use_container_width=True)
    else:
        st.info("Chưa có danh sách lô để hiển thị biểu đồ Phễu.")

    st.divider()
    st.divider()
    st.markdown("##### 📉 Tiến trình Tổng hợp theo Thời gian")
    st.caption("Biểu đồ gộp thể hiện biến động các công đoạn dọc theo trục ngày. Có thể filter để làm nổi bật.")

    # 2. Multi-line Chart Filters
    mlf_farm, mlf_vu, mlf_team, mlf_lot, mlf_date = render_chart_filters("mlf", include_date=True)
    filtered_ml_dfs = get_filtered_dfs(mlf_farm, mlf_vu, mlf_team, mlf_lot, mlf_date, {
        "lots": df_lots_all, "stg": df_stg_all, "des": df_des_all, "har": df_har_all
    })

    ml_lots_df = filtered_ml_dfs["lots"]
    ml_stg_df = filtered_ml_dfs["stg"]
    ml_har_df = filtered_ml_dfs["har"]
    ml_des_df = filtered_ml_dfs["des"]

    plot_dfs = []
    
    # 1. Trồng mới / Trồng dặm
    if not ml_lots_df.empty and "ngay_trong" in ml_lots_df.columns:
        ml_lots_merged = ml_lots_df
        
        cols_to_keep = ["ngay_trong", "so_luong", "loai_trong"]
        if "lot_id" in ml_lots_merged.columns: cols_to_keep.append("lot_id")
        
        # Missing 'loai_trong' maps to "Trồng mới"
        ml_lots_merged["loai_trong"] = ml_lots_merged["loai_trong"].fillna("Trồng mới")
        
        # Trồng mới
        df_new = ml_lots_merged[ml_lots_merged["loai_trong"] == "Trồng mới"].copy()
        if not df_new.empty:
            df_new.rename(columns={"ngay_trong": "Date"}, inplace=True)
            df_new["Giai đoạn"] = "1a. Trồng mới"
            plot_dfs.append(df_new)
            
        # Trồng dặm
        df_old = ml_lots_merged[ml_lots_merged["loai_trong"] == "Trồng dặm"].copy()
        if not df_old.empty:
            df_old.rename(columns={"ngay_trong": "Date"}, inplace=True)
            df_old["Giai đoạn"] = "1b. Trồng dặm"
            plot_dfs.append(df_old)

    # 2. Chích bắp & Cắt bắp
    if not ml_stg_df.empty and "ngay_thuc_hien" in ml_stg_df.columns:
        cols_to_keep = ["ngay_thuc_hien", "so_luong", "giai_doan"]
        if "lot_id" in ml_stg_df.columns: cols_to_keep.append("lot_id")
        df_p = ml_stg_df[cols_to_keep].copy()
        df_p.rename(columns={"ngay_thuc_hien": "Date"}, inplace=True)
        df_p["Giai đoạn"] = df_p["giai_doan"].apply(lambda x: f"2. {x}" if x == "Chích bắp" else f"3. {x}")
        df_p.drop(columns=["giai_doan"], inplace=True)
        plot_dfs.append(df_p)

    # 3. Thu hoạch
    if not ml_har_df.empty and "ngay_thu_hoach" in ml_har_df.columns:
        cols_to_keep = ["ngay_thu_hoach", "so_luong"]
        if "lot_id" in ml_har_df.columns: cols_to_keep.append("lot_id")
        df_p = ml_har_df[cols_to_keep].copy()
        df_p.rename(columns={"ngay_thu_hoach": "Date"}, inplace=True)
        df_p["Giai đoạn"] = "4. Thu hoạch"
        plot_dfs.append(df_p)

    # 4. Xuất hủy
    if not ml_des_df.empty and "ngay_xuat_huy" in ml_des_df.columns:
        cols_to_keep = ["ngay_xuat_huy", "so_luong"]
        if "lot_id" in ml_des_df.columns: cols_to_keep.append("lot_id")
        df_p = ml_des_df[cols_to_keep].copy()
        df_p.rename(columns={"ngay_xuat_huy": "Date"}, inplace=True)
        df_p["Giai đoạn"] = "5. Xuất hủy"
        plot_dfs.append(df_p)

    if plot_dfs:
        df_combined = pd.concat(plot_dfs)
        df_combined["Date"] = pd.to_datetime(df_combined["Date"])

        # --- Breakdown text: số lượng theo từng Lô, dùng cho tooltip ---
        # Lấy tên lô và diện tích từ lot_id nếu có (join base_lots)
        lot_name_map = {}
        lot_dt_map = {}
        if not ml_lots_df.empty and "lot_id" in ml_lots_df.columns:
            if "lo" in ml_lots_df.columns:
                lot_name_map = ml_lots_df.set_index("lot_id")["lo"].to_dict()
            if "dien_tich_trong" in ml_lots_df.columns:
                lot_dt_map = ml_lots_df.set_index("lot_id")["dien_tich_trong"].fillna(ml_lots_df.set_index("lot_id")["dien_tich"]).to_dict()
            elif "dien_tich" in ml_lots_df.columns:
                lot_dt_map = ml_lots_df.set_index("lot_id")["dien_tich"].to_dict()
        
        if "lot_id" in df_combined.columns:
            df_combined["Tên Lô"] = df_combined["lot_id"].map(lot_name_map).fillna(df_combined.get("lot_id", ""))
            df_combined["Diện tích"] = df_combined["lot_id"].map(lot_dt_map).fillna(0)
            df_breakdown = (
                df_combined.groupby(["Date", "Giai đoạn", "Tên Lô", "Diện tích"], as_index=False)["so_luong"]
                .sum()
                .sort_values("so_luong", ascending=False)
            )
            def make_breakdown(grp):
                lines = []
                for _, r in grp.iterrows():
                    dt = r['Diện tích']
                    dt_str = f" ({dt:.2f} ha)" if pd.notna(dt) and dt != 0 else ""
                    lines.append(f"&nbsp;&nbsp;• Lô {r['Tên Lô']}{dt_str}: {int(r['so_luong']):,}")
                return "<br>".join(lines)
            df_bd_text = (
                df_breakdown.groupby(["Date", "Giai đoạn"])
                .apply(make_breakdown)
                .reset_index(name="breakdown_text")
            )
        else:
            df_bd_text = None

        df_grouped = df_combined.groupby(["Date", "Giai đoạn"], as_index=False)["so_luong"].sum()
        df_grouped.sort_values(by="Date", inplace=True)

        if df_bd_text is not None:
            df_grouped = df_grouped.merge(df_bd_text, on=["Date", "Giai đoạn"], how="left")
            df_grouped["breakdown_text"] = df_grouped["breakdown_text"].fillna("")
        else:
            df_grouped["breakdown_text"] = ""

        all_stages = sorted(df_grouped["Giai đoạn"].unique().tolist())
        selected_stages = st.multiselect("Lọc và Nổi bật Giai đoạn", options=all_stages, default=all_stages, key="global_stage_hl")
        
        # Color mapping matching the funnel defaults
        color_map = {
            "1a. Trồng mới": "#4CAF50",    # Green
            "1b. Trồng dặm": "#8BC34A",    # Light Green
            "2. Chích bắp": "#FFC107",     # Amber
            "3. Cắt bắp": "#FF9800",       # Orange
            "4. Thu hoạch": "#2196F3",     # Blue
            "5. Xuất hủy": "#F44336"       # Red
        }
        
        fig = px.line(
            df_grouped, x="Date", y="so_luong", color="Giai đoạn",
            markers=True, line_shape="linear", color_discrete_map=color_map,
            custom_data=["breakdown_text"],
            labels={"Date": "Ngày Thực hiện", "so_luong": "Số lượng (Cây/Buồng)"}
        )
        
        # Custom tooltip
        fig.update_traces(
            hovertemplate=(
                "<b>%{fullData.name}</b><br>"
                "📅 %{x|%d/%m/%Y}<br>"
                "Tổng: <b>%{y:,.0f}</b><br>"
                "Chi tiết theo Lô:<br>%{customdata[0]}"
                "<extra></extra>"
            )
        )
        
        # Highlight logic - fade out unselected
        if selected_stages and len(selected_stages) < len(all_stages):
            for trace in fig.data:
                if trace.name not in selected_stages:
                    trace.opacity = 0.15
                    trace.line.width = 1
                else:
                    trace.opacity = 1.0
                    trace.line.width = 3
        
        fig.update_layout(plot_bgcolor="rgba(0,0,0,0)", yaxis={"showgrid": True, "gridcolor": "rgba(0,0,0,0.1)"}, hovermode="x unified")
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("Chưa có bất kỳ dữ liệu nào để vẽ biểu đồ dòng thời gian.")

    st.divider()
    st.markdown("##### 🌳 Sự số lượng cây thực tế (Kiểm kê)")
    st.caption("Theo dõi số lượng cây thực tế trên từng Lô qua các lần kiểm đếm.")
    
    # 3. Tree Inventory Filters
    ti_farm, ti_vu, ti_team, ti_lot, ti_date = render_chart_filters("ti", include_date=True, use_dynamic_lots=False)
    filtered_ti_dfs = get_filtered_dfs(ti_farm, ti_vu, ti_team, ti_lot, ti_date, {"inv": df_tree_inv_all})
    ti_inv_df = filtered_ti_dfs["inv"]
    
    if not ti_inv_df.empty and "ngay_kiem_ke" in ti_inv_df.columns:
        df_inv = ti_inv_df.copy()
        
        # Ánh xạ lot_id sang lo và dien_tich
        if not df_lots_all.empty:
            mapped_dict_lo = df_lots_all.set_index("lot_id")["lo"].to_dict()
            if "dien_tich_trong" in df_lots_all.columns:
                mapped_dict_dt = df_lots_all.set_index("lot_id")["dien_tich_trong"].fillna(df_lots_all.set_index("lot_id")["dien_tich"]).to_dict()
            else:
                mapped_dict_dt = df_lots_all.set_index("lot_id")["dien_tich"].to_dict()
        else:
            mapped_dict_lo, mapped_dict_dt = {}, {}
            
        df_inv["Tên Lô"] = df_inv["lot_id"].map(lambda x: mapped_dict_lo.get(x, x))
        df_inv["Diện tích"] = df_inv["lot_id"].map(lambda x: mapped_dict_dt.get(x, 0))
        
        df_inv["Ngày"] = pd.to_datetime(df_inv["ngay_kiem_ke"])
        df_inv_grouped = df_inv.groupby(["Ngày", "Tên Lô", "Diện tích"], as_index=False)["so_luong_cay_thuc_te"].sum()
        df_inv_grouped.sort_values(by="Ngày", inplace=True)
        # Tạo chuỗi hover tuỳ chỉnh
        df_inv_grouped["hover"] = df_inv_grouped.apply(lambda r: f"<b>Lô {r['Tên Lô']}</b><br>Diện tích: {r['Diện tích']:.2f} ha<br>Ngày: {r['Ngày'].strftime('%d/%m/%Y')}<br>Số lượng: {int(r['so_luong_cay_thuc_te']):,} cây", axis=1)

        fig_inv = px.line(
            df_inv_grouped, x="Ngày", y="so_luong_cay_thuc_te", color="Tên Lô", 
            markers=True, line_shape="linear",
            custom_data=["hover"],
            labels={"Ngày": "Ngày Kiểm Kê", "so_luong_cay_thuc_te": "Số lượng cây", "Tên Lô": "Lô"},
        )
        fig_inv.update_traces(hovertemplate="%{customdata[0]}<extra></extra>")
        fig_inv.update_layout(plot_bgcolor="rgba(0,0,0,0)", yaxis={"showgrid": True, "gridcolor": "rgba(0,0,0,0.1)"}, hovermode="x unified")
        st.plotly_chart(fig_inv, use_container_width=True)
    else:
        st.info("Chưa có dữ liệu Kiểm kê cây trên farm này.")

# =====================================================
# MÁY TÍNH PHÂN BỔ CONTAINER (KINH DOANH)
# =====================================================
CAT_FORECAST_WEEK_OPTIONS = [8, 9]


def build_weekly_cat_forecast(df_stg: pd.DataFrame, forecast_weeks_inclusive: int = 8) -> pd.DataFrame:
    """Gom dự báo thu hoạch từ cắt bắp theo ISO year/week."""
    if df_stg is None or df_stg.empty or "giai_doan" not in df_stg.columns:
        return pd.DataFrame(columns=["farm", "year", "week", "forecast_bunches"])

    df_cat = df_stg[df_stg["giai_doan"] == "Cắt bắp"].copy()
    if df_cat.empty:
        return pd.DataFrame(columns=["farm", "year", "week", "forecast_bunches"])

    try:
        forecast_weeks_inclusive = int(forecast_weeks_inclusive)
    except (TypeError, ValueError):
        forecast_weeks_inclusive = 8
    if forecast_weeks_inclusive not in CAT_FORECAST_WEEK_OPTIONS:
        forecast_weeks_inclusive = 8
    forecast_days_from_cut = (forecast_weeks_inclusive - 1) * 7

    micro_offsets = list(range(-MICRO_WINDOW_HALF, MICRO_WINDOW_HALF + 1))
    raw_weights = [pow(2.718281828459045, -0.5 * pow(offset / MICRO_SIGMA, 2)) for offset in micro_offsets]
    total_weight = sum(raw_weights) or 1
    micro_weights = [w / total_weight for w in raw_weights]

    rows = []
    for _, row in df_cat.iterrows():
        ngay_cat = pd.to_datetime(row.get("ngay_thuc_hien"), errors="coerce")
        if pd.isna(ngay_cat):
            continue
        try:
            qty = int(row.get("so_luong", 0) or 0)
        except (TypeError, ValueError):
            qty = 0
        if qty <= 0:
            continue

        farm_name = row.get("farm") or "Không rõ farm"
        midpoint = ngay_cat + timedelta(days=forecast_days_from_cut)
        for offset, weight in zip(micro_offsets, micro_weights):
            harvest_day = midpoint + timedelta(days=offset)
            iso = harvest_day.isocalendar()
            rows.append({
                "farm": farm_name,
                "year": int(iso.year),
                "week": int(iso.week),
                "_qty_float": qty * weight,
            })

    if not rows:
        return pd.DataFrame(columns=["farm", "year", "week", "forecast_bunches"])

    weekly = pd.DataFrame(rows).groupby(["farm", "year", "week"], as_index=False)["_qty_float"].sum()
    total_target = int(round(weekly["_qty_float"].sum()))
    weekly["_floor"] = weekly["_qty_float"].apply(lambda x: int(x))
    weekly["_remainder"] = weekly["_qty_float"] - weekly["_floor"]
    deficit = total_target - int(weekly["_floor"].sum())
    weekly["forecast_bunches"] = weekly["_floor"]
    if deficit > 0:
        idx = weekly.sort_values("_remainder", ascending=False).head(deficit).index
        weekly.loc[idx, "forecast_bunches"] += 1

    return weekly[["farm", "year", "week", "forecast_bunches"]].sort_values(["year", "week", "farm"])


CONTAINER_SKU_DEFINITIONS = OPTIMIZER_SKU_RULES
CONTAINER_MARKET_OPTIONS = ["Nhật", "Hàn"]
CONTAINER_EMPTY_OPTION = "Không chọn"
CONTAINER_MODE_ORDERS = "Tính số hàng từ số buồng"
CONTAINER_MODE_MAX_CONTS = "Tính số cont tối đa từ số buồng"
CONTAINER_MODE_CONTS_TO_BUNCHES = "Tính số buồng từ số cont"
CONTAINER_LEGACY_MODE_ORDERS = "Theo đơn hàng"
CONTAINER_LEGACY_MODE_ORDERS_ARROW = "Buồng -> Đơn hàng"
CONTAINER_LEGACY_MODE_MAX_CONTS = "Tối đa cont theo thị trường"
CONTAINER_LEGACY_MODE_MAX_CONTS_ARROW = "Buồng -> Tối đa cont"
CONTAINER_LEGACY_MODE_MAX_CONTS_BUNCHES_FIRST = "Buồng -> Số cont tối đa"
CONTAINER_LEGACY_MODE_CONTS_TO_BUNCHES = "Từ cont -> số buồng"
CONTAINER_LEGACY_MODE_CONTS_TO_BUNCHES_ARROW = "Cont -> Số buồng"
CONTAINER_LEGACY_MODE_CONTS_TO_BUNCHES_TARGET = "Cont mục tiêu -> Buồng cần xẻ"
CONTAINER_CUSTOMER_MARKETS = {
    "Wismettac (Nhật 1)": "Nhật",
    "Advance (Nhật 2)": "Nhật",
    "Uone": "Hàn",
}
CONTAINER_CUSTOMER_OPTIONS = list(CONTAINER_CUSTOMER_MARKETS.keys())


def _container_sku_definitions_for_hands(hands_per_bunch: int) -> dict:
    try:
        return get_optimizer_sku_rules(hands_per_bunch)
    except Exception:
        return CONTAINER_SKU_DEFINITIONS


def _container_sku_options_for_market(market: str, hands_per_bunch: int = 12) -> list:
    sku_definitions = _container_sku_definitions_for_hands(hands_per_bunch)
    options = [
        sku for sku, cfg in sku_definitions.items()
        if market in cfg.get("markets", [])
    ]
    return options or list(sku_definitions.keys())


def _container_customer_market(customer: str) -> str:
    return CONTAINER_CUSTOMER_MARKETS.get(_container_clean_text(customer), "")


def _container_is_cont_to_bunch_mode(mode: str) -> bool:
    return _container_clean_text(mode) in {
        CONTAINER_MODE_CONTS_TO_BUNCHES,
        CONTAINER_LEGACY_MODE_CONTS_TO_BUNCHES,
        CONTAINER_LEGACY_MODE_CONTS_TO_BUNCHES_ARROW,
        CONTAINER_LEGACY_MODE_CONTS_TO_BUNCHES_TARGET,
    }


def _container_display_mode_label(mode: str) -> str:
    mode = _container_clean_text(mode, CONTAINER_MODE_ORDERS)
    return {
        CONTAINER_LEGACY_MODE_ORDERS: CONTAINER_MODE_ORDERS,
        CONTAINER_LEGACY_MODE_ORDERS_ARROW: CONTAINER_MODE_ORDERS,
        CONTAINER_LEGACY_MODE_MAX_CONTS: CONTAINER_MODE_MAX_CONTS,
        CONTAINER_LEGACY_MODE_MAX_CONTS_ARROW: CONTAINER_MODE_MAX_CONTS,
        CONTAINER_LEGACY_MODE_MAX_CONTS_BUNCHES_FIRST: CONTAINER_MODE_MAX_CONTS,
        CONTAINER_LEGACY_MODE_CONTS_TO_BUNCHES: CONTAINER_MODE_CONTS_TO_BUNCHES,
        CONTAINER_LEGACY_MODE_CONTS_TO_BUNCHES_ARROW: CONTAINER_MODE_CONTS_TO_BUNCHES,
        CONTAINER_LEGACY_MODE_CONTS_TO_BUNCHES_TARGET: CONTAINER_MODE_CONTS_TO_BUNCHES,
    }.get(mode, mode)


def _default_container_sku_editor_rows() -> list:
    return [{
        "_row_id": "blank_0",
        "Khách hàng": "",
        "Thị trường": "",
        "Mã hàng": "",
        "Nhu cầu": 0,
    }]


def _default_container_target_rows() -> list:
    return [{
        "_cont_id": "target_cont_0",
        "Khách hàng": "",
        "Thị trường": "",
        "items": [{
            "_item_id": "target_item_0",
            "Mã hàng": "",
            "Số thùng": 0,
        }],
    }]


def _unique_keep_order(values: list) -> list:
    result = []
    for value in values:
        text = _container_clean_text(value)
        if text and text not in result:
            result.append(text)
    return result


def _container_clean_text(value, default: str = "") -> str:
    try:
        if value is None or pd.isna(value):
            return default
    except (TypeError, ValueError):
        if value is None:
            return default
    text = str(value).strip()
    return text if text else default


def _container_clean_int(value, default: int = 0) -> int:
    try:
        if value is None or pd.isna(value):
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _container_range_weight(hand_from: int, hand_to: int, hand_weights: dict) -> float:
    return sum(float(hand_weights.get(hand, 0.0)) for hand in range(int(hand_from), int(hand_to) + 1))


def _container_range_formula(hand_from: int, hand_to: int, hand_weights: dict) -> str:
    values = [float(hand_weights.get(hand, 0.0)) for hand in range(int(hand_from), int(hand_to) + 1)]
    parts = " + ".join(f"{value:.2f}" for value in values)
    total = sum(values)
    return f"{parts} = {total:.2f} kg/buồng"


def _container_priority_map(values: list, fallback_values: list) -> dict:
    ordered = _unique_keep_order(values)
    for value in fallback_values:
        text = _container_clean_text(value)
        if text and text not in ordered:
            ordered.append(text)
    return {value: index + 1 for index, value in enumerate(ordered)}


def _normalize_container_target_rows(container_rows: list, hands_per_bunch: int) -> list:
    normalized_conts = []
    for cont_idx, cont in enumerate(container_rows or []):
        cont_id = cont.get("_cont_id") or f"target_cont_{cont_idx}"
        customer = _container_clean_text(cont.get("Khách hàng"))
        if customer not in CONTAINER_CUSTOMER_OPTIONS:
            customer = ""
        market = _container_customer_market(customer)
        sku_options = _container_sku_options_for_market(market, hands_per_bunch) if market else []
        normalized_items = []
        for item_idx, item in enumerate(cont.get("items") or []):
            item_id = item.get("_item_id") or f"{cont_id}_item_{item_idx}"
            sku = _container_clean_text(item.get("Mã hàng")).upper()
            if sku not in sku_options:
                sku = ""
            normalized_items.append({
                "_item_id": item_id,
                "Mã hàng": sku,
                "Số thùng": max(0, _container_clean_int(item.get("Số thùng"))),
            })
        if not normalized_items:
            normalized_items = [{
                "_item_id": f"{cont_id}_item_0",
                "Mã hàng": "",
                "Số thùng": 0,
            }]
        normalized_conts.append({
            "_cont_id": cont_id,
            "Khách hàng": customer,
            "Thị trường": market,
            "items": normalized_items,
        })
    return normalized_conts or _default_container_target_rows()


def _prepare_container_rows_from_container_targets(container_rows: list, hands_per_bunch: int) -> tuple[list, list]:
    errors = []
    grouped_rows = {}
    row_order = {}
    order_counter = 0
    for cont_idx, cont in enumerate(container_rows or [], start=1):
        customer = _container_clean_text(cont.get("Khách hàng"))
        market = _container_clean_text(cont.get("Thị trường")) or _container_customer_market(customer)
        if not customer:
            errors.append(f"Cont #{cont_idx}: chưa chọn khách hàng.")
        if not market:
            errors.append(f"Cont #{cont_idx}: chưa xác định được thị trường.")

        allowed_skus = _container_sku_options_for_market(market, hands_per_bunch) if market else []
        item_total = 0
        has_positive_item = False
        for item_idx, item in enumerate(cont.get("items") or [], start=1):
            sku = _container_clean_text(item.get("Mã hàng")).upper()
            boxes = max(0, _container_clean_int(item.get("Số thùng")))
            item_total += boxes
            if boxes <= 0 and not sku:
                continue
            has_positive_item = True
            if boxes <= 0:
                errors.append(f"Cont #{cont_idx}, dòng {item_idx}: số thùng phải lớn hơn 0.")
                continue
            if not sku:
                errors.append(f"Cont #{cont_idx}, dòng {item_idx}: chưa chọn mã hàng.")
                continue
            if sku not in allowed_skus:
                errors.append(f"Cont #{cont_idx}, dòng {item_idx}: mã {sku} không hợp lệ cho thị trường {market}.")
                continue
            key = (customer, market, sku)
            if key not in grouped_rows:
                order_counter += 1
                grouped_rows[key] = {
                    "customer_priority": order_counter,
                    "customer": customer,
                    "market_priority": order_counter,
                    "market": market,
                    "sku_priority": order_counter,
                    "sku": sku,
                    "demand": 0,
                    "unit": "Thùng",
                }
                row_order[key] = order_counter
            grouped_rows[key]["demand"] += boxes

        if not has_positive_item:
            errors.append(f"Cont #{cont_idx}: chưa có dòng mã hàng hợp lệ.")
        if item_total != BOXES_PER_CONTAINER:
            diff = BOXES_PER_CONTAINER - item_total
            if diff > 0:
                errors.append(f"Cont #{cont_idx}: còn thiếu {diff:,} thùng để đủ 1 cont.")
            else:
                errors.append(f"Cont #{cont_idx}: đang vượt {-diff:,} thùng so với 1 cont.")

    allocation_rows = sorted(grouped_rows.values(), key=lambda row: row_order[(row["customer"], row["market"], row["sku"])])
    return allocation_rows, errors


def _prepare_container_rows_for_allocation(order_rows: list, customer_order: list, sku_orders_by_customer: dict) -> list:
    all_customers = [row.get("Khách hàng") for row in order_rows]
    customer_priority = _container_priority_map(customer_order, all_customers)
    allocation_rows = []

    for row_index, row in enumerate(order_rows):
        customer = _container_clean_text(row.get("Khách hàng"))
        market = _container_clean_text(row.get("Thị trường"))
        sku = _container_clean_text(row.get("Mã hàng")).upper()
        demand = _container_clean_int(row.get("Nhu cầu"))
        sku_def = CONTAINER_SKU_DEFINITIONS.get(sku)
        if not customer or not market or not sku_def or demand <= 0:
            continue

        customer_skus = [
            item.get("Mã hàng")
            for item in order_rows
            if _container_clean_text(item.get("Khách hàng")) == customer
        ]
        sku_priority = _container_priority_map(sku_orders_by_customer.get(customer, []), customer_skus)
        priority = customer_priority.get(customer, 999)
        allocation_rows.append({
            "customer_priority": priority,
            "customer": customer,
            "market_priority": priority,
            "market": market,
            "sku_priority": sku_priority.get(sku, 999),
            "sku": sku,
            "demand": demand,
            "unit": "Thùng",
            "_row_index": row_index,
        })

    return allocation_rows


CONTAINER_SAVED_PLANS_KEY = "container_calc_saved_plans"
CONTAINER_SAVED_PLANS_TABLE = "container_allocation_plans"


def _container_current_account() -> tuple[str, str]:
    return (
        _container_clean_text(st.session_state.get("current_farm"), "Phòng Kinh doanh"),
        _container_clean_text(st.session_state.get("current_team"), "Kinh doanh"),
    )


def _container_json_safe(value):
    try:
        return json.loads(json.dumps(value, default=str))
    except (TypeError, ValueError):
        return {}


def _container_normalize_hand_weights(hand_weights) -> dict[int, float]:
    if not isinstance(hand_weights, dict):
        return {}
    normalized = {}
    for hand, weight in hand_weights.items():
        try:
            normalized[int(hand)] = float(weight)
        except (TypeError, ValueError):
            continue
    return dict(sorted(normalized.items()))


def _container_plan_from_db(row: dict) -> dict:
    plan = row.get("full_plan") if isinstance(row.get("full_plan"), dict) else {}
    plan = dict(plan)
    result_data = row.get("result_data") if isinstance(row.get("result_data"), dict) else {}
    summary = row.get("summary") if isinstance(row.get("summary"), dict) else {}
    if result_data and not isinstance(result_data.get("summary"), dict):
        result_data["summary"] = summary

    plan.setdefault("result", result_data)
    plan.setdefault("mode", row.get("mode", ""))
    plan.setdefault("source_mode", row.get("source_mode", ""))
    plan.setdefault("source", row.get("source_label", ""))
    plan.setdefault("source_bunches", row.get("source_bunches", 0))
    plan.setdefault("hands_per_bunch", row.get("hands_per_bunch", 0))
    plan.setdefault("kg_per_bunch", row.get("kg_per_bunch", 0))
    plan["id"] = row.get("id")
    plan["name"] = row.get("plan_name") or plan.get("name") or _container_saved_plan_title(plan)
    plan["saved_at"] = row.get("created_at") or plan.get("saved_at")
    return plan


def _container_saved_plans() -> list:
    account_farm, account_team = _container_current_account()
    try:
        res = (
            supabase.table(CONTAINER_SAVED_PLANS_TABLE)
            .select("*")
            .eq("account_farm", account_farm)
            .eq("account_team", account_team)
            .eq("is_deleted", False)
            .order("created_at", desc=True)
            .execute()
        )
        plans = [_container_plan_from_db(row) for row in (res.data or [])]
        st.session_state[CONTAINER_SAVED_PLANS_KEY] = plans
        return plans
    except Exception as exc:
        if not st.session_state.get("container_saved_plan_db_warning_shown"):
            st.warning(f"Chưa tải được kế hoạch đã lưu từ DB: {exc}")
            st.session_state["container_saved_plan_db_warning_shown"] = True
        plans = st.session_state.get(CONTAINER_SAVED_PLANS_KEY)
        return plans if isinstance(plans, list) else []


def _container_format_saved_at(saved_at: str) -> str:
    try:
        saved_dt = datetime.fromisoformat(str(saved_at).replace("Z", "+00:00"))
        return saved_dt.strftime("%H:%M - %d/%m/%Y")
    except (TypeError, ValueError):
        return "Chưa rõ thời gian"


def _container_saved_plan_title(plan: dict, index: int = 0) -> str:
    mode = _container_display_mode_label(plan.get("mode"))
    saved_at = _container_format_saved_at(plan.get("saved_at"))
    return f"{mode} · {saved_at}" if saved_at else f"Kế hoạch {index + 1}"


def _save_container_plan(plan: dict) -> bool:
    account_farm, account_team = _container_current_account()
    plan = dict(plan)
    plan["saved_at"] = datetime.now().isoformat(timespec="seconds")
    plan["name"] = _container_saved_plan_title(plan)
    result = plan.get("result") if isinstance(plan.get("result"), dict) else {}
    summary = result.get("summary") if isinstance(result.get("summary"), dict) else {}
    input_data = {
        "source_mode": plan.get("source_mode"),
        "source": plan.get("source"),
        "source_bunches": plan.get("source_bunches"),
        "kg_per_bunch": plan.get("kg_per_bunch"),
        "hands_per_bunch": plan.get("hands_per_bunch"),
        "hand_weights": plan.get("hand_weights"),
        "sku_rows": plan.get("sku_rows"),
        "container_targets": plan.get("container_targets"),
        "customer_order": plan.get("customer_order"),
        "sku_orders_by_customer": plan.get("sku_orders_by_customer"),
        "market_order": plan.get("market_order"),
    }
    record = {
        "account_farm": account_farm,
        "account_team": account_team,
        "plan_name": plan["name"],
        "mode": plan.get("mode", ""),
        "source_mode": plan.get("source_mode", ""),
        "source_label": plan.get("source", ""),
        "source_bunches": int(plan.get("source_bunches", 0)),
        "hands_per_bunch": int(plan.get("hands_per_bunch", 0)),
        "kg_per_bunch": float(plan.get("kg_per_bunch", 0)),
        "input_data": _container_json_safe(input_data),
        "result_data": _container_json_safe(result),
        "summary": _container_json_safe(summary),
        "full_plan": _container_json_safe(plan),
    }
    try:
        supabase.table(CONTAINER_SAVED_PLANS_TABLE).insert(record).execute()
        st.session_state.pop("container_saved_plan_db_warning_shown", None)
        return True
    except Exception as exc:
        st.error(f"Không lưu được kế hoạch vào DB: {exc}")
        return False


def _delete_container_plan(plan_id: str) -> bool:
    account_farm, account_team = _container_current_account()
    try:
        (
            supabase.table(CONTAINER_SAVED_PLANS_TABLE)
            .update({"is_deleted": True, "updated_at": datetime.now(timezone.utc).isoformat()})
            .eq("id", plan_id)
            .eq("account_farm", account_farm)
            .eq("account_team", account_team)
            .execute()
        )
        return True
    except Exception as exc:
        st.error(f"Không xóa được kế hoạch: {exc}")
        return False


def _container_plan_summary_rows(plan: dict) -> list:
    result = plan.get("result", {}) if isinstance(plan.get("result"), dict) else {}
    summary = result.get("summary", {}) if isinstance(result.get("summary"), dict) else {}
    if _container_is_cont_to_bunch_mode(plan.get("mode")):
        rows = [
            {"Thông tin": "Chế độ", "Giá trị": _container_display_mode_label(plan.get("mode"))},
            {"Thông tin": "Cont mục tiêu", "Giá trị": f"{int(summary.get('target_containers', 0)):,}"},
            {"Thông tin": "Tổng thùng", "Giá trị": f"{int(summary.get('requested_boxes', summary.get('fulfilled_boxes', 0))):,}"},
            {"Thông tin": "Loại buồng", "Giá trị": f"{int(plan.get('hands_per_bunch', 0))} nải"},
            {"Thông tin": "Kg/buồng", "Giá trị": f"{float(plan.get('kg_per_bunch', 0)):.2f}"},
            {"Thông tin": "Buồng xẻ tối thiểu", "Giá trị": f"{int(summary.get('active_bunches_estimated', 0)):,}"},
            {"Thông tin": "Kg cần xẻ", "Giá trị": f"{float(summary.get('source_kg', 0)):,.1f}"},
            {"Thông tin": "Kg dư", "Giá trị": f"{float(summary.get('extra_kg_from_rounding', 0)):,.1f}"},
            {"Thông tin": "Thuật toán", "Giá trị": f"{summary.get('solver_status', '')} · {summary.get('solver_backend', '')}".strip(" ·")},
            {"Thông tin": "Thời điểm lưu", "Giá trị": _container_format_saved_at(plan.get("saved_at"))},
        ]
        return rows
    rows = [
        {"Thông tin": "Chế độ", "Giá trị": _container_display_mode_label(plan.get("mode"))},
        {"Thông tin": "Nguồn", "Giá trị": plan.get("source", "")},
        {"Thông tin": "Số buồng nguồn", "Giá trị": f"{int(plan.get('source_bunches', 0)):,}"},
        {"Thông tin": "Loại buồng", "Giá trị": f"{int(plan.get('hands_per_bunch', 0))} nải"},
        {"Thông tin": "Kg/buồng", "Giá trị": f"{float(plan.get('kg_per_bunch', 0)):.2f}"},
        {"Thông tin": "Buồng xẻ tối thiểu", "Giá trị": f"{int(summary.get('active_bunches_estimated', 0)):,}"},
        {"Thông tin": "Thùng đáp ứng", "Giá trị": f"{int(summary.get('fulfilled_boxes', 0)):,}"},
        {"Thông tin": "Cont đáp ứng", "Giá trị": f"{float(summary.get('fulfilled_containers', 0)):,.2f}"},
        {"Thông tin": "Thuật toán", "Giá trị": f"{summary.get('solver_status', '')} · {summary.get('solver_backend', '')}".strip(" ·")},
        {"Thông tin": "Thời điểm lưu", "Giá trị": _container_format_saved_at(plan.get("saved_at"))},
    ]
    return rows


def _container_plan_market_rows(plan: dict) -> pd.DataFrame:
    result = plan.get("result", {}) if isinstance(plan.get("result"), dict) else {}
    rows = result.get("rows", [])
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    if "market" not in df.columns:
        return pd.DataFrame()

    if "full_containers" in df.columns:
        return pd.DataFrame([
            {
                "Thị trường": row.get("market", ""),
                "Cont đủ": int(row.get("full_containers", 0)),
                "Thùng chốt": int(row.get("boxes_allocated", 0)),
                "Thùng lẻ": int(row.get("remaining_boxes_potential", 0)),
                "Buồng dùng": int(row.get("active_bunches_used", row.get("bundles_used", 0))),
                "Kg phân bổ": round(float(row.get("kg_allocated", 0)), 1),
            }
            for row in rows
        ])

    grouped = (
        df.groupby("market", as_index=False)
        .agg(
            requested_boxes=("requested_boxes", "sum"),
            boxes_fulfilled=("boxes_fulfilled", "sum"),
            short_boxes=("short_boxes", "sum"),
        )
        .sort_values("market")
    )
    grouped["full_containers"] = grouped["boxes_fulfilled"] // BOXES_PER_CONTAINER
    grouped["remaining_boxes"] = grouped["boxes_fulfilled"] % BOXES_PER_CONTAINER
    return grouped.rename(columns={
        "market": "Thị trường",
        "requested_boxes": "Thùng yêu cầu",
        "boxes_fulfilled": "Thùng đáp ứng",
        "short_boxes": "Thiếu thùng",
        "full_containers": "Cont đủ",
        "remaining_boxes": "Thùng lẻ",
    })


def _container_plan_output_rows(plan: dict) -> pd.DataFrame:
    result = plan.get("result", {}) if isinstance(plan.get("result"), dict) else {}
    rows = result.get("rows", [])
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)

    if "full_containers" in df.columns:
        output_rows = []
        detail_df = pd.DataFrame(result.get("detail_rows", []))
        for _, market_row in df.iterrows():
            market = market_row.get("market", "")
            output_rows.append({
                "Thị trường": market,
                "Mã hàng": "Tổng",
                "Nải dùng": "",
                "Cont đủ": int(market_row.get("full_containers", 0)),
                "Thùng chốt": int(market_row.get("boxes_allocated", 0)),
                "Thùng lẻ": int(market_row.get("remaining_boxes_potential", 0)),
                "Buồng dùng": int(market_row.get("active_bunches_used", market_row.get("bundles_used", 0))),
                "Kg phân bổ": f"{float(market_row.get('kg_allocated', 0)):,.0f}",
            })
            if not detail_df.empty:
                for _, detail_row in detail_df[detail_df["market"] == market].iterrows():
                    output_rows.append({
                        "Thị trường": market,
                        "Mã hàng": detail_row.get("sku", ""),
                        "Nải dùng": detail_row.get("range_label", ""),
                        "Cont đủ": "",
                        "Thùng chốt": int(detail_row.get("boxes_equivalent", 0)),
                        "Thùng lẻ": "",
                        "Buồng dùng": int(detail_row.get("bundles_used", 0)),
                        "Kg phân bổ": f"{float(detail_row.get('kg_allocated', 0)):,.0f}",
                    })
        return pd.DataFrame(output_rows)

    if "market" not in df.columns:
        return pd.DataFrame()
    market_summary = (
        df.groupby("market", as_index=False)
        .agg(
            requested_boxes=("requested_boxes", "sum"),
            boxes_fulfilled=("boxes_fulfilled", "sum"),
            short_boxes=("short_boxes", "sum"),
        )
        .sort_values("market")
    )
    market_summary["full_containers"] = market_summary["boxes_fulfilled"] // BOXES_PER_CONTAINER
    market_summary["remaining_boxes"] = market_summary["boxes_fulfilled"] % BOXES_PER_CONTAINER
    output_rows = []
    for _, market_row in market_summary.iterrows():
        output_rows.append({
            "Thị trường": market_row["market"],
            "Khách hàng": "",
            "Mã hàng": "Tổng",
            "Thùng yêu cầu": int(market_row["requested_boxes"]),
            "Thùng đáp ứng": int(market_row["boxes_fulfilled"]),
            "Thiếu thùng": int(market_row["short_boxes"]),
            "Cont đủ": int(market_row["full_containers"]),
            "Thùng lẻ": int(market_row["remaining_boxes"]),
            "Buồng xẻ tối thiểu": "",
        })
    sort_cols = [
        col for col in ("customer_priority", "sku_priority", "processing_order")
        if col in df.columns
    ]
    if sort_cols:
        df = df.sort_values(sort_cols)
    for _, sku_row in df.iterrows():
        output_rows.append({
            "Thị trường": sku_row.get("market", ""),
            "Khách hàng": sku_row.get("customer", ""),
            "Mã hàng": sku_row.get("sku", ""),
            "Thùng yêu cầu": int(sku_row.get("requested_boxes", 0)),
            "Thùng đáp ứng": int(sku_row.get("boxes_fulfilled", 0)),
            "Thiếu thùng": int(sku_row.get("short_boxes", 0)),
            "Cont đủ": "",
            "Thùng lẻ": "",
            "Buồng xẻ tối thiểu": int(sku_row.get("bunches_allocated", 0)),
        })
    return pd.DataFrame(output_rows)


def _container_plan_process_rows(plan: dict) -> pd.DataFrame:
    result = plan.get("result", {}) if isinstance(plan.get("result"), dict) else {}
    detail_rows = result.get("detail_rows") or result.get("rows") or []
    if not detail_rows:
        return pd.DataFrame()
    df = pd.DataFrame(detail_rows)
    hand_weights = _container_normalize_hand_weights(plan.get("hand_weights"))
    sort_cols = [
        col for col in ("market_priority", "customer_priority", "sku_priority", "processing_order")
        if col in df.columns
    ]
    if sort_cols:
        df = df.sort_values(sort_cols)

    output_rows = []
    for _, row in df.iterrows():
        try:
            hand_from = int(row.get("hand_from", 0))
            hand_to = int(row.get("hand_to", 0))
        except (TypeError, ValueError):
            hand_from = 0
            hand_to = 0
        bunches_used = int(row.get("bunches_allocated", row.get("bundles_used", 0)) or 0)
        if hand_from > 0 and hand_to >= hand_from:
            calc_text = (
                f"{bunches_used:,} buồng x "
                f"({_container_range_formula(hand_from, hand_to, hand_weights)}) "
                f"= {float(row.get('kg_allocated', 0)):,.1f} kg"
            )
        else:
            calc_text = "Chưa phân bổ"
        boxes_value = row.get("boxes_equivalent", row.get("boxes_fulfilled", 0))
        result_text = f"~{float(boxes_value or 0):,.2f} thùng quy đổi"
        if int(row.get("short_boxes", 0) or 0) > 0:
            result_text += f", thiếu {int(row.get('short_boxes', 0)):,} thùng"
        output_rows.append({
            "Bước": int(row.get("processing_order", row.get("market_priority", 0)) or 0) or "",
            "Thị trường": row.get("market", ""),
            "Khách hàng": row.get("customer", ""),
            "Mã hàng": row.get("sku", ""),
            "Nải chọn": row.get("range_label", ""),
            "Cách tính": calc_text,
            "Kết quả": result_text,
        })
    return pd.DataFrame(output_rows)


def _container_plan_order_input_rows(plan: dict) -> pd.DataFrame:
    rows = plan.get("sku_rows") or []
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    columns = [col for col in ["Khách hàng", "Thị trường", "Mã hàng", "Nhu cầu"] if col in df.columns]
    return df[columns] if columns else df


def _container_plan_container_input_rows(plan: dict) -> pd.DataFrame:
    container_targets = plan.get("container_targets") or []
    output_rows = []
    for cont_idx, cont in enumerate(container_targets, start=1):
        customer = _container_clean_text(cont.get("Khách hàng"))
        market = _container_clean_text(cont.get("Thị trường"))
        for item in cont.get("items") or []:
            sku = _container_clean_text(item.get("Mã hàng")).upper()
            boxes = _container_clean_int(item.get("Số thùng"))
            if not sku and boxes <= 0:
                continue
            output_rows.append({
                "Cont": f"Cont #{cont_idx}",
                "Khách hàng": customer,
                "Thị trường": market,
                "Mã hàng": sku,
                "Số thùng": boxes,
            })
    return pd.DataFrame(output_rows)


@dialog_decorator("Chi tiết kế hoạch")
def _render_container_saved_plan_dialog(plan_id: str):
    plan = next(
        (item for item in _container_saved_plans() if isinstance(item, dict) and item.get("id") == plan_id),
        None,
    )
    if not plan:
        st.warning("Kế hoạch này không còn tồn tại.")
        return

    st.markdown(f"#### {plan.get('name') or _container_saved_plan_title(plan)}")
    st.dataframe(pd.DataFrame(_container_plan_summary_rows(plan)), use_container_width=True, hide_index=True)

    hand_weights = _container_normalize_hand_weights(plan.get("hand_weights"))
    if hand_weights:
        with st.expander("Kg từng nải", expanded=False):
            st.dataframe(
                pd.DataFrame([
                    {"Nải": hand, "Kg sau quy đổi": round(float(weight), 3)}
                    for hand, weight in hand_weights.items()
                ]),
                use_container_width=True,
                hide_index=True,
            )

    if plan.get("sku_rows") and not plan.get("container_targets"):
        with st.expander("Input đơn hàng", expanded=True):
            st.dataframe(_container_plan_order_input_rows(plan), use_container_width=True, hide_index=True)

    if plan.get("container_targets"):
        with st.expander("Input cont", expanded=True):
            st.dataframe(_container_plan_container_input_rows(plan), use_container_width=True, hide_index=True)

    priority_rows = []
    for idx, customer in enumerate(plan.get("customer_order", []) or [], start=1):
        priority_rows.append({"Nhóm ưu tiên": "Khách hàng", "Ưu tiên": idx, "Giá trị": customer})
    for customer, skus in (plan.get("sku_orders_by_customer", {}) or {}).items():
        for idx, sku in enumerate(skus or [], start=1):
            priority_rows.append({"Nhóm ưu tiên": f"Loại hàng - {customer}", "Ưu tiên": idx, "Giá trị": sku})
    for idx, market in enumerate(plan.get("market_order", []) or [], start=1):
        priority_rows.append({"Nhóm ưu tiên": "Thị trường", "Ưu tiên": idx, "Giá trị": market})
    if priority_rows:
        with st.expander("Input ưu tiên", expanded=False):
            st.dataframe(pd.DataFrame(priority_rows), use_container_width=True, hide_index=True)

    output_df = _container_plan_output_rows(plan)
    if not output_df.empty:
        with st.expander("Kết quả phân bổ", expanded=True):
            st.dataframe(output_df, use_container_width=True, hide_index=True)

    process_df = _container_plan_process_rows(plan)
    if not process_df.empty:
        with st.expander("Quá trình chọn nải chi tiết", expanded=False):
            st.dataframe(process_df, use_container_width=True, hide_index=True)

    result = plan.get("result", {}) if isinstance(plan.get("result"), dict) else {}
    if result.get("remaining_hands"):
        with st.expander("Tồn nải còn lại", expanded=False):
            st.dataframe(
                pd.DataFrame([
                    {"Nải": hand, "Buồng còn lại": qty}
                    for hand, qty in result["remaining_hands"].items()
                ]),
                use_container_width=True,
                hide_index=True,
            )
    with st.expander("Dữ liệu đầy đủ", expanded=False):
        st.json(_container_json_safe(plan), expanded=False)


def _render_container_saved_plan_cards():
    plans = _container_saved_plans()
    if not plans:
        return

    st.divider()
    st.markdown("##### Kế hoạch đã lưu")
    for idx, plan in enumerate(plans):
        if not isinstance(plan, dict):
            continue
        plan_id = plan.get("id") or f"plan_{idx}"
        result = plan.get("result", {}) if isinstance(plan.get("result"), dict) else {}
        summary = result.get("summary", {}) if isinstance(result.get("summary"), dict) else {}
        with st.container(border=True):
            title_col, b_col, box_col, action_col, delete_col = st.columns([3, 1, 1, 0.8, 0.7])
            with title_col:
                st.markdown(f"**{plan.get('name') or _container_saved_plan_title(plan, idx)}**")
                if _container_is_cont_to_bunch_mode(plan.get("mode")):
                    st.caption(
                        f"{int(summary.get('target_containers', 0)):,} cont mục tiêu · "
                        f"{int(plan.get('hands_per_bunch', 0))} nải · "
                        f"{float(plan.get('kg_per_bunch', 0)):.1f} kg/buồng"
                    )
                else:
                    st.caption(f"{plan.get('source', '')} · {int(plan.get('source_bunches', 0)):,} buồng · {int(plan.get('hands_per_bunch', 0))} nải · {float(plan.get('kg_per_bunch', 0)):.1f} kg/buồng")
            with b_col:
                st.metric("Buồng xẻ", f"{int(summary.get('active_bunches_estimated', 0)):,}")
            with box_col:
                st.metric("Thùng", f"{int(summary.get('fulfilled_boxes', 0)):,}")
            with action_col:
                if st.button("Xem", key=f"container_plan_open_{plan_id}", use_container_width=True):
                    _render_container_saved_plan_dialog(plan_id)
            with delete_col:
                if st.button("Xóa", key=f"container_plan_delete_{plan_id}", use_container_width=True):
                    if _delete_container_plan(plan_id):
                        st.rerun()


def _render_container_target_controls(hands_per_bunch: int) -> list:
    state_key = "container_calc_target_conts"
    if state_key not in st.session_state:
        st.session_state[state_key] = _default_container_target_rows()

    current_conts = _normalize_container_target_rows(st.session_state.get(state_key), hands_per_bunch)
    rendered_conts = []
    delete_cont_index = None

    for cont_idx, cont in enumerate(current_conts):
        cont_id = cont["_cont_id"]
        with st.container(border=True):
            title_col, customer_col, delete_col = st.columns([1.2, 3, 0.8])
            with title_col:
                st.markdown(f"###### Cont #{cont_idx + 1}")
            with customer_col:
                customer_options = [CONTAINER_EMPTY_OPTION] + CONTAINER_CUSTOMER_OPTIONS
                customer_key = f"container_target_customer_{cont_id}"
                if st.session_state.get(customer_key) not in customer_options:
                    st.session_state[customer_key] = cont["Khách hàng"] if cont["Khách hàng"] in CONTAINER_CUSTOMER_OPTIONS else None
                customer = st.selectbox(
                    "Khách hàng",
                    options=customer_options,
                    index=None,
                    placeholder="Chọn khách hàng",
                    key=customer_key,
                )
                if customer == CONTAINER_EMPTY_OPTION:
                    customer = ""
                market = _container_customer_market(customer)
                if market:
                    st.caption(f"Thị trường: {market}")
            with delete_col:
                st.markdown("")
                if st.button("Xóa cont", key=f"container_target_delete_cont_{cont_id}", disabled=len(current_conts) <= 1, use_container_width=True):
                    delete_cont_index = cont_idx

            item_rows = []
            delete_item_index = None
            sku_options = _container_sku_options_for_market(market, hands_per_bunch) if market else []
            item_header = st.columns([2, 1.2, 0.7])
            item_header[0].caption("Mã hàng")
            item_header[1].caption("Số thùng")
            for item_idx, item in enumerate(cont.get("items") or []):
                item_id = item["_item_id"]
                item_cols = st.columns([2, 1.2, 0.7])
                with item_cols[0]:
                    sku_select_options = [CONTAINER_EMPTY_OPTION] + sku_options
                    sku_key = f"container_target_sku_{cont_id}_{item_id}"
                    if st.session_state.get(sku_key) not in sku_select_options:
                        st.session_state[sku_key] = item["Mã hàng"] if item["Mã hàng"] in sku_options else None
                    sku = st.selectbox(
                        "Mã hàng",
                        options=sku_select_options,
                        index=None,
                        placeholder="Chọn mã hàng" if market else "Chọn khách hàng trước",
                        key=sku_key,
                        label_visibility="collapsed",
                        disabled=not bool(market),
                    )
                    if sku == CONTAINER_EMPTY_OPTION:
                        sku = ""
                with item_cols[1]:
                    boxes = st.number_input(
                        "Số thùng",
                        min_value=0,
                        value=max(0, _container_clean_int(item.get("Số thùng"))),
                        step=1,
                        key=f"container_target_boxes_{cont_id}_{item_id}",
                        label_visibility="collapsed",
                    )
                with item_cols[2]:
                    if st.button("Xóa", key=f"container_target_delete_item_{cont_id}_{item_id}", disabled=len(cont.get("items") or []) <= 1, use_container_width=True):
                        delete_item_index = item_idx
                item_rows.append({
                    "_item_id": item_id,
                    "Mã hàng": sku or "",
                    "Số thùng": boxes,
                })

            if delete_item_index is not None:
                item_rows = [item for idx, item in enumerate(item_rows) if idx != delete_item_index]
                st.session_state[state_key] = rendered_conts + [{
                    "_cont_id": cont_id,
                    "Khách hàng": customer or "",
                    "Thị trường": market,
                    "items": item_rows,
                }] + current_conts[cont_idx + 1:]
                st.rerun()

            add_cols = st.columns([1, 4])
            with add_cols[0]:
                if st.button("Thêm mã", key=f"container_target_add_item_{cont_id}", use_container_width=True):
                    item_rows.append({
                        "_item_id": f"target_item_{datetime.now().timestamp()}",
                        "Mã hàng": "",
                        "Số thùng": 0,
                    })
                    st.session_state[state_key] = rendered_conts + [{
                        "_cont_id": cont_id,
                        "Khách hàng": customer or "",
                        "Thị trường": market,
                        "items": item_rows,
                    }] + current_conts[cont_idx + 1:]
                    st.rerun()

            total_boxes = sum(_container_clean_int(item.get("Số thùng")) for item in item_rows)
            if total_boxes == BOXES_PER_CONTAINER:
                st.success(f"Cont #{cont_idx + 1}: đủ {BOXES_PER_CONTAINER:,} thùng.")
            elif total_boxes < BOXES_PER_CONTAINER:
                st.warning(f"Cont #{cont_idx + 1}: còn thiếu {BOXES_PER_CONTAINER - total_boxes:,} thùng để đủ 1 cont.")
            else:
                st.error(f"Cont #{cont_idx + 1}: vượt {total_boxes - BOXES_PER_CONTAINER:,} thùng so với 1 cont.")

            rendered_conts.append({
                "_cont_id": cont_id,
                "Khách hàng": customer or "",
                "Thị trường": market,
                "items": item_rows,
            })

    if delete_cont_index is not None:
        rendered_conts = [cont for idx, cont in enumerate(rendered_conts) if idx != delete_cont_index]
        st.session_state[state_key] = rendered_conts or _default_container_target_rows()
        st.rerun()

    add_cont_cols = st.columns([1, 5])
    with add_cont_cols[0]:
        if st.button("Thêm cont", key="container_target_add_cont", use_container_width=True):
            rendered_conts.append({
                "_cont_id": f"target_cont_{datetime.now().timestamp()}",
                "Khách hàng": "",
                "Thị trường": "",
                "items": [{
                    "_item_id": f"target_item_{datetime.now().timestamp()}",
                    "Mã hàng": "",
                    "Số thùng": 0,
                }],
            })
            st.session_state[state_key] = rendered_conts
            st.rerun()

    st.session_state[state_key] = rendered_conts
    return rendered_conts


def _render_container_customer_priority_controls(key_prefix: str) -> list:
    priority_customer_options = [CONTAINER_EMPTY_OPTION] + CONTAINER_CUSTOMER_OPTIONS
    customer_order = []
    p_cols = st.columns(3)
    for idx, col in enumerate(p_cols, start=1):
        with col:
            customer_key = f"{key_prefix}_{idx}"
            if st.session_state.get(customer_key) not in priority_customer_options:
                st.session_state[customer_key] = None
            selected_customer = st.selectbox(
                f"Khách hàng ưu tiên {idx}",
                options=priority_customer_options,
                index=None,
                placeholder="Không chọn",
                key=customer_key,
            )
            if (
                selected_customer
                and selected_customer != CONTAINER_EMPTY_OPTION
                and selected_customer not in customer_order
            ):
                customer_order.append(selected_customer)
    return customer_order


def _render_container_market_priority_controls(key_prefix: str, default_all_markets: bool = True) -> list:
    priority_market_options = ["Không chọn"] + CONTAINER_MARKET_OPTIONS
    p_cols = st.columns(2)
    market_order = []
    for idx, col in enumerate(p_cols, start=1):
        with col:
            market_key = f"{key_prefix}_{idx}"
            if st.session_state.get(market_key) not in priority_market_options:
                st.session_state[market_key] = (
                    priority_market_options[idx]
                    if default_all_markets and idx < len(priority_market_options)
                    else "Không chọn"
                )
            selected_market = st.selectbox(
                f"Thị trường ưu tiên {idx}",
                options=priority_market_options,
                key=market_key,
            )
            if selected_market != "Không chọn" and selected_market not in market_order:
                market_order.append(selected_market)

    if default_all_markets:
        for market in CONTAINER_MARKET_OPTIONS:
            if market not in market_order:
                market_order.append(market)
    return market_order


def render_container_allocation_calculator():
    st.markdown("#### Máy tính phân bổ container theo nải")

    if st.session_state.get("container_calc_customer_schema_version") != 1:
        st.session_state["container_calc_sku_rows"] = _default_container_sku_editor_rows()
        st.session_state["container_calc_customer_schema_version"] = 1
    if "container_calc_sku_rows" not in st.session_state:
        st.session_state["container_calc_sku_rows"] = _default_container_sku_editor_rows()

    calc_mode_options = [
        CONTAINER_MODE_ORDERS,
        CONTAINER_MODE_MAX_CONTS,
        CONTAINER_MODE_CONTS_TO_BUNCHES,
    ]
    legacy_calc_mode_map = {
        CONTAINER_LEGACY_MODE_ORDERS: CONTAINER_MODE_ORDERS,
        CONTAINER_LEGACY_MODE_ORDERS_ARROW: CONTAINER_MODE_ORDERS,
        CONTAINER_LEGACY_MODE_MAX_CONTS: CONTAINER_MODE_MAX_CONTS,
        CONTAINER_LEGACY_MODE_MAX_CONTS_ARROW: CONTAINER_MODE_MAX_CONTS,
        CONTAINER_LEGACY_MODE_MAX_CONTS_BUNCHES_FIRST: CONTAINER_MODE_MAX_CONTS,
        CONTAINER_LEGACY_MODE_CONTS_TO_BUNCHES: CONTAINER_MODE_CONTS_TO_BUNCHES,
        CONTAINER_LEGACY_MODE_CONTS_TO_BUNCHES_ARROW: CONTAINER_MODE_CONTS_TO_BUNCHES,
        CONTAINER_LEGACY_MODE_CONTS_TO_BUNCHES_TARGET: CONTAINER_MODE_CONTS_TO_BUNCHES,
    }
    current_mode = st.session_state.get("container_calculation_mode")
    if current_mode in legacy_calc_mode_map:
        st.session_state["container_calculation_mode"] = legacy_calc_mode_map[current_mode]
    calc_mode = _persisted_segmented_control(
        "Chế độ tính",
        calc_mode_options,
        key="container_calculation_mode",
        query_key="container_calc_mode",
        slugs={
            calc_mode_options[0]: "orders",
            calc_mode_options[1]: "max_by_market",
            calc_mode_options[2]: "containers_to_bunches",
        },
        default=calc_mode_options[0],
        label_visibility="visible",
    )

    source_bunches = 0
    source_label = "Chưa chọn nguồn"
    if calc_mode == CONTAINER_MODE_CONTS_TO_BUNCHES:
        source_mode = CONTAINER_MODE_CONTS_TO_BUNCHES
        source_label = "Tính ngược từ cơ cấu cont"
    else:
        source_mode = st.radio(
            "Nguồn số buồng",
            ["Dự báo từ cắt bắp", "Nhập tay"],
            horizontal=True,
            key="container_source_mode",
        )

    if source_mode == "Dự báo từ cắt bắp":
        df_stg_all = fetch_table_data("stage_logs", "Phòng Kinh doanh")
        lead_col, f_col, y_col, w_col = st.columns([1.2, 1.6, 1.6, 1.6])
        with lead_col:
            selected_forecast_weeks = st.selectbox(
                "Cách dự báo",
                options=CAT_FORECAST_WEEK_OPTIONS,
                format_func=lambda value: f"Dự báo +{value} tuần",
                key="container_forecast_weeks_inclusive",
            )
        weekly_forecast = build_weekly_cat_forecast(df_stg_all, selected_forecast_weeks)
        if weekly_forecast.empty:
            st.warning("Chưa có dữ liệu cắt bắp để dự báo theo tuần.")
        else:
            with f_col:
                farm_options = ["Tất cả"] + sorted(weekly_forecast["farm"].dropna().unique().tolist())
                if st.session_state.get("container_forecast_farm") not in farm_options:
                    st.session_state["container_forecast_farm"] = farm_options[0]
                selected_farm = st.selectbox("Farm", options=farm_options, key="container_forecast_farm")
            df_week_opts = weekly_forecast.copy()
            if selected_farm != "Tất cả":
                df_week_opts = df_week_opts[df_week_opts["farm"] == selected_farm]
            year_options = sorted(df_week_opts["year"].dropna().astype(int).unique().tolist())
            default_year_idx = year_options.index(date.today().year) if date.today().year in year_options else 0
            with y_col:
                if st.session_state.get("container_forecast_year") not in year_options:
                    st.session_state["container_forecast_year"] = year_options[default_year_idx]
                selected_year = st.selectbox("Năm thu hoạch dự báo", options=year_options, index=default_year_idx, key="container_forecast_year")
            df_week_opts = df_week_opts[df_week_opts["year"] == int(selected_year)]
            week_options = sorted(df_week_opts["week"].dropna().astype(int).unique().tolist())
            default_week = date.today().isocalendar().week
            default_week_idx = week_options.index(default_week) if default_week in week_options else 0
            with w_col:
                week_key = f"container_forecast_harvest_week_{selected_farm}_{selected_year}_{selected_forecast_weeks}"
                if st.session_state.get(week_key) not in week_options:
                    st.session_state[week_key] = week_options[default_week_idx]
                selected_week = st.selectbox(
                    "Tuần thu hoạch dự báo",
                    options=week_options,
                    index=default_week_idx,
                    key=week_key,
                )

            df_selected = df_week_opts[df_week_opts["week"] == int(selected_week)]
            source_bunches = int(df_selected["forecast_bunches"].sum())
            source_label = f"{selected_farm} · Tuần thu hoạch dự báo {selected_week}/{selected_year} · +{selected_forecast_weeks} tuần"
            st.metric("Số buồng dự báo thu hoạch từ cắt bắp", f"{source_bunches:,} buồng", source_label)
            st.caption(
                f"+{selected_forecast_weeks} tuần tính cả tuần cắt bắp. Ví dụ +8: cắt tuần 20 thì rơi vào tuần thu hoạch dự báo 27."
            )
    else:
        if calc_mode != CONTAINER_MODE_CONTS_TO_BUNCHES:
            source_bunches = int(st.number_input("Số buồng thu hoạch / dự kiến", min_value=0, value=0, step=100, key="container_manual_bunches"))
            source_label = "Nhập tay"

    st.divider()
    cfg1, cfg2, cfg3, cfg4 = st.columns([1.2, 1.3, 1, 1])
    with cfg1:
        bunch_type = st.selectbox(
            "Loại buồng",
            options=[12, 9],
            format_func=lambda value: f"{value} nải",
            key="container_bunch_type",
        )
    with cfg2:
        if bunch_type == 12:
            target_kg_per_bunch = st.radio(
                "Kịch bản kg/buồng",
                options=[18.0, 20.0],
                format_func=lambda value: f"{value:g} kg",
                horizontal=True,
                key="container_12_hand_kg_scenario",
            )
        else:
            target_kg_per_bunch = 15.6
            st.metric("Kịch bản kg/buồng", "15.6 kg")
    with cfg3:
        hand_profile = build_hand_weight_profile(bunch_type, target_kg_per_bunch)
        hands_per_bunch = int(hand_profile["hands_per_bunch"])
        kg_per_bunch = float(hand_profile["kg_per_bunch"])
        hand_weights = hand_profile["hand_weights"]
        st.metric("Kg/buồng", f"{kg_per_bunch:.1f} kg")
    with cfg4:
        st.metric("Kg/nải TB", f"{hand_profile['kg_per_hand']:.2f}")

    with st.popover("Bảng kg từng nải"):
        st.caption("Kg/nải TB chỉ là trung bình; bảng này và solver dùng kg từng nải sau quy đổi.")
        st.dataframe(
            pd.DataFrame([
                {"Nải": hand, "Kg sau quy đổi": round(weight, 3)}
                for hand, weight in hand_weights.items()
            ]),
            use_container_width=True,
            hide_index=True,
        )

    if calc_mode == CONTAINER_MODE_CONTS_TO_BUNCHES:
        st.markdown("##### Cấu hình cont mục tiêu")
        container_targets = _render_container_target_controls(hands_per_bunch)
        allocation_rows, validation_errors = _prepare_container_rows_from_container_targets(
            container_targets,
            hands_per_bunch,
        )
        if validation_errors:
            for error in validation_errors:
                st.warning(error)
            st.info("Hoàn tất từng cont đúng 1,320 thùng và chọn đủ khách hàng/mã hàng để chạy tính số buồng.")
            _render_container_saved_plan_cards()
            return

        if not allocation_rows:
            st.info("Thêm ít nhất một cont hợp lệ để tính số buồng xẻ tối thiểu.")
            _render_container_saved_plan_cards()
            return

        status_box = st.status("Đang chứng minh số buồng xẻ tối thiểu...", expanded=True)
        progress_bar = st.progress(0)

        def _container_exact_progress(message, progress=None):
            if message:
                status_box.write(message)
            if progress is not None:
                progress_bar.progress(max(0, min(100, int(progress * 100))))

        try:
            result = calculate_min_bunches_for_container_plan(
                allocation_rows,
                kg_per_bunch,
                hands_per_bunch,
                kg_per_box=KG_PER_BOX,
                boxes_per_container=BOXES_PER_CONTAINER,
                hand_weights=hand_weights,
                progress_callback=_container_exact_progress,
            )
        finally:
            progress_bar.empty()
        summary = result["summary"]

        st.markdown("##### Kết quả tổng")
        solver_status = summary.get("solver_status", result.get("solver_status", "APPROXIMATE"))
        solver_backend = summary.get("solver_backend", result.get("solver_backend", "unknown"))
        if solver_status == "APPROXIMATE":
            status_box.update(label="Chưa có kết quả chứng minh chính xác", state="error", expanded=True)
            st.warning("Không dùng kết quả xấp xỉ cho mode này. Hãy cài OR-Tools hoặc chạy lại khi solver chính khả dụng.")
        elif solver_status == "NO_SOLUTION":
            status_box.update(label="Chưa chứng minh được số buồng tối thiểu", state="error", expanded=True)
            st.error("Không tìm được hoặc chưa chứng minh được số buồng xẻ tối thiểu cho cơ cấu cont đã nhập.")
        else:
            status_box.update(label="Đã chứng minh số buồng xẻ tối thiểu", state="complete", expanded=False)
            st.caption(f"Thuật toán tối ưu: {solver_status} · {solver_backend}")

        target_containers = len(container_targets)
        target_boxes = target_containers * BOXES_PER_CONTAINER
        kg_allocated = float(summary.get("kg_allocated", 0))
        extra_kg = float(summary.get("extra_kg_from_rounding", 0))
        m1, m2, m3, m4, m5, m6 = st.columns(6)
        with m1:
            st.metric("Cont mục tiêu", f"{target_containers:,}")
        with m2:
            st.metric("Tổng thùng", f"{target_boxes:,}")
        with m3:
            st.metric("Buồng xẻ tối thiểu", f"{int(summary.get('active_bunches_estimated', 0)):,}")
        with m4:
            st.metric("Kg cần xẻ", f"{float(summary.get('source_kg', 0)):,.0f}")
        with m5:
            st.metric("Kg phân bổ", f"{kg_allocated:,.0f}")
        with m6:
            st.metric("Kg dư", f"{extra_kg:,.1f}")

        if solver_status != "OPTIMAL":
            _render_container_saved_plan_cards()
            return

        if result["rows"]:
            st.success(
                f"Kết luận: cần xẻ tối thiểu {int(summary.get('active_bunches_estimated', 0)):,} buồng "
                f"để đáp ứng {target_containers:,} cont ({target_boxes:,} thùng)."
            )
            result_df = pd.DataFrame(result["rows"])
            sort_cols = [
                col for col in ("customer_priority", "sku_priority", "processing_order")
                if col in result_df.columns
            ]
            if sort_cols:
                result_df = result_df.sort_values(sort_cols)
            compact_rows = []
            for _, sku_row in result_df.iterrows():
                compact_rows.append({
                    "Khách hàng": sku_row.get("customer", ""),
                    "Thị trường": sku_row.get("market", ""),
                    "Mã hàng": sku_row.get("sku", ""),
                    "Thùng cần": int(sku_row.get("requested_boxes", 0)),
                    "Thùng đáp ứng": int(sku_row.get("boxes_fulfilled", 0)),
                    "Buồng dùng": int(sku_row.get("bunches_allocated", 0)),
                    "Kg phân bổ": f"{float(sku_row.get('kg_allocated', 0)):,.0f}",
                })
            st.dataframe(pd.DataFrame(compact_rows), use_container_width=True, hide_index=True)

            process_rows = []
            detail_result_df = pd.DataFrame(result.get("detail_rows") or result["rows"])
            if not detail_result_df.empty:
                detail_sort_cols = [
                    col for col in ("customer_priority", "sku_priority", "processing_order")
                    if col in detail_result_df.columns
                ]
                if detail_sort_cols:
                    detail_result_df = detail_result_df.sort_values(detail_sort_cols)
                for _, sku_row in detail_result_df.iterrows():
                    process_rows.append({
                        "Bước": int(sku_row["processing_order"]),
                        "Thị trường": sku_row["market"],
                        "Khách hàng": sku_row.get("customer", ""),
                        "Mã hàng": sku_row["sku"],
                        "Nải chọn": sku_row["range_label"],
                        "Cách tính": (
                            f"{int(sku_row['bunches_allocated']):,} buồng x "
                            f"({_container_range_formula(sku_row['hand_from'], sku_row['hand_to'], hand_weights)}) "
                            f"= {float(sku_row['kg_allocated']):,.1f} kg"
                        ),
                        "Kết quả": (
                            f"~{float(sku_row.get('boxes_equivalent', sku_row['kg_allocated'] / KG_PER_BOX)):,.2f} thùng quy đổi; "
                            "tổng thùng chốt xem ở dòng mã hàng"
                        ),
                    })
            if process_rows:
                with st.expander("Quá trình chọn nải chi tiết", expanded=False):
                    st.dataframe(pd.DataFrame(process_rows), use_container_width=True, hide_index=True)

        remaining_df = pd.DataFrame([
            {"Nải": hand, "Buồng còn lại": qty}
            for hand, qty in result.get("remaining_hands", {}).items()
        ])
        if not remaining_df.empty:
            with st.expander("Tồn nải còn lại sau phân bổ", expanded=False):
                st.dataframe(remaining_df, use_container_width=True, hide_index=True)

        if st.button("Lưu kế hoạch", use_container_width=True, type="secondary", key="container_save_min_bunch_session"):
            if _save_container_plan({
                "mode": calc_mode,
                "source_mode": source_mode,
                "source": f"{target_containers:,} cont mục tiêu",
                "source_bunches": int(summary.get("active_bunches_estimated", 0)),
                "kg_per_bunch": kg_per_bunch,
                "hands_per_bunch": hands_per_bunch,
                "hand_weights": hand_weights,
                "container_targets": container_targets,
                "sku_rows": allocation_rows,
                "result": result,
            }):
                st.success("Đã lưu kế hoạch vào account hiện tại.")

        _render_container_saved_plan_cards()
        return

    if calc_mode == CONTAINER_MODE_MAX_CONTS:
        st.markdown("##### Ưu tiên thị trường")
        market_order = _render_container_market_priority_controls("container_max_market_priority")
        max_result = calculate_max_containers_by_market(
            source_bunches,
            kg_per_bunch,
            hands_per_bunch,
            market_order,
            kg_per_box=KG_PER_BOX,
            boxes_per_container=BOXES_PER_CONTAINER,
            hand_weights=hand_weights,
        )
        summary = max_result["summary"]

        st.markdown("##### Kết quả tổng")
        solver_status = summary.get("solver_status", max_result.get("solver_status", "APPROXIMATE"))
        solver_backend = summary.get("solver_backend", max_result.get("solver_backend", "unknown"))
        if solver_status == "APPROXIMATE":
            st.warning("Kết quả xấp xỉ do thiếu solver tối ưu hoặc đang dùng fallback.")
        else:
            st.caption(f"Thuật toán tối ưu: {solver_status} · {solver_backend}")
        m1, m2, m3, m4, m5, m6 = st.columns(6)
        with m1:
            st.metric("Nguồn buồng", f"{summary['total_bunches']:,}")
        with m2:
            st.metric("Buồng xẻ tối thiểu", f"{summary.get('active_bunches_estimated', 0):,}")
        with m3:
            st.metric("Kg nguồn", f"{summary['source_kg']:,.0f}")
        with m4:
            st.metric("Kg/nải TB", f"{summary['kg_per_hand']:.2f}")
        with m5:
            st.metric("Cont lý thuyết", f"{summary['source_cont_capacity']:,.2f}")
        with m6:
            st.metric("Cont đủ tối đa", f"{summary['fulfilled_containers']:,}")

        if max_result["rows"]:
            max_market_df = pd.DataFrame(max_result["rows"])
            conclusion_parts = [
                f"{row['market']}: {int(row['full_containers'])} cont"
                for _, row in max_market_df.iterrows()
            ]
            if conclusion_parts:
                st.success("Kết luận tối đa: " + " · ".join(conclusion_parts))

            compact_max_rows = []
            detail_df = pd.DataFrame(max_result.get("detail_rows", []))
            for _, market_row in max_market_df.iterrows():
                market_name = market_row["market"]
                compact_max_rows.append({
                    "Thị trường": market_name,
                    "Mã hàng": "Tổng",
                    "Nải dùng": "",
                    "Cont đủ": int(market_row["full_containers"]),
                    "Thùng chốt": int(market_row["boxes_allocated"]),
                    "Thùng lẻ": int(market_row["remaining_boxes_potential"]),
                    "Buồng dùng": int(market_row.get("active_bunches_used", market_row["bundles_used"])),
                    "Kg phân bổ": f"{market_row['kg_allocated']:,.0f}",
                })
                if not detail_df.empty:
                    for _, detail_row in detail_df[detail_df["market"] == market_name].iterrows():
                        compact_max_rows.append({
                            "Thị trường": market_name,
                            "Mã hàng": detail_row["sku"],
                            "Nải dùng": detail_row["range_label"],
                            "Cont đủ": "",
                            "Thùng chốt": int(detail_row["boxes_equivalent"]),
                            "Thùng lẻ": "",
                            "Buồng dùng": int(detail_row["bundles_used"]),
                            "Kg phân bổ": f"{detail_row['kg_allocated']:,.0f}",
                        })
            st.dataframe(pd.DataFrame(compact_max_rows), use_container_width=True, hide_index=True)
            process_rows = []
            for _, market_row in max_market_df.iterrows():
                market_name = market_row["market"]
                process_rows.append({
                    "Bước": int(market_row["market_priority"]),
                    "Thị trường": market_name,
                    "Mã hàng": "Tổng",
                    "Nải chọn": "Theo bộ quy cách",
                    "Cách tính": (
                        f"{int(market_row['available_bundles']):,} buồng khả dụng; "
                        f"solver chọn các dải chi tiết bên dưới"
                    ),
                    "Kết quả": (
                        f"{int(market_row['capacity_boxes']):,} thùng khả dụng, "
                        f"chốt {int(market_row['full_containers'])} cont"
                    ),
                })
                if not detail_df.empty:
                    for _, detail_row in detail_df[detail_df["market"] == market_name].iterrows():
                        process_rows.append({
                            "Bước": "",
                            "Thị trường": market_name,
                            "Mã hàng": detail_row["sku"],
                            "Nải chọn": detail_row["range_label"],
                            "Cách tính": (
                                f"{int(detail_row['bundles_used']):,} buồng x "
                                f"({_container_range_formula(detail_row['hand_from'], detail_row['hand_to'], hand_weights)}) "
                                f"/ {KG_PER_BOX} kg"
                            ),
                            "Kết quả": f"{int(detail_row['boxes_equivalent']):,} thùng quy đổi",
                        })
            if process_rows:
                with st.expander("Quá trình chọn nải chi tiết", expanded=False):
                    st.dataframe(pd.DataFrame(process_rows), use_container_width=True, hide_index=True)
        else:
            st.info("Chọn nguồn buồng hợp lệ để tính tối đa cont theo thị trường.")

        remaining_df = pd.DataFrame([
            {"Nải": hand, "Buồng còn lại": qty}
            for hand, qty in max_result["remaining_hands"].items()
        ])
        if not remaining_df.empty:
            with st.expander("Tồn nải còn lại sau phân bổ", expanded=False):
                st.dataframe(remaining_df, use_container_width=True, hide_index=True)

        if st.button("Lưu kế hoạch", use_container_width=True, type="secondary", key="container_save_max_session"):
            if _save_container_plan({
                "mode": calc_mode,
                "source_mode": source_mode,
                "source": source_label,
                "source_bunches": source_bunches,
                "kg_per_bunch": kg_per_bunch,
                "hands_per_bunch": hands_per_bunch,
                "hand_weights": hand_weights,
                "market_order": market_order,
                "result": max_result,
            }):
                st.success("Đã lưu kế hoạch vào account hiện tại.")

        _render_container_saved_plan_cards()
        return

    title_col, help_col = st.columns([10, 1])
    with title_col:
        st.markdown("##### Cấu hình đơn hàng")
    with help_col:
        with st.popover("?"):
            sku_range_rows = []
            for sku, sku_def in _container_sku_definitions_for_hands(hands_per_bunch).items():
                ranges = ", ".join(f"{hand_from}-{hand_to}" for hand_from, hand_to in sku_def["ranges"])
                sku_range_rows.append({
                    "Mã hàng": sku,
                    "Thị trường": ", ".join(sku_def.get("markets", [])),
                    "Nhóm": sku_def["group"],
                    "Dải nải có thể dùng": ranges,
                })
            st.dataframe(
                pd.DataFrame(sku_range_rows),
                use_container_width=True,
                hide_index=True,
            )

    current_rows = st.session_state["container_calc_sku_rows"] or _default_container_sku_editor_rows()
    normalized_current_rows = []
    for idx, row in enumerate(current_rows):
        customer = _container_clean_text(row.get("Khách hàng"))
        if customer not in CONTAINER_CUSTOMER_OPTIONS:
            customer = ""
        market = _container_customer_market(customer)
        market_skus = _container_sku_options_for_market(market, hands_per_bunch) if market else []
        sku = _container_clean_text(row.get("Mã hàng")).upper()
        if not market_skus or sku not in market_skus:
            sku = ""
        normalized_current_rows.append({
            "_row_id": row.get("_row_id") or f"default_{idx}",
            "Khách hàng": customer,
            "Thị trường": market,
            "Mã hàng": sku,
            "Nhu cầu": _container_clean_int(row.get("Nhu cầu")),
        })
    current_rows = normalized_current_rows

    order_rows = []
    delete_row_index = None
    header_cols = st.columns([2, 1.4, 1.4, 0.7])
    header_cols[0].caption("Khách hàng")
    header_cols[1].caption("Mã hàng")
    header_cols[2].caption("Nhu cầu (thùng)")
    for idx, row in enumerate(current_rows):
        row_id = row["_row_id"]
        row_cols = st.columns([2, 1.4, 1.4, 0.7])
        with row_cols[0]:
            customer_options = [CONTAINER_EMPTY_OPTION] + CONTAINER_CUSTOMER_OPTIONS
            customer_key = f"container_order_customer_{row_id}"
            if st.session_state.get(customer_key) not in customer_options:
                st.session_state[customer_key] = row["Khách hàng"] if row["Khách hàng"] in CONTAINER_CUSTOMER_OPTIONS else None
            customer = st.selectbox(
                "Khách hàng",
                options=customer_options,
                index=None,
                placeholder="Chọn khách hàng",
                key=customer_key,
                label_visibility="collapsed",
            )
            if customer == CONTAINER_EMPTY_OPTION:
                customer = ""
            market = _container_customer_market(customer)
        with row_cols[1]:
            sku_options = _container_sku_options_for_market(market, hands_per_bunch) if market else []
            sku_select_options = [CONTAINER_EMPTY_OPTION] + sku_options
            sku_key = f"container_order_sku_{row_id}"
            if st.session_state.get(sku_key) not in sku_select_options:
                st.session_state[sku_key] = row["Mã hàng"] if row["Mã hàng"] in sku_options else None
            sku = st.selectbox(
                "Mã hàng",
                options=sku_select_options,
                index=None,
                placeholder="Chọn mã hàng" if market else "Chọn khách hàng trước",
                key=sku_key,
                label_visibility="collapsed",
                disabled=not bool(market),
            )
            if sku == CONTAINER_EMPTY_OPTION:
                sku = ""
        with row_cols[2]:
            demand = st.number_input(
                "Nhu cầu (thùng)",
                min_value=0,
                value=_container_clean_int(row.get("Nhu cầu")),
                step=1,
                key=f"container_order_demand_{row_id}",
                label_visibility="collapsed",
            )
        with row_cols[3]:
            if st.button("Xóa", key=f"container_order_delete_{row_id}", disabled=len(current_rows) <= 1):
                delete_row_index = idx

        order_rows.append({
            "_row_id": row_id,
            "Khách hàng": customer or "",
            "Thị trường": market,
            "Mã hàng": sku or "",
            "Nhu cầu": demand,
        })

    action_cols = st.columns([1, 5])
    with action_cols[0]:
        if st.button("Thêm dòng", key="container_order_add"):
            st.session_state["container_calc_sku_rows"] = order_rows + [{
                "_row_id": f"new_{datetime.now().timestamp()}",
                "Khách hàng": "",
                "Thị trường": "",
                "Mã hàng": "",
                "Nhu cầu": 0,
            }]
            st.rerun()

    if delete_row_index is not None:
        st.session_state["container_calc_sku_rows"] = [
            row for idx, row in enumerate(order_rows) if idx != delete_row_index
        ]
        st.rerun()

    st.session_state["container_calc_sku_rows"] = order_rows

    order_totals_by_market = {}
    for row in order_rows:
        market = _container_clean_text(row.get("Thị trường"))
        sku = _container_clean_text(row.get("Mã hàng"))
        demand = _container_clean_int(row.get("Nhu cầu"))
        if market and sku and demand > 0:
            order_totals_by_market[market] = order_totals_by_market.get(market, 0) + demand
    for market, requested_boxes in order_totals_by_market.items():
        if 0 < requested_boxes < BOXES_PER_CONTAINER:
            st.warning(
                f"Đơn hàng {market} đang có {requested_boxes:,} thùng, chưa đủ 1 cont "
                f"({BOXES_PER_CONTAINER:,} thùng). Có thể đáp ứng đủ số thùng này nhưng chưa chốt được cont nguyên."
            )

    customer_order = _render_container_customer_priority_controls("container_customer_priority")

    active_customers = _unique_keep_order([
        row.get("Khách hàng")
        for row in order_rows
        if _container_clean_text(row.get("Mã hàng"))
    ])
    sku_orders_by_customer = {}
    for customer in active_customers:
        market_skus = _unique_keep_order([
            row.get("Mã hàng")
            for row in order_rows
            if _container_clean_text(row.get("Khách hàng")) == customer
        ])
        if not market_skus:
            continue
        sku_options = [CONTAINER_EMPTY_OPTION] + market_skus
        st.markdown(f"###### Ưu tiên loại hàng - {customer}")
        sku_cols = st.columns(min(3, max(1, len(market_skus))))
        selected_skus = []
        for idx, col in enumerate(sku_cols, start=1):
            with col:
                sku_key = f"container_sku_priority_{customer}_{idx}"
                if st.session_state.get(sku_key) not in sku_options:
                    st.session_state[sku_key] = None
                selected_sku = st.selectbox(
                    f"Loại hàng ưu tiên {idx}",
                    options=sku_options,
                    index=None,
                    placeholder="Không chọn",
                    key=sku_key,
                )
                if selected_sku and selected_sku != CONTAINER_EMPTY_OPTION:
                    selected_skus.append(selected_sku)
        sku_orders_by_customer[customer] = selected_skus

    allocation_rows = _prepare_container_rows_for_allocation(order_rows, customer_order, sku_orders_by_customer)
    if not allocation_rows:
        st.info("Chọn khách hàng, mã hàng và nhập nhu cầu để bắt đầu tính phân bổ.")
        _render_container_saved_plan_cards()
        return

    result = allocate_bunches_optimized(
        source_bunches,
        kg_per_bunch,
        hands_per_bunch,
        allocation_rows,
        kg_per_box=KG_PER_BOX,
        boxes_per_container=BOXES_PER_CONTAINER,
        hand_weights=hand_weights,
    )
    summary = result["summary"]

    st.markdown("##### Kết quả tổng")
    solver_status = summary.get("solver_status", result.get("solver_status", "APPROXIMATE"))
    solver_backend = summary.get("solver_backend", result.get("solver_backend", "unknown"))
    if solver_status == "APPROXIMATE":
        st.warning("Kết quả xấp xỉ do thiếu solver tối ưu hoặc đang dùng fallback.")
    else:
        st.caption(f"Thuật toán tối ưu: {solver_status} · {solver_backend}")
    m1, m2, m3, m4, m5, m6 = st.columns(6)
    with m1:
        st.metric("Nguồn buồng", f"{summary['total_bunches']:,}")
    with m2:
        st.metric("Buồng xẻ tối thiểu", f"{summary.get('active_bunches_estimated', 0):,}")
    with m3:
        st.metric("Kg nguồn", f"{summary['source_kg']:,.0f}")
    with m4:
        st.metric("Kg/nải TB", f"{summary['kg_per_hand']:.2f}")
    with m5:
        st.metric("Cont lý thuyết", f"{summary['source_cont_capacity']:,.2f}")
    with m6:
        st.metric("Thùng đáp ứng", f"{summary['fulfilled_boxes']:,}", f"Thiếu {summary['short_boxes']:,} thùng")

    if result["rows"]:
        result_df = pd.DataFrame(result["rows"])
        market_summary = (
            result_df.groupby("market", as_index=False)
            .agg(
                market_priority=("market_priority", "min"),
                requested_boxes=("requested_boxes", "sum"),
                boxes_fulfilled=("boxes_fulfilled", "sum"),
                short_boxes=("short_boxes", "sum"),
            )
            .sort_values(["market_priority", "market"])
        )
        market_summary["full_containers"] = market_summary["boxes_fulfilled"] // BOXES_PER_CONTAINER
        market_summary["remaining_boxes"] = market_summary["boxes_fulfilled"] % BOXES_PER_CONTAINER
        conclusion_parts = []
        for _, row in market_summary.iterrows():
            market_name = row["market"]
            full_containers = int(row["full_containers"])
            remaining_boxes = int(row["remaining_boxes"])
            if remaining_boxes > 0:
                conclusion_parts.append(f"{market_name}: {full_containers} cont + {remaining_boxes:,} thùng lẻ")
            else:
                conclusion_parts.append(f"{market_name}: {full_containers} cont")
        if conclusion_parts:
            active_bunches = int(summary.get("active_bunches_estimated", 0))
            st.success(
                "Kết luận theo thị trường: "
                + " · ".join(conclusion_parts)
                + f" · Buồng xẻ tối thiểu: {active_bunches:,} buồng"
            )

        market_lookup = market_summary.set_index("market").to_dict("index")
        compact_rows = []
        order_sort_cols = [
            "customer_priority" if "customer_priority" in result_df.columns else "market_priority",
            "sku_priority",
            "processing_order",
        ]
        sorted_result_df = result_df.sort_values(order_sort_cols)
        for _, market_row in market_summary.iterrows():
            market_name = market_row["market"]
            market_info = market_lookup.get(market_name, {})
            compact_rows.append({
                "Thị trường": market_name,
                "Khách hàng": "",
                "Mã hàng": "Tổng",
                "Thùng yêu cầu": int(market_info.get("requested_boxes", 0)),
                "Thùng đáp ứng": int(market_info.get("boxes_fulfilled", 0)),
                "Thiếu thùng": int(market_info.get("short_boxes", 0)),
                "Cont đủ": int(market_info.get("full_containers", 0)),
                "Thùng lẻ": int(market_info.get("remaining_boxes", 0)),
                "Buồng xẻ tối thiểu": "",
            })
        for _, sku_row in sorted_result_df.iterrows():
            compact_rows.append({
                "Thị trường": sku_row["market"],
                "Khách hàng": sku_row.get("customer", ""),
                "Mã hàng": sku_row["sku"],
                "Thùng yêu cầu": int(sku_row["requested_boxes"]),
                "Thùng đáp ứng": int(sku_row["boxes_fulfilled"]),
                "Thiếu thùng": int(sku_row["short_boxes"]),
                "Cont đủ": "",
                "Thùng lẻ": "",
                "Buồng xẻ tối thiểu": int(sku_row["bunches_allocated"]),
            })

        st.dataframe(pd.DataFrame(compact_rows), use_container_width=True, hide_index=True)
        process_rows = []
        detail_result_df = pd.DataFrame(result.get("detail_rows") or result["rows"])
        if not detail_result_df.empty:
            detail_sort_cols = [
                "customer_priority" if "customer_priority" in detail_result_df.columns else "market_priority",
                "sku_priority",
                "processing_order",
            ]
            detail_result_df = detail_result_df.sort_values(detail_sort_cols)
        for _, sku_row in detail_result_df.iterrows():
            process_rows.append({
                "Bước": int(sku_row["processing_order"]),
                "Thị trường": sku_row["market"],
                "Khách hàng": sku_row.get("customer", ""),
                "Mã hàng": sku_row["sku"],
                "Nải chọn": sku_row["range_label"],
                "Cách tính": (
                    f"{int(sku_row['bunches_allocated']):,} buồng x "
                    f"({_container_range_formula(sku_row['hand_from'], sku_row['hand_to'], hand_weights)}) "
                    f"= {float(sku_row['kg_allocated']):,.1f} kg"
                ),
                "Kết quả": (
                    f"~{float(sku_row.get('boxes_equivalent', sku_row['kg_allocated'] / KG_PER_BOX)):,.2f} thùng quy đổi; "
                    "tổng thùng chốt xem ở dòng mã hàng"
                    + (
                        f", thiếu {int(sku_row['short_boxes']):,} thùng"
                        if int(sku_row["short_boxes"]) > 0
                        else ""
                    )
                ),
            })
        if process_rows:
            with st.expander("Quá trình chọn nải chi tiết", expanded=False):
                st.dataframe(pd.DataFrame(process_rows), use_container_width=True, hide_index=True)
    else:
        st.info("Nhập số buồng và ít nhất một dòng mã hàng hợp lệ để xem kết quả.")

    remaining_df = pd.DataFrame([
        {"Nải": hand, "Buồng còn lại": qty}
        for hand, qty in result["remaining_hands"].items()
    ])
    if not remaining_df.empty:
        with st.expander("Tồn nải còn lại sau phân bổ", expanded=False):
            st.dataframe(remaining_df, use_container_width=True, hide_index=True)

    if st.button("Lưu kế hoạch", use_container_width=True, type="secondary", key="container_save_session"):
        if _save_container_plan({
            "mode": calc_mode,
            "source_mode": source_mode,
            "source": source_label,
            "source_bunches": source_bunches,
            "kg_per_bunch": kg_per_bunch,
            "hands_per_bunch": hands_per_bunch,
            "hand_weights": hand_weights,
            "sku_rows": order_rows,
            "customer_order": customer_order,
            "sku_orders_by_customer": sku_orders_by_customer,
            "result": result,
        }):
            st.success("Đã lưu kế hoạch vào account hiện tại.")

    _render_container_saved_plan_cards()


# =====================================================
# GIAO DIỆN CHÍNH (MAIN APP) - ROLE BASED 
# =====================================================
def render_main_app():
    c_farm = st.session_state["current_farm"]
    c_team = st.session_state["current_team"]

    # --- SIDEBAR ---
    with st.sidebar:
        if os.path.exists("logo.png"): st.image("logo.png", use_container_width=True)
        else: st.markdown("### 🍌 Trường Tồn")
        
        st.divider()
        st.markdown(f'<span class="farm-badge">🏭 {c_farm}</span>', unsafe_allow_html=True)
        st.markdown(f'<span class="team-badge">👥 {c_team}</span>', unsafe_allow_html=True)
        st.caption(f"Đăng nhập lúc: {datetime.now().strftime('%H:%M - %d/%m/%Y')}")
        st.divider()

        if st.button("🚪 Đăng xuất", use_container_width=True, type="secondary"):
            logout()
            st.rerun()
        st.divider()


    # --- HEADER ---
    if "toast" in st.session_state:
        st.success(st.session_state.pop("toast"))
        


    # =================================================
    # MODULE ADMIN
    # =================================================
    if c_farm == "Admin" and c_team == "Quản trị viên":
        tab_opts = ["🌐 Dữ liệu toàn cục", "💰 Chi phí", "👑 Quản trị Mùa Vụ"]
        active_tab = st.segmented_control("Chức năng", tab_opts, label_visibility="collapsed", key="tab_admin_menu", default=tab_opts[0])
        if active_tab is None: active_tab = tab_opts[0]
        
        if active_tab == tab_opts[2]:
            st.info("👋 Chào mừng Quản trị viên. Tại đây bạn có thể quản lý lịch sử Vụ cho từng lô.")
            
            res = supabase.table("seasons").select("*, dim_lo!inner(lo_name, dim_farm!inner(farm_name))").eq("is_deleted", False).order("created_at", desc=True).execute()
            df_seasons = pd.DataFrame(res.data) if res.data else pd.DataFrame()
            
            if df_seasons.empty:
                st.warning("Hiện tại hệ thống chưa có dữ liệu Vụ (Seasons). Vui lòng tạo lô ở Nông trường trước!")
                return
            
            # Flatten dim_lo join
            df_seasons["farm"] = df_seasons["dim_lo"].apply(lambda x: x.get("dim_farm", {}).get("farm_name") if isinstance(x, dict) else None)
            df_seasons["lo"] = df_seasons["dim_lo"].apply(lambda x: x.get("lo_name") if isinstance(x, dict) else None)
                
            f_farm = st.selectbox("Lọc Farm", options=["Tất cả"] + list(df_seasons["farm"].dropna().unique()))
            if f_farm != "Tất cả":
                df_seasons = df_seasons[df_seasons["farm"] == f_farm]
                
            st.markdown("### 📋 Danh sách Vụ (Seasons)")
            st.dataframe(
                df_seasons[["farm", "lo", "vu", "loai_trong", "ngay_bat_dau", "ngay_ket_thuc_thuc_te"]], 
                use_container_width=True, hide_index=True, on_select="rerun", selection_mode="single-row", key="sel_admin_seasons"
            )
            
            idx_list = st.session_state.get(f"sel_admin_seasons", {}).get("selection", {}).get("rows", [])
            if idx_list and len(idx_list) > 0 and idx_list[0] < len(df_seasons):
                row = df_seasons.iloc[idx_list[0]].to_dict()
                
                with st.container(border=True):
                    st.markdown(f"#### 🛠️ Chốt vụ: `{row['farm']}` - Lô `{row['lo']}` (Hiện tại: `{row['vu']}`)")
                    col1, col2 = st.columns(2)
                    with col1:
                        cur_start = row.get("ngay_bat_dau")
                        def_start_date = pd.to_datetime(cur_start).date() if cur_start else date.today()
                        start_date = st.date_input("📆 Ngày bắt đầu", value=def_start_date)
                        
                        cur_end = row.get("ngay_ket_thuc_thuc_te")
                        def_end_date = pd.to_datetime(cur_end).date() if cur_end else date.today()
                        end_date = st.date_input("📆 Ngày kết thúc (Thực tế)", value=def_end_date)
                    with col2:
                        curr_v = str(row["vu"])
                        try:
                            next_v_num = int(curr_v.replace("F", "")) + 1
                        except:
                            next_v_num = 1
                        next_v = f"F{next_v_num}"
                        
                        auto_next = st.checkbox(f"🚀 Cho phép tự động tạo vụ nối tiếp: {next_v}", value=True)
                    
                    st.markdown("")
                    if st.button("💾 Lưu thay đổi & Chốt vụ", use_container_width=True, type="primary"):
                        try:
                            supabase.table("seasons").update({
                                "ngay_bat_dau": start_date.isoformat(),
                                "ngay_ket_thuc_thuc_te": end_date.isoformat()
                            }).eq("id", row["id"]).execute()
                            
                            if auto_next and not cur_end:
                                new_season = {
                                    "dim_lo_id": row["dim_lo_id"],
                                    "vu": next_v,
                                    "loai_trong": row["loai_trong"],
                                    "ngay_bat_dau": end_date.isoformat()
                                }
                                supabase.table("seasons").insert(new_season).execute()
                            
                            st.session_state["toast"] = f"✅ Đã chốt vụ {row['vu']} thành công!"
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ Lỗi khi chốt vụ: {e}")
        elif active_tab == tab_opts[1]:
            render_cost_dashboard(supabase, c_farm, c_team)
        elif active_tab == tab_opts[0]:
            render_global_data_tab("Admin")
        return

    # =================================================
    # MODULE KINH DOANH
    # =================================================
    if c_farm == "Phòng Kinh doanh" and c_team == "Kinh doanh":
        tab_opts = ["🌐 Dữ liệu toàn cục", "💰 Chi phí", "📦 Máy tính phân bổ cont"]
        active_tab = _persisted_segmented_control(
            "Chức năng",
            tab_opts,
            key="tab_sales_menu",
            query_key="sales_tab",
            slugs={
                tab_opts[0]: "global",
                tab_opts[1]: "cost",
                tab_opts[2]: "container_allocation",
            },
            default=tab_opts[0],
        )
        if active_tab == tab_opts[2]:
            render_container_allocation_calculator()
        elif active_tab == tab_opts[1]:
            render_cost_dashboard(supabase, c_farm, c_team)
        else:
            render_global_data_tab("Phòng Kinh doanh")
        return

    # =================================================
    # MODULE 4: QUẢN LÝ FARM
    # =================================================
    if c_team == "Quản lý farm":
        st.info("👋 Chế độ chỉ xem (Read-only). Tương tác với các biểu đồ bên dưới để phân tích dữ liệu.")
        tab_opts = ["🌐 Dữ liệu toàn cục", "💰 Chi phí"]
        active_tab = st.segmented_control("Chức năng", tab_opts, label_visibility="collapsed", key="tab_farm_manager_menu", default=tab_opts[0])
        if active_tab is None: active_tab = tab_opts[0]
        if active_tab == tab_opts[1]:
            render_cost_dashboard(supabase, c_farm, c_team)
        else:
            render_global_data_tab(c_farm)
        return

    # =================================================
    # MODULE 1: ĐỘI NÔNG TRƯỜNG (NT1, NT2)
    # =================================================
    # MODULE 1: ĐỘI NÔNG TRƯỜNG (NT1, NT2)
    # =================================================
    if c_team in ["NT1", "NT2", "Đội BVTV"]:
        if c_team == "Đội BVTV":
            tab_opts = ["🌐 Dữ liệu toàn cục", "💰 Chi phí", "📈 Cập nhật Tiến độ"]
        else:
            tab_opts = ["🌐 Dữ liệu toàn cục", "💰 Chi phí", "🌱 Khởi tạo Lô trồng", "📈 Cập nhật Tiến độ", "📏 Đo Size", "🗑️ Cập nhật Xuất hủy", "🌳 Kiểm kê cây", "🧪 Đo pH Đất", "🦠 Kiểm tra Fusarium"]
            
        active_tab = st.segmented_control("Chức năng", tab_opts, label_visibility="collapsed", key="tab_nt_menu", default=tab_opts[0])
        if active_tab is None: active_tab = tab_opts[0] # Prevent empty state

        # TAB 1: KHỞI TẠO LÔ
        if active_tab == "💰 Chi phí":
            render_cost_dashboard(supabase, c_farm, c_team)

        # TAB 1: KHỞI TẠO LÔ
        elif active_tab == "🌱 Khởi tạo Lô trồng":
            st.markdown("#### Đăng ký đợt xuống giống mới")
            df_lots = fetch_table_data("base_lots", c_farm)
            df_lots_team = df_lots[df_lots["team"] == c_team] if not df_lots.empty else pd.DataFrame()
            editing_row, is_within_48h = get_editing_row("base_lots", df_lots_team)
            is_editing = editing_row is not None
            
            with st.container(border=True):
                col_a, col_b = st.columns(2)
                with col_a:
                    lo = st.text_input("🏷️ Tên Lô (VD: A1, B3...)", placeholder="Nhập tên lô...", key="add_base_lo")
                    loai_trong = st.selectbox("🌱 Loại trồng", options=LOAI_TRONG_OPTIONS, key="add_base_loai")
                with col_b:
                    col_b1, col_b2 = st.columns([2, 1])
                    with col_b1:
                        ngay_trong = st.date_input("📆 Ngày trồng", value=date.today(), key="add_base_ngay")
                    with col_b2:
                        st.text_input("📍 Tuần", value=str(ngay_trong.isocalendar()[1]), disabled=True, key=f"main_w_base_{ngay_trong}")
                    so_luong = st.number_input("🔢 Số lượng trồng (cây)", min_value=0, step=100, key="add_base_sl")

                if st.button("✅ Tạo Lô Trồng", key="btn_add_base", use_container_width=True, type="primary"):
                    if not lo.strip(): st.error("❌ Nhập tên lô.")
                    elif so_luong <= 0: st.error("❌ Cần nhập số lượng.")
                    else:
                        ten_lo_goc = lo.strip().upper()
                        ngay_trong_str = ngay_trong.strftime('%d%m%Y')
                        lot_id = f"{ten_lo_goc}_{ngay_trong_str}"
                        # Auto-create lô trong dim_lo nếu chưa tồn tại
                        dim_lo_id = get_or_create_dim_lo(c_farm, ten_lo_goc, c_team)
                        if not dim_lo_id:
                            st.error(f"❌ Không thể khởi tạo Lô '{ten_lo_goc}'. Vui lòng thử lại hoặc liên hệ quản trị viên.")
                        else:
                            data_base = {
                                "dim_lo_id": dim_lo_id,
                                "ngay_trong": ngay_trong.isoformat(), "so_luong": so_luong,
                                "so_luong_con_lai": so_luong,
                                "tuan": ngay_trong.isocalendar()[1],
                                "loai_trong": loai_trong
                            }
                            data_season = {
                                "dim_lo_id": dim_lo_id, "vu": "F0",
                                "loai_trong": loai_trong,
                                "ngay_bat_dau": ngay_trong.isoformat()
                            }
                            confirm_action_dialog("INSERT_BASE", "base_lots", None, (data_base, data_season), f"✅ Tạo Lô {lot_id} thành công!")

            st.markdown("---")
            col_t, col_e, col_d = st.columns([5, 1.5, 1.5])
            with col_t:
                st.markdown('<p class="dataframe-header" style="margin-top:0.5rem;">Dữ liệu của đội bạn (Click 1 dòng để sửa/xóa)</p>', unsafe_allow_html=True)
            if is_editing and is_within_48h:
                with col_e:
                    if st.button("✏️ Chỉnh sửa", key="edit_base_nt", use_container_width=True):
                        edit_base_lot_dialog(editing_row)
                with col_d:
                    if st.button("🗑️ Xóa", key="del_base_nt", use_container_width=True):
                        confirm_action_dialog("DELETE", "base_lots", editing_row["id"], None, f"✅ Đã xóa thành công lô {editing_row.get('lot_id')}!")
            elif is_editing and not is_within_48h:
                with col_e: st.caption("🔒 Quá 48h")

            render_team_dataframe("base_lots", df_lots_team, ["lot_id", "ngay_trong", "so_luong", "created_at"])

        # TAB: ĐO SIZE (DÀNH CHO NT1/NT2)
        elif active_tab == "📏 Đo Size":
            st.markdown("#### Đo kích thước buồng mẫu")
            available_lots = get_lots_by_farm(c_farm)
            if not available_lots:
                st.warning("⚠️ Chưa có Lô trồng nào.")
            else:
                df_sm = fetch_table_data("size_measure_logs", c_farm)
                df_sm_team = df_sm[df_sm["team"] == c_team] if not df_sm.empty else pd.DataFrame()
                editing_row, is_within_48h = get_editing_row("size_measure_logs", df_sm_team)
                is_editing = editing_row is not None
                
                with st.container(border=True):
                    col_a, col_b = st.columns(2)
                    with col_a:
                        lot_id = st.selectbox("🏷️ Chọn Lô", options=available_lots, key="add_sm_lot")
                        mau_day_color = build_color_selectbox("add_sm")
                        lan_do = st.radio("📏 Lần đo", options=[1, 2], horizontal=True, key="add_sm_lando")
                    with col_b:
                        col_b1, col_b2 = st.columns([2, 1])
                        with col_b1:
                            ngay_do = st.date_input("📆 Ngày đo", value=date.today(), key="add_sm_ngay")
                        with col_b2:
                            st.text_input("📍 Tuần", value=str(ngay_do.isocalendar()[1]), disabled=True, key=f"main_w_sm_{ngay_do}")
                        col_b3, col_b4 = st.columns(2)
                        with col_b3:
                            hang_kiem_tra = st.text_input("📏 Hàng kiểm tra", placeholder="VD: H1-H5", key="add_sm_hkt")
                        with col_b4:
                            size_cal = st.number_input("📏 Size (Cal)", min_value=0.0, step=0.1, key="add_sm_cal")
                        sl = st.number_input("🔢 Số lượng buồng mẫu", min_value=0, step=10, key="add_sm_sl")
                    
                    if st.button("➕ Thêm vào Danh sách", key="btn_add_sm", use_container_width=True, type="secondary"):
                        if not mau_day_color: st.error("❌ Phải chọn màu dây")
                        elif sl <= 0: st.error("❌ Số lượng buồng > 0")
                        else:
                            # Validate/create ribbon
                            farm_id = get_farm_id_from_name(c_farm)
                            iso = ngay_do.isocalendar()
                            _, err = get_or_create_ribbon(farm_id, iso[0], iso[1], mau_day_color)
                            if err:
                                st.error(err)
                                st.stop()
                            # Validation: Nếu chọn là Lần 2, phải kiểm tra xem Lần 1 đã có chưa.
                            if lan_do == 2:
                                _dim_id = get_dim_lo_id(c_farm, lot_id)
                                if _dim_id:
                                    res_lan1 = supabase.table("size_measure_logs") \
                                        .select("id").eq("dim_lo_id", _dim_id).eq("mau_day", mau_day_color).eq("lan_do", 1).eq("is_deleted", False).execute()
                                    if not res_lan1.data:
                                        st.error(f"❌ Không thể đo Lần 2. Lô `{lot_id}` với màu dây `{mau_day_color}` chưa được đo Lần 1.")
                                        st.stop()
                                    
                            st.session_state["queue_sm"].append({
                                "Lô": lot_id, "Màu dây": mau_day_color, "Lần đo": lan_do, "Số lượng": sl,
                                "Ngày đo": ngay_do.isoformat(), "Tuần": ngay_do.isocalendar()[1],
                                "Hàng KT": hang_kiem_tra.strip(), "Size": size_cal
                            })
                            st.rerun()

                def process_sm_queue():
                    queue = st.session_state["queue_sm"]
                    success_count = 0
                    for item in queue:
                        data = {
                            "farm": c_farm, "team": c_team, "lot_id": item["Lô"],
                            "mau_day": item["Màu dây"], "lan_do": item["Lần đo"], "so_luong_mau": item["Số lượng"],
                            "ngay_do": item["Ngày đo"], "tuan": item["Tuần"],
                            "hang_kiem_tra": item["Hàng KT"], "size_cal": item["Size"]
                        }
                        if insert_to_db("size_measure_logs", data):
                            success_count += 1
                        else:
                            st.error(f"❌ Lỗi ghi lô {item['Lô']}")
                            return
                    st.session_state["queue_sm"] = []
                    st.session_state["toast"] = f"✅ Đã lưu {success_count} dòng Đo Size!"
                    st.cache_data.clear()
                    st.rerun()

                render_queue_ui("queue_sm", ["Lô", "Màu dây", "Lần đo", "Số lượng", "Hàng KT", "Size", "Ngày đo", "Tuần"], process_sm_queue)
                
                st.markdown("---")
                col_t, col_e, col_d = st.columns([5, 1.5, 1.5])
                with col_t:
                    st.markdown('<p class="dataframe-header" style="margin-top:0.5rem;">Dữ liệu của đội bạn</p>', unsafe_allow_html=True)
                if is_editing and is_within_48h:
                    with col_e:
                        if st.button("✏️ Sửa", key="edit_sm_nt", use_container_width=True):
                            edit_size_measure_dialog(editing_row, available_lots)
                    with col_d:
                        if st.button("🗑️ Xóa", key="del_sm_nt", use_container_width=True):
                            confirm_action_dialog("DELETE", "size_measure_logs", editing_row["id"], None, "✅ Đã xóa Đo Size!")
                elif is_editing and not is_within_48h:
                    with col_e: st.caption("🔒 Quá 48h")
                    
                render_team_dataframe("size_measure_logs", df_sm_team, ["lot_id", "mau_day", "lan_do", "hang_kiem_tra", "size_cal", "so_luong_mau", "ngay_do"])

        # TAB: CẬP NHẬT KIỂM KÊ CÂY
        elif active_tab == "🌳 Kiểm kê cây":
            st.markdown("#### Báo cáo số lượng cây thực tế (Ngẫu nhiên / Tháng)")
            available_lots = get_lots_by_farm(c_farm)
            if not available_lots:
                st.warning("⚠️ Chưa có Lô trồng nào.")
            else:
                df_inv = fetch_table_data("tree_inventory_logs", c_farm)
                df_inv_team = df_inv[df_inv["team"] == c_team] if not df_inv.empty else pd.DataFrame()
                editing_row, is_within_48h = get_editing_row("tree_inventory_logs", df_inv_team)
                is_editing = editing_row is not None
                
                with st.container(border=True):
                    col_a, col_b = st.columns(2)
                    with col_a:
                        lot_id = st.selectbox("🏷️ Chọn Lô", options=available_lots, key="add_inv_lot")
                    with col_b:
                        col_b1, col_b2 = st.columns([2, 1])
                        with col_b1:
                            ngay_kk = st.date_input("📆 Ngày kiểm kê", value=date.today(), key="add_inv_ngay")
                        with col_b2:
                            st.text_input("📍 Tuần", value=str(ngay_kk.isocalendar()[1]), disabled=True, key=f"main_w_inv_{ngay_kk}")
                        sl = st.number_input("🔢 Số lượng cây thực tế", min_value=0, step=100, key="add_inv_sl")
                    
                    if st.button("➕ Thêm vào Danh sách", key="btn_add_inv", use_container_width=True, type="secondary"):
                        if sl <= 0: st.error("❌ Số lượng phải lớn hơn 0")
                        else:
                            st.session_state["queue_inv"].append({
                                "Lô": lot_id, "Số lượng": sl, "Ngày": ngay_kk.isoformat(), "Tuần": ngay_kk.isocalendar()[1]
                            })
                            st.rerun()

                def process_inv_queue():
                    queue = st.session_state["queue_inv"]
                    success_count = 0
                    for item in queue:
                        data = {
                            "farm": c_farm, "team": c_team, "lot_id": item["Lô"],
                            "so_luong_cay_thuc_te": item["Số lượng"], "ngay_kiem_ke": item["Ngày"],
                            "tuan": item["Tuần"]
                        }
                        if insert_to_db("tree_inventory_logs", data):
                            success_count += 1
                        else:
                            st.error(f"❌ Lỗi ghi lô {item['Lô']}")
                            return
                    st.session_state["queue_inv"] = []
                    st.cache_data.clear()
                    st.session_state["toast"] = f"✅ Đã lưu {success_count} dòng Kiểm kê!"
                    st.rerun()

                render_queue_ui("queue_inv", ["Lô", "Số lượng", "Ngày", "Tuần"], process_inv_queue)
                
                st.markdown("---")
                col_t, col_e, col_d = st.columns([5, 1.5, 1.5])
                with col_t:
                    st.markdown('<p class="dataframe-header" style="margin-top:0.5rem;">Dữ liệu của đội bạn</p>', unsafe_allow_html=True)
                if is_editing and is_within_48h:
                    with col_e:
                        if st.button("✏️ Sửa", key="edit_inv_nt", use_container_width=True):
                            edit_tree_inventory_dialog(editing_row, available_lots)
                    with col_d:
                        if st.button("🗑️ Xóa", key="del_inv_nt", use_container_width=True):
                            confirm_action_dialog("DELETE", "tree_inventory_logs", editing_row["id"], None, "✅ Đã xóa nhật ký kiểm kê!")
                elif is_editing and not is_within_48h:
                    with col_e: st.caption("🔒 Quá 48h")
                    
                render_team_dataframe("tree_inventory_logs", df_inv_team, ["lot_id", "ngay_kiem_ke", "so_luong_cay_thuc_te"])

        # TAB: ĐO PH ĐẤT
        elif active_tab == "🧪 Đo pH Đất":
            st.markdown("#### Ghi nhận kết quả Đo pH Đất")
            available_lots = get_lots_by_farm(c_farm)
            
            if not available_lots:
                st.warning("⚠️ Chưa có Lô trồng nào trên hệ thống.")
            else:
                df_ph = fetch_table_data("soil_ph_logs", c_farm)
                df_ph_team = df_ph[df_ph["team"] == c_team] if not df_ph.empty else pd.DataFrame()
                
                with st.container(border=True):
                    col_a, col_b, col_c = st.columns([2, 1, 1])
                    with col_a:
                        lot_id = st.selectbox("🏷️ Chọn Lô", options=available_lots, key="add_ph_lot")
                    with col_b:
                        ngay_do = st.date_input("📆 Ngày đo", value=date.today(), key="add_ph_ngay")
                    with col_c:
                        tuan_do = st.text_input("📍 Tuần", value=str(ngay_do.isocalendar()[1]), disabled=True, key=f"add_ph_tuan_{ngay_do}")
                    
                    st.markdown("---")
                    
                    @st.fragment
                    def render_ph_inputs():
                        if "ph_measure_count" not in st.session_state:
                            st.session_state.ph_measure_count = 1
                        
                        ph_data = []
                        for i in range(1, st.session_state.ph_measure_count + 1):
                            col1, col2 = st.columns([1, 3])
                            with col1:
                                st.markdown(f"**Lần đo {i}**")
                            with col2:
                                val = st.number_input(f"pH", min_value=0.0, max_value=14.0, step=0.1, key=f"add_ph_val_{i}", label_visibility="collapsed")
                                ph_data.append(val)
                                
                        st.markdown("")
                        col_btn_add, col_btn_save = st.columns(2)
                        with col_btn_add:
                            if st.button("➕ Thêm lần đo", use_container_width=True, type="secondary"):
                                st.session_state.ph_measure_count += 1
                                st.rerun()
                                
                        with col_btn_save:
                            if st.button("🚀 Lưu kết quả", use_container_width=True, type="primary"):
                                success_count = 0
                                for lan, val in enumerate(ph_data, start=1):
                                    if val > 0:
                                        data = {
                                            "farm": c_farm, "team": c_team, "lot_id": lot_id,
                                            "ngay_do": ngay_do.isoformat(), "tuan": ngay_do.isocalendar()[1],
                                            "lan_do": lan, "ph_value": val
                                        }
                                        try:
                                            supabase.table("soil_ph_logs").insert(data).execute()
                                            success_count += 1
                                        except Exception as e:
                                            st.error(f"❌ Lỗi ghi Lần {lan}: {e}")
                                
                                if success_count > 0:
                                    st.session_state.ph_measure_count = 1
                                    st.cache_data.clear()
                                    st.session_state["toast"] = f"✅ Ghi nhận {success_count} kết quả Đo pH thành công!"
                                    st.rerun()

                    render_ph_inputs()
                
                st.markdown("---")
                col_t, col_e, col_d = st.columns([5, 1.5, 1.5])
                with col_t:
                    st.markdown('<p class="dataframe-header" style="margin-top:0.5rem;">Dữ liệu của đội bạn (Click 1 dòng để sửa/xóa)</p>', unsafe_allow_html=True)
                
                editing_row, is_within_48h = get_editing_row("soil_ph_logs", df_ph_team)
                is_editing = editing_row is not None
                
                if is_editing and is_within_48h:
                    with col_e:
                        if st.button("✏️ Chỉnh sửa", key="edit_ph_nt", use_container_width=True):
                            edit_soil_ph_dialog(editing_row, available_lots)
                    with col_d:
                        if st.button("🗑️ Xóa", key="del_ph_nt", use_container_width=True):
                            confirm_action_dialog("DELETE", "soil_ph_logs", editing_row["id"], None, "✅ Đã xóa kết quả đo pH!")
                elif is_editing and not is_within_48h:
                    with col_e: st.caption("🔒 Quá 48h")
                    with col_d: st.empty()
                else:
                    with col_e: st.empty()
                    with col_d: st.empty()
                    
                render_team_dataframe("soil_ph_logs", df_ph_team, ["lot_id", "ngay_do", "lan_do", "ph_value"])

        # TAB 2: CẬP NHẬT TIẾN ĐỘ NT
        elif active_tab == "📈 Cập nhật Tiến độ":
            st.markdown("#### Ghi nhận: Chích bắp / Cắt bắp")
            available_lots = get_lots_by_farm(c_farm)
            if not available_lots:
                st.warning("⚠️ Chưa có Lô trồng nào. Hãy tạo ở Tab 1.")
            else:
                df_stg = fetch_table_data("stage_logs", c_farm)
                df_stg_team = df_stg[df_stg["team"] == c_team] if not df_stg.empty else pd.DataFrame()
                editing_row, is_within_48h = get_editing_row("stage_logs", df_stg_team)
                is_editing = editing_row is not None
                
                with st.container(border=True):
                    col_a, col_b = st.columns(2)
                    with col_a:
                        lot_id = st.selectbox("🏷️ Chọn Lô", options=available_lots, key="add_stg_lot")
                        if c_team == "Đội BVTV":
                            giai_doan_opts = ["Chích bắp"]
                        else:
                            giai_doan_opts = ["Cắt bắp"]
                        giai_doan = st.radio("📌 Giai đoạn", options=giai_doan_opts, horizontal=True, key="add_stg_gd")
                        # Đội BVTV chỉ nhập Chích bắp → không cần màu dây
                        mau_day_color = None
                        if giai_doan == "Cắt bắp":
                            mau_day_color = build_color_selectbox("add_stg")
                    with col_b:
                        col_b1, col_b2 = st.columns([2, 1])
                        with col_b1:
                            ngay_th = st.date_input("📆 Ngày thực hiện", value=date.today(), key="add_stg_ngay")
                        with col_b2:
                            st.text_input("📍 Tuần", value=str(ngay_th.isocalendar()[1]), disabled=True, key=f"main_w_stg_{ngay_th}")
                        sl = st.number_input("🔢 Số lượng cây", min_value=0, step=100, key="add_stg_sl")

                    if st.button("➕ Thêm vào Danh sách", key="btn_add_stg", use_container_width=True, type="secondary"):
                        if sl <= 0: st.error("❌ Nhập số lượng > 0.")
                        elif giai_doan == "Cắt bắp" and not mau_day_color: st.error("❌ Phải chọn màu dây định danh lứa.")
                        else:
                            # Validate/create ribbon for Cắt bắp
                            if giai_doan == "Cắt bắp" and mau_day_color:
                                farm_id = get_farm_id_from_name(c_farm)
                                iso = ngay_th.isocalendar()
                                _, err = get_or_create_ribbon(farm_id, iso[0], iso[1], mau_day_color)
                                if err:
                                    st.error(err)
                                    st.stop()
                            is_time_valid, msg_time = validate_timeline_logic(c_farm, lot_id, ngay_th, giai_doan)
                            if not is_time_valid:
                                st.error(msg_time)
                            else:
                                st.session_state["queue_stg"].append({
                                    "Lô": lot_id, "Giai đoạn": giai_doan, "Màu dây": mau_day_color,
                                    "Ngày": ngay_th.isoformat(), "Số lượng": sl, "Tuần": ngay_th.isocalendar()[1]
                                })
                                st.rerun()

                def process_stg_queue():
                    queue = st.session_state["queue_stg"]
                    lot_reqs = {}
                    for item in queue:
                        k = (item["Lô"], item["Giai đoạn"])
                        lot_reqs[k] = lot_reqs.get(k, 0) + item["Số lượng"]
                    for (l_id, g_doan), req_sl in lot_reqs.items():
                        valid, msg = check_quantity_limit(c_farm, l_id, req_sl, "stage", giai_doan=g_doan)
                        if not valid:
                            st.error(f"❌ Lỗi tổng số lượng ở Lô {l_id} - {g_doan}: {msg}")
                            return
                    success_count = 0
                    for item in queue:
                        is_valid, msg, allocations = allocate_fifo_quantity(c_farm, item["Lô"], item["Số lượng"], "stage", item["Ngày"], item["Giai đoạn"], item["Giai đoạn"])
                        if is_valid:
                            for alloc in allocations:
                                data = {"farm": c_farm, "team": c_team, "giai_doan": item["Giai đoạn"], "ngay_thuc_hien": item["Ngày"], "tuan": item["Tuần"], "lot_id": alloc["lot_id"], "so_luong": alloc["so_luong"], "base_lot_id": alloc.get("base_lot_id")}
                                if not insert_to_db("stage_logs", data):
                                    st.error(f"❌ Lỗi ghi phân rã {alloc['lot_id']}")
                                    return
                            success_count += 1
                        else:
                            st.error(msg)
                            return
                    st.session_state["queue_stg"] = []
                    st.cache_data.clear()
                    st.session_state["toast"] = f"✅ Đã lưu {success_count} dòng Tiến độ!"
                    st.rerun()

                render_queue_ui("queue_stg", ["Lô", "Giai đoạn", "Màu dây", "Số lượng", "Ngày", "Tuần"], process_stg_queue)

                st.markdown("---")
                col_t, col_e, col_d = st.columns([5, 1.5, 1.5])
                with col_t:
                    st.markdown('<p class="dataframe-header" style="margin-top:0.5rem;">Dữ liệu của đội bạn (Click 1 dòng để sửa/xóa)</p>', unsafe_allow_html=True)
                if is_editing and is_within_48h:
                    with col_e:
                        if st.button("✏️ Chỉnh sửa", key="edit_stg_nt", use_container_width=True):
                            edit_stage_log_dialog(editing_row, available_lots, c_team)
                    with col_d:
                        if st.button("🗑️ Xóa", key="del_stg_nt", use_container_width=True):
                            confirm_action_dialog("DELETE", "stage_logs", editing_row["id"], None, "✅ Đã xóa thành công!")
                elif is_editing and not is_within_48h:
                    with col_e: st.caption("🔒 Quá 48h")
                
                render_team_dataframe("stage_logs", df_stg_team, ["lot_id", "giai_doan", "ngay_thuc_hien", "so_luong"])

        # TAB 3: XUẤT HỦY
        elif active_tab == "🗑️ Cập nhật Xuất hủy":
            st.markdown("#### Ghi nhận số lượng cây chết / hư hỏng")
            available_lots = get_lots_by_farm(c_farm)
            if not available_lots:
                st.warning("⚠️ Chưa có Lô trồng nào.")
            else:
                df_des = fetch_table_data("destruction_logs", c_farm)
                df_des_team = df_des[df_des["team"] == c_team] if not df_des.empty else pd.DataFrame()
                editing_row, is_within_48h = get_editing_row("destruction_logs", df_des_team)
                is_editing = editing_row is not None
                
                with st.container(border=True):
                    col_a, col_b = st.columns(2)
                    with col_a:
                        lot_id = st.selectbox("🏷️ Chọn Lô", options=available_lots, key="add_des_lot")
                        giai_doan_xuat_huy = st.selectbox("⏱️ Giai đoạn xuất hủy", options=DESTRUCTION_STAGE_OPTIONS, key="add_des_gxh")
                        
                        # Ribbon color — only for "Trước thu hoạch"
                        mau_day_color = None
                        if giai_doan_xuat_huy == "Trước thu hoạch":
                            farm_id = get_farm_id_from_name(c_farm)
                            if farm_id:
                                ribbons = get_all_ribbons_for_farm(farm_id)
                                ribbon_opts = sorted({r["color_name"] for r in ribbons})
                                if ribbon_opts:
                                    mau_day_color = st.selectbox("🎨 Màu dây", options=ribbon_opts, key="add_des_mau_sel")
                                else:
                                    st.warning("⚠️ Farm chưa có màu dây nào trong ribbon_schedule.")
                        
                        predefined_reasons = ["Bệnh", "Đổ Ngã", "Khác"]
                        if hasattr(st, "pills"):
                            selected_reason = st.pills("📝 Nhóm lý do", options=predefined_reasons, key="add_des_reason_group")
                            if not selected_reason:
                                selected_reason = "Khác"
                        else:
                            selected_reason = st.radio("📝 Nhóm lý do", options=predefined_reasons, horizontal=True, key="add_des_reason_group")
                            
                        if selected_reason == "Khác":
                            ly_do = st.text_area("📝 Chi tiết lý do (Bắt buộc)", height=80, key="add_des_lydo")
                        else:
                            ly_do = selected_reason
                            
                    with col_b:
                        col_b1, col_b2 = st.columns([2, 1])
                        with col_b1:
                            ngay = st.date_input("📆 Ngày xuất hủy", value=date.today(), key="add_des_ngay")
                        with col_b2:
                            st.text_input("📍 Tuần", value=str(ngay.isocalendar()[1]), disabled=True, key=f"main_w_des_{ngay}")
                        sl = st.number_input("🔢 Số lượng cây xuất hủy", min_value=0, step=10, key="add_des_sl")

                    if st.button("➕ Thêm vào Danh sách", key="btn_add_des", use_container_width=True, type="secondary"):
                        if sl <= 0: st.error("❌ Nhập số lượng > 0.")
                        elif selected_reason == "Khác" and not ly_do.strip(): st.error("❌ Cần ghi rõ chi tiết lý do (khi chọn Khác).")
                        else:
                            st.session_state["queue_des"].append({
                                "Lô": lot_id, "Giai đoạn": giai_doan_xuat_huy, "Màu dây": mau_day_color or "", "Lý do": ly_do.strip(),
                                "Ngày": ngay.isoformat(), "Số lượng": sl, "Tuần": ngay.isocalendar()[1]
                            })
                            st.rerun()

                def process_des_queue():
                    queue = st.session_state["queue_des"]
                    lot_reqs = {}
                    for item in queue:
                        lot_reqs[item["Lô"]] = lot_reqs.get(item["Lô"], 0) + item["Số lượng"]
                    for l_id, req_sl in lot_reqs.items():
                        valid, msg = check_quantity_limit(c_farm, l_id, req_sl, "destruction")
                        if not valid:
                            st.error(f"❌ Lỗi tổng số lượng ở Lô {l_id}: {msg}")
                            return
                    success_count = 0
                    for item in queue:
                        is_valid, msg, allocations = allocate_fifo_quantity(c_farm, item["Lô"], item["Số lượng"], "destruction", item["Ngày"], "Xuất hủy", giai_doan=item.get("Giai đoạn"), mau_day=item.get("Màu dây"))
                        if is_valid:
                            for alloc in allocations:
                                data = {"farm": c_farm, "team": c_team, "ngay_xuat_huy": item["Ngày"], "giai_doan": item["Giai đoạn"], "ly_do": item["Lý do"], "tuan": item["Tuần"], "lot_id": alloc["lot_id"], "so_luong": alloc["so_luong"], "base_lot_id": alloc.get("base_lot_id")}
                                if not insert_to_db("destruction_logs", data):
                                    st.error(f"❌ Lỗi ghi phân rã {alloc['lot_id']}")
                                    return
                            success_count += 1
                        else:
                            st.error(msg)
                            return
                    st.session_state["queue_des"] = []
                    st.cache_data.clear()
                    st.session_state["toast"] = f"✅ Đã lưu xuất hủy {success_count} dòng!"
                    st.rerun()

                render_queue_ui("queue_des", ["Lô", "Giai đoạn", "Màu dây", "Lý do", "Số lượng", "Ngày", "Tuần"], process_des_queue)

                st.markdown("---")
                col_t, col_e, col_d = st.columns([5, 1.5, 1.5])
                with col_t:
                    st.markdown('<p class="dataframe-header" style="margin-top:0.5rem;">Dữ liệu của đội bạn (Click 1 dòng để sửa/xóa)</p>', unsafe_allow_html=True)
                if is_editing and is_within_48h:
                    with col_e:
                        if st.button("✏️ Chỉnh sửa", key="edit_des_nt", use_container_width=True):
                            edit_destruction_log_dialog(editing_row, available_lots)
                    with col_d:
                        if st.button("🗑️ Xóa", key="del_des_nt", use_container_width=True):
                            confirm_action_dialog("DELETE", "destruction_logs", editing_row["id"], None, "✅ Đã xóa xuất hủy!")
                elif is_editing and not is_within_48h:
                    with col_e: st.caption("🔒 Quá 48h")
                    
                render_team_dataframe("destruction_logs", df_des_team, ["lot_id", "ngay_xuat_huy", "giai_doan", "so_luong", "ly_do"])

        # TAB 8: KIỂM TRA FUSARIUM
        elif active_tab == "🦠 Kiểm tra Fusarium":
            st.markdown("#### Ghi nhận số lượng cây bị bệnh Fusarium")
            available_lots = get_lots_by_farm(c_farm)
            if not available_lots:
                st.warning("⚠️ Chưa có Lô trồng nào. Hãy tạo ở Tab 1.")
            else:
                df_fus = fetch_table_data("fusarium_logs", c_farm)
                df_fus_team = df_fus[df_fus["team"] == c_team] if not df_fus.empty else pd.DataFrame()
                editing_row, is_within_48h = get_editing_row("fusarium_logs", df_fus_team)
                is_editing = editing_row is not None
                
                with st.container(border=True):
                    col_a, col_b = st.columns(2)
                    with col_a:
                        lot_id = st.selectbox("🏷️ Chọn Lô kiểm tra", options=available_lots, key="add_fus_lot")
                    with col_b:
                        col_b1, col_b2 = st.columns([2, 1])
                        with col_b1:
                            ngay_kiem_tra = st.date_input("📆 Ngày kiểm tra", value=date.today(), key="add_fus_ngay")
                        with col_b2:
                            st.text_input("📍 Tuần", value=str(ngay_kiem_tra.isocalendar()[1]), disabled=True, key=f"main_w_fus_{ngay_kiem_tra}")
                        so_cay = st.number_input("🔢 Số lượng cây bị Fusarium", min_value=0, step=1, key="add_fus_cay")
                        
                    st.markdown("")
                    if st.button("➕ Thêm vào Danh sách", key="btn_add_fus", use_container_width=True, type="secondary"):
                        if so_cay < 0: st.error("❌ Số lượng cây phải >= 0.")
                        else:
                            st.session_state["queue_fus"].append({
                                "Lô": lot_id, "Ngày": ngay_kiem_tra.isoformat(), "Số cây": so_cay, "Tuần": ngay_kiem_tra.isocalendar()[1]
                            })
                            st.rerun()

                def process_fus_queue():
                    queue = st.session_state["queue_fus"]
                    success_count = 0
                    for item in queue:
                        data = {
                            "farm": c_farm, "team": c_team, "lot_id": item["Lô"],
                            "ngay_kiem_tra": item["Ngày"], "so_cay_fusarium": item["Số cây"], "tuan": item["Tuần"]
                        }
                        try:
                            supabase.table("fusarium_logs").insert(data).execute()
                            success_count += 1
                        except Exception as e:
                            st.error(f"❌ Lỗi ghi Lô {item['Lô']}: {e}")
                            return
                            
                    st.session_state["queue_fus"] = []
                    st.cache_data.clear()
                    st.session_state["toast"] = f"✅ Ghi nhận {success_count} kết quả kiểm tra Fusarium thành công!"
                    st.rerun()

                render_queue_ui("queue_fus", ["Lô", "Số cây", "Ngày", "Tuần"], process_fus_queue)
                
                st.markdown("---")
                col_t, col_e, col_d = st.columns([5, 1.5, 1.5])
                with col_t:
                    st.markdown('<p class="dataframe-header" style="margin-top:0.5rem;">Dữ liệu của đội bạn (Click 1 dòng để sửa/xóa)</p>', unsafe_allow_html=True)
                
                if is_editing and is_within_48h:
                    with col_e:
                        if st.button("✏️ Chỉnh sửa", key="edit_fus_nt", use_container_width=True):
                            edit_fusarium_log_dialog(editing_row, available_lots)
                    with col_d:
                        if st.button("🗑️ Xóa", key="del_fus_nt", use_container_width=True):
                            confirm_action_dialog("DELETE", "fusarium_logs", editing_row["id"], None, "✅ Đã xóa bản ghi kiểm tra Fusarium!")
                elif is_editing and not is_within_48h:
                    with col_e: st.caption("🔒 Quá 48h")
                    with col_d: st.empty()
                else:
                    with col_e: st.empty()
                    with col_d: st.empty()
                    
                render_team_dataframe("fusarium_logs", df_fus_team, ["lot_id", "ngay_kiem_tra", "so_cay_fusarium", "tuan"])

        # TAB 4: DỮ LIỆU TOÀN CỤC
        elif active_tab == "🌐 Dữ liệu toàn cục":
            render_global_data_tab(c_farm)

    # =================================================
    # MODULE 2: ĐỘI THU HOẠCH
    # =================================================
    elif c_team == "Đội Thu Hoạch":
        tab_opts = ["🌐 Dữ liệu toàn cục", "💰 Chi phí", "🍌 Nhật ký Thu Hoạch"]
        active_tab = st.segmented_control("Chức năng", tab_opts, label_visibility="collapsed", key="tab_har_menu", default=tab_opts[0])
        if active_tab is None: active_tab = tab_opts[0]
        
        if active_tab == tab_opts[2]:
            st.markdown("#### Ghi nhận Sản lượng Thu hoạch hàng ngày")
            available_lots = get_lots_by_farm(c_farm)
            if not available_lots:
                st.warning("⚠️ Chưa có Lô trồng nào trên hệ thống.")
            else:
                df_har = fetch_table_data("harvest_logs", c_farm)
                df_har_team = df_har[df_har["team"] == c_team] if not df_har.empty else pd.DataFrame()
                editing_row, is_within_48h = get_editing_row("harvest_logs", df_har_team)
                is_editing = editing_row is not None
                
                with st.container(border=True):
                    col_a, col_b = st.columns(2)
                    with col_a:
                        lot_id = st.selectbox("🏷️ Chọn Lô thu hoạch", options=available_lots, key="add_har_lot")
                        # Ribbon color from schedule
                        farm_id_har = get_farm_id_from_name(c_farm)
                        ribbons_har = get_all_ribbons_for_farm(farm_id_har) if farm_id_har else []
                        ribbon_opts_har = sorted({r["color_name"] for r in ribbons_har})
                        if ribbon_opts_har:
                            mau_day_color = st.selectbox("🎨 Màu dây", options=ribbon_opts_har, key="add_har_mau_sel")
                        else:
                            st.warning("⚠️ Farm chưa có màu dây trong ribbon_schedule.")
                            mau_day_color = None
                        hinh_thuc_thu_hoach = st.selectbox("🚜 Hình thức thu hoạch", options=["Bằng xe cày", "Bằng ròng rọc"], key="add_har_hinh_thuc")
                    with col_b:
                        col_b1, col_b2 = st.columns([2, 1])
                        with col_b1:
                            ngay = st.date_input("📆 Ngày thu hoạch", value=date.today(), key="add_har_dt")
                        with col_b2:
                            st.text_input("📍 Tuần", value=str(ngay.isocalendar()[1]), disabled=True, key=f"main_w_har_{ngay}")
                        sl = st.number_input("🍌 Số lượng buồng thu hoạch", min_value=0, step=50, key="add_har_sl")
    
                    st.markdown("")
                    if st.button("➕ Thêm vào Danh sách", key="btn_add_har", use_container_width=True, type="secondary"):
                        if not mau_day_color: st.error("❌ Cần chọn Màu dây.")
                        elif sl <= 0: st.error("❌ Số lượng buồng phải > 0")
                        else:
                            # Validation: Bắt buộc đã đo size lần 1 cho lô và màu dây này
                            _dim_id_har = get_dim_lo_id(c_farm, lot_id)
                            if _dim_id_har:
                                res_sm = supabase.table("size_measure_logs").select("id") \
                                    .eq("dim_lo_id", _dim_id_har).eq("mau_day", mau_day_color).eq("lan_do", 1).eq("is_deleted", False).execute()
                                if not res_sm.data:
                                    st.error(f"❌ Cảnh báo: Lô `{lot_id}` với màu dây `{mau_day_color}` chưa qua Đo Size lần 1. Hãy nhắc NT.")
                                    st.stop()
                                
                            st.session_state["queue_har"].append({
                                "Lô": lot_id, "Màu dây": mau_day_color, "Hình thức": hinh_thuc_thu_hoach,
                                "Số lượng": sl, "Ngày": ngay.isoformat(), "Tuần": ngay.isocalendar()[1]
                            })
                            st.rerun()

                def process_har_queue():
                    queue = st.session_state["queue_har"]
                    lot_reqs = {}
                    for item in queue:
                        lot_reqs[item["Lô"]] = lot_reqs.get(item["Lô"], 0) + item["Số lượng"]
                    for l_id, req_sl in lot_reqs.items():
                        valid, msg = check_quantity_limit(c_farm, l_id, req_sl, "harvest")
                        if not valid:
                            st.error(f"❌ Lỗi tổng số lượng ở Lô {l_id}: {msg}")
                            return
                    success_count = 0
                    for item in queue:
                        is_valid, msg, allocations = allocate_fifo_quantity(c_farm, item["Lô"], item["Số lượng"], "harvest", item["Ngày"], "Thu hoạch", mau_day=item["Màu dây"])
                        if is_valid:
                            for alloc in allocations:
                                data = {
                                    "farm": c_farm, "team": c_team, "ngay_thu_hoach": item["Ngày"],
                                    "hinh_thuc_thu_hoach": item["Hình thức"], "tuan": item["Tuần"], "lot_id": alloc["lot_id"], "so_luong": alloc["so_luong"],
                                    "base_lot_id": alloc.get("base_lot_id"), "mau_day": item["Màu dây"]
                                }
                                if not insert_to_db("harvest_logs", data):
                                    st.error(f"❌ Lỗi ghi phân rã {alloc['lot_id']}")
                                    return
                            success_count += 1
                        else:
                            st.error(msg)
                            return
                    st.session_state["queue_har"] = []
                    st.cache_data.clear()
                    st.session_state["toast"] = f"✅ Đã lưu thu hoạch {success_count} nhóm Lô thành công!"
                    st.rerun()

                render_queue_ui("queue_har", ["Lô", "Màu dây", "Hình thức", "Số lượng", "Ngày", "Tuần"], process_har_queue)
                
                st.markdown("---")
                col_t, col_e, col_d = st.columns([5, 1.5, 1.5])
                with col_t:
                    st.markdown('<p class="dataframe-header" style="margin-top:0.5rem;">Dữ liệu của đội bạn (Click 1 dòng để sửa/xóa)</p>', unsafe_allow_html=True)
                if is_editing and is_within_48h:
                    with col_e:
                        if st.button("✏️ Chỉnh sửa", key="edit_har", use_container_width=True):
                            edit_harvest_log_dialog(editing_row, available_lots)
                    with col_d:
                        if st.button("🗑️ Xóa", key="del_har", use_container_width=True):
                            confirm_action_dialog("DELETE", "harvest_logs", editing_row["id"], None, "✅ Xóa thành công!")
                elif is_editing and not is_within_48h:
                    with col_e: st.caption("🔒 Quá 48h")
                    
                render_team_dataframe("harvest_logs", df_har_team, ["lot_id", "mau_day", "ngay_thu_hoach", "so_luong", "hinh_thuc_thu_hoach", "created_at"])

        elif active_tab == tab_opts[1]:
            render_cost_dashboard(supabase, c_farm, c_team)
        elif active_tab == tab_opts[0]:
            render_global_data_tab(c_farm)

    # =================================================
    # MODULE 3: XƯỞNG ĐÓNG GÓI
    # =================================================
    elif c_team == "Xưởng Đóng Gói":
        tab_opts = ["🌐 Dữ liệu toàn cục", "💰 Chi phí", "📦 Cập nhật BSR"]
        active_tab = st.segmented_control("Chức năng", tab_opts, label_visibility="collapsed", key="tab_bsr_menu", default=tab_opts[0])
        if active_tab is None: active_tab = tab_opts[0]
        
        if active_tab == tab_opts[2]:
            st.markdown("#### Ghi nhận Tỷ lệ BSR thành phẩm")
            available_lots = get_lots_by_farm(c_farm)
            if not available_lots:
                st.warning("⚠️ Chưa có Lô trồng nào trên hệ thống.")
            else:
                df_bsr = fetch_table_data("bsr_logs", c_farm)
                df_bsr_team = df_bsr[df_bsr["team"] == c_team] if not df_bsr.empty else pd.DataFrame()
                editing_row, is_within_48h = get_editing_row("bsr_logs", df_bsr_team)
                is_editing = editing_row is not None
                
                with st.container(border=True):
                    col_a, col_b = st.columns(2)
                    with col_a:
                        lot_id = st.selectbox("🏷️ Chọn Lô đóng gói", options=available_lots, key="add_bsr_lot")
                    with col_b:
                        col_b1, col_b2 = st.columns([2, 1])
                        with col_b1:
                            ngay = st.date_input("📆 Ngày đóng gói", value=date.today(), key="add_bsr_dt")
                        with col_b2:
                            st.text_input("📍 Tuần", value=str(ngay.isocalendar()[1]), disabled=True, key=f"main_w_bsr_{ngay}")
                        bsr_val = st.number_input("📐 Nhập tỷ lệ BSR (Buồng / Sản Rạ)", min_value=0.0, value=0.0, step=0.1, format="%.2f", key="add_bsr_val")
    
                    st.markdown("")
                    if st.button("➕ Thêm vào Danh sách", key="btn_add_bsr", use_container_width=True, type="secondary"):
                        if bsr_val <= 0: st.error("❌ Tỷ lệ BSR phải > 0")
                        else:
                            st.session_state["queue_bsr"].append({
                                "Lô": lot_id, "BSR": bsr_val, "Ngày": ngay.isoformat(), "Tuần": ngay.isocalendar()[1]
                            })
                            st.rerun()

                def process_bsr_queue():
                    queue = st.session_state["queue_bsr"]
                    success_count = 0
                    for item in queue:
                        data = {
                            "farm": c_farm, "team": c_team, "lot_id": item["Lô"],
                            "ngay_nhap": item["Ngày"], "bsr": item["BSR"], "tuan": item["Tuần"]
                        }
                        if insert_to_db("bsr_logs", data):
                            success_count += 1
                        else:
                            st.error(f"❌ Lỗi ghi lô {item['Lô']}")
                            return
                    st.session_state["queue_bsr"] = []
                    st.cache_data.clear()
                    st.session_state["toast"] = f"✅ Ghi nhận {success_count} bản ghi BSR thành công!"
                    st.rerun()

                render_queue_ui("queue_bsr", ["Lô", "BSR", "Ngày", "Tuần"], process_bsr_queue)
                
                st.markdown("---")
                col_t, col_e, col_d = st.columns([5, 1.5, 1.5])
                with col_t:
                    st.markdown('<p class="dataframe-header" style="margin-top:0.5rem;">Dữ liệu của đội bạn (Click 1 dòng để sửa/xóa)</p>', unsafe_allow_html=True)
                if is_editing and is_within_48h:
                    with col_e:
                        if st.button("✏️ Chỉnh sửa", key="edit_bsr", use_container_width=True):
                            edit_bsr_log_dialog(editing_row, available_lots)
                    with col_d:
                        if st.button("🗑️ Xóa", key="del_bsr", use_container_width=True):
                            confirm_action_dialog("DELETE", "bsr_logs", editing_row["id"], None, "✅ Xóa thành công!")
                elif is_editing and not is_within_48h:
                    with col_e: st.caption("🔒 Quá 48h")
                    
                render_team_dataframe("bsr_logs", df_bsr_team, ["lot_id", "ngay_nhap", "bsr", "created_at"])

        elif active_tab == tab_opts[1]:
            render_cost_dashboard(supabase, c_farm, c_team)
        elif active_tab == tab_opts[0]:
            render_global_data_tab(c_farm)

# =====================================================
# MAIN ROUTING
# =====================================================
if __name__ == "__main__":
    init_session_state()
    if not st.session_state["logged_in"]:
        render_login()
    else:
        render_main_app()
