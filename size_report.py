from __future__ import annotations

import io
import re
import unicodedata
from datetime import date, datetime, timedelta
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HARVEST_READY_CAL = 36.0
HARVEST_WARNING_CAL = 35.0
COMPLETED_GROUP_RETENTION_DAYS = 30
FARM_HARVEST_OFFSETS = {
    "Farm 126": 8,
    "Farm 157": 9,
}


def _normalize_text(value: Any) -> str:
    text = str(value or "").strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).lower()


def _to_float(value: Any) -> float | None:
    try:
        if value is None or pd.isna(value):
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _to_int(value: Any) -> int | None:
    number = _to_float(value)
    if number is None:
        return None
    return int(round(number))


def _to_date(value: Any) -> date | None:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def compute_size_forecast(measurements: dict[int, float]) -> dict[str, Any]:
    ordered = [
        (int(measurement_no), float(size))
        for measurement_no, size in measurements.items()
        if measurement_no in (1, 2, 3) and _to_float(size) is not None
    ]
    ordered.sort(key=lambda item: item[0])
    if not ordered:
        return {
            "latest_size": None,
            "avg_growth": None,
            "projected_next": None,
            "highlight": False,
            "note": "",
        }

    sizes = [size for _, size in ordered]
    latest_size = sizes[-1]
    avg_growth = None
    projected_next = None
    if len(sizes) >= 2:
        changes = [sizes[index] - sizes[index - 1] for index in range(1, len(sizes))]
        avg_growth = sum(changes) / len(changes)
        projected_next = latest_size + avg_growth

    if latest_size >= HARVEST_READY_CAL:
        note = "Đã đạt size thu hoạch"
        highlight = True
    elif latest_size >= HARVEST_WARNING_CAL:
        note = "Gần đạt size, tuần sau có thể thu"
        highlight = True
    elif projected_next is not None and projected_next >= HARVEST_READY_CAL:
        note = "Dự kiến lần/tuần tới đạt size thu hoạch"
        highlight = True
    else:
        note = "Chưa tới size"
        highlight = False

    return {
        "latest_size": latest_size,
        "avg_growth": avg_growth,
        "projected_next": projected_next,
        "highlight": highlight,
        "note": note,
    }


def validate_size_measurement_sequence(
    existing_logs: pd.DataFrame,
    requested_measurement: int,
    test_row: str,
    *,
    exclude_id: Any = None,
) -> str | None:
    requested_measurement = int(requested_measurement)
    data = existing_logs.copy() if existing_logs is not None else pd.DataFrame()
    if not data.empty:
        if "is_deleted" in data.columns:
            data = data[data["is_deleted"].fillna(False) == False]  # noqa: E712
        if exclude_id is not None and "id" in data.columns:
            data = data[data["id"].astype(str) != str(exclude_id)]

    if requested_measurement > 1:
        previous_measurement = requested_measurement - 1
        existing_measurements = (
            set(data["lan_do"].map(_to_int).dropna().astype(int))
            if not data.empty and "lan_do" in data.columns
            else set()
        )
        if previous_measurement not in existing_measurements:
            return (
                f"Không thể đo Lần {requested_measurement} "
                f"khi chưa có Lần {previous_measurement}."
            )

    expected_rows: list[str] = []
    if not data.empty and "hang_kiem_tra" in data.columns:
        expected_rows = [
            str(value).strip()
            for value in data["hang_kiem_tra"].tolist()
            if str(value or "").strip()
        ]
    expected_test_row = expected_rows[0] if expected_rows else ""
    current_test_row = str(test_row or "").strip()
    if expected_test_row and current_test_row != expected_test_row:
        return (
            "Các lần đo cùng lô và màu dây phải dùng cùng "
            f"Hàng kiểm tra: {expected_test_row}."
        )
    return None


def _group_key(row: pd.Series) -> tuple[Any, ...]:
    dim_lo_id = row.get("dim_lo_id")
    if dim_lo_id is not None and not pd.isna(dim_lo_id):
        lot_key = ("dim_lo_id", int(dim_lo_id))
    else:
        lot_key = ("lot", str(row.get("farm") or ""), str(row.get("lo") or row.get("lot_id") or ""))
    return (*lot_key, str(row.get("mau_day") or "").strip())


def _latest_rows_by_measurement(group: pd.DataFrame) -> dict[int, pd.Series]:
    sortable = group.copy()
    sortable["_ngay_do_sort"] = pd.to_datetime(sortable.get("ngay_do"), errors="coerce")
    if "created_at" in sortable.columns:
        sortable["_created_sort"] = pd.to_datetime(sortable["created_at"], errors="coerce")
    else:
        sortable["_created_sort"] = pd.NaT
    sortable = sortable.sort_values(
        ["_ngay_do_sort", "_created_sort"],
        na_position="first",
    )

    result: dict[int, pd.Series] = {}
    for _, row in sortable.iterrows():
        measurement_no = _to_int(row.get("lan_do"))
        if measurement_no in (1, 2, 3):
            result[measurement_no] = row
    return result


def _attach_ribbon_colors(
    stage_logs: pd.DataFrame,
    ribbon_schedule: pd.DataFrame | None,
) -> pd.DataFrame:
    if stage_logs is None or stage_logs.empty:
        return pd.DataFrame()
    result = stage_logs.copy()
    if "mau_day" not in result.columns:
        result["mau_day"] = ""
    if ribbon_schedule is None or ribbon_schedule.empty:
        return result

    schedule = ribbon_schedule.copy()
    if "is_deleted" in schedule.columns:
        schedule = schedule[schedule["is_deleted"].fillna(False) == False]  # noqa: E712
    lookup: dict[tuple[str, int, int], str] = {}
    for _, row in schedule.iterrows():
        farm = str(row.get("farm") or row.get("farm_name") or "").strip()
        year = _to_int(row.get("year"))
        week = _to_int(row.get("week_number"))
        color = str(row.get("color_name") or "").strip()
        if farm and year is not None and week is not None and color:
            lookup[(farm, year, week)] = color

    def resolve_color(row: pd.Series) -> str:
        existing = str(row.get("mau_day") or "").strip()
        if existing:
            return existing
        stage_date = _to_date(row.get("ngay_thuc_hien"))
        year = stage_date.isocalendar().year if stage_date else _to_int(row.get("year"))
        week = _to_int(row.get("tuan"))
        farm = str(row.get("farm") or "").strip()
        if year is None or week is None:
            return ""
        return lookup.get((farm, year, week), "")

    result["mau_day"] = result.apply(resolve_color, axis=1)
    return result


def _find_cut_bap_week(
    stage_logs: pd.DataFrame,
    group_row: pd.Series,
    latest_measure_date: date | None,
) -> tuple[int | None, int | None]:
    if stage_logs is None or stage_logs.empty:
        return None, None

    candidates = stage_logs.copy()
    if "giai_doan" in candidates.columns:
        candidates = candidates[
            candidates["giai_doan"].map(_normalize_text).str.contains("cat bap", na=False)
        ]
    if candidates.empty:
        return None, None

    dim_lo_id = group_row.get("dim_lo_id")
    if dim_lo_id is not None and not pd.isna(dim_lo_id) and "dim_lo_id" in candidates.columns:
        candidates = candidates[pd.to_numeric(candidates["dim_lo_id"], errors="coerce") == int(dim_lo_id)]
    else:
        if "farm" in candidates.columns:
            candidates = candidates[candidates["farm"] == group_row.get("farm")]
        lot_name = group_row.get("lo") or group_row.get("lot_id")
        if "lo" in candidates.columns:
            candidates = candidates[candidates["lo"] == lot_name]

    color = str(group_row.get("mau_day") or "").strip()
    if color and "mau_day" in candidates.columns:
        normalized_color = _normalize_text(color)
        same_color = candidates[
            candidates["mau_day"].map(_normalize_text) == normalized_color
        ]
        candidates = same_color
    if candidates.empty:
        return None, None

    candidates["_stage_date"] = pd.to_datetime(candidates.get("ngay_thuc_hien"), errors="coerce")
    if latest_measure_date is not None:
        before_measurement = candidates[
            candidates["_stage_date"].dt.date <= latest_measure_date
        ]
        if not before_measurement.empty:
            candidates = before_measurement
    candidates = candidates.sort_values("_stage_date", na_position="first")
    selected = candidates.iloc[-1]

    cut_date = _to_date(selected.get("ngay_thuc_hien"))
    cut_week = _to_int(selected.get("tuan"))
    cut_year = cut_date.isocalendar().year if cut_date else None
    if cut_week is None and cut_date:
        cut_week = cut_date.isocalendar().week
    return cut_year, cut_week


def _forecast_harvest_week(farm: str, cut_year: int | None, cut_week: int | None) -> int | None:
    if cut_year is None or cut_week is None:
        return None
    offset = FARM_HARVEST_OFFSETS.get(str(farm))
    if offset is None:
        return None
    try:
        cut_week_start = datetime.fromisocalendar(int(cut_year), int(cut_week), 1).date()
    except ValueError:
        return None
    harvest_date = cut_week_start + timedelta(weeks=offset - 1)
    return harvest_date.isocalendar().week


def build_size_report_rows(
    lots: pd.DataFrame,
    size_logs: pd.DataFrame,
    stage_logs: pd.DataFrame | None = None,
    ribbon_schedule: pd.DataFrame | None = None,
    *,
    today: date | None = None,
) -> pd.DataFrame:
    columns = [
        "lo",
        "team",
        "hang_kiem_tra",
        "so_buong_do",
        "mau_day",
        "tuan_cat_bap",
        "tuan_thu_hoach_du_kien",
        "lan_do_1",
        "lan_do_2",
        "lan_do_3",
        "muc_tang_size_tb",
        "chu_thich",
        "highlight",
    ]
    if size_logs is None or size_logs.empty:
        return pd.DataFrame(columns=columns)

    today = today or date.today()
    data = size_logs.copy()
    if "is_deleted" in data.columns:
        data = data[data["is_deleted"].fillna(False) == False]  # noqa: E712
    data = data[data["lan_do"].map(_to_int).isin([1, 2, 3])]
    if data.empty:
        return pd.DataFrame(columns=columns)

    lot_lookup: dict[int, dict[str, Any]] = {}
    if lots is not None and not lots.empty:
        for _, lot_row in lots.iterrows():
            dim_lo_id = lot_row.get("dim_lo_id")
            if dim_lo_id is None or pd.isna(dim_lo_id):
                dim_lo_id = lot_row.get("lo_id")
            if dim_lo_id is not None and not pd.isna(dim_lo_id):
                lot_lookup[int(dim_lo_id)] = lot_row.to_dict()

    data["_group_key"] = data.apply(_group_key, axis=1)
    resolved_stage_logs = _attach_ribbon_colors(
        stage_logs if stage_logs is not None else pd.DataFrame(),
        ribbon_schedule,
    )
    rows: list[dict[str, Any]] = []
    for _, group in data.groupby("_group_key", sort=False):
        measurements = _latest_rows_by_measurement(group)
        if not measurements:
            continue

        third_measurement_date = _to_date(measurements[3].get("ngay_do")) if 3 in measurements else None
        if (
            third_measurement_date is not None
            and (today - third_measurement_date).days > COMPLETED_GROUP_RETENTION_DAYS
        ):
            continue

        latest_no = max(measurements)
        latest_row = measurements[latest_no]
        latest_measure_date = _to_date(latest_row.get("ngay_do"))
        dim_lo_id = latest_row.get("dim_lo_id")
        lot_meta = {}
        if dim_lo_id is not None and not pd.isna(dim_lo_id):
            lot_meta = lot_lookup.get(int(dim_lo_id), {})

        sizes = {
            measurement_no: _to_float(row.get("size_cal"))
            for measurement_no, row in measurements.items()
        }
        sizes = {number: size for number, size in sizes.items() if size is not None}
        forecast = compute_size_forecast(sizes)
        cut_year, cut_week = _find_cut_bap_week(
            resolved_stage_logs,
            latest_row,
            latest_measure_date,
        )

        non_empty_test_rows = [
            str(row.get("hang_kiem_tra") or "").strip()
            for _, row in sorted(measurements.items())
            if str(row.get("hang_kiem_tra") or "").strip()
        ]
        test_row = non_empty_test_rows[-1] if non_empty_test_rows else ""
        farm = str(latest_row.get("farm") or lot_meta.get("farm") or "")
        rows.append(
            {
                "lo": latest_row.get("lo") or latest_row.get("lot_id") or lot_meta.get("lo") or "",
                "team": latest_row.get("team") or lot_meta.get("team") or "",
                "hang_kiem_tra": test_row,
                "so_buong_do": _to_int(latest_row.get("so_luong_mau")) or 0,
                "mau_day": str(latest_row.get("mau_day") or "").strip(),
                "tuan_cat_bap": cut_week,
                "tuan_thu_hoach_du_kien": _forecast_harvest_week(farm, cut_year, cut_week),
                "lan_do_1": sizes.get(1),
                "lan_do_2": sizes.get(2),
                "lan_do_3": sizes.get(3),
                "muc_tang_size_tb": forecast["avg_growth"],
                "chu_thich": forecast["note"],
                "highlight": forecast["highlight"],
            }
        )

    result = pd.DataFrame(rows, columns=columns)
    if not result.empty:
        result = result.sort_values(["team", "lo", "mau_day"], na_position="last").reset_index(drop=True)
    return result


def generate_size_measure_excel(
    lots: pd.DataFrame,
    size_logs: pd.DataFrame,
    stage_logs: pd.DataFrame | None = None,
    ribbon_schedule: pd.DataFrame | None = None,
    *,
    farm_name: str = "",
    today: date | None = None,
) -> bytes:
    today = today or date.today()
    rows = build_size_report_rows(
        lots,
        size_logs,
        stage_logs,
        ribbon_schedule,
        today=today,
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Báo cáo đo size"
    sheet.sheet_view.showGridLines = False

    thin_gray = Side(style="thin", color="808080")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    header_fill = PatternFill("solid", fgColor="FFE699")
    note_fill = PatternFill("solid", fgColor="FCE4D6")
    ready_fill = PatternFill("solid", fgColor="FFF2CC")
    total_fill = PatternFill("solid", fgColor="FFE699")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.merge_cells("A1:M1")
    sheet["A1"] = "BÁO CÁO KIỂM TRA ĐO SIZE THU HOẠCH CHUỐI"
    sheet["A1"].font = Font(name="Times New Roman", size=16, bold=True)
    sheet["A1"].alignment = center

    sheet.merge_cells("A2:M2")
    iso = today.isocalendar()
    farm_suffix = f" - {farm_name}" if farm_name else ""
    sheet["A2"] = f"Tuần: {iso.week} - Ngày kiểm tra: {today.strftime('%d/%m/%Y')}{farm_suffix}"
    sheet["A2"].font = Font(name="Times New Roman", size=12)
    sheet["A2"].alignment = center

    sheet.merge_cells("A3:M3")
    sheet["A3"] = (
        "* Size thu hoạch mục tiêu: 36/37 cal, đo nải thứ hai từ dưới lên. "
        "Lưu ý đo cùng hàng trong các lần đo."
    )
    sheet["A3"].font = Font(name="Times New Roman", size=11, italic=True, color="FF0000")
    sheet["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    sheet["A3"].fill = note_fill

    fixed_headers = [
        "STT",
        "Lô",
        "Đội",
        "Hàng kiểm tra",
        "Số buồng đo",
        "Màu dây",
        "Tuần cắt bắp",
        "Tuần thu hoạch dự kiến",
    ]
    for column, label in enumerate(fixed_headers, 1):
        sheet.merge_cells(start_row=4, start_column=column, end_row=5, end_column=column)
        cell = sheet.cell(row=4, column=column, value=label)
        cell.fill = header_fill
        cell.font = Font(name="Times New Roman", size=11, bold=True)
        cell.alignment = center
        cell.border = border
        sheet.cell(row=5, column=column).border = border

    sheet.merge_cells("I4:K4")
    sheet["I4"] = "TB size (cal) nải 2 từ dưới lên"
    sheet["I4"].fill = header_fill
    sheet["I4"].font = Font(name="Times New Roman", size=11, bold=True)
    sheet["I4"].alignment = center
    for column in range(9, 12):
        sheet.cell(row=4, column=column).border = border

    for column, label in zip(range(9, 12), ["Lần đo 1", "Lần đo 2", "Lần đo 3"]):
        cell = sheet.cell(row=5, column=column, value=label)
        cell.fill = header_fill
        cell.font = Font(name="Times New Roman", size=11, bold=True)
        cell.alignment = center
        cell.border = border

    for column, label in [(12, "Mức tăng size TB"), (13, "Chú thích")]:
        cell = sheet.cell(row=5, column=column, value=label)
        cell.fill = header_fill
        cell.font = Font(name="Times New Roman", size=11, bold=True)
        cell.alignment = center
        cell.border = border
        sheet.cell(row=4, column=column).fill = header_fill
        sheet.cell(row=4, column=column).border = border

    first_data_row = 6
    for index, row in rows.iterrows():
        excel_row = first_data_row + index
        values = [
            index + 1,
            row["lo"],
            row["team"],
            row["hang_kiem_tra"],
            row["so_buong_do"],
            row["mau_day"],
            row["tuan_cat_bap"],
            row["tuan_thu_hoach_du_kien"],
            row["lan_do_1"],
            row["lan_do_2"],
            row["lan_do_3"],
            row["muc_tang_size_tb"],
            row["chu_thich"],
        ]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row=excel_row, column=column, value=value)
            cell.font = Font(name="Times New Roman", size=11)
            cell.alignment = center
            cell.border = border
            if row["highlight"]:
                cell.fill = ready_fill
        for column in range(9, 13):
            sheet.cell(row=excel_row, column=column).number_format = "0.00"

    total_row = first_data_row + len(rows)
    sheet.cell(row=total_row, column=1, value="Tổng")
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    sheet.cell(row=total_row, column=5, value=int(rows["so_buong_do"].sum()) if not rows.empty else 0)
    for column in range(1, 14):
        cell = sheet.cell(row=total_row, column=column)
        cell.fill = total_fill
        cell.font = Font(name="Times New Roman", size=11, bold=True)
        cell.alignment = center
        cell.border = border

    widths = {
        1: 7,
        2: 12,
        3: 14,
        4: 18,
        5: 14,
        6: 18,
        7: 15,
        8: 22,
        9: 13,
        10: 13,
        11: 13,
        12: 18,
        13: 42,
    }
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.row_dimensions[1].height = 26
    sheet.row_dimensions[2].height = 22
    sheet.row_dimensions[3].height = 38
    sheet.row_dimensions[4].height = 28
    sheet.row_dimensions[5].height = 25
    sheet.freeze_panes = "A6"
    sheet.auto_filter.ref = f"A5:M{max(total_row - 1, 5)}"
    sheet.print_title_rows = "1:5"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_area = f"A1:M{total_row}"

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
